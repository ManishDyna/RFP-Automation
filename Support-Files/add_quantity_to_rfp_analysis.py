"""
Enrich an RFP-Analysis xlsx (RFP-Material_List sheet) by pulling Quantity --
and filling missing Material_Code / Material_Description -- from each row's
source RFP Excel.

For every analysis row we try to identify the matching source row using a
four-pass strategy (mirrors the project's main matching logic but more
forgiving than strict positional pairing):

  Pass 1 (code)        : analysis Material_Code -> 9-digit code in source Name
  Pass 2 (keyword)     : analysis Matched_Keywords found as substring in
                         source Name or Description
  Pass 3 (description) : meaningful tokens from analysis Material_Description
                         found in source Name or Description
  Pass 4 (positional)  : remaining analysis rows zipped to remaining source
                         rows in original sheet order (last resort)

Source RFP Excel lookup chain:
  1. C:\\python\\RFP-automation\\TempRFP-Files (flat)
  2. C:\\python\\RFP-automation\\ALLRFPs\\<Company>\\<RFP_Title>\\downloaded-rfp\\
  3. SharePoint: RFP-logs/ALLRFPs/<Company>/<RFP_Title>/downloaded-rfp/
     (downloaded copies are cached into TempRFP-Files for re-use)

Outputs (next to the input analysis xlsx):
  - <name>_with_quantity.xlsx
  - <name>_with_quantity_report.txt

Usage (from project root):
  python Support-Files/add_quantity_to_rfp_analysis.py
  python Support-Files/add_quantity_to_rfp_analysis.py --analysis "<path>"
  python Support-Files/add_quantity_to_rfp_analysis.py --no-download
"""

import argparse
import os
import re
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.styles import Font

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import core.common_imports  # noqa: F401
from helpers.core_helper import (
    _find_other_content_sheet_name,
    find_column_name,
)
from config.config import (
    CLIENT_ID, CLIENT_SECRET, TENANT_ID,
    SHAREPOINT_HOSTNAME, SITE_PATH, DRIVE_NAME, SP_BASE_FOLDER,
)

DEFAULT_ANALYSIS = r"C:\Users\Manish.Soni\Downloads\RFP-Analysis_refreshed_20260515_135031.xlsx"
DEFAULT_TEMP_DIR = r"C:\python\RFP-automation\TempRFP-Files"
DEFAULT_ALLRFPS = r"C:\python\RFP-automation\ALLRFPs"

# Instruction-row hints we want to skip when reading Other Content
INSTRUCTION_HINTS = (
    "click on the +",
    "question, item, or lot name",
    "header text",
    "required action",
    "submit the",
    "for example,",
)


# ---------------------------------------------------------------------------
# Analysis loader
# ---------------------------------------------------------------------------
def load_analysis(analysis_path: str):
    """
    Returns:
      per_file_rows : { excel_file: [analysis_row_dict, ...] in sheet order }
      file_meta     : { excel_file: (company_name, rfp_title) }
      col_idx       : { column_name: 1-based column index }
      total_rows    : total data rows
    Each analysis_row_dict has:
      row_idx, excel_file, code, matched_kw, desc, qty_blank,
      code_blank, desc_blank
    """
    wb = openpyxl.load_workbook(analysis_path, read_only=True, data_only=True)
    ws = wb["RFP-Material_List"]

    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx0 = {h: i for i, h in enumerate(header)}
    required = [
        "Company_Name", "RFP_Title", "Excel_File", "Material_Code",
        "Material_Description", "Matched_Keywords", "Keyword_Matched",
        "Material_Matched", "Quantity",
    ]
    for col in required:
        if col not in idx0:
            wb.close()
            raise RuntimeError(f"Required column missing: {col}")

    per_file_rows = defaultdict(list)
    file_meta = {}
    total = 0

    for r_idx, row in enumerate(rows, start=2):
        total += 1
        xf = row[idx0["Excel_File"]]
        if not xf:
            continue
        code = row[idx0["Material_Code"]]
        desc = row[idx0["Material_Description"]]
        mk = row[idx0["Matched_Keywords"]]
        qty = row[idx0["Quantity"]]

        code_str = str(code).strip() if code is not None else ""
        desc_str = str(desc).strip() if desc is not None else ""
        mk_str = str(mk).strip() if mk is not None else ""

        per_file_rows[xf].append({
            "row_idx": r_idx,
            "code": code_str,
            "matched_kw": mk_str,
            "desc": desc_str,
            "qty_blank": qty is None or (isinstance(qty, str) and not qty.strip()),
            "code_blank": not bool(code_str),
            "desc_blank": not bool(desc_str),
        })
        if xf not in file_meta:
            file_meta[xf] = (row[idx0["Company_Name"]], row[idx0["RFP_Title"]])

    # Column indexes for write-back (1-based)
    col_idx_1b = {h: i + 1 for i, h in enumerate(header)}
    wb.close()
    return per_file_rows, file_meta, col_idx_1b, total


# ---------------------------------------------------------------------------
# Source RFP reader
# ---------------------------------------------------------------------------
def _read_other_content_df(excel_path: str):
    sheet = _find_other_content_sheet_name(excel_path)
    df = None
    if sheet:
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet)
        except Exception:
            df = None
    if df is None:
        EXPECTED = ["intend to respond", "currency", "material number", "price", "quantity"]
        try:
            for s in pd.ExcelFile(excel_path).sheet_names:
                try:
                    cand = pd.read_excel(excel_path, sheet_name=s)
                    cols_lc = [str(c).lower().strip() for c in cand.columns]
                    hits = sum(1 for ec in EXPECTED if any(ec in c for c in cols_lc))
                    if hits >= 2:
                        df = cand
                        break
                except Exception:
                    continue
        except Exception:
            return None
    return df


def collect_source_rows(excel_path: str):
    """
    Return a list of source-row dicts (in source order) for every 'real' data
    row in Other Content. Each dict has:
      pos, code, name, description, quantity
    Instruction/header rows are filtered out.
    """
    df = _read_other_content_df(excel_path)
    if df is None:
        return []

    name_col = find_column_name(df.columns, "name")
    if not name_col:
        return []
    desc_col = find_column_name(df.columns, "description")
    qty_col = find_column_name(df.columns, "quantity")
    code_col = find_column_name(df.columns, "material code")

    out = []
    for r_pos, rec in enumerate(df.to_dict("records"), start=2):
        name_val = rec.get(name_col)
        if pd.isna(name_val):
            continue
        name_text = str(name_val).strip()
        if not name_text:
            continue
        low = name_text.lower()
        if any(hint in low for hint in INSTRUCTION_HINTS):
            continue

        desc_val = rec.get(desc_col) if desc_col else None
        desc_text = "" if (desc_val is None or pd.isna(desc_val)) else str(desc_val).strip()

        qty_val = rec.get(qty_col) if qty_col else None
        if qty_val is None or (isinstance(qty_val, float) and pd.isna(qty_val)):
            qty_clean = ""
        else:
            qty_clean = qty_val

        # Resolve a 9-digit material code:
        # prefer one extracted from Name, fall back to "Material Code" column
        codes_in_name = re.findall(r"\d{9}", name_text)
        if codes_in_name:
            code_clean = codes_in_name[0]
        else:
            code_val = rec.get(code_col) if code_col else None
            if code_val is None or (isinstance(code_val, float) and pd.isna(code_val)):
                code_clean = ""
            else:
                m = re.search(r"\d{9}", str(code_val))
                code_clean = m.group(0) if m else ""

        out.append({
            "pos": r_pos,
            "code": code_clean,
            "name": name_text,
            "description": desc_text,
            "quantity": qty_clean,
        })
    return out


# ---------------------------------------------------------------------------
# Multi-pass matcher
# ---------------------------------------------------------------------------
_TOKEN_SPLIT = re.compile(r"[\s,;:./\(\)\[\]\-]+")


def _significant_tokens(text: str, min_len: int = 4, limit: int = 8):
    if not text:
        return []
    seen = []
    for t in _TOKEN_SPLIT.split(text.upper()):
        t = t.strip()
        if len(t) >= min_len and t not in seen:
            seen.append(t)
        if len(seen) >= limit:
            break
    return seen


def _group_key(a):
    """
    Identity key for fan-out grouping. Rows that share (code, matched_kw, desc)
    describe the same logical material and can share a source row's Quantity.
    Rows with all-blank identity become singletons so the positional fallback
    still pairs them 1:1.
    """
    if not (a["code"] or a["matched_kw"] or a["desc"]):
        return ("__singleton__", a["row_idx"])
    return (a["code"], a["matched_kw"], a["desc"])


def match_analysis_to_source(analysis_rows, source_rows):
    """
    Group-then-match-then-fan-out:
      1. Group analysis rows by (code, matched_kw, desc).
      2. Run the 4-pass matcher on one representative per group.
      3. Copy the rep's matched source row to EVERY member of the group.

    This recovers cases where the analysis xlsx contains duplicate rows for
    the same physical source row (the dominant cause of leftover blanks).

    Returns:
      mapping   : { analysis_row_idx: source_row_dict }
      pass_used : { analysis_row_idx: 'code'|'keyword'|'desc'|'positional' }
      unmatched : list of analysis_row_dict whose group found no source row
    """
    # 1. Group by identity (preserving first-seen order)
    groups = defaultdict(list)
    group_order = []
    for a in analysis_rows:
        k = _group_key(a)
        if k not in groups:
            group_order.append(k)
        groups[k].append(a)

    reps = [groups[k][0] for k in group_order]

    # 2. Run the 4-pass matcher on group representatives
    used_pos = set()
    rep_mapping = {}
    rep_pass = {}

    # Pass 1: code-based (exact 9-digit equality)
    by_code = {}
    for s in source_rows:
        if s["code"] and s["code"] not in by_code:
            by_code[s["code"]] = s
    pending = []
    for a in reps:
        if a["code"] and a["code"] in by_code and by_code[a["code"]]["pos"] not in used_pos:
            s = by_code[a["code"]]
            rep_mapping[a["row_idx"]] = s
            rep_pass[a["row_idx"]] = "code"
            used_pos.add(s["pos"])
        else:
            pending.append(a)

    # Pass 2: keyword (Matched_Keywords substring in source name+desc)
    pending_next = []
    for a in pending:
        mk = a["matched_kw"]
        placed = False
        if mk:
            mk_up = mk.upper()
            for s in source_rows:
                if s["pos"] in used_pos:
                    continue
                hay = (s["name"] + " " + s["description"]).upper()
                if mk_up in hay:
                    rep_mapping[a["row_idx"]] = s
                    rep_pass[a["row_idx"]] = "keyword"
                    used_pos.add(s["pos"])
                    placed = True
                    break
        if not placed:
            pending_next.append(a)

    # Pass 3: description tokens
    pending = pending_next
    pending_next = []
    for a in pending:
        tokens = _significant_tokens(a["desc"])
        placed = False
        if tokens:
            for s in source_rows:
                if s["pos"] in used_pos:
                    continue
                hay = (s["name"] + " " + s["description"]).upper()
                if any(t in hay for t in tokens):
                    rep_mapping[a["row_idx"]] = s
                    rep_pass[a["row_idx"]] = "desc"
                    used_pos.add(s["pos"])
                    placed = True
                    break
        if not placed:
            pending_next.append(a)

    # Pass 4: positional fallback for remaining reps
    pending = pending_next
    free_sources = [s for s in source_rows if s["pos"] not in used_pos]
    used_in_pass4 = 0
    for a, s in zip(pending, free_sources):
        rep_mapping[a["row_idx"]] = s
        rep_pass[a["row_idx"]] = "positional"
        used_pos.add(s["pos"])
        used_in_pass4 += 1

    unmatched_reps = pending[used_in_pass4:]
    unmatched_rep_ids = {r["row_idx"] for r in unmatched_reps}

    # 3. Fan out the rep's matched source row to every group member
    mapping = {}
    pass_used = {}
    unmatched = []
    for k in group_order:
        members = groups[k]
        rep = members[0]
        if rep["row_idx"] in unmatched_rep_ids:
            unmatched.extend(members)
            continue
        s = rep_mapping[rep["row_idx"]]
        p = rep_pass[rep["row_idx"]]
        for m in members:
            mapping[m["row_idx"]] = s
            pass_used[m["row_idx"]] = p

    return mapping, pass_used, unmatched


# ---------------------------------------------------------------------------
# File location: TempRFP-Files -> ALLRFPs -> SharePoint
# ---------------------------------------------------------------------------
def locate_local_file(excel_file: str, temp_dir: str, allrfps_dir: str,
                       company: str, rfp_title: str):
    """Search local sources in order. Returns absolute path or None."""
    # 1. Flat TempRFP-Files
    if os.path.isdir(temp_dir):
        target_lc = excel_file.lower()
        for name in os.listdir(temp_dir):
            if name.lower() == target_lc:
                return os.path.join(temp_dir, name)

    # 2. ALLRFPs/<Company>/<RFP_Title>/downloaded-rfp/
    if company and rfp_title and os.path.isdir(allrfps_dir):
        candidate_dirs = [
            os.path.join(allrfps_dir, company, rfp_title, "downloaded-rfp"),
            os.path.join(allrfps_dir, company, rfp_title),
        ]
        target_lc = excel_file.lower()
        for cdir in candidate_dirs:
            if not os.path.isdir(cdir):
                continue
            for name in os.listdir(cdir):
                if name.lower() == target_lc:
                    return os.path.join(cdir, name)
            # also accept any xls/xlsx in that folder
            for name in os.listdir(cdir):
                if name.lower().endswith((".xls", ".xlsx")) and not name.startswith("~$"):
                    return os.path.join(cdir, name)
    return None


def _sp_client():
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from download_from_sharepoint import SharePointDownloader
    c = SharePointDownloader(
        CLIENT_ID, CLIENT_SECRET, TENANT_ID,
        SHAREPOINT_HOSTNAME, SITE_PATH, DRIVE_NAME,
    )
    c.auth()
    c.resolve_site_and_drive()
    return c


def download_from_sharepoint(client, company: str, rfp_title: str,
                              excel_file: str, dest_dir: str):
    import requests
    if not company or not rfp_title:
        return None
    folder = f"{SP_BASE_FOLDER}/ALLRFPs/{company}/{rfp_title}/downloaded-rfp"
    url = (
        f"https://graph.microsoft.com/v1.0/sites/{client.site_id}"
        f"/drives/{client.drive_id}/root:/{folder}:/children"
    )
    client.ensure_token()
    r = requests.get(url, headers=client.headers)
    if r.status_code != 200:
        print(f"      [SP] folder not accessible: {folder} ({r.status_code})")
        return None
    items = r.json().get("value", [])
    target_lc = excel_file.lower()
    chosen = None
    for it in items:
        if "file" in it and it.get("name", "").lower() == target_lc:
            chosen = it
            break
    if chosen is None:
        for it in items:
            n = it.get("name", "")
            if "file" in it and n.lower().endswith((".xls", ".xlsx")) and not n.startswith("~$"):
                chosen = it
                break
    if chosen is None:
        print(f"      [SP] no Excel under {folder}")
        return None
    sp_path = f"{folder}/{chosen['name']}"
    dl_url = (
        f"https://graph.microsoft.com/v1.0/sites/{client.site_id}"
        f"/drives/{client.drive_id}/root:/{sp_path}:/content"
    )
    client.ensure_token()
    fr = requests.get(dl_url, headers=client.headers, stream=True)
    if fr.status_code != 200:
        print(f"      [SP] download failed: {sp_path} ({fr.status_code})")
        return None
    os.makedirs(dest_dir, exist_ok=True)
    local_path = os.path.join(dest_dir, excel_file)
    with open(local_path, "wb") as f:
        for chunk in fr.iter_content(chunk_size=8192):
            f.write(chunk)
    print(f"      [SP OK] downloaded {chosen['name']} ({os.path.getsize(local_path)/1024:.1f} KB)")
    return local_path


# ---------------------------------------------------------------------------
# Write back
# ---------------------------------------------------------------------------
def write_back(analysis_path: str, out_path: str, fills_qty: dict,
               fills_code: dict, fills_desc: dict, col_idx: dict):
    wb = openpyxl.load_workbook(analysis_path)
    ws = wb["RFP-Material_List"]

    qty_c = col_idx.get("Quantity")
    code_c = col_idx.get("Material_Code")
    desc_c = col_idx.get("Material_Description")
    if qty_c is None:
        qty_c = ws.max_column + 1
        ws.cell(row=1, column=qty_c, value="Quantity").font = Font(bold=True)
        ws.column_dimensions[ws.cell(row=1, column=qty_c).column_letter].width = 14.0

    for r, v in fills_qty.items():
        ws.cell(row=r, column=qty_c, value=v)
    if code_c:
        for r, v in fills_code.items():
            ws.cell(row=r, column=code_c, value=v)
    if desc_c:
        for r, v in fills_desc.items():
            ws.cell(row=r, column=desc_c, value=v)

    try:
        wb.save(out_path)
    except PermissionError:
        base, ext = os.path.splitext(out_path)
        out_path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
        print(f"  WARNING: target locked; writing to {out_path}")
        wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--analysis", default=DEFAULT_ANALYSIS)
    ap.add_argument("--temp-dir", default=DEFAULT_TEMP_DIR)
    ap.add_argument("--allrfps-dir", default=DEFAULT_ALLRFPS)
    ap.add_argument("--out", default=None)
    ap.add_argument("--no-download", action="store_true",
                    help="Skip SharePoint fallback")
    args = ap.parse_args()

    if not os.path.exists(args.analysis):
        print(f"ERROR: Analysis file not found: {args.analysis}", file=sys.stderr)
        sys.exit(1)

    out_path = args.out or (os.path.splitext(args.analysis)[0] + "_with_quantity.xlsx")
    report_path = os.path.splitext(out_path)[0] + "_report.txt"

    print(f"Analysis    : {args.analysis}")
    print(f"Temp dir    : {args.temp_dir}")
    print(f"ALLRFPs dir : {args.allrfps_dir}")
    print(f"Output      : {out_path}")
    print(f"SharePoint  : {'OFF' if args.no_download else 'ON (fallback)'}")
    print()

    print("Reading analysis...")
    per_file_rows, file_meta, col_idx, total_rows = load_analysis(args.analysis)
    files_needed = sorted(per_file_rows.keys())
    rows_with_blank_qty = sum(
        1 for rs in per_file_rows.values() for r in rs if r["qty_blank"]
    )
    print(f"  Total data rows                : {total_rows}")
    print(f"  Rows with blank Quantity       : {rows_with_blank_qty}")
    print(f"  Unique source Excel files      : {len(files_needed)}")
    print()

    sp_client = None
    file_status = {}   # excel_file -> 'local'|'sp'|'missing'|'read_error'
    file_error = {}
    fills_qty, fills_code, fills_desc = {}, {}, {}
    pass_counts = defaultdict(int)        # 'code', 'keyword', 'desc', 'positional'
    files_no_source = []                  # files whose source had 0 usable rows
    mismatch_files = []                   # (xf, n_analysis_rows, n_source_rows)
    unmatched_rows_by_file = defaultdict(list)  # rows that couldn't be paired at all

    for i, xf in enumerate(files_needed, start=1):
        company, rfp_title = file_meta.get(xf, (None, None))
        rows_for_file = per_file_rows[xf]
        n_blank_qty = sum(1 for r in rows_for_file if r["qty_blank"])
        print(f"[{i}/{len(files_needed)}] {xf}  rows={len(rows_for_file)} (blank-qty={n_blank_qty})")

        local_path = locate_local_file(xf, args.temp_dir, args.allrfps_dir, company, rfp_title)
        source = None
        if local_path:
            source = "local"

        if not local_path and not args.no_download:
            if sp_client is None:
                try:
                    sp_client = _sp_client()
                except Exception as e:
                    print(f"    [FATAL] SharePoint auth failed: {e}")
                    print("    Continuing in local-only mode.")
                    args.no_download = True
                    sp_client = None
            if sp_client is not None:
                local_path = download_from_sharepoint(
                    sp_client, company, rfp_title, xf, args.temp_dir
                )
                if local_path:
                    source = "sp"

        if not local_path:
            print(f"    [MISS] file not found locally or on SharePoint")
            file_status[xf] = "missing"
            continue

        try:
            source_rows = collect_source_rows(local_path)
        except Exception as e:
            print(f"    [READ ERROR] {e}")
            file_status[xf] = "read_error"
            file_error[xf] = str(e)
            continue

        file_status[xf] = source
        if not source_rows:
            print(f"    [NO DATA] 0 usable source rows")
            files_no_source.append(xf)
            continue

        if len(rows_for_file) != len(source_rows):
            mismatch_files.append((xf, len(rows_for_file), len(source_rows)))

        mapping, pass_used, unmatched = match_analysis_to_source(rows_for_file, source_rows)

        # Record per-pass counts
        for p in pass_used.values():
            pass_counts[p] += 1
        for u in unmatched:
            unmatched_rows_by_file[xf].append((u["row_idx"], u["code"], u["matched_kw"]))

        # Build fills (only fill blanks; never overwrite existing values)
        for arow in rows_for_file:
            s = mapping.get(arow["row_idx"])
            if not s:
                continue
            if arow["qty_blank"]:
                q = s.get("quantity")
                if q is not None and not (isinstance(q, str) and not q.strip()):
                    fills_qty[arow["row_idx"]] = q
            if arow["code_blank"] and s.get("code"):
                fills_code[arow["row_idx"]] = s["code"]
            if arow["desc_blank"]:
                d = s.get("description") or s.get("name")
                if d:
                    fills_desc[arow["row_idx"]] = d

        print(f"    matched: code={sum(1 for p in pass_used.values() if p=='code')}, "
              f"kw={sum(1 for p in pass_used.values() if p=='keyword')}, "
              f"desc={sum(1 for p in pass_used.values() if p=='desc')}, "
              f"pos={sum(1 for p in pass_used.values() if p=='positional')}, "
              f"unmatched={len(unmatched)}")

    print(f"\nWriting filled analysis -> {out_path}")
    out_path = write_back(args.analysis, out_path, fills_qty, fills_code, fills_desc, col_idx)

    files_local = [f for f, s in file_status.items() if s == "local"]
    files_sp = [f for f, s in file_status.items() if s == "sp"]
    files_missing = [f for f, s in file_status.items() if s == "missing"]
    files_error = [f for f, s in file_status.items() if s == "read_error"]
    unmatched_total = sum(len(v) for v in unmatched_rows_by_file.values())

    print("\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"  Total data rows                 : {total_rows}")
    print(f"  Rows with blank Quantity        : {rows_with_blank_qty}")
    print(f"  Quantity values filled          : {len(fills_qty)}")
    print(f"  Material_Code values filled     : {len(fills_code)}")
    print(f"  Material_Description filled     : {len(fills_desc)}")
    print()
    print(f"  Matches by pass:")
    print(f"    code        : {pass_counts['code']}")
    print(f"    keyword     : {pass_counts['keyword']}")
    print(f"    desc        : {pass_counts['desc']}")
    print(f"    positional  : {pass_counts['positional']}")
    print(f"  Rows unmatched (no source row available): {unmatched_total}")
    print()
    print(f"  Files from local                : {len(files_local)}")
    print(f"  Files downloaded from SP        : {len(files_sp)}")
    print(f"  Files missing everywhere        : {len(files_missing)}")
    print(f"  Files read-failed               : {len(files_error)}")
    print(f"  Files with 0 usable source rows : {len(files_no_source)}")
    print(f"  Files w/ row-count mismatch     : {len(mismatch_files)}")
    print()
    print(f"  Output     : {out_path}")
    print(f"  Report     : {report_path}")

    # Text report
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"RFP Quantity Fill Report\n")
        f.write(f"Generated : {datetime.now().isoformat()}\n")
        f.write(f"Analysis  : {args.analysis}\n")
        f.write(f"Output    : {out_path}\n")
        f.write(f"Temp dir  : {args.temp_dir}\n")
        f.write(f"ALLRFPs   : {args.allrfps_dir}\n\n")

        f.write(f"Total data rows                 : {total_rows}\n")
        f.write(f"Rows with blank Quantity        : {rows_with_blank_qty}\n")
        f.write(f"Quantity filled                 : {len(fills_qty)}\n")
        f.write(f"Material_Code filled            : {len(fills_code)}\n")
        f.write(f"Material_Description filled     : {len(fills_desc)}\n\n")
        f.write(f"Matches by pass:\n")
        f.write(f"  code        : {pass_counts['code']}\n")
        f.write(f"  keyword     : {pass_counts['keyword']}\n")
        f.write(f"  desc        : {pass_counts['desc']}\n")
        f.write(f"  positional  : {pass_counts['positional']}\n")
        f.write(f"Rows unmatched (no source row): {unmatched_total}\n\n")
        f.write(f"Files from local                : {len(files_local)}\n")
        f.write(f"Files downloaded from SP        : {len(files_sp)}\n")
        f.write(f"Files missing                   : {len(files_missing)}\n")
        f.write(f"Files read-failed               : {len(files_error)}\n")
        f.write(f"Files with 0 usable source rows : {len(files_no_source)}\n\n")

        if files_sp:
            f.write("=== Files Downloaded from SharePoint ===\n")
            for x in sorted(files_sp):
                f.write(f"  {x}\n")
            f.write("\n")

        if files_missing:
            f.write("=== Files NOT Processed (not in local dirs, not on SharePoint) ===\n")
            for x in sorted(files_missing):
                co, ti = file_meta.get(x, (None, None))
                f.write(f"  {x}\n     Company: {co}\n     RFP    : {ti}\n")
            f.write("\n")

        if files_error:
            f.write("=== Files NOT Processed (read error) ===\n")
            for x in sorted(files_error):
                f.write(f"  {x}\n     Error: {file_error.get(x)}\n")
            f.write("\n")

        if files_no_source:
            f.write("=== Files With 0 Usable Source Rows ===\n")
            for x in sorted(files_no_source):
                f.write(f"  {x}\n")
            f.write("\n")

        if mismatch_files:
            f.write("=== Row Count Mismatches (analysis vs source) ===\n")
            f.write("(Multi-pass matcher still tries to pair these; rows that\n")
            f.write(" remained unmatched after all 4 passes are listed below.)\n\n")
            for xf, an, sn in sorted(mismatch_files):
                f.write(f"  {xf}\n     analysis rows: {an}    source rows: {sn}\n")
            f.write("\n")

        if unmatched_rows_by_file:
            f.write("=== Analysis Rows Not Paired To Any Source Row ===\n")
            f.write(f"({unmatched_total} rows — usually means analysis has more rows than the source RFP currently provides)\n\n")
            for xf in sorted(unmatched_rows_by_file):
                rows = unmatched_rows_by_file[xf]
                f.write(f"  {xf}  ({len(rows)} row(s))\n")
                for r_idx, code, mk in rows[:20]:
                    f.write(f"      analysis row {r_idx}    code='{code}'  kw='{mk}'\n")
                if len(rows) > 20:
                    f.write(f"      ... and {len(rows)-20} more\n")
            f.write("\n")

    print("\nDone.")


if __name__ == "__main__":
    main()

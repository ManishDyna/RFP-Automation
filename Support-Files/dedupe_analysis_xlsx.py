"""
Dedupe RFP-Material_List and RFP-List in the analysis xlsx after the
whitespace-normalize pass collapsed near-duplicate RFP_Title spellings into
identical rows. Uses pandas drop_duplicates (vectorized) for speed.

RFP-Material_List: dedup key = all 12 columns
RFP-List         : dedup key = RFP_Title (keep first)

Keeps the first occurrence. Safe to re-run (idempotent).
"""

import argparse
import os
import sys
from datetime import datetime

import openpyxl
import pandas as pd

DEFAULT_ANALYSIS = r"C:\Users\Manish.Soni\Downloads\RFP-Analysis_refreshed_20260515_135031_with_quantity.xlsx"


def _df_from_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header, *body = rows
    return pd.DataFrame(body, columns=list(header))


def _write_df_to_sheet(ws, df):
    """Clear ws (keep header) and rewrite rows from df."""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    cols = list(df.columns)
    for r_offset, rec in enumerate(df.itertuples(index=False, name=None), start=2):
        for c_offset, val in enumerate(rec, start=1):
            ws.cell(row=r_offset, column=c_offset, value=val)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--analysis", default=DEFAULT_ANALYSIS)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()
    if not os.path.exists(args.analysis):
        print(f"ERROR: {args.analysis} not found", file=sys.stderr)
        sys.exit(1)
    out_path = args.out or args.analysis

    print(f"Analysis: {args.analysis}")
    print(f"Output  : {out_path}\n")

    print("Loading workbook (writable, may take 30-60s)...")
    wb = openpyxl.load_workbook(args.analysis)

    # ---------------- RFP-Material_List ----------------
    ws = wb["RFP-Material_List"]
    print("Reading RFP-Material_List into DataFrame...")
    df = _df_from_sheet(ws)
    before = len(df)
    print(f"  rows in: {before}")
    df_dedup = df.drop_duplicates(keep="first")
    after = len(df_dedup)
    removed = before - after
    print(f"  duplicates removed : {removed}")
    print(f"  rows out           : {after}")
    if removed > 0:
        print("Writing deduped rows back...")
        _write_df_to_sheet(ws, df_dedup)

    # ---------------- RFP-List ----------------
    ws2 = wb["RFP-List"]
    print("\nReading RFP-List into DataFrame...")
    df2 = _df_from_sheet(ws2)
    before2 = len(df2)
    print(f"  rows in: {before2}")
    if "RFP_Title" not in df2.columns:
        print("ERROR: RFP_Title not in RFP-List", file=sys.stderr)
        sys.exit(1)
    df2_dedup = df2.drop_duplicates(subset=["RFP_Title"], keep="first")
    after2 = len(df2_dedup)
    removed2 = before2 - after2
    print(f"  duplicates removed : {removed2}")
    print(f"  rows out           : {after2}")
    if removed2 > 0:
        print("Writing deduped rows back...")
        _write_df_to_sheet(ws2, df2_dedup)

    print("\nSaving workbook...")
    try:
        wb.save(out_path)
    except PermissionError:
        base, ext = os.path.splitext(out_path)
        out_path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
        print(f"  WARNING: target locked; writing to {out_path}")
        wb.save(out_path)

    print("\nDone.")
    print(f"  Output: {out_path}")


if __name__ == "__main__":
    main()

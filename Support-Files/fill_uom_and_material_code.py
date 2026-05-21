"""
Fill 'Unit of Measurement' (new column) and missing Material_Code values in
RFP-Analysis-Overall.xlsx by looking up the matching row in each RFP's
original Excel file.

Lookup order for each row's source Excel:
  1) Local folder: C:\\python\\RFP-automation\\ALLRFPs (recursive walk)
  2) SharePoint:   RFP-logs/ALLRFPs/<Company_Name>/<RFP_Title>/downloaded-rfp/<Excel_File>

Matching strategy (per row in analysis, against rows in original RFP Excel):
  a) By 9-digit material code (from analysis Material_Code, falling back to
     digits embedded in Material_Description) -> match original's
     material_code (extracted from Name / Material Number / Material Code).
  b) By normalized Description equality.
  c) By normalized Name equality.
  d) By normalized "Number Name" combined equality (SEC layout).
  e) Prefix containment as a last resort.

Existing Material_Code values are preserved (only empty cells are filled).
Output is written to a new file: RFP-Analysis-Overall_with_UoM.xlsx.
"""

import os
import re
import sys
import warnings
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font

warnings.filterwarnings("ignore")

# Make project imports work when invoked from anywhere
HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


ANALYSIS_PATH = r"C:\Users\Manish.Soni\Downloads\RFP-fille-Analysis\RFP-Analysis-Overall.xlsx"
OUTPUT_PATH = r"C:\Users\Manish.Soni\Downloads\RFP-fille-Analysis\RFP-Analysis-Overall_with_UoM.xlsx"
ALLRFPS_BASE = r"C:\python\RFP-automation\ALLRFPs"
SP_TMP_DIR = r"C:\tmp\rfp_sharepoint_cache"

NINE_DIGIT = re.compile(r"\d{9}")
_WS_RE = re.compile(r"\s+")


def norm_text(s) -> str:
    if s is None:
        return ""
    return _WS_RE.sub(" ", str(s)).strip().lower()


def find_col(cols, *targets):
    for target in targets:
        t = target.lower().replace(" ", "").replace("_", "")
        for c in cols:
            if t in str(c).lower().replace(" ", "").replace("_", ""):
                return c
    return None


def find_material_sheet(xl_path: str):
    """Pick the sheet that actually holds material rows with Unit of Measure data.

    Some RFP workbooks have BOTH 'Other Content' (often empty) and a numbered
    section like '6 Commercial Envelope' that contains the real data. Always
    prefer a sheet that has a 'Unit of Measure' column AND non-empty rows in
    that column (skipping the first two header/description rows).
    """
    try:
        xl = pd.ExcelFile(xl_path)
    except Exception as e:
        print(f"  [WARN] Cannot open '{xl_path}': {e}")
        return None

    candidates_with_uom = []
    for s in xl.sheet_names:
        try:
            df = pd.read_excel(xl_path, sheet_name=s)
        except Exception:
            continue
        cols = list(df.columns)
        cols_norm = [str(c).lower().replace(" ", "") for c in cols]
        if not any("unitofmeasure" in c for c in cols_norm):
            continue
        uom_col = next(c for c, n in zip(cols, cols_norm) if "unitofmeasure" in n)
        # Count data rows after the first 2 (header description rows): a real
        # data row has either a non-empty UoM value or a Name/Description value.
        data_slice = df.iloc[2:] if len(df) > 2 else df
        non_empty_uom = data_slice[uom_col].dropna().shape[0]
        candidates_with_uom.append((s, non_empty_uom, len(data_slice)))

    if candidates_with_uom:
        # Pick the one with most non-empty UoM values; tie-break on row count
        candidates_with_uom.sort(key=lambda t: (t[1], t[2]), reverse=True)
        return candidates_with_uom[0][0]

    # No UoM sheet at all — fall back to 'Other Content' if present
    for name in xl.sheet_names:
        if name.replace(" ", "").lower() == "othercontent":
            return name
    # Last resort: any sheet with both 'name' and 'description'
    for s in xl.sheet_names:
        try:
            df = pd.read_excel(xl_path, sheet_name=s, nrows=5)
            cols_norm = [str(c).lower() for c in df.columns]
            if any("name" == c.strip() for c in cols_norm) and any(
                "description" == c.strip() for c in cols_norm
            ):
                return s
        except Exception:
            continue
    return None


def parse_rfp_excel(xl_path: str):
    sheet = find_material_sheet(xl_path)
    if not sheet:
        return []
    try:
        df = pd.read_excel(xl_path, sheet_name=sheet)
    except Exception as e:
        print(f"  [WARN] Cannot read sheet '{sheet}' from '{xl_path}': {e}")
        return []

    cols = list(df.columns)
    name_col = find_col(cols, "name")
    desc_col = find_col(cols, "description")
    num_col = find_col(cols, "number")
    uom_col = find_col(cols, "unitofmeasure", "unitofmeasurement", "uom")
    mat_num_col = find_col(cols, "materialnumber")
    mat_code_col = find_col(cols, "materialcode")
    qty_col = find_col(cols, "quantity")

    rows = []
    for r in df.to_dict("records"):
        name_v = r.get(name_col) if name_col else None
        desc_v = r.get(desc_col) if desc_col else None
        if (name_v is None or pd.isna(name_v)) and (desc_v is None or pd.isna(desc_v)):
            continue

        # Material code: try Name -> Material Number -> Material Code
        mat_codes = []
        if name_v is not None and not pd.isna(name_v):
            mat_codes = NINE_DIGIT.findall(str(name_v))
        if not mat_codes and mat_num_col:
            v = r.get(mat_num_col)
            if v is not None and not pd.isna(v):
                mat_codes = NINE_DIGIT.findall(str(v))
        if not mat_codes and mat_code_col:
            v = r.get(mat_code_col)
            if v is not None and not pd.isna(v):
                mat_codes = NINE_DIGIT.findall(str(v))
        material_code = mat_codes[0] if mat_codes else ""

        uom = ""
        if uom_col:
            v = r.get(uom_col)
            if v is not None and not pd.isna(v):
                uom = str(v).strip()

        number = ""
        if num_col:
            v = r.get(num_col)
            if v is not None and not pd.isna(v):
                number = str(v).strip()
                # Strip trailing ".0" from float-coerced integer numbers
                if re.fullmatch(r"-?\d+\.0", number):
                    number = number[:-2]

        qty = None
        if qty_col:
            v = r.get(qty_col)
            if v is not None and not pd.isna(v):
                qty = v

        name_s = str(name_v).strip() if name_v is not None and not pd.isna(name_v) else ""
        desc_s = str(desc_v).strip() if desc_v is not None and not pd.isna(desc_v) else ""

        rows.append(
            {
                "number": number,
                "name": name_s,
                "name_norm": norm_text(name_s),
                "description": desc_s,
                "description_norm": norm_text(desc_s),
                "number_name_norm": norm_text(f"{number} {name_s}") if number else "",
                "material_code": material_code,
                "uom": uom,
                "quantity": qty,
            }
        )
    return rows


def match_row(analysis_desc, analysis_mat_code, analysis_qty, candidates):
    if not candidates:
        return None

    code_str = str(analysis_mat_code).strip() if analysis_mat_code else ""
    if code_str:
        for c in candidates:
            if c["material_code"] and c["material_code"] == code_str:
                return c

    desc_str = str(analysis_desc) if analysis_desc else ""
    codes_in_desc = NINE_DIGIT.findall(desc_str)
    for code in codes_in_desc:
        for c in candidates:
            if c["material_code"] == code:
                return c

    norm_q = norm_text(analysis_desc)
    if norm_q:
        for c in candidates:
            if c["description_norm"] and c["description_norm"] == norm_q:
                return c
        for c in candidates:
            if c["name_norm"] and c["name_norm"] == norm_q:
                return c
        for c in candidates:
            if c["number_name_norm"] and c["number_name_norm"] == norm_q:
                return c
        # Prefix / containment fallback — pick longest reciprocal overlap
        best = None
        best_len = 0
        for c in candidates:
            for ref in (c["description_norm"], c["name_norm"], c["number_name_norm"]):
                if not ref or len(ref) < 12:
                    continue
                if norm_q.startswith(ref) or ref.startswith(norm_q):
                    if len(ref) > best_len:
                        best, best_len = c, len(ref)
        if best:
            return best

    return None


def build_local_index(base: str) -> dict:
    idx = {}
    for root, _dirs, files in os.walk(base):
        for fn in files:
            if fn.lower().endswith((".xls", ".xlsx")):
                idx.setdefault(fn, []).append(os.path.join(root, fn))
    return idx


class MinimalGraphClient:
    """Minimal Graph client (auth + download) that avoids the heavy
    helpers.sharepoint_helper import chain (which triggers a circular import
    because that module pulls in core.common_imports)."""

    def __init__(self, client_id, client_secret, tenant_id, hostname, site_path, drive_name):
        import requests
        import msal

        self._requests = requests
        self._msal = msal
        self.client_id = client_id
        self.client_secret = client_secret
        self.tenant_id = tenant_id
        self.hostname = hostname
        self.site_path = site_path
        self.drive_name = drive_name
        self.token = None
        self.headers = None
        self.site_id = None
        self.drive_id = None

    def auth(self):
        app = self._msal.ConfidentialClientApplication(
            self.client_id,
            authority=f"https://login.microsoftonline.com/{self.tenant_id}",
            client_credential=self.client_secret,
        )
        result = app.acquire_token_for_client(["https://graph.microsoft.com/.default"])
        if "access_token" not in result:
            raise RuntimeError(f"Token acquire failed: {result}")
        self.token = result["access_token"]
        self.headers = {"Authorization": f"Bearer {self.token}"}

    def resolve_site_and_drive(self):
        r = self._requests.get(
            f"https://graph.microsoft.com/v1.0/sites/{self.hostname}:{self.site_path}",
            headers=self.headers,
        )
        if r.status_code != 200:
            raise RuntimeError(f"Resolve site failed: {r.status_code} {r.text[:200]}")
        self.site_id = r.json().get("id")
        r = self._requests.get(
            f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives",
            headers=self.headers,
        )
        if r.status_code != 200:
            raise RuntimeError(f"List drives failed: {r.status_code} {r.text[:200]}")
        for d in r.json().get("value", []):
            if d.get("name") == self.drive_name:
                self.drive_id = d.get("id")
                break
        if not self.drive_id:
            raise RuntimeError(f"Drive '{self.drive_name}' not found on site")

    def download(self, sp_path: str, local_path: str):
        url = (
            f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives/"
            f"{self.drive_id}/root:/{sp_path}:/content"
        )
        r = self._requests.get(url, headers=self.headers, stream=True)
        if r.status_code != 200:
            raise RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")
        os.makedirs(os.path.dirname(local_path), exist_ok=True)
        with open(local_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)
        return local_path


def init_graph_client():
    try:
        # Direct import from config module — avoids helpers chain
        sys.path.insert(0, str(PROJECT_ROOT))
        from config.config import (
            CLIENT_ID,
            CLIENT_SECRET,
            TENANT_ID,
            SHAREPOINT_HOSTNAME,
            SITE_PATH,
            DRIVE_NAME,
        )
    except Exception as e:
        print(f"[WARN] Cannot import config: {e}")
        return None
    try:
        gc = MinimalGraphClient(
            CLIENT_ID, CLIENT_SECRET, TENANT_ID, SHAREPOINT_HOSTNAME, SITE_PATH, DRIVE_NAME
        )
        gc.auth()
        gc.resolve_site_and_drive()
        print("[OK] SharePoint client initialized")
        return gc
    except Exception as e:
        print(f"[WARN] SharePoint auth failed: {e}")
        return None


def fetch_from_sharepoint(graph_client, company_name: str, rfp_title: str, excel_file: str) -> str:
    safe_company = re.sub(r'[<>:"/\\|?*]', "_", company_name).strip().rstrip(".")
    clean_title = re.sub(r"\s+", " ", rfp_title).strip()
    sp_path = f"RFP-logs/ALLRFPs/{safe_company}/{clean_title}/downloaded-rfp/{excel_file}"
    os.makedirs(SP_TMP_DIR, exist_ok=True)
    local_path = os.path.join(SP_TMP_DIR, excel_file)
    try:
        graph_client.download(sp_path, local_path)
        return local_path
    except Exception as e:
        print(f"  [SP-MISS] {excel_file}: {e}")
        return ""


def main():
    print(f"Analysis Excel: {ANALYSIS_PATH}")
    print(f"Output Excel:   {OUTPUT_PATH}\n")

    print("Loading analysis Excel...")
    wb = openpyxl.load_workbook(ANALYSIS_PATH)
    ws = wb["RFP-Material_List"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Existing headers: {headers}")

    def col(name):
        if name not in headers:
            raise SystemExit(f"Required column '{name}' missing in RFP-Material_List")
        return headers.index(name) + 1

    col_company = col("Company_Name")
    col_rfp_title = col("RFP_Title")
    col_excel_file = col("Excel_File")
    col_material_code = col("Material_Code")
    col_material_desc = col("Material_Description")
    col_quantity = headers.index("Quantity") + 1 if "Quantity" in headers else None

    # Append the new column if not already present
    if "Unit of Measurement" not in headers:
        new_col_idx = ws.max_column + 1
        cell = ws.cell(row=1, column=new_col_idx, value="Unit of Measurement")
        cell.font = Font(bold=True)
    else:
        new_col_idx = headers.index("Unit of Measurement") + 1

    print("Building local file index...")
    local_idx = build_local_index(ALLRFPS_BASE)
    print(f"  {sum(len(v) for v in local_idx.values())} Excel files across {len(local_idx)} unique names\n")

    total = ws.max_row - 1
    print(f"Pass 1: scan {total} rows to collect unique (company, title, filename) triples...")
    needed = {}  # excel_file -> (company, title)
    for row_idx in range(2, ws.max_row + 1):
        excel_file = ws.cell(row=row_idx, column=col_excel_file).value
        if not excel_file:
            continue
        if excel_file not in needed:
            needed[excel_file] = (
                ws.cell(row=row_idx, column=col_company).value or "",
                ws.cell(row=row_idx, column=col_rfp_title).value or "",
            )
    print(f"  {len(needed)} unique RFP files referenced\n")

    file_cache = {}
    sp_attempted = set()
    sp_fetched = set()
    sp_failed = set()
    graph_client = None
    graph_init_tried = False

    print("Pass 2: parse referenced Excel files...")
    parsed = 0
    parsed_local = 0
    parsed_sp = 0
    for excel_file, (company, title) in needed.items():
        path = local_idx.get(excel_file, [None])[0]
        if not path:
            if not graph_init_tried:
                print("  Initializing SharePoint client (one-time)...")
                graph_client = init_graph_client()
                graph_init_tried = True
            if graph_client is not None:
                sp_attempted.add(excel_file)
                fetched = fetch_from_sharepoint(graph_client, company, title, excel_file)
                if fetched:
                    path = fetched
                    sp_fetched.add(excel_file)
                else:
                    sp_failed.add(excel_file)
            else:
                sp_failed.add(excel_file)

        if not path:
            file_cache[excel_file] = []
            continue

        rows = parse_rfp_excel(path)
        file_cache[excel_file] = rows
        parsed += 1
        if excel_file in sp_fetched:
            parsed_sp += 1
        else:
            parsed_local += 1
        if parsed % 100 == 0:
            print(f"  {parsed}/{len(needed)} files parsed ({parsed_local} local, {parsed_sp} SP)")

    print(f"  Parsed {parsed} files (local={parsed_local}, sharepoint={parsed_sp})")
    if sp_failed:
        print(f"  SharePoint fetch failed for {len(sp_failed)} files")
    print()

    print("Pass 3: match analysis rows -> fill UoM and missing Material_Code...")
    matched_uom = 0
    matched_code = 0
    no_match_rows = 0
    no_file_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        excel_file = ws.cell(row=row_idx, column=col_excel_file).value
        if not excel_file:
            continue
        candidates = file_cache.get(excel_file, [])
        if not candidates:
            no_file_rows += 1
            continue

        mat_code = ws.cell(row=row_idx, column=col_material_code).value
        mat_desc = ws.cell(row=row_idx, column=col_material_desc).value
        qty = ws.cell(row=row_idx, column=col_quantity).value if col_quantity else None

        match = match_row(mat_desc, mat_code, qty, candidates)
        if not match:
            no_match_rows += 1
            continue

        if match["uom"]:
            ws.cell(row=row_idx, column=new_col_idx, value=match["uom"])
            matched_uom += 1
        if (mat_code is None or str(mat_code).strip() == "") and match["material_code"]:
            ws.cell(row=row_idx, column=col_material_code, value=match["material_code"])
            matched_code += 1

        if row_idx % 2000 == 0:
            print(f"  {row_idx-1}/{total} rows scanned")

    print("\nSaving workbook...")
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"  Wrote: {OUTPUT_PATH}\n")

    print("=== Summary ===")
    print(f"  Total data rows:                       {total}")
    print(f"  Rows with UoM filled:                  {matched_uom}")
    print(f"  Rows with missing Material_Code filled:{matched_code}")
    print(f"  Rows skipped (file unavailable):       {no_file_rows}")
    print(f"  Rows scanned but no match in file:     {no_match_rows}")
    if sp_attempted:
        print(f"  SharePoint attempted: {len(sp_attempted)}, fetched: {len(sp_fetched)}, failed: {len(sp_failed)}")
    if sp_failed:
        print("\nFiles still missing after SharePoint fallback:")
        for f in sorted(sp_failed):
            print(f"  {f}")


if __name__ == "__main__":
    main()

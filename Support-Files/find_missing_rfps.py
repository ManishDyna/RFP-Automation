"""
Find RFPs Missing from Dataverse
=================================
Compares an "All RFPs" CSV/Excel file against the Dataverse database
to find which RFPs have NOT been added to the system yet.

Supported CSV formats:
    Format 1 (Portal export):  Title, ID, End Time, Event Type, Participated
        - "Title" is used as RFP_ID
        - Company is detected from title prefix (Aramco_, SEC RFP, SABIC_, HADEED_)

    Format 2 (Simple):  RFP_ID, Company_Name
        - Direct column mapping

Output:
    - Console summary showing matched / missing counts
    - CSV file with missing RFPs:   missing_rfps_<timestamp>.csv
    - CSV file with existing RFPs:  existing_rfps_<timestamp>.csv (optional, with --show-existing)

Usage:
    python find_missing_rfps.py --file All-RFPs.xls
    python find_missing_rfps.py --file my_rfps.csv --output results/
    python find_missing_rfps.py --file my_rfps.csv --show-existing
    python find_missing_rfps.py --file my_rfps.csv --company "Aramco e-Marketplace"
"""

import os
import re
import sys
import csv
import argparse
from datetime import datetime

# Add project root to sys.path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config.config import (
    TENANT_ID, CLIENT_ID, CLIENT_SECRET, RESOURCE_URL,
    RFP_ACTIVITY_LOG_TABLE_API, RFP_ACTIVITY_LOG_TABLE_LOGICAL,
)
from helpers.dataverse_helper import DataverseClient


# ─────────────────────────────────────────────────────────────────────────────
# Company detection from RFP title prefix
# ─────────────────────────────────────────────────────────────────────────────

TITLE_PREFIX_TO_COMPANY = {
    "aramco":  "Aramco e-Marketplace",
    "sec":     "Saudi Energy",
    "sabic":   "SABIC - Saudi Basic Industries Corp.",
    "hadeed":  "HADEED - RAJHI STEEL",
}


def detect_company_from_title(title: str) -> str:
    """
    Detect company name from the RFP title prefix.

    Examples:
        "Aramco_4203233143_CABLE..."  → "Aramco e-Marketplace"
        "SEC RFP-C001718985"          → "Saudi Energy"
        "SABIC_12345_PRODUCT..."      → "SABIC - Saudi Basic Industries Corp."
        "HADEED_99999_STEEL..."       → "HADEED - RAJHI STEEL"
    """
    title_lower = title.strip().lower()

    for prefix, company in TITLE_PREFIX_TO_COMPANY.items():
        # Match prefix followed by underscore, space, or hyphen
        if title_lower.startswith(prefix + "_") or \
           title_lower.startswith(prefix + " ") or \
           title_lower.startswith(prefix + "-"):
            return company

    return "Unknown"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def _normalize(s: str) -> str:
    """Strip all non-alphanumeric chars and lowercase for fuzzy matching."""
    return re.sub(r'[^a-z0-9]', '', s.lower())


def _resolve_col(headers: list, target: str):
    """Case-insensitive column lookup (spaces/underscores treated the same)."""
    normalized_target = target.lower().replace(" ", "_")
    for h in headers:
        if h.strip().lower().replace(" ", "_") == normalized_target:
            return h
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Load CSV / Excel
# ─────────────────────────────────────────────────────────────────────────────

def _load_raw_rows(file_path: str) -> list[dict]:
    """Read raw rows from CSV or Excel file."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        try:
            if ext == ".xls":
                # Use xlrd for old .xls format
                import xlrd
                wb = xlrd.open_workbook(file_path)
                ws = wb.sheet_by_index(0)
                headers = [str(v).strip() for v in ws.row_values(0)]
                raw_rows = []
                for i in range(1, ws.nrows):
                    vals = [str(v).strip() for v in ws.row_values(i)]
                    raw_rows.append(dict(zip(headers, vals)))
                return raw_rows
            else:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                raw_headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                raw_rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    raw_rows.append(dict(zip(raw_headers, [str(v or "").strip() for v in row])))
                wb.close()
                return raw_rows
        except Exception as exc:
            print(f"[ERROR] Could not read Excel file: {exc}")
            sys.exit(1)

    elif ext == ".csv":
        raw_rows = []
        with open(file_path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                raw_rows.append(dict(row))
        return raw_rows

    else:
        print(f"[ERROR] Unsupported file type '{ext}'. Use .csv, .xlsx, or .xls")
        sys.exit(1)


def load_csv_rfps(file_path: str, company_override: str = None) -> list[dict]:
    """
    Load RFPs from a CSV or Excel file.

    Auto-detects format:
        Format 1 (Portal): has "Title" column → Title = RFP_ID, company detected from prefix
        Format 2 (Simple): has "RFP_ID" + "Company_Name" columns

    Returns list of dicts: [{'rfp_id': ..., 'company_name': ..., 'end_time': ..., ...}, ...]
    """
    if not os.path.isfile(file_path):
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)

    raw_rows = _load_raw_rows(file_path)

    if not raw_rows:
        print(f"[ERROR] No data rows found in {file_path}")
        sys.exit(1)

    headers = list(raw_rows[0].keys())
    _log(f"CSV columns found: {headers}")

    # Detect format
    col_title = _resolve_col(headers, "title")
    col_rfp_id = _resolve_col(headers, "rfp_id")
    col_company = _resolve_col(headers, "company_name")
    col_end_time = _resolve_col(headers, "end_time")
    col_event_type = _resolve_col(headers, "event_type")
    col_participated = _resolve_col(headers, "participated")
    col_doc_id = _resolve_col(headers, "id")

    # Determine which column has the RFP ID
    if col_title:
        id_col = col_title
        format_name = "Portal format (Title column)"
    elif col_rfp_id:
        id_col = col_rfp_id
        format_name = "Simple format (RFP_ID column)"
    else:
        print(f"[ERROR] CSV must have a 'Title' or 'RFP_ID' column. Found: {headers}")
        sys.exit(1)

    _log(f"Detected: {format_name}")

    rows = []
    skipped = 0
    for raw in raw_rows:
        rfp_id = str(raw.get(id_col) or "").strip()

        if not rfp_id or rfp_id.lower() in ("nan", "none", ""):
            skipped += 1
            continue

        # Determine company name
        if company_override:
            company = company_override
        elif col_company:
            company = str(raw.get(col_company) or "").strip()
        else:
            company = detect_company_from_title(rfp_id)

        row = {
            "rfp_id": rfp_id,
            "company_name": company,
        }

        # Carry over extra columns if present
        if col_end_time:
            row["end_time"] = str(raw.get(col_end_time) or "").strip()
        if col_event_type:
            row["event_type"] = str(raw.get(col_event_type) or "").strip()
        if col_participated:
            row["participated"] = str(raw.get(col_participated) or "").strip()
        if col_doc_id:
            row["doc_id"] = str(raw.get(col_doc_id) or "").strip()

        rows.append(row)

    _log(f"Loaded {len(rows)} RFPs from file ({skipped} blank rows skipped).")
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Fetch from Dataverse
# ─────────────────────────────────────────────────────────────────────────────

def fetch_dataverse_rfps() -> list[dict]:
    """
    Fetch all RFP_ID and Company_Name from Dataverse.
    Returns list of dicts: [{'rfp_id': ..., 'company_name': ...}, ...]
    """
    _log("Connecting to Dataverse...")
    client = DataverseClient(
        tenant_id=TENANT_ID,
        client_id=CLIENT_ID,
        client_secret=CLIENT_SECRET,
        resource_url=RESOURCE_URL,
    )

    _log("Fetching all RFPs from Dataverse (this may take a moment)...")
    db_rows = client.get_all_rows(
        table_api_name=RFP_ACTIVITY_LOG_TABLE_API,
        select_columns=["RFP_ID", "Company_Name"],
        table_logical_name=RFP_ACTIVITY_LOG_TABLE_LOGICAL,
        use_display_names=True,
    )

    results = []
    for row in db_rows:
        rfp_id = str(row.get("RFP_ID") or "").strip()
        company = str(row.get("Company_Name") or "").strip()
        if rfp_id:
            results.append({"rfp_id": rfp_id, "company_name": company})

    _log(f"Fetched {len(results)} RFPs from Dataverse.")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Compare
# ─────────────────────────────────────────────────────────────────────────────

def compare_rfps(csv_rfps: list[dict], db_rfps: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Compare CSV RFPs against Dataverse RFPs using fuzzy matching
    (strips all non-alphanumeric chars before comparing RFP IDs).

    Returns:
        (missing_rfps, existing_rfps)
        - missing_rfps: RFPs in CSV but NOT in Dataverse
        - existing_rfps: RFPs in CSV that ARE in Dataverse
    """
    # Build a set of normalized RFP IDs from Dataverse
    db_normalized = set()
    for row in db_rfps:
        db_normalized.add(_normalize(row["rfp_id"]))

    missing = []
    existing = []

    for row in csv_rfps:
        norm = _normalize(row["rfp_id"])
        if norm in db_normalized:
            existing.append(row)
        else:
            missing.append(row)

    return missing, existing


# ─────────────────────────────────────────────────────────────────────────────
# Output
# ─────────────────────────────────────────────────────────────────────────────

def write_csv(rows: list[dict], file_path: str):
    """Write rows to a CSV file with all available columns."""
    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)

    # Build fieldnames from all keys across rows, in consistent order
    base_fields = ["Company_Name", "RFP_ID"]
    extra_fields = ["End_Time", "Event_Type", "Participated", "Doc_ID"]
    # Map internal keys to CSV header names
    key_to_header = {
        "company_name": "Company_Name",
        "rfp_id": "RFP_ID",
        "end_time": "End_Time",
        "event_type": "Event_Type",
        "participated": "Participated",
        "doc_id": "Doc_ID",
    }

    # Determine which extra fields are actually present
    fieldnames = list(base_fields)
    for ef_key, ef_header in [("end_time", "End_Time"), ("event_type", "Event_Type"),
                               ("participated", "Participated"), ("doc_id", "Doc_ID")]:
        if any(ef_key in row for row in rows):
            fieldnames.append(ef_header)

    with open(file_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            csv_row = {}
            for key, value in row.items():
                header = key_to_header.get(key)
                if header and header in fieldnames:
                    csv_row[header] = value
            writer.writerow(csv_row)

    _log(f"Saved {len(rows)} rows to: {file_path}")


def print_summary(csv_rfps, db_rfps, missing, existing):
    """Print a nice summary to the console."""
    print()
    print("=" * 65)
    print("  RFP Comparison Report: CSV vs Dataverse")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65)
    print(f"  Total RFPs in CSV file     : {len(csv_rfps)}")
    print(f"  Total RFPs in Dataverse    : {len(db_rfps)}")
    print(f"  Already in DB (matched)    : {len(existing)}")
    print(f"  MISSING from DB            : {len(missing)}")
    print("=" * 65)

    if missing:
        # Group missing by company
        by_company = {}
        for row in missing:
            by_company.setdefault(row["company_name"], []).append(row["rfp_id"])

        print(f"\n  Missing RFPs by Company:")
        print(f"  {'-' * 55}")
        for company in sorted(by_company.keys()):
            ids = by_company[company]
            print(f"  {company}: {len(ids)} missing")
            for rfp_id in ids[:10]:  # Show first 10 per company
                print(f"    - {rfp_id}")
            if len(ids) > 10:
                print(f"    ... and {len(ids) - 10} more")
        print(f"  {'-' * 55}")
    else:
        print("\n  All RFPs from CSV are already in Dataverse!")

    print()


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Find RFPs from a CSV/Excel that are missing in Dataverse",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Supported input formats:
  Format 1 (Portal export):  Title, ID, End Time, Event Type, Participated
  Format 2 (Simple):         RFP_ID, Company_Name

Examples:
  python find_missing_rfps.py --file All-RFPs.xls
  python find_missing_rfps.py --file all_rfps.csv --output results/
  python find_missing_rfps.py --file all_rfps.csv --show-existing
  python find_missing_rfps.py --file all_rfps.csv --company "Aramco e-Marketplace"
        """,
    )
    parser.add_argument("--file", required=True, help="Path to CSV/Excel file with all RFPs")
    parser.add_argument("--output", default=".", help="Output directory for result CSVs (default: current directory)")
    parser.add_argument("--show-existing", action="store_true", help="Also save a CSV of RFPs that already exist in DB")
    parser.add_argument("--company", default=None, help="Override company name for all rows (useful when CSV has no Company_Name column)")

    args = parser.parse_args()

    # Step 1: Load CSV
    _log(f"Loading file: {args.file}")
    csv_rfps = load_csv_rfps(args.file, company_override=args.company)

    if not csv_rfps:
        print("[ERROR] No valid RFPs found in file.")
        sys.exit(1)

    # Show detected companies
    companies = sorted(set(r["company_name"] for r in csv_rfps))
    _log(f"Companies detected: {', '.join(companies)}")
    for co in companies:
        count = sum(1 for r in csv_rfps if r["company_name"] == co)
        _log(f"  {co}: {count} RFPs")

    # Step 2: Fetch from Dataverse
    db_rfps = fetch_dataverse_rfps()

    # Step 3: Compare
    _log("Comparing CSV against Dataverse...")
    missing, existing = compare_rfps(csv_rfps, db_rfps)

    # Step 4: Print summary
    print_summary(csv_rfps, db_rfps, missing, existing)

    # Step 5: Save results
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = args.output

    if missing:
        missing_path = os.path.join(output_dir, f"missing_rfps_{timestamp}.csv")
        write_csv(missing, missing_path)

    if args.show_existing and existing:
        existing_path = os.path.join(output_dir, f"existing_rfps_{timestamp}.csv")
        write_csv(existing, existing_path)

    if not missing:
        _log("Nothing to do - all RFPs are already in the database.")


if __name__ == "__main__":
    main()

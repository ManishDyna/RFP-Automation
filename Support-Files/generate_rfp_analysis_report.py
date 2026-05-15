"""
Generate RFP-Analysis.xlsx report from a Dataverse CSV export of cr673_bahra_rfps_v2s.

Produces a 3-sheet workbook matching the format of the original RFP-Analysis_original.xlsx:
  - RFP-Material_List : one row per material item in each RFP
  - RFP-List          : one row per RFP with match summary + participation
  - RFP-Count         : pivot-style nested counts grouped by company > material > keyword > participant

Usage:
  python Support-Files/generate_rfp_analysis_report.py \
      --csv "C:\\path\\to\\cr673_bahra_rfps_v2s.csv" \
      --out "C:\\path\\to\\RFP-Analysis_refreshed_YYYYMMDD_HHMMSS.xlsx"

If --out is omitted, writes to the user's Downloads folder with a timestamped name.
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font


CSV_FIELD_LIMIT = 10_000_000

MATERIAL_LIST_HEADERS = [
    "Company_Name", "RFP_Title", "RFP_ID", "End_Time", "Excel_File",
    "Material_Code", "Material_Description", "Material_Matched",
    "Matched_Keywords", "Keyword_Matched", "Participant",
]
RFP_LIST_HEADERS = [
    "Company_Name", "RFP_Title", "RFP_ID", "End_Time", "Participant",
    "is_material_match", "is_keyword_match",
    "no_of_matched_materials", "no_of_matched_keywords",
]
RFP_COUNT_HEADERS = [
    "Row Labels", "Count of RFP_Title",
    "Sum of no_of_matched_materials", "Sum of no_of_matched_keywords",
]

MATERIAL_LIST_WIDTHS = {
    "A": 24.28515625, "B": 89.5703125, "C": 14.42578125, "D": 20.7109375,
    "E": 92.7109375, "F": 14.42578125, "G": 69.7109375, "H": 18.0,
    "I": 46.85546875, "J": 18.42578125, "K": 10.85546875,
}
RFP_LIST_WIDTHS = {
    "B": 49.28515625, "D": 17.140625, "E": 21.0, "F": 17.85546875,
    "G": 80.42578125, "H": 23.42578125, "I": 23.7109375,
}
RFP_COUNT_WIDTHS = {
    "A": 26.85546875, "B": 16.7109375, "C": 30.0, "D": 30.28515625,
}


def normalize_participated(val: str) -> str:
    if not val:
        return "Not Participated"
    v = val.strip().lower()
    if v in ("yes", "submitted"):
        return "Participated"
    if v in ("declined",):
        return "Declined"
    return "Not Participated"


def format_end_time(val: str) -> str:
    if not val:
        return ""
    s = val.strip()
    if "." in s:
        date_part, _, frac = s.partition(".")
        frac_digits = "".join(c for c in frac if c.isdigit())[:6]
        s_for_parse = f"{date_part}.{frac_digits}" if frac_digits else date_part
    else:
        s_for_parse = s

    fmts = [
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M",
    ]
    dt = None
    for fmt in fmts:
        try:
            dt = datetime.strptime(s_for_parse, fmt)
            break
        except ValueError:
            continue
    if dt is None:
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return s
    return dt.strftime("%m/%d/%Y %I:%M %p")


def load_csv_rows(csv_path: str):
    csv.field_size_limit(CSV_FIELD_LIMIT)
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def parse_matched_data(raw: str):
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return None


def derive_rfp_list_row(row: dict) -> dict:
    matched = parse_matched_data(row.get("cr673_matched_data", ""))
    summary = (matched or {}).get("summary") or {}
    exact_count = int(summary.get("exact_match_count") or 0)
    keyword_count = int(summary.get("keyword_match_count") or 0)
    participant = normalize_participated(row.get("cr673_participated", ""))
    return {
        "Company_Name": row.get("cr673_company_name") or "",
        "RFP_Title": row.get("cr673_rfp_id") or "",
        "RFP_ID": row.get("cr673_rfp_id") or "",
        "End_Time": format_end_time(row.get("cr673_rfp_end_date", "")),
        "Participant": participant,
        "is_material_match": "Material matched" if exact_count > 0 else "Material not matched",
        "is_keyword_match": "Keyword matched" if keyword_count > 0 else "Keyword not matched",
        "no_of_matched_materials": exact_count,
        "no_of_matched_keywords": keyword_count,
    }


def derive_material_list_rows(row: dict):
    matched = parse_matched_data(row.get("cr673_matched_data", ""))
    if not matched:
        return []
    company = row.get("cr673_company_name") or ""
    rfp_id_val = row.get("cr673_rfp_id") or ""
    end_time = format_end_time(row.get("cr673_rfp_end_date", ""))
    excel_file = matched.get("source_file") or ""
    participant_full = normalize_participated(row.get("cr673_participated", ""))
    participant_yn = "Yes" if participant_full == "Participated" else "No"

    out = []

    def base():
        return {
            "Company_Name": company,
            "RFP_Title": rfp_id_val,
            "RFP_ID": rfp_id_val,
            "End_Time": end_time,
            "Excel_File": excel_file,
            "Participant": participant_yn,
        }

    for item in matched.get("exact_matches") or []:
        r = base()
        r["Material_Code"] = item.get("material_code") or ""
        r["Material_Description"] = item.get("material_description") or item.get("excel_description") or item.get("excel_name") or ""
        r["Material_Matched"] = "Yes"
        r["Matched_Keywords"] = ""
        r["Keyword_Matched"] = "No"
        out.append(r)

    for item in matched.get("keyword_matches") or []:
        r = base()
        r["Material_Code"] = item.get("material_code") or ""
        r["Material_Description"] = item.get("material_description") or item.get("excel_description") or item.get("excel_name") or ""
        r["Material_Matched"] = "No"
        r["Matched_Keywords"] = item.get("matched_keyword") or ""
        r["Keyword_Matched"] = "Yes"
        out.append(r)

    for item in matched.get("not_matched") or []:
        r = base()
        r["Material_Code"] = item.get("material_code") or ""
        r["Material_Description"] = item.get("material_description") or item.get("excel_description") or item.get("excel_name") or ""
        r["Material_Matched"] = "No"
        r["Matched_Keywords"] = ""
        r["Keyword_Matched"] = "No"
        out.append(r)

    return out


def apply_header_style(ws, n_cols: int):
    bold = Font(bold=True)
    for col_idx in range(1, n_cols + 1):
        ws.cell(row=1, column=col_idx).font = bold


def apply_column_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_material_list_sheet(ws, material_rows):
    ws.append(MATERIAL_LIST_HEADERS)
    for r in material_rows:
        ws.append([r.get(h, "") for h in MATERIAL_LIST_HEADERS])
    apply_header_style(ws, len(MATERIAL_LIST_HEADERS))
    apply_column_widths(ws, MATERIAL_LIST_WIDTHS)


def write_rfp_list_sheet(ws, rfp_rows):
    ws.append(RFP_LIST_HEADERS)
    for r in rfp_rows:
        ws.append([r.get(h, "") for h in RFP_LIST_HEADERS])
    apply_header_style(ws, len(RFP_LIST_HEADERS))
    apply_column_widths(ws, RFP_LIST_WIDTHS)


def write_rfp_count_sheet(ws, rfp_rows):
    """Pivot: Company > is_material_match > is_keyword_match > Participant."""
    tree = {}
    grand_count = 0
    grand_mat = 0
    grand_kw = 0
    for r in rfp_rows:
        company = r["Company_Name"]
        mat = r["is_material_match"]
        kw = r["is_keyword_match"]
        part = r["Participant"]
        mat_n = r["no_of_matched_materials"]
        kw_n = r["no_of_matched_keywords"]
        c = tree.setdefault(company, {})
        m = c.setdefault(mat, {})
        k = m.setdefault(kw, {})
        leaf = k.setdefault(part, {"count": 0, "mat": 0, "kw": 0})
        leaf["count"] += 1
        leaf["mat"] += mat_n
        leaf["kw"] += kw_n
        grand_count += 1
        grand_mat += mat_n
        grand_kw += kw_n

    ws.append(RFP_COUNT_HEADERS)

    material_order = ["Material matched", "Material not matched"]
    keyword_order = ["Keyword matched", "Keyword not matched"]
    participant_order = ["Declined", "Not Participated", "Participated"]

    for company in sorted(tree.keys()):
        ws.append([company, None, None, None])
        c = tree[company]
        for mat in material_order:
            if mat not in c:
                continue
            ws.append([mat, None, None, None])
            for kw in keyword_order:
                if kw not in c[mat]:
                    continue
                ws.append([kw, None, None, None])
                for part in participant_order:
                    if part not in c[mat][kw]:
                        continue
                    leaf = c[mat][kw][part]
                    ws.append([part, leaf["count"], leaf["mat"], leaf["kw"]])

    ws.append(["Grand Total", grand_count, grand_mat, grand_kw])
    apply_column_widths(ws, RFP_COUNT_WIDTHS)


def main():
    ap = argparse.ArgumentParser(description="Generate RFP-Analysis.xlsx from Dataverse CSV export")
    ap.add_argument("--csv", required=True, help="Path to cr673_bahra_rfps_v2s.csv")
    ap.add_argument("--out", default=None, help="Output xlsx path (defaults to Downloads with timestamp)")
    args = ap.parse_args()

    csv_path = args.csv
    if not os.path.exists(csv_path):
        print(f"ERROR: CSV not found: {csv_path}", file=sys.stderr)
        sys.exit(1)

    if args.out:
        out_path = args.out
    else:
        downloads = Path.home() / "Downloads"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = str(downloads / f"RFP-Analysis_refreshed_{ts}.xlsx")

    print(f"Reading CSV: {csv_path}")
    rows = load_csv_rows(csv_path)
    print(f"  {len(rows)} CSV rows loaded")

    print("Building RFP-List rows ...")
    rfp_list_rows = [derive_rfp_list_row(r) for r in rows]
    print(f"  {len(rfp_list_rows)} RFP rows")

    print("Building RFP-Material_List rows ...")
    material_rows = []
    for r in rows:
        material_rows.extend(derive_material_list_rows(r))
    print(f"  {len(material_rows)} material rows")

    print("Writing workbook ...")
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "RFP-Material_List"
    write_material_list_sheet(ws1, material_rows)

    ws2 = wb.create_sheet("RFP-List")
    write_rfp_list_sheet(ws2, rfp_list_rows)

    ws3 = wb.create_sheet("RFP-Count")
    write_rfp_count_sheet(ws3, rfp_list_rows)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    print(f"\nWrote: {out_path}")

    total_mat = sum(r["no_of_matched_materials"] for r in rfp_list_rows)
    total_kw = sum(r["no_of_matched_keywords"] for r in rfp_list_rows)
    print("\nSummary check:")
    print(f"  RFP-List rows                  : {len(rfp_list_rows)}")
    print(f"  RFP-Material_List rows         : {len(material_rows)}")
    print(f"  Sum no_of_matched_materials    : {total_mat}")
    print(f"  Sum no_of_matched_keywords     : {total_kw}")


if __name__ == "__main__":
    main()

"""
Get RFP Owner & Publish Date — Local Only
==========================================
Reads the same two-file input as download_from_csv.py, but instead of
downloading anything it simply opens each RFP's portal page and scrapes
the Owner name and Publish date/time.

No downloads.  No log entries.  No report files.

Uses TWO input files (identical format to download_from_csv.py):
  File 1 — RFP ID list  (--file argument)
      CSV or Excel with columns: RFP_ID, Company_Name [, Link]
  File 2 — Master links file  (same hardcoded path)
      Excel (.xls): Title, ID, End Time, Event Type, Participated
      Path: ALLRFPs/Portal-Rfps/All-RFPs.xls

Usage:
    python get_rfp_info.py --file rfps.csv
    python get_rfp_info.py --file rfps.xlsx
    python get_rfp_info.py --file rfps.xlsx --username user@example.com --password MyPass
    python get_rfp_info.py --file rfps.xlsx --headless

Credentials (priority order):
    1. --username / --password CLI flags
    2. BAHRA_SAP_USERNAME / BAHRA_SAP_PASSWORD env vars

Output:
    Prints a table to stdout with columns:
        RFP_ID | Company | Owner | Publish_Date
    Also saves a CSV: RFP_Info_<timestamp>.csv in the current directory.
"""

import os
import re
import csv
import sys
import asyncio
import argparse
from datetime import datetime


# ─────────────────────────────────────────────────────────────────────────────
# Constants  (same as download_from_csv.py)
# ─────────────────────────────────────────────────────────────────────────────

PORTAL_URL        = "https://service.ariba.com/Sourcing.aw/109582016/aw?awh=r&awssk=u9fNiSxN&dard=1#b0"
ARIBA_BASE_URL    = "https://service.ariba.com/Sourcing.aw/109582016/aw?awh=r&awssk=u9fNiSxN&dard=1"
MASTER_LINKS_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "ALLRFPs", "Portal-Rfps", "All-RFPs.xls",
)
LOGIN_MAX_RETRIES = 3

# ─── Hardcoded credentials (override with CLI flags or env vars if needed) ───
DEFAULT_USERNAME = "Loai.Albar@bahra-cables.com"
DEFAULT_PASSWORD = "Bahra@2026"


# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────

def _log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def _die(msg: str):
    print(f"[ERROR] {msg}")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# Input file loading  (copied from download_from_csv.py)
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_for_match(s: str) -> str:
    return re.sub(r'[^a-z0-9]', '', s.lower())


def _clean_id(rfp_id: str) -> str:
    return re.sub(r'\s+', ' ', rfp_id).strip()


def _resolve_col(headers: list, target: str):
    for h in headers:
        if h.strip().lower().replace(" ", "_") == target.lower():
            return h
    return None


def _company_from_title(title: str) -> str:
    """Extract company name from a title like 'Aramco_4203233143_CABLE...' → 'Aramco'."""
    part = title.split('_')[0].strip()
    return part if part else "Unknown"


def _parse_rows(raw_rows: list, source_label: str, hyperlinks: dict = None) -> list:
    """
    Parse rows from either format:
      Format A (standard):    RFP_ID, Company_Name [, Link]
      Format B (All-RFPs):    Title, ID [, End Time, Event Type, Participated]
    """
    if not raw_rows:
        _die(f"No data rows found in {source_label}")

    headers  = list(raw_rows[0].keys())
    col_id   = _resolve_col(headers, "rfp_id")
    col_co   = _resolve_col(headers, "company_name")
    col_link = _resolve_col(headers, "link")

    # ── Format B detection: has Title + ID but no RFP_ID ─────────────────
    col_title = _resolve_col(headers, "title")
    col_docid = _resolve_col(headers, "id")
    use_master_format = (col_title and col_docid and not col_id)

    if not use_master_format and not col_id:
        _die(
            f"Input file must have either:\n"
            f"  • 'RFP_ID' and 'Company_Name' columns  (standard format), or\n"
            f"  • 'Title' and 'ID' columns              (All-RFPs format)\n"
            f"Found: {headers}"
        )

    rows = []
    for lineno, raw in enumerate(raw_rows, start=2):
        if use_master_format:
            rfp_id  = str(raw.get(col_title)  or "").strip()
            doc_id  = str(raw.get(col_docid)  or "").strip()
            company = _company_from_title(rfp_id)
            # Prefer hyperlink embedded in the Title cell; fall back to Doc ID URL
            hl_key  = lineno - 1   # row index used when reading with hyperlink map
            link    = (hyperlinks or {}).get(hl_key, "")
            if not link and doc_id:
                link = f"{ARIBA_BASE_URL}&an={doc_id}"
        else:
            rfp_id  = str(raw.get(col_id)  or "").strip()
            company = str(raw.get(col_co)  or "").strip()
            link    = str(raw.get(col_link) or "").strip() if col_link else ""
            if link.lower() in ("nan", "none", "n/a", "-"):
                link = ""

        if not rfp_id:
            print(f"  [WARN] Row {lineno}: empty title/rfp_id — skipped")
            continue
        rows.append({"rfp_id": rfp_id, "company_name": company, "link": link})

    return rows


def load_input_file(file_path: str) -> list:
    if not os.path.isfile(file_path):
        _die(f"Input file not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            # Load with data_only=False so hyperlinks are preserved
            wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
            ws = wb.active

            raw_headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

            # Detect All-RFPs format to capture hyperlinks from the Title cell
            headers_lower = [h.lower() for h in raw_headers]
            is_master_fmt = ("title" in headers_lower and "id" in headers_lower
                             and "rfp_id" not in headers_lower)

            hyperlinks = {}
            raw_rows   = []
            for row_i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                raw_rows.append(dict(zip(raw_headers, [str(c.value or "").strip() for c in row])))
                if is_master_fmt:
                    title_col_idx = headers_lower.index("title")
                    cell = row[title_col_idx]
                    if cell.hyperlink and cell.hyperlink.target:
                        hyperlinks[row_i] = cell.hyperlink.target.strip()

            wb.close()
        except Exception as exc:
            _die(f"Could not read Excel file '{file_path}': {exc}")

    elif ext == ".csv":
        hyperlinks = {}
        raw_rows   = []
        with open(file_path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                raw_rows.append(dict(row))
    else:
        _die(f"Unsupported file type '{ext}'. Use .csv, .xlsx, or .xls")

    return _parse_rows(raw_rows, file_path, hyperlinks=hyperlinks)


# ─────────────────────────────────────────────────────────────────────────────
# Master links file  (copied from download_from_csv.py)
# ─────────────────────────────────────────────────────────────────────────────

def load_master_rfp_links(master_file: str) -> dict:
    if not os.path.isfile(master_file):
        _log(f"[WARN] Master links file not found: {master_file}")
        return {}
    try:
        import xlrd
        wb = xlrd.open_workbook(master_file)
        ws = wb.sheet_by_index(0)
        headers = [str(v).strip() for v in ws.row_values(0)]
        title_col     = next((i for i, h in enumerate(headers) if h.lower() == 'title'),    None)
        id_col        = next((i for i, h in enumerate(headers) if h.lower() == 'id'),       None)
        end_time_col  = next((i for i, h in enumerate(headers)
                              if h.lower() in ('end time', 'end_time', 'endtime', 'end date', 'close date')), None)
        event_col     = next((i for i, h in enumerate(headers)
                              if h.lower() in ('event type', 'event_type', 'eventtype', 'type')), None)
        part_col      = next((i for i, h in enumerate(headers)
                              if h.lower() in ('participated', 'participation')), None)
        if title_col is None or id_col is None:
            _log(f"[WARN] Master file missing 'Title' or 'ID' column. Found: {headers}")
            return {}
        hyperlink_map = getattr(ws, 'hyperlink_map', {})
        result = {}
        for row_i in range(1, ws.nrows):
            row    = ws.row_values(row_i)
            title  = str(row[title_col] if title_col < len(row) else "").strip()
            doc_id = str(row[id_col]    if id_col    < len(row) else "").strip()
            if not title or not doc_id:
                continue
            hl = hyperlink_map.get((row_i, title_col))
            if hl and getattr(hl, 'url_or_path', ''):
                link = hl.url_or_path.strip()
            else:
                link = f"{ARIBA_BASE_URL}&an={doc_id}"

            def _cell(col):
                if col is not None and col < len(row):
                    v = row[col]
                    # xlrd may return floats for date serial numbers — keep as string
                    return str(v).strip() if not isinstance(v, float) or v != int(v) else str(int(v))
                return ""

            norm = _normalize_for_match(title)
            result[norm] = {
                "id":          doc_id,
                "link":        link,
                "title":       title,
                "end_time":    _cell(end_time_col),
                "event_type":  _cell(event_col),
                "participated": _cell(part_col),
            }
        _log(f"Master links file loaded: {len(result)} RFPs indexed.")
        return result
    except ImportError:
        _log("[WARN] 'xlrd' not installed. Run: pip install xlrd==1.2.0")
        return {}
    except Exception as exc:
        _log(f"[WARN] Could not read master links file: {exc}")
        return {}


def enrich_with_master_links(rows: list, master_links: dict) -> list:
    for row in rows:
        norm = _normalize_for_match(row["rfp_id"])
        if norm in master_links:
            master = master_links[norm]
            if not row["link"]:
                row["link"] = master["link"]
            # Always pull extra fields from master (fill blanks only)
            row.setdefault("doc_id",      master.get("id", ""))
            row.setdefault("end_time",    master.get("end_time", ""))
            row.setdefault("event_type",  master.get("event_type", ""))
            row.setdefault("participated", master.get("participated", ""))
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Portal session helpers  (copied from download_from_csv.py)
# ─────────────────────────────────────────────────────────────────────────────

async def _wait_ready(page, timeout: int = 30000):
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout)
    except Exception:
        pass
    try:
        await page.wait_for_load_state("domcontentloaded", timeout=5000)
    except Exception:
        pass


async def _page_has_login_form(page) -> bool:
    LOGIN_SELECTORS = [
        'xpath=//*[@id="_boebpb"]/div[1]/input',
        'input[name="UserName"]',
        'input[id="UserName"]',
        '#UserName',
        'input[name="username"]',
        'input[type="password"]',
    ]
    for sel in LOGIN_SELECTORS:
        try:
            if await page.locator(sel).count() > 0:
                return True
        except Exception:
            pass
    return False


async def is_logged_in(page) -> bool:
    try:
        url = page.url.lower()
        if any(kw in url for kw in ("login", "signin", "logon", "sso", "auth?")):
            return False
        if await _page_has_login_form(page):
            return False
        return True
    except Exception:
        return False


async def do_login(page, username: str, password: str) -> bool:
    USER_SELECTORS = [
        'xpath=//*[@id="_boebpb"]/div[1]/input',
        '#UserName',
        'input[name="UserName"]',
        'input[id="UserName"]',
        'input[type="email"]',
        'input[type="text"]',
    ]
    PASS_SELECTORS = [
        '#Password',
        'input[name="Password"]',
        'input[id="Password"]',
        'input[type="password"]',
    ]
    SUBMIT_SELECTORS = [
        'input[type="submit"]',
        'button[type="submit"]',
        '#loginButton',
        'button:has-text("Log in")',
        'button:has-text("Sign in")',
    ]

    for attempt in range(1, LOGIN_MAX_RETRIES + 1):
        try:
            _log(f"Login attempt {attempt}/{LOGIN_MAX_RETRIES} …")
            await page.goto(PORTAL_URL, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(3)
            await _wait_ready(page, 30000)

            if await is_logged_in(page):
                _log("Already authenticated.")
                return True

            user_loc = None
            for sel in USER_SELECTORS:
                loc = page.locator(sel)
                if await loc.count() > 0:
                    user_loc = loc.first
                    break
            if user_loc is None:
                await asyncio.sleep(5)
                continue

            await user_loc.click()
            await user_loc.fill(username)

            pass_loc = None
            for sel in PASS_SELECTORS:
                loc = page.locator(sel)
                if await loc.count() > 0:
                    pass_loc = loc.first
                    break
            if pass_loc is None:
                await asyncio.sleep(5)
                continue

            await pass_loc.click()
            await pass_loc.fill(password)

            submitted = False
            for sel in SUBMIT_SELECTORS:
                loc = page.locator(sel)
                if await loc.count() > 0:
                    try:
                        async with page.expect_navigation(wait_until="networkidle", timeout=60000):
                            await loc.first.click()
                    except Exception:
                        await loc.first.click()
                        await _wait_ready(page, 60000)
                    submitted = True
                    break

            if not submitted:
                await asyncio.sleep(5)
                continue

            await asyncio.sleep(3)
            if await is_logged_in(page):
                _log("Logged in successfully.")
                return True

            await asyncio.sleep(5)

        except Exception as exc:
            _log(f"  Login attempt {attempt} error: {exc}")
            if attempt < LOGIN_MAX_RETRIES:
                await asyncio.sleep(5)

    _log("All login attempts failed.")
    return False


async def ensure_logged_in(page, username: str, password: str) -> bool:
    if await is_logged_in(page):
        return True
    _log("Session expired — re-logging in …")
    return await do_login(page, username, password)


# ─────────────────────────────────────────────────────────────────────────────
# Owner + publish date scraper  (based on download_rfp.py logic)
# ─────────────────────────────────────────────────────────────────────────────

async def _click_if_visible(page, selector: str, timeout: int = 5000) -> bool:
    try:
        loc = page.locator(selector)
        if await loc.count() > 0:
            await loc.first.wait_for(state="visible", timeout=timeout)
            await loc.first.click()
            return True
    except Exception:
        pass
    return False


async def extract_owner_and_publish(page) -> dict:
    """
    Scrape all available details from the RFP detail page.
    Returns {'owner', 'publish_time', 'end_date', 'event_type', 'status', 'description'}.
    """
    owner_name   = None
    publish_time = None
    end_date     = None
    event_type   = None
    status       = None
    description  = None

    # ── Selector list for table cells on the details page ──────────────────
    SELECTORS = [
        'div.wideLabels table td',
        'table.wideLabels td',
        'div.wideLabels td',
        'table td',
        '.w-tbl-cell',
        'div[class*="label"] table td',
    ]

    all_cells  = None
    cell_count = 0

    for selector in SELECTORS:
        try:
            loc   = page.locator(selector)
            count = await loc.count()
            if count > 0:
                all_cells  = loc
                cell_count = count
                break
        except Exception:
            continue

    if not all_cells or cell_count == 0:
        # Fallback: search raw body text for patterns
        try:
            body_text = await page.locator('body').inner_text()
            lines = body_text.split('\n')
            for i, line in enumerate(lines):
                line_lower = line.strip().lower()

                if not owner_name and any(kw in line_lower for kw in ['owner', 'owned by', 'created by']):
                    for j in range(max(0, i - 2), min(len(lines), i + 3)):
                        candidate = lines[j].strip()
                        if (len(candidate) > 3
                                and re.match(r'^[A-Za-z\s\.\',\-]+$', candidate)
                                and not any(kw in candidate.lower() for kw in [
                                    'owner', 'publish', 'time', 'date', 'currency',
                                    'commodity', 'event', 'type', 'ariba', 'rfp'])):
                            owner_name = candidate
                            break

                if not publish_time and any(kw in line_lower for kw in ['publish', 'published', 'posted']):
                    if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}.*\d{1,2}:\d{2}', line):
                        publish_time = line.strip()

                if not end_date and any(kw in line_lower for kw in ['end', 'close', 'deadline', 'due']):
                    if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}.*\d{1,2}:\d{2}', line):
                        end_date = line.strip()

        except Exception:
            pass
        return {
            'owner': owner_name, 'publish_time': publish_time,
            'end_date': end_date, 'event_type': event_type,
            'status': status, 'description': description,
        }

    DATE_PATTERNS = [
        r'\d{1,2}/\d{1,2}/\d{4}.*\d{1,2}:\d{2}.*[AP]M',
        r'\d{1,2}-\d{1,2}-\d{4}.*\d{1,2}:\d{2}',
        r'\d{4}-\d{2}-\d{2}.*\d{2}:\d{2}',
        r'\d{1,2}/\d{1,2}/\d{4}',
    ]

    for i in range(cell_count):
        try:
            cell_text  = (await all_cells.nth(i).inner_text()).strip()
            cell_lower = cell_text.lower()

            # ── Owner ─────────────────────────────────────────────────────
            if not owner_name and any(kw in cell_lower for kw in ['owner', 'owner:', 'owned by', 'created by']):
                for j in range(max(0, i - 2), min(cell_count, i + 3)):
                    try:
                        candidate = (await all_cells.nth(j).inner_text()).strip()
                        if not candidate or len(candidate) <= 3:
                            continue
                        if (re.match(r'^[A-Za-z\s\.\',\-]+$', candidate)
                                and not any(kw in candidate.lower() for kw in [
                                    'owner', 'publish', 'time', 'date', 'currency',
                                    'commodity', 'event', 'type', 'loading',
                                    'ariba', 'supplier', 'portal', 'rfp'])):
                            owner_name = candidate
                            break
                    except Exception:
                        continue

            # ── Publish / Created date ────────────────────────────────────
            if not publish_time and any(kw in cell_lower for kw in [
                    'publish', 'published', 'created', 'posted', 'open date', 'start date']):
                for j in range(max(0, i - 2), min(cell_count, i + 3)):
                    try:
                        candidate = (await all_cells.nth(j).inner_text()).strip()
                        for pattern in DATE_PATTERNS:
                            if re.search(pattern, candidate, re.IGNORECASE):
                                publish_time = candidate
                                break
                        if publish_time:
                            break
                    except Exception:
                        continue

            # ── End / Close / Deadline date ───────────────────────────────
            if not end_date and any(kw in cell_lower for kw in [
                    'end date', 'end time', 'close date', 'deadline', 'due date',
                    'submission', 'close time', 'closing']):
                for j in range(max(0, i - 2), min(cell_count, i + 3)):
                    try:
                        candidate = (await all_cells.nth(j).inner_text()).strip()
                        for pattern in DATE_PATTERNS:
                            if re.search(pattern, candidate, re.IGNORECASE):
                                end_date = candidate
                                break
                        if end_date:
                            break
                    except Exception:
                        continue

            # ── Event Type ────────────────────────────────────────────────
            if not event_type and any(kw in cell_lower for kw in [
                    'event type', 'type:', 'sourcing type', 'process type']):
                for j in range(i + 1, min(cell_count, i + 3)):
                    try:
                        candidate = (await all_cells.nth(j).inner_text()).strip()
                        if candidate and len(candidate) > 1 and 'type' not in candidate.lower():
                            event_type = candidate
                            break
                    except Exception:
                        continue

            # ── Status ────────────────────────────────────────────────────
            if not status and any(kw in cell_lower for kw in ['status:', 'state:']):
                for j in range(i + 1, min(cell_count, i + 3)):
                    try:
                        candidate = (await all_cells.nth(j).inner_text()).strip()
                        if candidate and len(candidate) > 1:
                            status = candidate
                            break
                    except Exception:
                        continue

            # ── Description (first non-empty long text block) ─────────────
            if not description and any(kw in cell_lower for kw in ['description', 'scope', 'overview']):
                for j in range(i + 1, min(cell_count, i + 5)):
                    try:
                        candidate = (await all_cells.nth(j).inner_text()).strip()
                        if candidate and len(candidate) > 10:
                            description = candidate[:500]  # cap at 500 chars
                            break
                    except Exception:
                        continue

            if owner_name and publish_time and end_date and event_type:
                break

        except Exception:
            continue

    return {
        'owner':       owner_name,
        'publish_time': publish_time,
        'end_date':    end_date,
        'event_type':  event_type,
        'status':      status,
        'description': description,
    }


async def get_rfp_info(page, rfp_id: str, link: str, username: str, password: str) -> dict:
    """
    Open the RFP portal page in a new tab, click the details tab,
    and return all available fields.
    Does NOT download anything.
    """
    result = {
        'rfp_id':      rfp_id,
        'owner':       None,
        'publish_time': None,
        'end_date':    None,
        'event_type':  None,
        'status':      None,
        'description': None,
        'error':       None,
    }

    if not link:
        result['error'] = "No link available"
        return result

    new_page = await page.context.new_page()
    try:
        _log(f"  Opening: {link[:80]}…")
        await new_page.goto(link, wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(2)

        # Re-login if session dropped
        url_lower = new_page.url.lower()
        if (any(kw in url_lower for kw in ("login", "signin", "logon", "sso", "auth?"))
                or await _page_has_login_form(new_page)):
            _log("  Session expired mid-request — re-logging in …")
            await new_page.close()
            if not await do_login(page, username, password):
                result['error'] = "Session expired, re-login failed"
                return result
            new_page = await page.context.new_page()
            await new_page.goto(link, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(2)

        # Click the "Review Event details" tab to reveal the details table
        clicked = await _click_if_visible(new_page, "#_c8_tuc", timeout=8000)
        if clicked:
            _log("  Clicked details tab (#_c8_tuc)")
            await asyncio.sleep(5)
        else:
            _log("  Details tab not found — trying to read current page state")
            await asyncio.sleep(3)

        info = await extract_owner_and_publish(new_page)
        result['owner']        = info.get('owner')
        result['publish_time'] = info.get('publish_time')
        result['end_date']     = info.get('end_date')
        result['event_type']   = info.get('event_type')
        result['status']       = info.get('status')
        result['description']  = info.get('description')

        _log(f"  End Date     : {result['end_date']   or '(not found)'}")
        _log(f"  Event Type   : {result['event_type'] or '(not found)'}")
        _log(f"  Status       : {result['status']     or '(not found)'}")

    except Exception as exc:
        result['error'] = str(exc)
        _log(f"  Error scraping {rfp_id}: {exc}")
    finally:
        try:
            await new_page.close()
        except Exception:
            pass

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Output helpers
# ─────────────────────────────────────────────────────────────────────────────

def _print_table(results: list):
    """Print a neat aligned table to stdout."""
    COL_W   = [30, 20, 25, 30, 30, 20, 25]
    headers = ["RFP_ID", "Company", "Owner", "Publish_Date", "End_Date", "Event_Type", "Note"]

    def row_str(vals):
        return "  ".join(str(v or "").ljust(w) for v, w in zip(vals, COL_W))

    sep = "  ".join("-" * w for w in COL_W)
    print()
    print(row_str(headers))
    print(sep)
    for r in results:
        note = r.get('error') or ("" if r.get('owner') or r.get('publish_time') else "Not found")
        print(row_str([
            r['rfp_id'],
            r.get('company_name', ''),
            r.get('owner') or "",
            r.get('publish_time') or "",
            r.get('end_date') or r.get('end_time') or "",
            r.get('event_type') or "",
            note,
        ]))
    print()


# ─────────────────────────────────────────────────────────────────────────────
# Main runner
# ─────────────────────────────────────────────────────────────────────────────

async def run(input_file: str, username: str, password: str, headless: bool):
    from playwright.async_api import async_playwright

    # Load rows
    rows = load_input_file(input_file)
    if not rows:
        _die("No valid rows in input file.")

    # Enrich with master links file
    _log(f"Loading master links file: {MASTER_LINKS_FILE}")
    master_links = load_master_rfp_links(MASTER_LINKS_FILE)
    if master_links:
        rows = enrich_with_master_links(rows, master_links)
    else:
        _log("[WARN] Master links file not loaded — links from input file only.")

    # ── Output CSV: named after the input file so resume always uses same file ─
    input_basename = os.path.splitext(os.path.basename(input_file))[0]
    csv_path   = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              f"RFP_Info_{input_basename}.csv")
    fieldnames = [
        "RFP_ID", "Company_Name", "Link",
        "Owner", "Publish_Date", "End_Date",
        "Event_Type", "Status", "Participated",
        "Description", "Note",
    ]

    # ── Resume: load already-processed RFP IDs ────────────────────────────────
    done_ids = set()
    if os.path.isfile(csv_path):
        with open(csv_path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                rid = row.get("RFP_ID", "").strip()
                if rid:
                    done_ids.add(rid)
        _log(f"Resume mode — {len(done_ids)} RFP(s) already done, will skip them.")
        csv_fh = open(csv_path, "a", newline="", encoding="utf-8-sig")  # append
        csv_writer = csv.DictWriter(csv_fh, fieldnames=fieldnames)
    else:
        csv_fh = open(csv_path, "w", newline="", encoding="utf-8-sig")  # fresh
        csv_writer = csv.DictWriter(csv_fh, fieldnames=fieldnames)
        csv_writer.writeheader()
        csv_fh.flush()

    _log(f"Live output file: {csv_path}")

    def _append_row(r: dict):
        note = r.get('error') or ("" if r.get('owner') or r.get('publish_time') else "Not found")
        # End_Date: prefer scraped value, fall back to master-links value
        end_date = r.get('end_date') or r.get('end_time') or ""
        csv_writer.writerow({
            "RFP_ID":       r['rfp_id'],
            "Company_Name": r.get('company_name', ''),
            "Link":         r.get('link', ''),
            "Owner":        r.get('owner') or "",
            "Publish_Date": r.get('publish_time') or "",
            "End_Date":     end_date,
            "Event_Type":   r.get('event_type') or "",
            "Status":       r.get('status') or "",
            "Participated": r.get('participated') or "",
            "Description":  r.get('description') or "",
            "Note":         note,
        })
        csv_fh.flush()  # write to disk immediately

    no_link = [r for r in rows if not r["link"]]
    to_process = [r for r in rows if r["link"]]

    # Filter out already-done RFPs
    pending = [r for r in to_process if _clean_id(r['rfp_id']) not in done_ids]
    # Also skip no-link rows already recorded
    no_link_pending = [r for r in no_link if _clean_id(r['rfp_id']) not in done_ids]

    print()
    print("=" * 65)
    print("  RFP Owner & Publish Date Fetcher")
    print("=" * 65)
    print(f"  Input file  : {input_file}")
    print(f"  Total RFPs  : {len(rows)}")
    print(f"  Already done: {len(done_ids)}")
    print(f"  Remaining   : {len(pending)}")
    print(f"  No link     : {len(no_link_pending)} pending")
    print(f"  Headless    : {headless}")
    print("=" * 65)
    print()

    if no_link_pending:
        _log(f"[WARN] {len(no_link_pending)} RFP(s) have no link and will be skipped:")
        for r in no_link_pending:
            _log(f"       - {r['rfp_id']} ({r['company_name']})")
        print()

    if not pending and not no_link_pending:
        csv_fh.close()
        _log("All RFPs already processed. Nothing to do.")
        _log(f"Results file: {csv_path}")
        return

    results = []

    if pending:
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(
                headless=headless,
                args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"],
            )
            context = await browser.new_context(viewport={"width": 1280, "height": 1024})
            page    = await context.new_page()

            _log("Logging in to portal …")
            if not await do_login(page, username, password):
                await browser.close()
                csv_fh.close()
                _die("Login failed. Check your credentials.")

            for idx, row in enumerate(pending, start=1):
                rfp_id  = _clean_id(row['rfp_id'])
                company = row['company_name']
                link    = row['link']

                print()
                _log(f"[{idx}/{len(pending)}]  {rfp_id}  |  {company}")

                # Refresh session if needed
                if not await ensure_logged_in(page, username, password):
                    _log("  Cannot log in — skipping.")
                    entry = {
                        'rfp_id':      rfp_id,
                        'company_name': company,
                        'link':        link,
                        'owner':       None,
                        'publish_time': None,
                        'end_date':    None,
                        'end_time':    row.get('end_time', ''),
                        'event_type':  row.get('event_type', ''),
                        'status':      None,
                        'participated': row.get('participated', ''),
                        'description': None,
                        'error':       "Login failed",
                    }
                    results.append(entry)
                    _append_row(entry)
                    continue

                info = await get_rfp_info(page, rfp_id, link, username, password)
                info['company_name'] = company
                info['link']         = link
                # Merge master-links extras (scraped value wins; master is fallback)
                info['end_time']    = row.get('end_time', '')
                if not info.get('event_type'):
                    info['event_type'] = row.get('event_type', '')
                info['participated'] = row.get('participated', '')
                results.append(info)
                _append_row(info)  # saved to disk immediately

                _log(f"  Owner        : {info.get('owner') or '(not found)'}")
                _log(f"  Publish Date : {info.get('publish_time') or '(not found)'}")

            await browser.close()

    # Write any pending no-link rows (carry master-links data even without a URL)
    for r in no_link_pending:
        entry = {
            'rfp_id':      _clean_id(r['rfp_id']),
            'company_name': r['company_name'],
            'link':        '',
            'owner':       None,
            'publish_time': None,
            'end_date':    None,
            'end_time':    r.get('end_time', ''),
            'event_type':  r.get('event_type', ''),
            'status':      None,
            'participated': r.get('participated', ''),
            'description': None,
            'error':       "No portal link",
        }
        results.append(entry)
        _append_row(entry)

    csv_fh.close()

    if results:
        _print_table(results)
    print(f"  Saved to: {csv_path}")
    print()


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Fetch RFP owner name and publish date from the Ariba portal (no downloads, no logs)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Required input columns:   RFP_ID,  Company_Name
Optional column:          Link   (direct portal URL)

When Link is present  → navigates directly to that URL.
When Link is blank    → looks up the master All-RFPs.xls file for a URL.
If still no link      → RFP is skipped (reported as 'No portal link').

Examples:
  python get_rfp_info.py --file my_rfps.xlsx
  python get_rfp_info.py --file my_rfps.csv  --headless
  python get_rfp_info.py --file my_rfps.xlsx --username me@co.com --password secret
        """,
    )
    parser.add_argument("--file",     required=True, help="Path to input CSV or Excel file")
    parser.add_argument("--username", default=None,  help="Portal username (or BAHRA_SAP_USERNAME env var)")
    parser.add_argument("--password", default=None,  help="Portal password (or BAHRA_SAP_PASSWORD env var)")
    parser.add_argument("--headless", action="store_true", default=False, help="Run browser in headless mode")

    args = parser.parse_args()

    # Priority: CLI flag → env var → hardcoded default
    username = (args.username
                or os.getenv("BAHRA_SAP_USERNAME", "").strip()
                or DEFAULT_USERNAME)
    password = (args.password
                or os.getenv("BAHRA_SAP_PASSWORD", "").strip()
                or DEFAULT_PASSWORD)

    if not username or not password:
        _die(
            "Portal credentials required.\n"
            "  Use --username / --password flags, or set:\n"
            "  BAHRA_SAP_USERNAME and BAHRA_SAP_PASSWORD environment variables."
        )

    asyncio.run(run(args.file, username, password, args.headless))


if __name__ == "__main__":
    main()

"""
Master RFP Portal Sync
======================
End-to-end script that:
  1. Reads the RFP list from Dataverse (no portal listing download)
  2. Opens each RFP's portal page via the Link column already in DB
  3. Scrapes Owner / Publish_Date / End_Date / Event_Type  (no file downloads,
     no Download-button clicks)
  4. Diffs scraped values against DB and classifies each field as
        OK / MISSING / WRONG
  5. Writes a preview CSV. Phase A does NOT write to Dataverse.
  6. With --apply, reads the preview CSV and updates only the rows the
     user approved (default approve = YES; type NO in the user_approve
     column to veto a row).

Two-phase workflow
------------------
  Phase A (preview, no DB writes):
    python Support-Files\\master_rfp_sync.py
    python Support-Files\\master_rfp_sync.py --company "Saudi Energy" --limit 5

  Edit the preview CSV in Excel. Set user_approve = NO on any row you do
  NOT want applied. Blank / YES / empty = approved.

  Phase B (apply approved updates):
    python Support-Files\\master_rfp_sync.py --apply
    python Support-Files\\master_rfp_sync.py --apply --preview path\\to\\preview.csv

Date format
-----------
All date fields are normalized to 'M/D/YYYY H:MM AM/PM' in Asia/Riyadh
via sync_rfp_publish_date.normalize_date(). Both sides of every diff and
every Dataverse write go through this function. Never invents a new
format; never changes what is already stored.

Out of scope
------------
  * 'participated' column (RFP submission flow owns it)
  * Status / Description (scraped for reference only; no DB column)
  * Audit-log writes to Dataverse
  * File downloads from the portal
"""

import argparse
import asyncio
import csv
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# ─── Path bootstrap so helpers/ and config/ resolve when run from Support-Files ───
HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))

from helpers.dataverse_helper import DataverseClient
from config.config import (
    TENANT_ID, CLIENT_ID, CLIENT_SECRET, RESOURCE_URL,
    RFP_ACTIVITY_LOG_TABLE_API, RFP_ACTIVITY_LOG_TABLE_LOGICAL,
    COMPANY_OPTIONS,
)

# Reuse portal scraping + login from get_rfp_info.py
from get_rfp_info import (                       # noqa: E402
    do_login,
    ensure_logged_in,
    get_rfp_info as scrape_rfp_page,
    load_master_rfp_links,
    _normalize_for_match,
    DEFAULT_USERNAME,
    DEFAULT_PASSWORD,
)

# Reuse the locked DB-standard date formatter
from sync_rfp_publish_date import normalize_date  # noqa: E402


def load_links_xlsx(path: str) -> dict:
    """openpyxl-based loader for .xlsx master listings.
    Returns the same shape as get_rfp_info.load_master_rfp_links():
        {normalized_title: {'id', 'link', 'title', 'end_time', 'event_type', 'participated'}}
    The xlrd-based loader in get_rfp_info.py only supports .xls. This wrapper
    handles .xlsx without touching the shared function."""
    try:
        import openpyxl
    except ImportError:
        print("[WARN] 'openpyxl' not installed. Run: pip install openpyxl")
        return {}

    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        headers_lower = [h.lower() for h in headers]

        def _find_col(*candidates):
            for cand in candidates:
                if cand.lower() in headers_lower:
                    return headers_lower.index(cand.lower())
            return None

        title_col = _find_col("title")
        id_col = _find_col("id")
        end_col = _find_col("end time", "end_time", "endtime", "end date", "close date")
        event_col = _find_col("event type", "event_type", "eventtype", "type")
        part_col = _find_col("participated", "participation")

        if title_col is None or id_col is None:
            print(f"[WARN] xlsx missing 'Title' or 'ID' column. Found: {headers}")
            wb.close()
            return {}

        # Build the same ARIBA_BASE_URL fallback get_rfp_info.py uses
        try:
            from get_rfp_info import ARIBA_BASE_URL
        except ImportError:
            ARIBA_BASE_URL = ""

        result = {}
        for row in ws.iter_rows(min_row=2):
            if title_col >= len(row) or id_col >= len(row):
                continue
            title_cell = row[title_col]
            id_cell = row[id_col]
            title = str(title_cell.value or "").strip()
            doc_id = str(id_cell.value or "").strip()
            if not title or not doc_id:
                continue

            link = ""
            if title_cell.hyperlink and title_cell.hyperlink.target:
                link = title_cell.hyperlink.target.strip()
            if not link and ARIBA_BASE_URL:
                link = f"{ARIBA_BASE_URL}&an={doc_id}"

            def _cell_val(col):
                if col is None or col >= len(row):
                    return ""
                v = row[col].value
                return "" if v is None else str(v).strip()

            norm = _normalize_for_match(title)
            result[norm] = {
                "id":           doc_id,
                "link":         link,
                "title":        title,
                "end_time":     _cell_val(end_col),
                "event_type":   _cell_val(event_col),
                "participated": _cell_val(part_col),
            }
        wb.close()
        return result
    except Exception as exc:
        print(f"[WARN] Could not read xlsx links file: {exc}")
        return {}


# ─── Constants ───────────────────────────────────────────────────────────────

PROGRESS_FILE = HERE / ".master_rfp_sync_progress.json"
OUTPUT_DIR = HERE / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

PRIMARY_KEY_LOGICAL = f"{RFP_ACTIVITY_LOG_TABLE_LOGICAL}id"  # cr673_bahra_rfps_v2id

# DB column (display name)  ->  key returned by scrape_rfp_page()
FIELD_MAP = {
    "owner_name":    "owner",
    "publish_time":  "publish_time",
    "RFP_End_Date":  "end_date",
    "rfp_type":      "event_type",
}
DATE_FIELDS = {"publish_time", "RFP_End_Date"}

# Fields that may appear in the preview CSV and that Phase B is allowed to write.
# 'Link' is not scraped from the portal — it can only come from a --links-file
# enrichment, but it's still applied to Dataverse like any other approved row.
WRITABLE_FIELDS = set(FIELD_MAP.keys()) | {"Link"}

PREVIEW_HEADERS = [
    "timestamp", "rfp_id", "company_name", "record_id", "link",
    "field", "db_value", "portal_value", "status", "user_approve", "notes",
]

APPLIED_HEADERS = [
    "timestamp", "rfp_id", "record_id", "field",
    "old_value", "new_value", "diff_status", "result",
]

ERROR_RETRY_CAP = 2          # max scrape attempts per RFP across runs
RELOGIN_EVERY_N_RFPS = 25    # preemptive ensure_logged_in cadence
SLEEP_BETWEEN_RFPS = 1.0     # seconds — be polite to the portal


# ─── Progress file ───────────────────────────────────────────────────────────

def load_progress() -> dict:
    if not PROGRESS_FILE.exists():
        return {"started_at": None, "completed": set(), "errored": {}}
    try:
        with open(PROGRESS_FILE, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return {
            "started_at": data.get("started_at"),
            "completed":  set(data.get("completed", [])),
            "errored":    data.get("errored", {}),
        }
    except Exception:
        return {"started_at": None, "completed": set(), "errored": {}}


def save_progress(progress: dict) -> None:
    payload = {
        "started_at": progress.get("started_at"),
        "completed":  sorted(progress.get("completed", set())),
        "errored":    progress.get("errored", {}),
    }
    tmp = PROGRESS_FILE.with_suffix(".json.tmp")
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)
    os.replace(tmp, PROGRESS_FILE)


def reset_progress() -> None:
    if PROGRESS_FILE.exists():
        PROGRESS_FILE.unlink()
        print(f"  Progress file removed: {PROGRESS_FILE}")


# ─── Normalisation & diff classification ─────────────────────────────────────

def clean_id(rfp_id: str) -> str:
    return re.sub(r"\s+", " ", str(rfp_id or "")).strip()


def _is_blank(val) -> bool:
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s.lower() in ("nan", "none", "-")


def _norm_for_compare(val, field: str) -> str:
    """Canonical compared form: normalize_date() for date fields,
    case-insensitive collapsed-whitespace for everything else."""
    if _is_blank(val):
        return ""
    s = str(val).strip()
    if field in DATE_FIELDS:
        return normalize_date(s)
    return re.sub(r"\s+", " ", s).lower()


def classify(db_val, portal_val, field: str):
    """Return (status, db_display, portal_display, note).
    Rules locked in plan §Diff classification rules."""
    db_blank = _is_blank(db_val)
    portal_blank = _is_blank(portal_val)
    db_disp = "" if db_blank else str(db_val).strip()
    portal_disp = "" if portal_blank else str(portal_val).strip()

    # Both blank
    if db_blank and portal_blank:
        return "OK", db_disp, portal_disp, ""
    # Portal blank, DB has value → never overwrite with blank
    if portal_blank and not db_blank:
        return "OK", db_disp, portal_disp, "portal blank; kept DB value"
    # DB blank, portal has value → MISSING
    if db_blank and not portal_blank:
        return "MISSING", db_disp, portal_disp, ""

    # Both have values — compare normalised
    db_norm = _norm_for_compare(db_val, field)
    portal_norm = _norm_for_compare(portal_val, field)

    if field in DATE_FIELDS:
        # If either side failed to parse, normalize_date returns input unchanged.
        # Heuristic: if a date couldn't be parsed (no slashes / no AM/PM), surface ERROR.
        def _parse_failed(orig, norm):
            return norm == orig.strip() and not re.search(r"\d{1,2}/\d{1,2}/\d{4}", norm)
        if _parse_failed(db_disp, db_norm) or _parse_failed(portal_disp, portal_norm):
            return "ERROR", db_disp, portal_disp, "date parse failure — raw values shown"

    if db_norm == portal_norm:
        return "OK", db_disp, portal_disp, ""
    return "WRONG", db_disp, portal_disp, ""


# ─── Preview CSV writer ──────────────────────────────────────────────────────

def write_preview_row(writer, fh, *, rfp_id, company, record_id, link,
                      field, db, portal, status, note):
    writer.writerow({
        "timestamp":    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rfp_id":       rfp_id,
        "company_name": company,
        "record_id":    record_id,
        "link":         link,
        "field":        field,
        "db_value":     db,
        "portal_value": portal,
        "status":       status,
        "user_approve": "",
        "notes":        note,
    })
    fh.flush()


# ─── Phase A: scan + preview ─────────────────────────────────────────────────

async def phase_a_scan(args) -> None:
    print("[INIT] Connecting to Dataverse...")
    dv = DataverseClient(TENANT_ID, CLIENT_ID, CLIENT_SECRET, RESOURCE_URL)

    # Resolve primary key display name (so we can grab the record id out of rows)
    try:
        colmap = dv.get_column_mapping(RFP_ACTIVITY_LOG_TABLE_LOGICAL)
        logical_to_display = {v: k for k, v in colmap.items()}
        pk_display = logical_to_display.get(PRIMARY_KEY_LOGICAL, PRIMARY_KEY_LOGICAL)
    except Exception as exc:
        print(f"[WARN] Could not load column mapping: {exc}")
        pk_display = PRIMARY_KEY_LOGICAL

    print("[INIT] Fetching RFPs from Dataverse...")
    select_cols = [
        "RFP_ID", "Company_Name", "Link",
        "owner_name", "publish_time", "RFP_End_Date", "rfp_type",
    ]
    rows = dv.get_all_rows(
        table_api_name=RFP_ACTIVITY_LOG_TABLE_API,
        select_columns=select_cols,
        table_logical_name=RFP_ACTIVITY_LOG_TABLE_LOGICAL,
        use_display_names=True,
    )
    print(f"[INIT] Fetched {len(rows)} RFP rows.")

    # Filters
    if args.company:
        rows = [r for r in rows if (r.get("Company_Name") or "").strip() == args.company]
        print(f"[FILTER] --company={args.company!r}: {len(rows)} rows.")

    # Default: only process RFPs where owner_name OR publish_time is blank in DB.
    # Both fields are critical and must be filled. Use --all-rfps to scan everything.
    if not args.all_rfps:
        before = len(rows)
        rows = [
            r for r in rows
            if _is_blank(r.get("owner_name")) or _is_blank(r.get("publish_time"))
        ]
        print(f"[FILTER] missing owner_name OR publish_time: {len(rows)} rows "
              f"(dropped {before - len(rows)} already-complete rows; use --all-rfps to override).")

    if args.limit:
        rows = rows[: args.limit]
        print(f"[FILTER] --limit={args.limit}: {len(rows)} rows.")

    if not rows:
        print("[INIT] No RFPs to process. Exiting.")
        return

    # Optional: enrich missing DB Links from a master listing file.
    # Index format: {normalized_title: {link, id, ...}}
    # .xls -> reuse load_master_rfp_links() (xlrd)
    # .xlsx -> use load_links_xlsx() (openpyxl, defined above)
    file_links_index: dict = {}
    if args.links_file:
        links_path = Path(args.links_file)
        if not links_path.exists():
            print(f"[ERROR] --links-file not found: {args.links_file}")
            sys.exit(1)
        print(f"[INIT] Loading link enrichment file: {args.links_file}")
        ext = links_path.suffix.lower()
        if ext == ".xlsx":
            file_links_index = load_links_xlsx(str(links_path))
        elif ext == ".xls":
            file_links_index = load_master_rfp_links(str(links_path))
        else:
            print(f"[ERROR] Unsupported --links-file extension {ext!r}. Use .xls or .xlsx.")
            sys.exit(1)
        print(f"[INIT] Indexed {len(file_links_index)} entries from links file.")

    # Progress
    if args.reset:
        reset_progress()
    progress = load_progress()
    if not progress["started_at"]:
        progress["started_at"] = datetime.now().isoformat()
    save_progress(progress)

    # Filter completed
    pending = []
    for r in rows:
        rid = clean_id(r.get("RFP_ID", ""))
        if not rid:
            continue
        if rid in progress["completed"]:
            continue
        # Cap retries on errored RFPs
        attempts = progress["errored"].get(rid, {}).get("attempts", 0)
        if attempts >= ERROR_RETRY_CAP:
            continue
        pending.append(r)
    skipped = len(rows) - len(pending)
    print(f"[INIT] {skipped} already done / over retry cap; {len(pending)} pending.")
    if not pending:
        print("[INIT] Nothing pending. Use --reset to start over.")
        return

    # Preview CSV
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    preview_path = OUTPUT_DIR / f"master_rfp_preview_{timestamp}.csv"
    csv_fh = open(preview_path, "w", newline="", encoding="utf-8-sig")
    writer = csv.DictWriter(csv_fh, fieldnames=PREVIEW_HEADERS)
    writer.writeheader()
    csv_fh.flush()
    print(f"[INIT] Preview CSV: {preview_path}")

    counters = {k: 0 for k in ("OK", "MISSING", "WRONG", "ERROR", "SKIPPED_NO_LINK", "INFO")}

    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=args.headless)
        context = await browser.new_context()
        page = await context.new_page()

        print("[AUTH] Logging in...")
        if not await do_login(page, args.username, args.password):
            print("[AUTH] Login failed. Aborting Phase A.")
            csv_fh.close()
            await browser.close()
            return

        for idx, row in enumerate(pending, 1):
            rfp_id = clean_id(row.get("RFP_ID", ""))
            company = (row.get("Company_Name") or "").strip()
            db_link = (row.get("Link") or "").strip()
            record_id = row.get(pk_display) or row.get(PRIMARY_KEY_LOGICAL) or ""

            print(f"\n[{idx}/{len(pending)}] {rfp_id} ({company})")

            # Preemptive session check
            if idx % RELOGIN_EVERY_N_RFPS == 0:
                await ensure_logged_in(page, args.username, args.password)

            # Resolve effective link: DB Link first; fall back to --links-file.
            # If file fills an empty DB Link, queue a Link 'MISSING' row so the
            # value is also written back to Dataverse on --apply.
            link = db_link
            file_link = ""
            if not db_link and file_links_index:
                norm = _normalize_for_match(rfp_id)
                entry = file_links_index.get(norm, {})
                file_link = (entry.get("link") or "").strip()
                if file_link:
                    link = file_link
                    print(f"  [LINK] DB empty -> using --links-file link")
                    write_preview_row(
                        writer, csv_fh,
                        rfp_id=rfp_id, company=company, record_id=record_id, link=link,
                        field="Link", db="", portal=file_link,
                        status="MISSING", note="enriched from --links-file",
                    )
                    counters["MISSING"] += 1

            if not link:
                write_preview_row(
                    writer, csv_fh,
                    rfp_id=rfp_id, company=company, record_id=record_id, link="",
                    field="-", db="", portal="",
                    status="SKIPPED_NO_LINK",
                    note="no Link in Dataverse and no match in --links-file" if args.links_file
                         else "no Link in Dataverse",
                )
                counters["SKIPPED_NO_LINK"] += 1
                progress["completed"].add(rfp_id)
                save_progress(progress)
                continue

            # Scrape — scrape_rfp_page() opens a new tab, handles mid-request
            # relogin, and closes the tab.
            try:
                info = await scrape_rfp_page(page, rfp_id, link, args.username, args.password)
            except Exception as exc:
                print(f"  [ERROR] scrape exception: {exc}")
                write_preview_row(
                    writer, csv_fh,
                    rfp_id=rfp_id, company=company, record_id=record_id, link=link,
                    field="-", db="", portal="",
                    status="ERROR", note=f"scrape exception: {exc}",
                )
                counters["ERROR"] += 1
                prev = progress["errored"].get(rfp_id, {})
                progress["errored"][rfp_id] = {
                    "attempts": prev.get("attempts", 0) + 1,
                    "last_error": str(exc),
                }
                save_progress(progress)
                continue

            if info.get("error"):
                err = info["error"]
                write_preview_row(
                    writer, csv_fh,
                    rfp_id=rfp_id, company=company, record_id=record_id, link=link,
                    field="-", db="", portal="",
                    status="ERROR", note=err,
                )
                counters["ERROR"] += 1
                prev = progress["errored"].get(rfp_id, {})
                progress["errored"][rfp_id] = {
                    "attempts": prev.get("attempts", 0) + 1,
                    "last_error": err,
                }
                save_progress(progress)
                continue

            # Echo the scraped values so the user can see what came back
            # from the portal for the two key fields. get_rfp_info() only
            # logs End Date / Event Type / Status, so Owner + Publish Date
            # would otherwise be silent.
            print(f"  Owner        : {info.get('owner')        or '(not found)'}")
            print(f"  Publish Date : {info.get('publish_time') or '(not found)'}")

            # Diff each mapped field
            for db_field, scraped_key in FIELD_MAP.items():
                db_val = row.get(db_field)
                portal_val = info.get(scraped_key)
                status, db_disp, portal_disp, note = classify(db_val, portal_val, db_field)
                write_preview_row(
                    writer, csv_fh,
                    rfp_id=rfp_id, company=company, record_id=record_id, link=link,
                    field=db_field, db=db_disp, portal=portal_disp,
                    status=status, note=note,
                )
                counters[status] = counters.get(status, 0) + 1

            # Reference-only Status + Description (no DB column)
            extras = []
            if info.get("status"):
                extras.append(f"Status: {info['status']}")
            if info.get("description"):
                desc = str(info["description"])[:200]
                extras.append(f"Description: {desc}")
            if extras:
                write_preview_row(
                    writer, csv_fh,
                    rfp_id=rfp_id, company=company, record_id=record_id, link=link,
                    field="_info", db="", portal=" | ".join(extras),
                    status="INFO", note="reference only; not synced",
                )
                counters["INFO"] += 1

            progress["completed"].add(rfp_id)
            progress["errored"].pop(rfp_id, None)
            save_progress(progress)

            await asyncio.sleep(SLEEP_BETWEEN_RFPS)

        await browser.close()

    csv_fh.close()

    print(f"\n{'=' * 65}")
    print("  PHASE A COMPLETE")
    print(f"{'=' * 65}")
    for key in ("OK", "MISSING", "WRONG", "ERROR", "SKIPPED_NO_LINK", "INFO"):
        print(f"  {key:<18}: {counters[key]}")
    print(f"\n  Preview CSV : {preview_path}")
    print(f"  Progress    : {PROGRESS_FILE}")
    print("\n  Open the preview CSV in Excel. Set user_approve=NO on any rows")
    print("  you do NOT want applied. Blank / YES = approved.")
    print(f"\n  Then run:  python {Path(__file__).name} --apply --preview \"{preview_path}\"")
    print(f"{'=' * 65}")


# ─── Phase B: apply ──────────────────────────────────────────────────────────

def phase_b_apply(args) -> None:
    # Resolve preview CSV path
    if args.preview:
        preview_path = Path(args.preview)
    else:
        candidates = sorted(OUTPUT_DIR.glob("master_rfp_preview_*.csv"))
        if not candidates:
            print("[ERROR] No preview CSV found in Support-Files\\output\\. Run Phase A first.")
            sys.exit(1)
        preview_path = candidates[-1]
        print(f"[INIT] Using newest preview: {preview_path.name}")

    if not preview_path.exists():
        print(f"[ERROR] Preview not found: {preview_path}")
        sys.exit(1)

    only_status = (args.only_status or "ALL").upper()
    if only_status == "ALL":
        eligible = {"MISSING", "WRONG"}
    else:
        eligible = {only_status}

    to_apply = []
    vetoed = 0
    with open(preview_path, "r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            status = (row.get("status") or "").strip().upper()
            if status not in eligible:
                continue
            approve = (row.get("user_approve") or "").strip().upper()
            if approve == "NO":
                vetoed += 1
                continue
            field = (row.get("field") or "").strip()
            if field in ("", "-", "_info"):
                continue
            if field not in WRITABLE_FIELDS:
                # Unknown / out-of-scope column — never write
                continue
            to_apply.append({
                "rfp_id":       (row.get("rfp_id") or "").strip(),
                "record_id":    (row.get("record_id") or "").strip(),
                "field":        field,
                "portal_value": row.get("portal_value", ""),
                "db_value":     row.get("db_value", ""),
                "status":       status,
            })

    print(f"[INIT] {len(to_apply)} field-updates to apply | {vetoed} vetoed (user_approve=NO).")
    if not to_apply:
        print("[DONE] Nothing approved to apply. Exiting.")
        return

    # Group by record_id so we write all approved fields for an RFP in one PATCH
    grouped: dict = {}
    for r in to_apply:
        if not r["record_id"]:
            continue
        grouped.setdefault(r["record_id"], []).append(r)

    print("[INIT] Connecting to Dataverse...")
    dv = DataverseClient(TENANT_ID, CLIENT_ID, CLIENT_SECRET, RESOURCE_URL)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    applied_path = OUTPUT_DIR / f"master_rfp_applied_{timestamp}.csv"
    applied_fh = open(applied_path, "w", newline="", encoding="utf-8-sig")
    applied_writer = csv.DictWriter(applied_fh, fieldnames=APPLIED_HEADERS)
    applied_writer.writeheader()

    applied_count = 0
    failed_count = 0

    for record_id, fields in grouped.items():
        rfp_id_display = fields[0]["rfp_id"]
        # Build payload — re-normalize dates so the canonical form is what lands in DB
        payload = {}
        for f in fields:
            val = f["portal_value"]
            if f["field"] in DATE_FIELDS:
                val = normalize_date(val)
            payload[f["field"]] = val

        try:
            ok = dv.update_row(
                table_api_name=RFP_ACTIVITY_LOG_TABLE_API,
                record_id=record_id,
                data=payload,
                table_logical_name=RFP_ACTIVITY_LOG_TABLE_LOGICAL,
                use_display_names=True,
            )
            result = "OK" if ok else "FAIL"
            if ok:
                applied_count += len(fields)
                print(f"  [OK] {rfp_id_display}: {len(fields)} field(s)")
            else:
                failed_count += len(fields)
                print(f"  [FAIL] {rfp_id_display}")
        except Exception as exc:
            result = f"ERROR: {exc}"
            failed_count += len(fields)
            print(f"  [ERROR] {rfp_id_display}: {exc}")

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for f in fields:
            applied_writer.writerow({
                "timestamp":   now,
                "rfp_id":      rfp_id_display,
                "record_id":   record_id,
                "field":       f["field"],
                "old_value":   f["db_value"],
                "new_value":   payload[f["field"]],
                "diff_status": f["status"],
                "result":      result,
            })
        applied_fh.flush()

    applied_fh.close()

    print(f"\n{'=' * 65}")
    print("  PHASE B COMPLETE")
    print(f"{'=' * 65}")
    print(f"  Field updates applied : {applied_count}")
    print(f"  Failed                : {failed_count}")
    print(f"  Vetoed (user_approve=NO): {vetoed}")
    print(f"  Applied log: {applied_path}")
    print(f"{'=' * 65}")


# ─── CLI ─────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Master RFP portal sync — scrape Owner / Publish_Date / End_Date / "
            "Event_Type, diff against Dataverse, preview, then apply with approval."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--apply", action="store_true",
        help="Phase B: apply approved updates from a preview CSV to Dataverse. "
             "Without this flag the script runs Phase A (preview only — no DB writes).",
    )
    parser.add_argument(
        "--company", default=None,
        help=f"Filter to one company. Must be one of: {COMPANY_OPTIONS}",
    )
    parser.add_argument(
        "--username", default=None,
        help="Portal username. Defaults to BAHRA_SAP_USERNAME env var or the hardcoded value.",
    )
    parser.add_argument(
        "--password", default=None,
        help="Portal password. Defaults to BAHRA_SAP_PASSWORD env var or the hardcoded value.",
    )
    parser.add_argument(
        "--headless", action="store_true", default=False,
        help="Run the browser headless (default: visible).",
    )
    parser.add_argument(
        "--limit", type=int, default=None,
        help="Process at most N RFPs (handy for smoke testing).",
    )
    parser.add_argument(
        "--all-rfps", action="store_true", default=False,
        help="Scan every RFP in Dataverse. Default behavior is to scan ONLY "
             "RFPs where owner_name OR publish_time is blank in the DB.",
    )
    parser.add_argument(
        "--links-file", default=None,
        help="Path to All-RFPs.xls master listing. When an RFP in Dataverse has "
             "an empty Link column, the script looks the RFP up in this file by "
             "title (matched normalized) and uses the file's hyperlink. The Link "
             "value is also queued as a MISSING diff so it gets written back to "
             "Dataverse on --apply.",
    )
    parser.add_argument(
        "--reset", action="store_true",
        help="Delete the Phase A progress file before starting.",
    )
    parser.add_argument(
        "--preview", default=None,
        help="Phase B only: path to a specific preview CSV. "
             "Default: newest master_rfp_preview_*.csv in Support-Files\\output\\.",
    )
    parser.add_argument(
        "--only-status", default="ALL", choices=["MISSING", "WRONG", "ALL"],
        help="Phase B: restrict which diff statuses to apply (default: ALL = MISSING + WRONG).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    # Credentials: CLI > env > defaults
    if not args.username:
        args.username = os.environ.get("BAHRA_SAP_USERNAME") or DEFAULT_USERNAME
    if not args.password:
        args.password = os.environ.get("BAHRA_SAP_PASSWORD") or DEFAULT_PASSWORD

    if args.company and args.company not in COMPANY_OPTIONS:
        print(f"[ERROR] --company {args.company!r} is not in COMPANY_OPTIONS.")
        print(f"        Valid values: {COMPANY_OPTIONS}")
        sys.exit(1)

    if args.apply:
        phase_b_apply(args)
    else:
        asyncio.run(phase_a_scan(args))


if __name__ == "__main__":
    main()

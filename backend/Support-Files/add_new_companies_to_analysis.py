"""
Append RFP-Material_List and RFP-List rows for two companies that are missing
from the analysis xlsx:
  - HADEED - RAJHI STEEL
  - Saudi Aramco Mobil Refinery Company Limited

Scope is a hard-coded list of RFP folders (taken from the user's screenshots).
For each RFP:
  1. Locate the source Excel under
       C:/python/RFP-automation/ALLRFPs/<Company>/<RFP>/downloaded-rfp/
     If missing, download it from SharePoint (RFP-logs/ALLRFPs/...).
  2. Run the same matching as Support-Files/prefill_matched_data.py
     (master materials + active keywords loaded from Dataverse) to produce
     exact_matches / keyword_matches / not_matched items.
  3. Read the source row's `* Quantity` to fill Quantity inline.
  4. Optionally enrich End_Time + Participant from cr673_bahra_rfps_v2 in
     Dataverse (only if a record exists for the same RFP_ID).
  5. Append rows to the RFP-Material_List sheet and a summary row to
     RFP-List, keeping the exact column layout used elsewhere in the xlsx.

Usage (from project root):
  python Support-Files/add_new_companies_to_analysis.py
  python Support-Files/add_new_companies_to_analysis.py --analysis "<path>"
  python Support-Files/add_new_companies_to_analysis.py --no-download
"""

import argparse
import os
import re
import sys
from datetime import datetime

import math

import openpyxl
import pandas as pd

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import core.common_imports  # noqa: F401
from helpers.core_helper import (
    _find_other_content_sheet_name,
    find_column_name,
    extract_keywords_from_text,
)
from config.config import (
    CLIENT_ID, CLIENT_SECRET, TENANT_ID,
    SHAREPOINT_HOSTNAME, SITE_PATH, DRIVE_NAME, SP_BASE_FOLDER,
    RESOURCE_URL,
)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DEFAULT_ANALYSIS = r"C:\Users\Manish.Soni\Downloads\RFP-Analysis_refreshed_20260515_135031_with_quantity.xlsx"
ALLRFPS_DIR = os.path.join(PROJECT_ROOT, "ALLRFPs")
TEMP_DIR = os.path.join(PROJECT_ROOT, "TempRFP-Files")

# RFP folder names visible in the user's two screenshots
RFPS_BY_COMPANY = {
    "HADEED - RAJHI STEEL": [
        "RFP 6600009370-conduits",
        "RFP 6600009370-cable,light",
        "RFP 6600009370-cable",
        "RFP 6600008947-cables (RM1-26)",
        "9960000588-30 2",
        "9960000588-30",
        "9960000476-40",
        "9910001299 333036 SIEMENS CABLES",
        "9900017678 10 295741",
        "9900008669 365664 SIEMENS CABLE",
        "9900008667 352723 SIEMENS ELECTRICAL CABLE",
        "9900008666 347221 SIEMENS POWER CABLE",
        "9900004882 1539653 Fiber Optic Cable",
        "9900003750 390835",
        "9900003635 - 295341",
        "9900000334 730882 HV Shielded CABLE",
        "6600000035",
        "402623 343154 Siemens Cables",
        "390245",
        "346015 339998 330222 322479 319146 333215 Cables",
        "509994 8 PAIR STRANDED CONDUCTORS CABLE",
    ],
    "Saudi Aramco Mobil Refinery Company Limited": [
        "6000222509",
        "6000221652",
        "6000220272",
        "6000220118",
        "6000219917-11024276",
        "6000219871",
        "6000218524",
    ],
}


# ---------------------------------------------------------------------------
# Source file lookup
# ---------------------------------------------------------------------------
def locate_excel(company: str, rfp_title: str):
    """Search ALLRFPs/<Co>/<Rfp>/downloaded-rfp/ then TempRFP-Files. Return (path, basename)."""
    dl = os.path.join(ALLRFPS_DIR, company, rfp_title, "downloaded-rfp")
    if os.path.isdir(dl):
        for f in os.listdir(dl):
            if f.lower().endswith((".xls", ".xlsx")) and not f.startswith("~$"):
                return os.path.join(dl, f), f
    if os.path.isdir(TEMP_DIR):
        for ext in (".xls", ".xlsx"):
            cand = rfp_title + ext
            p = os.path.join(TEMP_DIR, cand)
            if os.path.exists(p):
                return p, cand
    return None, None


def _sp_client():
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from download_from_sharepoint import SharePointDownloader
    c = SharePointDownloader(
        CLIENT_ID, CLIENT_SECRET, TENANT_ID,
        SHAREPOINT_HOSTNAME, SITE_PATH, DRIVE_NAME,
    )
    c.auth()
    c.resolve_site_and_drive()
    return c


def download_excel(client, company: str, rfp_title: str):
    """Download the RFP's Excel from SharePoint into ALLRFPs/<Co>/<Rfp>/downloaded-rfp/."""
    import requests
    folder = f"{SP_BASE_FOLDER}/ALLRFPs/{company}/{rfp_title}/downloaded-rfp"
    url = (
        f"https://graph.microsoft.com/v1.0/sites/{client.site_id}"
        f"/drives/{client.drive_id}/root:/{folder}:/children"
    )
    client.ensure_token()
    r = requests.get(url, headers=client.headers)
    if r.status_code != 200:
        print(f"      [SP] folder not accessible ({r.status_code}): {folder}")
        return None, None
    items = r.json().get("value", [])
    chosen = None
    for it in items:
        n = it.get("name", "")
        if "file" in it and n.lower().endswith((".xls", ".xlsx")) and not n.startswith("~$"):
            chosen = it
            break
    if chosen is None:
        print(f"      [SP] no Excel found under {folder}")
        return None, None
    sp_path = f"{folder}/{chosen['name']}"
    dl_url = (
        f"https://graph.microsoft.com/v1.0/sites/{client.site_id}"
        f"/drives/{client.drive_id}/root:/{sp_path}:/content"
    )
    client.ensure_token()
    fr = requests.get(dl_url, headers=client.headers, stream=True)
    if fr.status_code != 200:
        print(f"      [SP] download failed ({fr.status_code}): {sp_path}")
        return None, None
    dest_dir = os.path.join(ALLRFPS_DIR, company, rfp_title, "downloaded-rfp")
    os.makedirs(dest_dir, exist_ok=True)
    local = os.path.join(dest_dir, chosen["name"])
    with open(local, "wb") as f:
        for chunk in fr.iter_content(chunk_size=8192):
            f.write(chunk)
    print(f"      [SP OK] {chosen['name']} ({os.path.getsize(local)/1024:.1f} KB)")
    return local, chosen["name"]


# ---------------------------------------------------------------------------
# Smart sheet picker: chooses the sheet that actually contains material rows.
# Saudi Aramco Mobil RFPs ship an empty "Other Content" plus a populated
# "1 Commercial Line Item" sheet — the default picker grabs the empty one.
# ---------------------------------------------------------------------------
EXPECTED_COLS = ("intend to respond", "currency", "material number",
                 "price", "quantity", "name", "description")


def _score_sheet_df(df):
    """Return (col_hits, valid_rows) for a candidate sheet DataFrame."""
    if df.empty:
        return 0, 0
    cols_lc = [str(c).lower().strip() for c in df.columns]
    col_hits = sum(1 for ec in EXPECTED_COLS if any(ec in c for c in cols_lc))
    if col_hits < 2:
        return col_hits, 0
    name_col = find_column_name(df.columns, "name")
    if not name_col:
        return col_hits, 0
    valid = 0
    for v in df[name_col]:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip().lower()
        if not s:
            continue
        if any(h in s for h in (
            "click on the +", "question, item, or lot name",
            "header text", "required action", "submit the",
        )):
            continue
        valid += 1
    return col_hits, valid


def pick_best_material_sheet(excel_path: str, xl=None, cache=None):
    """
    Pick the sheet most likely to contain real material rows.

    Fast path: try the default "Other Content" sheet first. If it has at
    least 1 valid material row AND >=2 expected columns, accept it
    immediately (covers ~all Saudi Energy/Aramco RFPs).

    Slow path: only when the fast path produces 0 valid rows, fall back to
    scanning every sheet (handles Saudi Aramco Mobil's "1 Commercial Line
    Item" layout).

    `xl` may be passed in by a caller that already holds the ExcelFile
    (avoids re-opening). `cache` is a dict that, if provided, will be
    populated with {sheet_name: dataframe} for the sheets we read, so the
    caller can reuse them without another disk hit.
    """
    own_xl = False
    try:
        if xl is None:
            xl = pd.ExcelFile(excel_path)
            own_xl = True
    except Exception:
        return None

    def _read(sn):
        if cache is not None and sn in cache:
            return cache[sn]
        try:
            df = xl.parse(sn)
        except Exception:
            return None
        if cache is not None:
            cache[sn] = df
        return df

    try:
        # Fast path: try "Other Content" by case/space-insensitive name match
        oc_name = None
        for sn in xl.sheet_names:
            if sn.replace(" ", "").lower() == "othercontent":
                oc_name = sn
                break
        if oc_name is not None:
            df_oc = _read(oc_name)
            if df_oc is not None:
                col_hits, valid = _score_sheet_df(df_oc)
                if col_hits >= 2 and valid >= 1:
                    return oc_name

        # Slow path: scan all sheets and pick the best
        best = oc_name
        best_score = -1
        for sn in xl.sheet_names:
            df = _read(sn)
            if df is None:
                continue
            col_hits, valid = _score_sheet_df(df)
            if col_hits < 2:
                continue
            score = valid * 100 + col_hits
            if score > best_score:
                best_score = score
                best = sn
        return best
    finally:
        if own_xl:
            try:
                xl.close()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Inline matching (mirrors Support-Files/prefill_matched_data.py
# match_materials_for_rfp but uses pick_best_material_sheet so it picks up
# alternate-layout RFPs like Saudi Aramco Mobil's "1 Commercial Line Item").
# ---------------------------------------------------------------------------
def _try_keyword_match(name_text, description_text, keywords_list):
    name_kw = extract_keywords_from_text(name_text)
    desc_kw = extract_keywords_from_text(description_text)
    all_kw = set(name_kw + desc_kw)
    for csv_kw in keywords_list:
        for mat_kw in all_kw:
            if csv_kw and mat_kw and (csv_kw in mat_kw or mat_kw in csv_kw):
                return csv_kw
    return None


def _find_description_by_keyword(name_text, description_text, master,
                                   master_col, desc_col_master, keywords_list):
    name_kw = extract_keywords_from_text(name_text)
    desc_kw = extract_keywords_from_text(description_text)
    all_kw = set(name_kw + desc_kw)
    kw_rows = pd.DataFrame()
    for mat_kw in all_kw:
        if not mat_kw:
            continue
        temp = master[master[master_col].astype(str).str.contains(mat_kw, case=False, na=False, regex=False)]
        if not temp.empty:
            kw_rows = pd.concat([kw_rows, temp]).drop_duplicates()
        for col in master.columns:
            if col != master_col and master[col].dtype == "object":
                try:
                    temp = master[master[col].astype(str).str.contains(mat_kw, case=False, na=False, regex=False)]
                    if not temp.empty:
                        kw_rows = pd.concat([kw_rows, temp]).drop_duplicates()
                except Exception:
                    pass
    if not kw_rows.empty and desc_col_master:
        val = kw_rows.iloc[0].get(desc_col_master, "")
        return "" if (isinstance(val, float) and math.isnan(val)) else str(val)
    return ""


def match_materials_smart(excel_path, rfp_id, company, master, master_col,
                          master_code_set, keywords_list):
    """Smart matcher. Same logic as prefill's match_materials_for_rfp but
    chooses the best material sheet via pick_best_material_sheet.

    Opens the workbook exactly once (shared between picker and matcher) and
    short-circuits to 'Other Content' as a fast path.
    """
    try:
        xl = pd.ExcelFile(excel_path)
    except Exception as e:
        print(f"      [WARN] cannot open {os.path.basename(excel_path)}: {e}")
        return None
    cache = {}
    try:
        sheet = pick_best_material_sheet(excel_path, xl=xl, cache=cache)
        if not sheet:
            return None
        df = cache.get(sheet)
        if df is None:
            try:
                df = xl.parse(sheet)
            except Exception as e:
                print(f"      [WARN] cannot read sheet '{sheet}': {e}")
                return None
    finally:
        try:
            xl.close()
        except Exception:
            pass

    col_name = find_column_name(df.columns, "name")
    if not col_name:
        return None
    col_desc = find_column_name(df.columns, "description")
    desc_col_master = (
        find_column_name(master.columns, "description")
        or find_column_name(master.columns, "material description")
    )

    exact_matches, keyword_matches, not_matched = [], [], []

    for idx, value in df[col_name].items():
        if pd.isna(value):
            continue
        name_text = str(value)
        description_text = (
            str(df.iloc[idx][col_desc])
            if col_desc and not pd.isna(df.iloc[idx][col_desc])
            else ""
        )
        material_codes = re.findall(r"\d{9}", name_text)

        if material_codes:
            for mat in material_codes:
                base_item = {
                    "material_code": mat,
                    "excel_name": name_text,
                    "excel_description": description_text,
                    "row_number": idx + 2,
                    "column_name": col_name,
                }
                matched_rows = master[master[master_col].astype(str) == mat]
                if not matched_rows.empty:
                    mat_desc = ""
                    if desc_col_master:
                        v = matched_rows.iloc[0].get(desc_col_master, "")
                        mat_desc = "" if (isinstance(v, float) and math.isnan(v)) else str(v)
                    base_item["material_description"] = mat_desc
                    exact_matches.append(base_item)
                    continue
                mk = _try_keyword_match(name_text, description_text, keywords_list)
                if mk:
                    base_item["matched_keyword"] = mk
                    base_item["material_description"] = _find_description_by_keyword(
                        name_text, description_text, master, master_col,
                        desc_col_master, keywords_list,
                    )
                    keyword_matches.append(base_item)
                else:
                    not_matched.append(base_item)
        else:
            if not keywords_list:
                continue
            mk = _try_keyword_match(name_text, description_text, keywords_list)
            if mk:
                base_item = {
                    "material_code": "",
                    "excel_name": name_text,
                    "excel_description": description_text,
                    "row_number": idx + 2,
                    "column_name": col_name,
                    "matched_keyword": mk,
                    "material_description": _find_description_by_keyword(
                        name_text, description_text, master, master_col,
                        desc_col_master, keywords_list,
                    ),
                }
                keyword_matches.append(base_item)

    total = len(exact_matches) + len(keyword_matches) + len(not_matched)
    matched_total = len(exact_matches) + len(keyword_matches)
    match_pct = round((matched_total / total * 100) if total > 0 else 0, 1)
    return {
        "rfp_id": rfp_id,
        "source_file": os.path.basename(excel_path),
        "rfp_end_date": "",
        "total_items": total,
        "summary": {
            "exact_match_count": len(exact_matches),
            "keyword_match_count": len(keyword_matches),
            "not_matched_count": len(not_matched),
            "match_percentage": match_pct,
        },
        "exact_matches": exact_matches,
        "keyword_matches": keyword_matches,
        "not_matched": not_matched,
        "_sheet_used": sheet,
        "_df": df,
    }


def build_qty_from_df(df):
    """Like build_row_number_to_quantity but reuses an already-loaded df."""
    if df is None:
        return {}
    qty_col = find_column_name(df.columns, "quantity")
    if not qty_col:
        return {}
    out = {}
    for idx, val in df[qty_col].items():
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        out[idx + 2] = val
    return out


# ---------------------------------------------------------------------------
# Quantity per source row_number
# ---------------------------------------------------------------------------
def build_row_number_to_quantity(excel_path: str, sheet_hint: str = None):
    """Read the chosen material sheet and return {row_number (idx+2): quantity}."""
    sheet = sheet_hint or pick_best_material_sheet(excel_path)
    if not sheet:
        return {}
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet)
    except Exception:
        return {}
    qty_col = find_column_name(df.columns, "quantity")
    if not qty_col:
        return {}
    out = {}
    for idx, val in df[qty_col].items():
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        out[idx + 2] = val
    return out


# ---------------------------------------------------------------------------
# Build analysis rows from a matched_data dict (mirrors derive_material_list_rows)
# ---------------------------------------------------------------------------
def _normalize_participated(val):
    if not val:
        return "Not Participated"
    v = str(val).strip().lower()
    if v in ("yes", "submitted"):
        return "Participated"
    if v == "declined":
        return "Declined"
    return "Not Participated"


def _format_end_time(val):
    if not val:
        return ""
    s = str(val).strip()
    if "." in s:
        date_part, _, frac = s.partition(".")
        frac_digits = "".join(c for c in frac if c.isdigit())[:6]
        s_for_parse = f"{date_part}.{frac_digits}" if frac_digits else date_part
    else:
        s_for_parse = s
    fmts = ["%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %I:%M %p", "%m/%d/%Y %H:%M"]
    dt = None
    for fmt in fmts:
        try:
            dt = datetime.strptime(s_for_parse, fmt)
            break
        except ValueError:
            continue
    if dt is None:
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        except Exception:
            return s
    return dt.strftime("%m/%d/%Y %I:%M %p")


def build_analysis_rows(matched, company, rfp_id, end_time, participant_full,
                        excel_file, qty_by_row):
    """
    Returns (material_list_rows, rfp_list_row) using the same column layout
    as the existing analysis xlsx (see derive_material_list_rows in
    Support-Files/generate_rfp_analysis_report.py).
    """
    participant_yn = "Yes" if participant_full == "Participated" else "No"

    def base():
        return {
            "Company_Name": company,
            "RFP_Title": rfp_id,
            "RFP_ID": rfp_id,
            "End_Time": end_time or "",
            "Excel_File": excel_file,
            "Participant": participant_yn,
        }

    rows = []
    for item in matched.get("exact_matches") or []:
        r = base()
        r["Material_Code"] = item.get("material_code") or ""
        r["Material_Description"] = (
            item.get("material_description")
            or item.get("excel_description")
            or item.get("excel_name")
            or ""
        )
        r["Material_Matched"] = "Yes"
        r["Matched_Keywords"] = ""
        r["Keyword_Matched"] = "No"
        r["Quantity"] = qty_by_row.get(item.get("row_number"), "")
        rows.append(r)
    for item in matched.get("keyword_matches") or []:
        r = base()
        r["Material_Code"] = item.get("material_code") or ""
        r["Material_Description"] = (
            item.get("material_description")
            or item.get("excel_description")
            or item.get("excel_name")
            or ""
        )
        r["Material_Matched"] = "No"
        r["Matched_Keywords"] = item.get("matched_keyword") or ""
        r["Keyword_Matched"] = "Yes"
        r["Quantity"] = qty_by_row.get(item.get("row_number"), "")
        rows.append(r)
    for item in matched.get("not_matched") or []:
        r = base()
        r["Material_Code"] = item.get("material_code") or ""
        r["Material_Description"] = (
            item.get("material_description")
            or item.get("excel_description")
            or item.get("excel_name")
            or ""
        )
        r["Material_Matched"] = "No"
        r["Matched_Keywords"] = ""
        r["Keyword_Matched"] = "No"
        r["Quantity"] = qty_by_row.get(item.get("row_number"), "")
        rows.append(r)

    summary = (matched or {}).get("summary") or {}
    exact_count = int(summary.get("exact_match_count") or 0)
    keyword_count = int(summary.get("keyword_match_count") or 0)
    rfp_list_row = {
        "Company_Name": company,
        "RFP_Title": rfp_id,
        "RFP_ID": rfp_id,
        "End_Time": end_time or "",
        "Participant": participant_full,
        "is_material_match": "Material matched" if exact_count > 0 else "Material not matched",
        "is_keyword_match": "Keyword matched" if keyword_count > 0 else "Keyword not matched",
        "no_of_matched_materials": exact_count,
        "no_of_matched_keywords": keyword_count,
    }
    return rows, rfp_list_row


# ---------------------------------------------------------------------------
# Dataverse lookup for End_Time / Participant
# ---------------------------------------------------------------------------
def fetch_rfp_metadata(dv_client, rfp_id):
    """Returns (end_time_iso, participated_raw) or (None, None) on miss."""
    try:
        rfp_id_safe = rfp_id.replace("'", "''")
        rows = dv_client.get_all_rows(
            table_api_name="cr673_bahra_rfps_v2s",
            select_columns=["cr673_rfp_id", "cr673_rfp_end_date", "cr673_participated"],
            table_logical_name="cr673_bahra_rfps_v2",
            use_display_names=True,
            filter_expr=f"cr673_rfp_id eq '{rfp_id_safe}'",
        )
        if rows:
            r = rows[0]
            return r.get("cr673_rfp_end_date"), r.get("cr673_participated")
    except Exception as e:
        print(f"    [DV WARN] {e}")
    return None, None


# ---------------------------------------------------------------------------
# Append rows to xlsx
# ---------------------------------------------------------------------------
def _purge_company_rows(ws, company_set):
    """Remove rows whose Company_Name is in company_set (idempotent re-runs)."""
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        co_col = header.index("Company_Name") + 1
    except ValueError:
        return 0
    to_delete = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=co_col).value
        if val in company_set:
            to_delete.append(r)
    # Delete from bottom up to keep indices valid
    for r in reversed(to_delete):
        ws.delete_rows(r, 1)
    return len(to_delete)


def append_to_analysis(analysis_path, out_path, ml_rows_new, rfp_list_rows_new,
                       company_set):
    wb = openpyxl.load_workbook(analysis_path)

    # Idempotent cleanup
    ws = wb["RFP-Material_List"]
    purged_ml = _purge_company_rows(ws, company_set)
    ws2 = wb["RFP-List"]
    purged_rl = _purge_company_rows(ws2, company_set)
    if purged_ml or purged_rl:
        print(f"  Removed {purged_ml} stale RFP-Material_List rows and "
              f"{purged_rl} stale RFP-List rows for the target companies before appending.")

    # RFP-Material_List
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    start_row = ws.max_row + 1
    for offset, r in enumerate(ml_rows_new):
        for c_idx, h in enumerate(header, start=1):
            ws.cell(row=start_row + offset, column=c_idx, value=r.get(h, ""))

    # RFP-List
    header2 = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    start_row2 = ws2.max_row + 1
    for offset, r in enumerate(rfp_list_rows_new):
        for c_idx, h in enumerate(header2, start=1):
            ws2.cell(row=start_row2 + offset, column=c_idx, value=r.get(h, ""))

    try:
        wb.save(out_path)
    except PermissionError:
        base, ext = os.path.splitext(out_path)
        out_path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
        print(f"  WARNING: target locked; writing to {out_path}")
        wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--analysis", default=DEFAULT_ANALYSIS)
    ap.add_argument("--out", default=None,
                    help="Output xlsx path (default: overwrite the input)")
    ap.add_argument("--no-download", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.analysis):
        print(f"ERROR: Analysis file not found: {args.analysis}", file=sys.stderr)
        sys.exit(1)

    out_path = args.out or args.analysis

    print(f"Analysis  : {args.analysis}")
    print(f"Output    : {out_path}")
    print(f"Download  : {'OFF' if args.no_download else 'ON'}")
    print()

    # Load master + keywords (same path as prefill)
    print("Loading master + keywords from Dataverse...")
    from services.master_data_service import (
        get_all_materials_for_matching, get_all_keywords_for_matching,
    )
    materials = get_all_materials_for_matching() or []
    keywords_list = get_all_keywords_for_matching() or []
    if not materials or not keywords_list:
        print(f"  WARN: materials={len(materials)} keywords={len(keywords_list)}")
    master = pd.DataFrame({"material": materials})
    master_col = "material"
    master_code_set = set(master[master_col].astype(str))
    print(f"  Loaded {len(materials)} materials, {len(keywords_list)} keywords\n")

    # Dataverse client for end-time/participant lookup
    from helpers.dataverse_helper import DataverseClient
    dv_client = DataverseClient(
        tenant_id=TENANT_ID, client_id=CLIENT_ID,
        client_secret=CLIENT_SECRET, resource_url=RESOURCE_URL,
    )

    sp_client = None
    ml_rows_new = []
    rfp_list_rows_new = []
    files_missing = []
    files_read_error = []
    skipped_no_materials = []

    total = sum(len(v) for v in RFPS_BY_COMPANY.values())
    i = 0
    for company, rfps in RFPS_BY_COMPANY.items():
        for rfp_title in rfps:
            i += 1
            print(f"[{i}/{total}] {company} / {rfp_title}")

            local_path, excel_file = locate_excel(company, rfp_title)
            if not local_path and not args.no_download:
                if sp_client is None:
                    try:
                        print("    Auth to SharePoint...")
                        sp_client = _sp_client()
                    except Exception as e:
                        print(f"    [FATAL] SP auth failed: {e}")
                        args.no_download = True
                        sp_client = None
                if sp_client is not None:
                    local_path, excel_file = download_excel(sp_client, company, rfp_title)

            if not local_path:
                print(f"    [MISS] file not found")
                files_missing.append((company, rfp_title))
                continue

            try:
                matched = match_materials_smart(
                    excel_path=local_path,
                    rfp_id=rfp_title,
                    company=company,
                    master=master,
                    master_col=master_col,
                    master_code_set=master_code_set,
                    keywords_list=keywords_list,
                )
            except Exception as e:
                print(f"    [READ ERROR] {e}")
                files_read_error.append((company, rfp_title, str(e)))
                continue

            if not matched:
                print(f"    [SKIP] no usable material sheet found")
                skipped_no_materials.append((company, rfp_title))
                continue

            sheet_used = matched.get("_sheet_used")
            qty_by_row = build_row_number_to_quantity(local_path, sheet_used)

            end_iso, part_raw = fetch_rfp_metadata(dv_client, rfp_title)
            end_time = _format_end_time(end_iso)
            participant_full = _normalize_participated(part_raw)

            rows, rfp_row = build_analysis_rows(
                matched, company, rfp_title, end_time, participant_full,
                excel_file, qty_by_row,
            )
            ml_rows_new.extend(rows)
            rfp_list_rows_new.append(rfp_row)

            summ = matched.get("summary") or {}
            print(f"    items={matched.get('total_items')}  "
                  f"exact={summ.get('exact_match_count')}  "
                  f"kw={summ.get('keyword_match_count')}  "
                  f"not_matched={summ.get('not_matched_count')}  "
                  f"qty_filled={sum(1 for r in rows if r['Quantity'] not in (None, ''))}")

    print(f"\nAppending {len(ml_rows_new)} material rows and "
          f"{len(rfp_list_rows_new)} RFP-List rows -> {out_path}")
    out_path = append_to_analysis(
        args.analysis, out_path, ml_rows_new, rfp_list_rows_new,
        company_set=set(RFPS_BY_COMPANY.keys()),
    )

    print("\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"  RFPs processed                : {len(rfp_list_rows_new)}")
    print(f"  Material rows appended        : {len(ml_rows_new)}")
    print(f"  Files missing (not on SP)     : {len(files_missing)}")
    print(f"  Files w/ read errors          : {len(files_read_error)}")
    print(f"  Files w/ no Other Content     : {len(skipped_no_materials)}")
    print(f"\n  Output : {out_path}")

    if files_missing:
        print("\nMissing files:")
        for co, ti in files_missing:
            print(f"  {co} / {ti}")
    if files_read_error:
        print("\nRead errors:")
        for co, ti, err in files_read_error:
            print(f"  {co} / {ti}  -> {err}")
    if skipped_no_materials:
        print("\nSkipped (no Other Content sheet):")
        for co, ti in skipped_no_materials:
            print(f"  {co} / {ti}")

    print("\nDone.")


if __name__ == "__main__":
    main()

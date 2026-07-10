"""
Post-processing pass on the analysis xlsx:

1. **Quantity stored as text-with-commas** -> convert to a real number.
   "10,000" -> 10000 (int), "1,500.5" -> 1500.5 (float). Cells already
   numeric or genuinely blank are left alone.

2. **RFP_Title whitespace normalization** -> collapse runs of whitespace
   to a single space, then trim. Folds near-duplicates like
   "SEC  RFP-C001423635" and "SEC RFP-C001423635" into one canonical title
   in both RFP-Material_List and RFP-List.

This script does NOT change Material_Description, Material_Matched, Quantity
values that were already numbers, or any other column. It is safe to run
multiple times.

Usage (from project root):
  python Support-Files/normalize_analysis_xlsx.py
  python Support-Files/normalize_analysis_xlsx.py --analysis "<path>"
"""

import argparse
import os
import re
import sys
from datetime import datetime

import openpyxl

DEFAULT_ANALYSIS = r"C:\Users\Manish.Soni\Downloads\RFP-Analysis_refreshed_20260515_135031_with_quantity.xlsx"

# Matches "10,000", "1,500.5", "1234.5", "1234" — anything that, after
# stripping commas, is a valid number.
_QTY_NUMERIC = re.compile(r"^[\d,]+(\.\d+)?$")
_WS_RUN = re.compile(r"\s+")


def parse_quantity(v):
    """Return (new_value, was_changed)."""
    if v is None:
        return v, False
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return v, False
    s = str(v).strip()
    if not s:
        return v, False
    if not _QTY_NUMERIC.match(s):
        return v, False
    cleaned = s.replace(",", "")
    try:
        if "." in cleaned:
            return float(cleaned), True
        return int(cleaned), True
    except ValueError:
        return v, False


def collapse_ws(v):
    """Return (new_value, was_changed)."""
    if v is None:
        return v, False
    if not isinstance(v, str):
        return v, False
    new = _WS_RUN.sub(" ", v).strip()
    return (new, new != v) if new != v else (v, False)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--analysis", default=DEFAULT_ANALYSIS)
    ap.add_argument("--out", default=None,
                    help="Output path (default: overwrite the input)")
    args = ap.parse_args()

    if not os.path.exists(args.analysis):
        print(f"ERROR: Analysis file not found: {args.analysis}", file=sys.stderr)
        sys.exit(1)
    out_path = args.out or args.analysis

    print(f"Analysis: {args.analysis}")
    print(f"Output  : {out_path}\n")

    print("Loading workbook (writable)...")
    wb = openpyxl.load_workbook(args.analysis)

    qty_fixed = 0
    title_fixed_ml = 0
    title_fixed_rl = 0

    # --- RFP-Material_List ---
    ws = wb["RFP-Material_List"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        qty_c = header.index("Quantity") + 1
        title_c = header.index("RFP_Title") + 1
        rfpid_c = header.index("RFP_ID") + 1
    except ValueError as e:
        print(f"ERROR: required column missing in RFP-Material_List: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Walking RFP-Material_List ({ws.max_row - 1} data rows)...")
    for r in range(2, ws.max_row + 1):
        # Quantity normalize
        v = ws.cell(row=r, column=qty_c).value
        new_v, changed = parse_quantity(v)
        if changed:
            ws.cell(row=r, column=qty_c, value=new_v)
            qty_fixed += 1

        # RFP_Title + RFP_ID whitespace collapse
        for c in (title_c, rfpid_c):
            v = ws.cell(row=r, column=c).value
            new_v, changed = collapse_ws(v)
            if changed:
                ws.cell(row=r, column=c, value=new_v)
                title_fixed_ml += 1
    print(f"  Quantity values converted to numeric : {qty_fixed}")
    print(f"  Title/ID whitespace fixes            : {title_fixed_ml}")

    # --- RFP-List ---
    ws2 = wb["RFP-List"]
    header2 = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    try:
        title_c2 = header2.index("RFP_Title") + 1
        rfpid_c2 = header2.index("RFP_ID") + 1
    except ValueError as e:
        print(f"ERROR: required column missing in RFP-List: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"\nWalking RFP-List ({ws2.max_row - 1} rows)...")
    for r in range(2, ws2.max_row + 1):
        for c in (title_c2, rfpid_c2):
            v = ws2.cell(row=r, column=c).value
            new_v, changed = collapse_ws(v)
            if changed:
                ws2.cell(row=r, column=c, value=new_v)
                title_fixed_rl += 1
    print(f"  Title/ID whitespace fixes            : {title_fixed_rl}")

    try:
        wb.save(out_path)
    except PermissionError:
        base, ext = os.path.splitext(out_path)
        out_path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
        print(f"  WARNING: target locked; writing to {out_path}")
        wb.save(out_path)

    print("\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"  Quantity cells converted to numeric : {qty_fixed}")
    print(f"  RFP_Title/ID whitespace fixes (ML)  : {title_fixed_ml}")
    print(f"  RFP_Title/ID whitespace fixes (RL)  : {title_fixed_rl}")
    print(f"  Output                              : {out_path}")
    print("\nDone.")


if __name__ == "__main__":
    main()

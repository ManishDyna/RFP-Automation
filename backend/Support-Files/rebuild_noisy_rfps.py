"""
Targeted rebuild for Excel_Files in the analysis xlsx that carry
instruction-text noise rows (Material_Description starts with phrases like
"Item or lot description", "Local vendors MUST bid in SAR only", etc.).

For each *affected* Excel_File, this script:
  1. Locates the source RFP Excel (TempRFP-Files -> ALLRFPs -> SharePoint).
  2. Re-runs the standard match logic via match_materials_smart() against
     the active master + keywords in Dataverse.
  3. Captures Quantity per source row.
  4. Preserves each existing RFP_Title spelling (the same Excel_File can be
     referenced by multiple Dataverse RFPs) and emits one fresh set of
     analysis rows per spelling, carrying over End_Time and Participant from
     the FIRST existing row for that spelling.
  5. Deletes the old analysis rows for that Excel_File and appends the new
     ones to RFP-Material_List.
  6. Replaces the per-spelling row in RFP-List with refreshed match counts.

Files that have NO instruction-text rows are not touched.

Usage (from project root):
  python Support-Files/rebuild_noisy_rfps.py
  python Support-Files/rebuild_noisy_rfps.py --analysis "<path>"
  python Support-Files/rebuild_noisy_rfps.py --no-download
"""

import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl
import pandas as pd

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)
SUPPORT_DIR = os.path.dirname(os.path.abspath(__file__))
if SUPPORT_DIR not in sys.path:
    sys.path.insert(0, SUPPORT_DIR)

import core.common_imports  # noqa: F401

# Reuse helpers from the sibling script
from add_new_companies_to_analysis import (
    locate_excel,
    _sp_client,
    download_excel,
    match_materials_smart,
    build_qty_from_df,
    build_analysis_rows,
    _format_end_time,
    _normalize_participated,
)
from config.config import (
    TENANT_ID, CLIENT_ID, CLIENT_SECRET, RESOURCE_URL,
)

DEFAULT_ANALYSIS = r"C:\Users\Manish.Soni\Downloads\RFP-Analysis_refreshed_20260515_135031_with_quantity.xlsx"

INSTRUCTION_DESC_PATTERNS = (
    "item or lot description",
    "question, item, or lot name",
    "local vendors must bid",
    "click on the +",
    "header text",
    "required action",
    "submit the",
    "for example,",
)


def is_instruction_row(desc, mm, km):
    if not desc:
        return False
    d_lc = str(desc).strip().lower()
    if (mm or "").strip().lower() != "no":
        return False
    if (km or "").strip().lower() != "no":
        return False
    return any(d_lc.startswith(p) for p in INSTRUCTION_DESC_PATTERNS)


# ---------------------------------------------------------------------------
# Snapshot the workbook (one open) and identify affected files + spellings.
# ---------------------------------------------------------------------------
def load_workbook_snapshot(analysis_path: str):
    wb = openpyxl.load_workbook(analysis_path)
    ws_ml = wb["RFP-Material_List"]
    ws_rl = wb["RFP-List"]

    ml_hdr = [ws_ml.cell(row=1, column=c).value for c in range(1, ws_ml.max_column + 1)]
    rl_hdr = [ws_rl.cell(row=1, column=c).value for c in range(1, ws_rl.max_column + 1)]

    ml_idx = {h: i + 1 for i, h in enumerate(ml_hdr)}  # 1-based for openpyxl
    rl_idx = {h: i + 1 for i, h in enumerate(rl_hdr)}

    # First pass: scan RFP-Material_List
    affected = set()
    title_spellings = defaultdict(set)         # excel_file -> {rfp_title, ...}
    first_meta = {}                            # (excel_file, rfp_title) -> dict(company, end_time, participant_yn)
    file_co = {}                               # excel_file -> company (most common)

    for r in range(2, ws_ml.max_row + 1):
        xf = ws_ml.cell(row=r, column=ml_idx["Excel_File"]).value
        if not xf:
            continue
        rfp_title = ws_ml.cell(row=r, column=ml_idx["RFP_Title"]).value
        co = ws_ml.cell(row=r, column=ml_idx["Company_Name"]).value
        et = ws_ml.cell(row=r, column=ml_idx["End_Time"]).value
        pt_yn = ws_ml.cell(row=r, column=ml_idx["Participant"]).value
        desc = ws_ml.cell(row=r, column=ml_idx["Material_Description"]).value
        mm = ws_ml.cell(row=r, column=ml_idx["Material_Matched"]).value
        km = ws_ml.cell(row=r, column=ml_idx["Keyword_Matched"]).value

        if is_instruction_row(desc, mm, km):
            affected.add(xf)

        if rfp_title:
            title_spellings[xf].add(rfp_title)
            key = (xf, rfp_title)
            if key not in first_meta:
                first_meta[key] = {
                    "company": co or "",
                    "end_time": et or "",
                    "participant_yn": pt_yn or "No",
                }
        if xf not in file_co and co:
            file_co[xf] = co

    # RFP-List: capture End_Time + Participant per RFP_Title to use as fallback
    rfp_list_meta = {}      # rfp_title -> {end_time, participant_full, company}
    for r in range(2, ws_rl.max_row + 1):
        rfp_title = ws_rl.cell(row=r, column=rl_idx["RFP_Title"]).value
        if not rfp_title:
            continue
        rfp_list_meta[rfp_title] = {
            "end_time": ws_rl.cell(row=r, column=rl_idx["End_Time"]).value or "",
            "participant_full": ws_rl.cell(row=r, column=rl_idx["Participant"]).value or "Not Participated",
            "company": ws_rl.cell(row=r, column=rl_idx["Company_Name"]).value or "",
        }

    return {
        "wb": wb, "ws_ml": ws_ml, "ws_rl": ws_rl,
        "ml_hdr": ml_hdr, "rl_hdr": rl_hdr,
        "ml_idx": ml_idx, "rl_idx": rl_idx,
        "affected": affected, "title_spellings": title_spellings,
        "first_meta": first_meta, "file_co": file_co,
        "rfp_list_meta": rfp_list_meta,
    }


# ---------------------------------------------------------------------------
# Delete + append helpers (operating on the same open workbook)
# ---------------------------------------------------------------------------
def delete_rows_for_files(ws_ml, ml_idx, files_set):
    xf_col = ml_idx["Excel_File"]
    to_delete = []
    for r in range(2, ws_ml.max_row + 1):
        v = ws_ml.cell(row=r, column=xf_col).value
        if v in files_set:
            to_delete.append(r)
    for r in reversed(to_delete):
        ws_ml.delete_rows(r, 1)
    return len(to_delete)


def append_ml_rows(ws_ml, ml_hdr, ml_idx, rows):
    start = ws_ml.max_row + 1
    for offset, r in enumerate(rows):
        for col_name, col_idx in ml_idx.items():
            ws_ml.cell(row=start + offset, column=col_idx, value=r.get(col_name, ""))
    return len(rows)


def replace_rfp_list_rows(ws_rl, rl_idx, replacements):
    """replacements: { rfp_title: new_row_dict_with_RFP-List-cols }"""
    title_col = rl_idx["RFP_Title"]
    replaced = 0
    matched_titles = set()
    for r in range(2, ws_rl.max_row + 1):
        t = ws_rl.cell(row=r, column=title_col).value
        if t in replacements:
            new = replacements[t]
            for col_name, col_idx in rl_idx.items():
                if col_name in new:
                    ws_rl.cell(row=r, column=col_idx, value=new[col_name])
            matched_titles.add(t)
            replaced += 1
    # Titles we wanted to replace but couldn't find — append them
    missing = set(replacements.keys()) - matched_titles
    if missing:
        start = ws_rl.max_row + 1
        for offset, t in enumerate(sorted(missing)):
            new = replacements[t]
            for col_name, col_idx in rl_idx.items():
                ws_rl.cell(row=start + offset, column=col_idx, value=new.get(col_name, ""))
    return replaced, len(missing)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--analysis", default=DEFAULT_ANALYSIS)
    ap.add_argument("--out", default=None, help="Output path (default: overwrite the input)")
    ap.add_argument("--no-download", action="store_true",
                    help="Skip SharePoint fallback; only use local files")
    args = ap.parse_args()

    if not os.path.exists(args.analysis):
        print(f"ERROR: Analysis file not found: {args.analysis}", file=sys.stderr)
        sys.exit(1)

    out_path = args.out or args.analysis
    print(f"Analysis : {args.analysis}")
    print(f"Output   : {out_path}")
    print(f"Download : {'OFF' if args.no_download else 'ON (SharePoint fallback)'}")
    print()

    print("Snapshotting workbook...")
    snap = load_workbook_snapshot(args.analysis)
    affected = snap["affected"]
    print(f"  Excel_Files in analysis             : {len(snap['title_spellings'])}")
    print(f"  Affected Excel_Files (have noise)   : {len(affected)}")
    print()

    if not affected:
        print("Nothing to rebuild. Exiting.")
        return

    # Load master + keywords once
    print("Loading master + keywords from Dataverse...")
    from services.master_data_service import (
        get_all_materials_for_matching, get_all_keywords_for_matching,
    )
    materials = get_all_materials_for_matching() or []
    keywords_list = get_all_keywords_for_matching() or []
    master = pd.DataFrame({"material": materials})
    master_col = "material"
    master_code_set = set(master[master_col].astype(str))
    print(f"  {len(materials)} materials, {len(keywords_list)} keywords\n")

    sp_client = None
    files_missing = []
    files_read_error = []
    rebuilt_files = 0
    new_ml_rows = []
    new_rl_rows = {}    # rfp_title -> new RFP-List dict

    TEMP_DIR_LOCAL = os.path.join(PROJECT_ROOT, "TempRFP-Files")

    for i, xf in enumerate(sorted(affected), start=1):
        spellings = sorted(snap["title_spellings"][xf])
        company = snap["file_co"].get(xf) or ""
        rfp_folder = spellings[0] if spellings else None
        print(f"[{i}/{len(affected)}] {xf}  spellings={len(spellings)}")

        # 1. Exact basename lookup in flat TempRFP-Files (handles whitespace
        #    mismatches between RFP_Title and Excel_File)
        local_path = None
        basename = None
        cand = os.path.join(TEMP_DIR_LOCAL, xf)
        if os.path.exists(cand):
            local_path, basename = cand, xf

        # 2. Try each spelling via ALLRFPs/<Co>/<RFP>/downloaded-rfp/ + flat
        if not local_path:
            for sp in spellings:
                local_path, basename = locate_excel(company, sp)
                if local_path:
                    break
        if not local_path and not args.no_download:
            if sp_client is None:
                try:
                    print("    Auth to SharePoint...")
                    sp_client = _sp_client()
                except Exception as e:
                    print(f"    [FATAL] SP auth failed: {e}")
                    args.no_download = True
                    sp_client = None
            if sp_client is not None:
                # Try each spelling on SharePoint until one works
                for sp in spellings:
                    local_path, basename = download_excel(sp_client, company, sp)
                    if local_path:
                        break

        if not local_path:
            print(f"    [MISS] no source file found")
            files_missing.append((company, xf))
            continue

        try:
            matched = match_materials_smart(
                excel_path=local_path,
                rfp_id=rfp_folder,
                company=company,
                master=master,
                master_col=master_col,
                master_code_set=master_code_set,
                keywords_list=keywords_list,
            )
        except Exception as e:
            print(f"    [READ ERROR] {e}")
            files_read_error.append((company, xf, str(e)))
            continue

        if not matched:
            print(f"    [SKIP] no usable material sheet")
            files_read_error.append((company, xf, "no usable material sheet"))
            continue

        qty_by_row = build_qty_from_df(matched.get("_df"))

        # Emit one fresh set of rows per spelling
        for spelling in spellings:
            meta_key = (xf, spelling)
            meta = snap["first_meta"].get(meta_key, {})
            end_time = meta.get("end_time") or snap["rfp_list_meta"].get(spelling, {}).get("end_time", "")
            participant_full = snap["rfp_list_meta"].get(spelling, {}).get(
                "participant_full",
                "Participated" if meta.get("participant_yn") == "Yes" else "Not Participated",
            )

            rows, rfp_row = build_analysis_rows(
                matched=matched,
                company=company,
                rfp_id=spelling,
                end_time=end_time or "",
                participant_full=participant_full,
                excel_file=xf,
                qty_by_row=qty_by_row,
            )
            new_ml_rows.extend(rows)
            new_rl_rows[spelling] = rfp_row

        summ = matched.get("summary") or {}
        print(f"    rebuilt: items={matched.get('total_items')}  "
              f"exact={summ.get('exact_match_count')}  "
              f"kw={summ.get('keyword_match_count')}  "
              f"not_matched={summ.get('not_matched_count')}  "
              f"spellings={len(spellings)}")
        rebuilt_files += 1

    # Apply mutations: delete old rows for rebuilt files, then append new ones
    print(f"\nDeleting old RFP-Material_List rows for {rebuilt_files} rebuilt file(s)...")
    rebuilt_set = set(snap["title_spellings"].keys()) & affected - set(
        xf for (_, xf) in files_missing
    ) - set(xf for (_, xf, _) in files_read_error)
    deleted = delete_rows_for_files(snap["ws_ml"], snap["ml_idx"], rebuilt_set)
    print(f"  Deleted {deleted} rows")

    appended = append_ml_rows(snap["ws_ml"], snap["ml_hdr"], snap["ml_idx"], new_ml_rows)
    print(f"  Appended {appended} rows")

    replaced, added = replace_rfp_list_rows(snap["ws_rl"], snap["rl_idx"], new_rl_rows)
    print(f"  RFP-List: replaced {replaced} rows, appended {added} new")

    try:
        snap["wb"].save(out_path)
    except PermissionError:
        base, ext = os.path.splitext(out_path)
        out_path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
        print(f"  WARNING: target locked; writing to {out_path}")
        snap["wb"].save(out_path)

    print("\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    print(f"  Affected files                : {len(affected)}")
    print(f"  Rebuilt                       : {rebuilt_files}")
    print(f"  Files missing (no source)     : {len(files_missing)}")
    print(f"  Files read-error              : {len(files_read_error)}")
    print(f"  RFP-Material_List rows deleted: {deleted}")
    print(f"  RFP-Material_List rows added  : {appended}")
    print(f"  Net change                    : {appended - deleted:+d}")
    print(f"  RFP-List rows replaced        : {replaced}")
    print(f"  Output                        : {out_path}")

    if files_missing:
        print("\nMissing files:")
        for co, xf in files_missing[:20]:
            print(f"  {co} / {xf}")
        if len(files_missing) > 20:
            print(f"  ... and {len(files_missing)-20} more")
    if files_read_error:
        print("\nRead errors:")
        for co, xf, err in files_read_error[:20]:
            print(f"  {co} / {xf}  -> {err}")
        if len(files_read_error) > 20:
            print(f"  ... and {len(files_read_error)-20} more")

    print("\nDone.")


if __name__ == "__main__":
    main()

"""
RFP Analysis Report Generator

Cross-references Portal RFPs with Dataverse database, local file system, and SharePoint.
Generates an Excel report with:
  - RFP Company Name
  - RFP Name / ID
  - End Date
  - Publish Date
  - Participant Status in Portal
  - Participant Status in Dataverse
  - File Exist in Local (Yes/No)
  - Local Files Path
  - File Exist in SharePoint (Yes/No)
  - SharePoint Path

Usage:
    python rfp_analysis_report.py
"""

import sys
import os
import io
import re
import time
import logging
import requests
from datetime import datetime

# Fix Windows console encoding
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from bs4 import BeautifulSoup

from config.config import (
    OUTPUT_DIR, TENANT_ID, CLIENT_ID, CLIENT_SECRET,
    SHAREPOINT_HOSTNAME, SITE_PATH, DRIVE_NAME, SP_BASE_FOLDER,
)
from helpers.core_helper import (
    get_rfp_activity_data_from_db,
    clean_rfp_title,
    normalize_filename,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

PORTAL_RFPS_DIR = os.path.join(OUTPUT_DIR, "Portal-Rfps")
SP_ALLRFPS_PATH = f"{SP_BASE_FOLDER}/ALLRFPs"
EXCEL_EXTENSIONS = ('.xls', '.xlsx')

# Company name mapping: portal file name -> local folder name / Dataverse name
COMPANY_MAP = {
    "Saudi Energy": {
        "local_folder": "Saudi Energy",
        "dataverse_name": "Saudi Energy",
        "sp_folder": "Saudi Energy",
    },
    "Aramco E-Marketplace": {
        "local_folder": "Aramco e-Marketplace",
        "dataverse_name": "Aramco e-Marketplace",
        "sp_folder": "Aramco e-Marketplace",
    },
    "HADEED - RAJHI STEEL": {
        "local_folder": "HADEED - RAJHI STEEL",
        "dataverse_name": "HADEED - RAJHI STEEL",
        "sp_folder": "HADEED - RAJHI STEEL",
    },
    "SABIC - Saudi Basic Industries Corp.": {
        "local_folder": "SABIC - Saudi Basic Industries Corp_",
        "dataverse_name": "SABIC - Saudi Basic Industries Corp.",
        "sp_folder": "SABIC - Saudi Basic Industries Corp_",
    },
}


def normalize_rfp_id(text: str) -> str:
    """Normalize RFP title/ID for matching by removing extra spaces and lowercasing."""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text).strip().lower()


def extract_contract_number(title: str) -> str:
    """Extract the contract number (e.g., C001738316) from an RFP title."""
    if not title:
        return ""
    # SEC-style: C followed by 7-10 digits
    match = re.search(r'C\d{7,10}', title)
    if match:
        return match.group(0)
    # Aramco-style: 10-digit number (e.g., 4203238896)
    match = re.search(r'\d{10}', title)
    if match:
        return match.group(0)
    return ""


def extract_aramco_folder_id(rfp_id: str) -> str:
    """
    Extract the Aramco folder name from a Dataverse RFP_ID.
    Dataverse: 'Aramco_4203238896_CABLE, POWER, 5KV THROUGH 35 KV'
    Local folder: 'Aramco_4203238896'
    Returns: 'Aramco_4203238896' or '' if not Aramco format.
    """
    if not rfp_id:
        return ""
    match = re.match(r'(Aramco_\d{10})', rfp_id)
    if match:
        return match.group(1)
    # Also try just the number for folders like '6201152020'
    match = re.search(r'(\d{10})', rfp_id)
    if match:
        return match.group(1)
    return ""


# ==================== PORTAL FILE PARSING ====================

def parse_sec_portal_file(filepath: str) -> list[dict]:
    """
    Parse Saudi Energy portal .xls file (HTML format).
    Returns list of dicts with: company, title, id, end_time, event_type, participated, portal_status
    """
    logger.info(f"Parsing SEC portal file: {filepath}")
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    soup = BeautifulSoup(content, 'lxml')
    tables = soup.find_all('table')

    if not tables:
        logger.warning("No tables found in SEC portal file")
        return []

    table = tables[0]
    all_cells = []
    for row in table.find_all('tr'):
        for cell in row.find_all(['td', 'th']):
            text = cell.get_text(strip=True)
            colspan = int(cell.get('colspan', 1))
            all_cells.append({'text': text, 'colspan': colspan})

    current_status = 'Unknown'
    rfps = []
    data_cells = []
    header_texts = {'Title', 'ID', 'End Time', 'Event Type', 'Participated'}

    for cell in all_cells:
        if cell['colspan'] >= 5 and re.match(r'Status:\s*\w+', cell['text']):
            current_status = cell['text']
            data_cells = []
            continue
        if cell['text'].startswith('TitleID'):
            continue
        if cell['colspan'] == 1 and cell['text'] in header_texts:
            continue

        if cell['colspan'] == 1 and cell['text']:
            data_cells.append(cell['text'])
            if len(data_cells) == 5:
                rfps.append({
                    'company': 'Saudi Energy',
                    'title': data_cells[0],
                    'doc_id': data_cells[1],
                    'end_time': data_cells[2],
                    'event_type': data_cells[3],
                    'portal_participated': data_cells[4],
                    'portal_status': current_status,
                })
                data_cells = []

    logger.info(f"Parsed {len(rfps)} RFPs from SEC portal file")
    return rfps


def parse_aramco_portal_file(filepath: str) -> list[dict]:
    """
    Parse Aramco E-Marketplace portal .xls file.
    Handles frameset files that reference external sheet files.
    """
    logger.info(f"Parsing Aramco portal file: {filepath}")

    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    # Check if it's a frameset referencing external files
    if 'frameset' in content.lower() and 'frame src=' in content.lower():
        soup = BeautifulSoup(content, 'lxml')
        frames = soup.find_all('frame')
        for frame in frames:
            src = frame.get('src', '')
            if src and 'sheet' in src.lower():
                base_dir = os.path.dirname(filepath)
                from urllib.parse import unquote
                decoded_src = unquote(src)
                ref_path = os.path.join(base_dir, decoded_src)

                if os.path.exists(ref_path):
                    logger.info(f"Found referenced sheet file: {ref_path}")
                    return _parse_aramco_html_table(ref_path)
                else:
                    logger.warning(
                        f"Aramco file references '{decoded_src}' but not found at '{ref_path}'. "
                        f"Place the data files folder next to the .xls file."
                    )

        logger.warning("Aramco portal file is a frameset with missing data files. Skipping.")
        return []

    return _parse_aramco_html_table(filepath)


def _parse_aramco_html_table(filepath: str) -> list[dict]:
    """Parse Aramco data from an HTML table file."""
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    soup = BeautifulSoup(content, 'lxml')
    tables = soup.find_all('table')
    if not tables:
        logger.warning("No tables found in Aramco file")
        return []

    rfps = []
    for table in tables:
        rows = table.find_all('tr')
        if len(rows) < 2:
            continue
        header_row = rows[0]
        headers = [cell.get_text(strip=True).lower() for cell in header_row.find_all(['th', 'td'])]
        if 'title' not in headers and 'name' not in headers:
            continue
        for row in rows[1:]:
            cells = [cell.get_text(strip=True) for cell in row.find_all(['td'])]
            if len(cells) >= 4:
                rfps.append({
                    'company': 'Aramco e-Marketplace',
                    'title': cells[0] if len(cells) > 0 else '',
                    'doc_id': cells[1] if len(cells) > 1 else '',
                    'end_time': cells[2] if len(cells) > 2 else '',
                    'event_type': cells[3] if len(cells) > 3 else 'RFP',
                    'portal_participated': cells[4] if len(cells) > 4 else '',
                    'portal_status': '',
                })

    logger.info(f"Parsed {len(rfps)} RFPs from Aramco portal file")
    return rfps


def load_portal_rfps() -> list[dict]:
    """Load all portal RFPs from XLS files in Portal-Rfps folder."""
    all_rfps = []

    if not os.path.exists(PORTAL_RFPS_DIR):
        logger.error(f"Portal-Rfps directory not found: {PORTAL_RFPS_DIR}")
        return all_rfps

    for filename in os.listdir(PORTAL_RFPS_DIR):
        if not filename.lower().endswith('.xls'):
            continue

        filepath = os.path.join(PORTAL_RFPS_DIR, filename)
        base_name = os.path.splitext(filename)[0]

        if 'saudi' in base_name.lower() or 'sec' in base_name.lower():
            rfps = parse_sec_portal_file(filepath)
        elif 'aramco' in base_name.lower():
            rfps = parse_aramco_portal_file(filepath)
        else:
            logger.warning(f"Unknown portal file: {filename}, attempting generic parse")
            rfps = parse_sec_portal_file(filepath)

        all_rfps.extend(rfps)

    logger.info(f"Total portal RFPs loaded: {len(all_rfps)}")
    return all_rfps


# ==================== DATAVERSE ====================

def load_dataverse_rfps() -> tuple[dict, dict, list]:
    """
    Load RFP data from Dataverse and index by normalized RFP_ID.
    Also builds a secondary index by contract number for fuzzy matching.
    Returns (by_norm_id, by_contract_num, raw_rows).
    """
    logger.info("Fetching RFP data from Dataverse...")
    rows = get_rfp_activity_data_from_db()
    logger.info(f"Fetched {len(rows)} RFP records from Dataverse")

    by_norm_id = {}
    by_contract_num = {}
    for row in rows:
        rfp_id = row.get('RFP_ID', '')
        if rfp_id:
            key = normalize_rfp_id(rfp_id)
            by_norm_id[key] = row
            c_num = extract_contract_number(rfp_id)
            if c_num:
                by_contract_num[c_num] = row

    return by_norm_id, by_contract_num, rows


# ==================== LOCAL FILE CHECK ====================

def build_local_file_index() -> dict:
    """
    Pre-scan all local ALLRFPs folders and build an index.
    Returns dict: (company_folder, normalized_folder_name) -> {has_excel, file_path, folder_path}
    Also indexes by contract number for fuzzy matching.
    """
    logger.info("Scanning local ALLRFPs folders...")
    index_by_name = {}  # (company, clean_name) -> info
    index_by_cnum = {}  # (company, contract_number) -> info

    if not os.path.exists(OUTPUT_DIR):
        return index_by_name, index_by_cnum

    for company_folder in os.listdir(OUTPUT_DIR):
        company_path = os.path.join(OUTPUT_DIR, company_folder)
        if not os.path.isdir(company_path) or company_folder == "Portal-Rfps":
            continue

        for rfp_folder in os.listdir(company_path):
            rfp_path = os.path.join(company_path, rfp_folder)
            if not os.path.isdir(rfp_path):
                continue

            # Check for Excel files
            has_excel = False
            file_path = ""

            # Check downloaded-rfp subfolder first
            dl_path = os.path.join(rfp_path, "downloaded-rfp")
            if os.path.exists(dl_path):
                for f in os.listdir(dl_path):
                    if f.lower().endswith(EXCEL_EXTENSIONS):
                        has_excel = True
                        file_path = os.path.join(dl_path, f)
                        break

            # Fallback: check RFP folder directly
            if not has_excel:
                for f in os.listdir(rfp_path):
                    if f.lower().endswith(EXCEL_EXTENSIONS):
                        has_excel = True
                        file_path = os.path.join(rfp_path, f)
                        break

            info = {
                'has_excel': has_excel,
                'file_path': file_path,
                'folder_path': rfp_path,
            }

            # Index by clean name
            clean_name = clean_rfp_title(rfp_folder)
            index_by_name[(company_folder, clean_name)] = info

            # Also index by normalized name (all lowercase, no special chars)
            norm_name = normalize_filename(rfp_folder)
            index_by_name[(company_folder, norm_name)] = info

            # Index by contract number
            c_num = extract_contract_number(rfp_folder)
            if c_num:
                index_by_cnum[(company_folder, c_num)] = info

            # Index by Aramco folder ID (e.g., 'Aramco_4203238896')
            aramco_id = extract_aramco_folder_id(rfp_folder)
            if aramco_id:
                index_by_name[(company_folder, aramco_id)] = info

    total_folders = len(set(v['folder_path'] for v in index_by_name.values()))
    logger.info(f"Indexed {total_folders} local RFP folders")
    return index_by_name, index_by_cnum


def check_local_files(title: str, company: str, local_name_idx: dict, local_cnum_idx: dict) -> tuple[bool, str]:
    """
    Check if RFP files exist locally using pre-built index.
    Tries: exact clean name -> Aramco folder ID -> normalized name -> contract number.
    """
    local_folder = COMPANY_MAP.get(company, {}).get('local_folder', company)
    clean_title = clean_rfp_title(title)

    # Try exact clean name match
    info = local_name_idx.get((local_folder, clean_title))
    if info:
        return info['has_excel'], info['file_path'] or info['folder_path']

    # Try Aramco folder ID match (e.g., 'Aramco_4203238896')
    aramco_id = extract_aramco_folder_id(title)
    if aramco_id:
        info = local_name_idx.get((local_folder, aramco_id))
        if info:
            return info['has_excel'], info['file_path'] or info['folder_path']

    # Try normalized name match
    norm_name = normalize_filename(title)
    info = local_name_idx.get((local_folder, norm_name))
    if info:
        return info['has_excel'], info['file_path'] or info['folder_path']

    # Try contract number match
    c_num = extract_contract_number(title)
    if c_num:
        info = local_cnum_idx.get((local_folder, c_num))
        if info:
            return info['has_excel'], info['file_path'] or info['folder_path']

    return False, ""


# ==================== SHAREPOINT CHECK ====================

def init_sharepoint_client():
    """Initialize and authenticate SharePoint Graph client."""
    logger.info("Authenticating with SharePoint...")
    import helpers.core_helper  # noqa: resolve circular imports
    from helpers.sharepoint_helper import GraphClient

    client = GraphClient(CLIENT_ID, CLIENT_SECRET, TENANT_ID, SHAREPOINT_HOSTNAME, SITE_PATH, DRIVE_NAME)
    client.auth()
    client.resolve_site_and_drive()
    logger.info("SharePoint authenticated.")
    return client


def build_sharepoint_index(graph_client) -> dict:
    """
    Scan SharePoint ALLRFPs folder and build an index.
    Returns dict: (company_folder, contract_number_or_name) -> {has_excel, sp_path}
    """
    logger.info(f"Scanning SharePoint folder: {SP_ALLRFPS_PATH} ...")

    sp_name_idx = {}  # (company, clean_name) -> info
    sp_cnum_idx = {}  # (company, contract_number) -> info

    # List company folders
    company_folders = _sp_list_children(graph_client, SP_ALLRFPS_PATH, folders_only=True)
    if not company_folders:
        logger.warning("No company folders found in SharePoint.")
        return sp_name_idx, sp_cnum_idx

    logger.info(f"Found {len(company_folders)} company folder(s) in SharePoint.")
    rfp_count = 0

    for ci, company in enumerate(company_folders, 1):
        company_name = company["name"]
        company_path = company["path"]
        logger.info(f"[{ci}/{len(company_folders)}] SharePoint scan: {company_name}")

        rfp_folders = _sp_list_children(graph_client, company_path, folders_only=True)
        logger.info(f"  Found {len(rfp_folders)} RFP folder(s)")

        for rfp in rfp_folders:
            rfp_name = rfp["name"]
            rfp_path = rfp["path"]

            has_excel, excel_path = _sp_check_excel(graph_client, rfp_path)

            info = {
                'has_excel': has_excel,
                'sp_path': excel_path if excel_path else rfp_path,
            }

            # Index by clean name
            clean_name = clean_rfp_title(rfp_name)
            sp_name_idx[(company_name, clean_name)] = info

            # Index by contract number
            c_num = extract_contract_number(rfp_name)
            if c_num:
                sp_cnum_idx[(company_name, c_num)] = info

            rfp_count += 1
            if rfp_count % 50 == 0:
                graph_client.ensure_token()

            time.sleep(0.03)

    logger.info(f"SharePoint scan complete. Indexed {rfp_count} RFP folders.")
    return sp_name_idx, sp_cnum_idx


def _sp_list_children(graph_client, sp_path: str, folders_only: bool = False) -> list:
    """List children in a SharePoint directory with pagination."""
    graph_client.ensure_token()
    url = (
        f"https://graph.microsoft.com/v1.0/sites/{graph_client.site_id}"
        f"/drives/{graph_client.drive_id}/root:/{sp_path}:/children"
    )

    results = []
    while url:
        try:
            response = requests.get(url, headers=graph_client.headers, timeout=30)
        except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectionError) as e:
            logger.warning(f"Request failed for {sp_path}: {e}")
            return results

        if response.status_code == 404:
            return []
        if response.status_code == 429:
            retry_after = int(response.headers.get("Retry-After", 5))
            logger.warning(f"Rate limited. Sleeping {retry_after}s...")
            time.sleep(retry_after)
            continue
        if response.status_code != 200:
            logger.warning(f"Could not list {sp_path}: HTTP {response.status_code}")
            return results

        data = response.json()
        for item in data.get("value", []):
            is_folder = "folder" in item
            name = item.get("name", "")
            if folders_only and not is_folder:
                continue
            results.append({"name": name, "path": f"{sp_path}/{name}"})

        url = data.get("@odata.nextLink")

    return results


def _sp_check_excel(graph_client, rfp_folder_path: str) -> tuple[bool, str]:
    """Check if downloaded-rfp subfolder in SharePoint has Excel files."""
    dl_path = f"{rfp_folder_path}/downloaded-rfp"
    graph_client.ensure_token()
    url = (
        f"https://graph.microsoft.com/v1.0/sites/{graph_client.site_id}"
        f"/drives/{graph_client.drive_id}/root:/{dl_path}:/children"
    )

    try:
        response = requests.get(url, headers=graph_client.headers, timeout=30)
    except (requests.exceptions.ReadTimeout, requests.exceptions.ConnectionError):
        return False, ""

    if response.status_code != 200:
        return False, ""

    for item in response.json().get("value", []):
        if "folder" in item:
            continue
        name = item.get("name", "")
        if name.lower().endswith(EXCEL_EXTENSIONS):
            return True, f"{dl_path}/{name}"

    return False, ""


def check_sharepoint_files(title: str, company: str, sp_name_idx: dict, sp_cnum_idx: dict) -> tuple[bool, str]:
    """Check if RFP files exist in SharePoint using pre-built index."""
    sp_folder = COMPANY_MAP.get(company, {}).get('sp_folder', company)
    clean_title = clean_rfp_title(title)

    # Try exact clean name match
    info = sp_name_idx.get((sp_folder, clean_title))
    if info:
        return info['has_excel'], info['sp_path']

    # Try Aramco folder ID match
    aramco_id = extract_aramco_folder_id(title)
    if aramco_id:
        info = sp_name_idx.get((sp_folder, aramco_id))
        if info:
            return info['has_excel'], info['sp_path']

    # Try contract number match
    c_num = extract_contract_number(title)
    if c_num:
        info = sp_cnum_idx.get((sp_folder, c_num))
        if info:
            return info['has_excel'], info['sp_path']

    return False, ""


# ==================== REPORT BUILDING ====================

def _format_publish_date(publish_date: str) -> str:
    """Format ISO publish_time to readable date string."""
    if publish_date and 'T' in str(publish_date):
        try:
            dt = datetime.fromisoformat(publish_date.replace('Z', '+00:00'))
            return dt.strftime('%m/%d/%Y %I:%M %p')
        except (ValueError, AttributeError):
            pass
    return publish_date or ''


def build_report() -> pd.DataFrame:
    """Build the cross-reference analysis report."""
    # Step 1: Load portal RFPs
    portal_rfps = load_portal_rfps()

    # Step 2: Load Dataverse data (with contract number index for fast lookup)
    dv_norm_idx, dv_cnum_idx, dv_all_rows = load_dataverse_rfps()

    # Step 3: Determine which companies came from portal files
    portal_companies = set(rfp['company'] for rfp in portal_rfps)
    logger.info(f"Companies from portal files: {portal_companies or '(none)'}")

    # Step 4: For companies missing from portal, use Dataverse as source
    # Map Dataverse company names to our COMPANY_MAP keys
    dv_company_to_key = {}
    for key, val in COMPANY_MAP.items():
        dv_company_to_key[val['dataverse_name']] = key

    # Find companies in Dataverse that are NOT in portal data
    dv_companies_present = set()
    for row in dv_all_rows:
        cn = row.get('Company_Name', '') or ''
        if cn:
            dv_companies_present.add(cn)

    missing_companies = dv_companies_present - portal_companies
    if missing_companies:
        logger.info(f"Companies missing from portal files (using Dataverse): {missing_companies}")
        for row in dv_all_rows:
            cn = row.get('Company_Name', '') or ''
            if cn not in missing_companies:
                continue
            rfp_id = row.get('RFP_ID', '') or ''
            if not rfp_id:
                continue
            portal_rfps.append({
                'company': cn,
                'title': rfp_id,
                'doc_id': '',
                'end_time': row.get('RFP_End_Date', '') or '',
                'event_type': 'RFP',
                'portal_participated': '',  # no portal data
                'portal_status': '(from Dataverse)',
                '_from_dataverse': True,  # flag to avoid re-lookup
                '_dv_row': row,
            })

    if not portal_rfps:
        logger.error("No RFPs loaded from portal or Dataverse.")
        return pd.DataFrame()

    logger.info(f"Total RFPs to analyze: {len(portal_rfps)}")

    # Step 5: Build local file index (fast pre-scan)
    local_name_idx, local_cnum_idx = build_local_file_index()

    # Step 6: Build SharePoint index
    try:
        sp_client = init_sharepoint_client()
        sp_name_idx, sp_cnum_idx = build_sharepoint_index(sp_client)
    except Exception as e:
        logger.error(f"SharePoint scan failed: {e}. SharePoint columns will be empty.")
        sp_name_idx, sp_cnum_idx = {}, {}

    # Step 7: Cross-reference and build report rows
    logger.info("Building cross-reference report...")
    report_rows = []

    for rfp in portal_rfps:
        company = rfp['company']
        title = rfp['title']
        portal_participated = rfp.get('portal_participated', '')
        end_time = rfp['end_time']
        portal_status = rfp.get('portal_status', '')

        # --- Dataverse lookup ---
        if rfp.get('_from_dataverse'):
            # Already from Dataverse, no need to re-lookup
            dv_row = rfp['_dv_row']
        else:
            norm_title = normalize_rfp_id(title)
            dv_row = dv_norm_idx.get(norm_title)
            # Fallback: try contract number
            if not dv_row:
                c_num = extract_contract_number(title)
                if c_num:
                    dv_row = dv_cnum_idx.get(c_num)

        # --- Extract all Dataverse fields ---
        dv_participated = ''
        publish_date = ''
        dv_link = ''
        dv_owner = ''
        dv_email_status = ''
        dv_material_matched = ''
        dv_keyword_matched = ''
        if dv_row:
            dv_participated = dv_row.get('participated', '') or ''
            publish_date = _format_publish_date(dv_row.get('publish_time', '') or '')
            dv_link = dv_row.get('Link', '') or ''
            dv_owner = dv_row.get('owner_name', '') or ''
            dv_email_status = dv_row.get('Email_Status', '') or ''
            dv_material_matched = dv_row.get('Material_Matched', '') or ''
            dv_keyword_matched = dv_row.get('Keyword_Matched', '') or ''

        # --- Local file check ---
        file_exists, local_path = check_local_files(title, company, local_name_idx, local_cnum_idx)

        # --- SharePoint check ---
        sp_exists, sp_path = check_sharepoint_files(title, company, sp_name_idx, sp_cnum_idx)

        report_rows.append({
            'Company Name': company,
            'RFP Name / Title': title,
            'End Date': end_time,
            'Publish Date': publish_date,
            'Portal Status Group': portal_status,
            'Participant Status (Portal)': portal_participated,
            'Participant Status (Dataverse)': dv_participated,
            'Found in Dataverse': 'Yes' if dv_row else 'No',
            'RFP Link': dv_link,
            'Owner Name': dv_owner,
            'Email Status': dv_email_status,
            'Material Matched': dv_material_matched,
            'Keyword Matched': dv_keyword_matched,
            'File Exist in Local': 'Yes' if file_exists else 'No',
            'Local File Path': local_path,
            'File Exist in SharePoint': 'Yes' if sp_exists else 'No',
            'SharePoint Path': sp_path,
        })

    df = pd.DataFrame(report_rows)
    return df


def print_summary(df: pd.DataFrame):
    """Print summary statistics of the report."""
    if df.empty:
        print("\nNo data to report.")
        return

    total = len(df)
    in_dataverse = len(df[df['Found in Dataverse'] == 'Yes'])
    not_in_dataverse = total - in_dataverse
    with_local = len(df[df['File Exist in Local'] == 'Yes'])
    without_local = total - with_local
    with_sp = len(df[df['File Exist in SharePoint'] == 'Yes'])

    # "Need to download" = in Dataverse but no local file
    in_dv_no_local = len(df[(df['Found in Dataverse'] == 'Yes') & (df['File Exist in Local'] == 'No')])

    print(f"\n{'=' * 80}")
    print(f"  RFP Cross-Reference Analysis Report")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'=' * 80}")

    print(f"\n  OVERVIEW")
    print(f"  {'─' * 50}")
    print(f"  Total Portal/Source RFPs      : {total}")
    print(f"  Found in Dataverse            : {in_dataverse} ({in_dataverse/total*100:.1f}%)")
    print(f"  NOT in Dataverse              : {not_in_dataverse} ({not_in_dataverse/total*100:.1f}%)")
    print(f"  With Local Files              : {with_local} ({with_local/total*100:.1f}%)")
    print(f"  With SharePoint Files         : {with_sp} ({with_sp/total*100:.1f}%)")

    print(f"\n  DOWNLOAD GAP ANALYSIS")
    print(f"  {'─' * 50}")
    print(f"  In Dataverse but NO local file: {in_dv_no_local}  << Need to Download")
    print(f"  NOT in Dataverse at all       : {not_in_dataverse}  << Need to Add to DB first")
    print(f"  Already downloaded            : {with_local}")

    # Per company detailed breakdown
    print(f"\n  PER COMPANY BREAKDOWN")
    print(f"  {'─' * 50}")
    for company in sorted(df['Company Name'].unique()):
        c_df = df[df['Company Name'] == company]
        c_total = len(c_df)
        c_dv = len(c_df[c_df['Found in Dataverse'] == 'Yes'])
        c_not_dv = c_total - c_dv
        c_local = len(c_df[c_df['File Exist in Local'] == 'Yes'])
        c_sp = len(c_df[c_df['File Exist in SharePoint'] == 'Yes'])
        c_need_dl = len(c_df[(c_df['Found in Dataverse'] == 'Yes') & (c_df['File Exist in Local'] == 'No')])

        print(f"\n  {company}:")
        print(f"    Total RFPs          : {c_total}")
        print(f"    In Dataverse        : {c_dv}")
        print(f"    NOT in Dataverse    : {c_not_dv}")
        print(f"    Local Files         : {c_local}")
        print(f"    SharePoint Files    : {c_sp}")
        print(f"    Need to Download    : {c_need_dl}")

        # Participation breakdown for this company
        print(f"    Participation (Portal):")
        for status, count in c_df['Participant Status (Portal)'].value_counts().items():
            label = status if status else '(no portal data)'
            print(f"      {label}: {count}")
        print(f"    Participation (Dataverse):")
        dv_found = c_df[c_df['Found in Dataverse'] == 'Yes']
        if not dv_found.empty:
            for status, count in dv_found['Participant Status (Dataverse)'].value_counts().items():
                label = status if status else '(empty)'
                print(f"      {label}: {count}")

    # Portal participation breakdown
    print(f"\n  OVERALL PORTAL PARTICIPATION")
    print(f"  {'─' * 50}")
    for status, count in df['Participant Status (Portal)'].value_counts().items():
        label = status if status else '(no portal data)'
        print(f"    {label}: {count}")

    # Mismatches (only where both portal and dataverse have data)
    both = df[(df['Found in Dataverse'] == 'Yes') & (df['Participant Status (Portal)'] != '')]
    if not both.empty:
        mismatches = both[
            both['Participant Status (Portal)'].str.lower() != both['Participant Status (Dataverse)'].str.lower()
        ]
        print(f"\n  STATUS MISMATCHES (Portal vs Dataverse)")
        print(f"  {'─' * 50}")
        print(f"    Mismatched records: {len(mismatches)} out of {len(both)} comparable")

    print(f"\n{'=' * 80}\n")


def main():
    output_filename = f"RFP_Analysis_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(os.getcwd(), output_filename)

    logger.info("Starting RFP Analysis Report generation...")

    df = build_report()

    if df.empty:
        logger.error("Report is empty. No data to export.")
        return

    # Print summary
    print_summary(df)

    # Export to Excel with formatting
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='All RFPs')

        ws = writer.sheets['All RFPs']

        # Auto-adjust column widths
        for col_idx, column in enumerate(df.columns, 1):
            max_length = max(
                len(str(column)),
                df[column].astype(str).map(len).max() if len(df) > 0 else 0
            )
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Style header row
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Freeze top row + auto-filter
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # Color Yes/No cells
        green_font = Font(color='006100', bold=True)
        red_font = Font(color='9C0006', bold=True)
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        yes_no_columns = ['File Exist in Local', 'Found in Dataverse', 'File Exist in SharePoint']
        col_indices = {col: list(df.columns).index(col) + 1 for col in yes_no_columns if col in df.columns}

        for row_idx in range(2, len(df) + 2):
            for col_name, col_idx in col_indices.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value == 'Yes':
                    cell.font = green_font
                    cell.fill = green_fill
                elif cell.value == 'No':
                    cell.font = red_font
                    cell.fill = red_fill

        # --- Summary sheet ---
        summary_data = []
        for company in sorted(df['Company Name'].unique()):
            c_df = df[df['Company Name'] == company]
            c_total = len(c_df)
            c_dv = len(c_df[c_df['Found in Dataverse'] == 'Yes'])
            c_local = len(c_df[c_df['File Exist in Local'] == 'Yes'])
            c_sp = len(c_df[c_df['File Exist in SharePoint'] == 'Yes'])
            c_need_dl = len(c_df[(c_df['Found in Dataverse'] == 'Yes') & (c_df['File Exist in Local'] == 'No')])
            c_not_dv = c_total - c_dv
            summary_data.append({
                'Company': company,
                'Total Portal RFPs': c_total,
                'In Dataverse': c_dv,
                'NOT in Dataverse': c_not_dv,
                'Downloaded (Local)': c_local,
                'Downloaded (SharePoint)': c_sp,
                'Need to Download': c_need_dl,
            })
        # Add totals row
        summary_data.append({
            'Company': 'TOTAL',
            'Total Portal RFPs': len(df),
            'In Dataverse': len(df[df['Found in Dataverse'] == 'Yes']),
            'NOT in Dataverse': len(df[df['Found in Dataverse'] == 'No']),
            'Downloaded (Local)': len(df[df['File Exist in Local'] == 'Yes']),
            'Downloaded (SharePoint)': len(df[df['File Exist in SharePoint'] == 'Yes']),
            'Need to Download': len(df[(df['Found in Dataverse'] == 'Yes') & (df['File Exist in Local'] == 'No')]),
        })

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, index=False, sheet_name='Summary')

        ws2 = writer.sheets['Summary']
        for col_idx in range(1, len(summary_df.columns) + 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 22
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        # Bold the totals row
        total_row = len(summary_data) + 1
        bold_font = Font(bold=True, size=11)
        for col_idx in range(1, len(summary_df.columns) + 1):
            ws2.cell(row=total_row, column=col_idx).font = bold_font

    logger.info(f"Report saved to: {output_path}")
    print(f"\nReport saved: {output_path}")


if __name__ == "__main__":
    main()

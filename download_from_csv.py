"""
Download RFPs from Portal — Two-File, Local Only
=================================================
Uses TWO input files:

  File 1 — RFP ID list  (provided via --file argument)
      CSV or Excel with columns:  RFP_ID,  Company_Name  [, Link]
      Example: ALLRFPs/Portal-Rfps/cr673_requestforproposals.csv

  File 2 — Master links file  (hardcoded path, always the same)
      Excel (.xls) with columns:  Title,  ID,  End Time,  Event Type,  Participated
      Path: ALLRFPs/Portal-Rfps/All-RFPs.xls
      The Title column matches the RFP_ID from File 1.
      The ID column (e.g. "Doc134918839") is used to build the Ariba portal URL.

How it works:
  1. Load RFP IDs from File 1  (your download list)
  2. Look up each RFP_ID in the Master links file (File 2) by matching Title
  3. Build an Ariba portal URL from the Doc ID found in the master file
  4. If a match is found → use that URL to download directly (no portal scraping needed)
  5. If no match    → fall back to portal login + company-page scraping

No SharePoint. No Dataverse. Just local.

Input file formats supported:
    CSV  (.csv)   – RFP_ID, Company_Name [, Link]
    Excel (.xlsx / .xls) – same columns, any sheet name (first sheet used)

Minimum required columns in File 1:
    RFP_ID       – must match the Title in the master links file
    Company_Name – used for folder structure and portal navigation fallback

Usage:
    python download_from_csv.py --file rfps.csv
    python download_from_csv.py --file rfps.xlsx
    python download_from_csv.py --file rfps.xlsx --username user@example.com --password MyPass
    python download_from_csv.py --file rfps.xlsx --output "D:/CustomFolder"
    python download_from_csv.py --file rfps.xlsx --headless

Credentials (priority order):
    1. --username / --password CLI flags
    2. BAHRA_SAP_USERNAME / BAHRA_SAP_PASSWORD env vars

Retry behaviour:
    - On any network / system error: waits 10 s then retries (up to 3 times)
    - On session expiry / logout: re-logs in automatically then continues
"""

import os
import re
import csv
import sys
import asyncio
import argparse
import traceback
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# Report statuses  (used as values in the Status column of the CSV report)
# ─────────────────────────────────────────────────────────────────────────────
STATUS_DOWNLOADED       = "Downloaded"
STATUS_SKIPPED_EXISTS   = "Skipped - Already Exists"
STATUS_NOT_FOUND_PORTAL = "Not Found on Portal"
STATUS_PORTAL_ERROR     = "Portal Unavailable"
STATUS_DOWNLOAD_FAILED  = "Download Failed"


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

PORTAL_URL        = "https://service.ariba.com/Sourcing.aw/109582016/aw?awh=r&awssk=u9fNiSxN&dard=1#b0"
ARIBA_BASE_URL    = "https://service.ariba.com/Sourcing.aw/109582016/aw?awh=r&awssk=u9fNiSxN&dard=1"
DEFAULT_OUTPUT    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ALLRFPs")
MASTER_LINKS_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "ALLRFPs", "Portal-Rfps", "All-RFPs.xls",
)

MAX_DOWNLOAD_ATTEMPTS = 3    # total attempts per RFP before giving up
RETRY_WAIT_SEC        = 10   # seconds to wait between error retries
LOGIN_MAX_RETRIES     = 3    # how many times to try re-logging in


# ─────────────────────────────────────────────────────────────────────────────
# Path helpers  (mirrors helpers/core_helper.py logic)
# ─────────────────────────────────────────────────────────────────────────────

def _sanitize(name: str) -> str:
    """Strip chars that are illegal in folder/file names."""
    return re.sub(r'[<>:"/\\|?*]', '_', name).strip().rstrip('.')


def _clean_id(rfp_id: str) -> str:
    """Normalise whitespace in an RFP ID (matches clean_rfp_title)."""
    return re.sub(r'\s+', ' ', rfp_id).strip()


def get_downloaded_rfp_folder(rfp_id: str, company_name: str, output_dir: str) -> str:
    """Return  <output_dir>/<Company>/<RFP_ID>/downloaded-rfp/"""
    return os.path.join(
        output_dir,
        _sanitize(company_name),
        _clean_id(rfp_id),
        "downloaded-rfp",
    )


def file_exists_locally(rfp_id: str, company_name: str, output_dir: str) -> tuple[bool, str | None]:
    """True if any .xls/.xlsx already lives in the downloaded-rfp folder."""
    folder = get_downloaded_rfp_folder(rfp_id, company_name, output_dir)
    if not os.path.isdir(folder):
        return False, None
    for f in os.listdir(folder):
        if f.lower().endswith(('.xls', '.xlsx')):
            return True, os.path.join(folder, f)
    return False, None


# ─────────────────────────────────────────────────────────────────────────────
# Master links file  (All-RFPs.xls)
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_for_match(s: str) -> str:
    """Strip all non-alphanumeric chars for fuzzy RFP ID matching.

    Handles variations like:
        'SEC RFP-C001718985'  →  'secrfpc001718985'
        'SEC RFP C001718985'  →  'secrfpc001718985'
        'SEC RFPC001718985'   →  'secrfpc001718985'
    """
    return re.sub(r'[^a-z0-9]', '', s.lower())


def load_master_rfp_links(master_file: str) -> dict:
    """
    Read the master All-RFPs.xls file and build a lookup dict.

    Expected columns: Title, ID  (plus optional End Time / Event Type / Participated)
    The Title column cells contain embedded hyperlinks — those are the real portal URLs.
    The ID column (e.g. 'Doc134918839') is only used as a fallback if no hyperlink exists.

    Returns:
        { normalized_title : {'id': DocID, 'link': URL, 'title': original_title} }
        Empty dict if the file cannot be read.
    """
    if not os.path.isfile(master_file):
        _log(f"[WARN] Master links file not found: {master_file}")
        return {}

    try:
        import xlrd
        wb = xlrd.open_workbook(master_file)
        ws = wb.sheet_by_index(0)
        headers = [str(v).strip() for v in ws.row_values(0)]

        title_col = next((i for i, h in enumerate(headers) if h.lower() == 'title'), None)
        id_col    = next((i for i, h in enumerate(headers) if h.lower() == 'id'),    None)

        if title_col is None or id_col is None:
            _log(f"[WARN] Master file missing 'Title' or 'ID' column. Found: {headers}")
            return {}

        # Build hyperlink lookup from the sheet: (row, col) → URL
        hyperlink_map = getattr(ws, 'hyperlink_map', {})

        result = {}
        no_link_count = 0
        for row_i in range(1, ws.nrows):
            row    = ws.row_values(row_i)
            title  = str(row[title_col] if title_col < len(row) else "").strip()
            doc_id = str(row[id_col]    if id_col    < len(row) else "").strip()
            if not title or not doc_id:
                continue

            # Prefer the real hyperlink embedded in the Title cell
            hl = hyperlink_map.get((row_i, title_col))
            if hl and getattr(hl, 'url_or_path', ''):
                link = hl.url_or_path.strip()
            else:
                # Fallback: construct from Doc ID (less reliable but better than nothing)
                link = f"{ARIBA_BASE_URL}&an={doc_id}"
                no_link_count += 1

            norm = _normalize_for_match(title)
            result[norm] = {"id": doc_id, "link": link, "title": title}

        _log(
            f"Master links file loaded: {len(result)} RFPs indexed from '{master_file}' "
            f"({len(result) - no_link_count} with real hyperlinks, {no_link_count} fallback URLs)."
        )
        return result

    except ImportError:
        _log("[WARN] 'xlrd' package not installed. Run: pip install xlrd==1.2.0")
        return {}
    except Exception as exc:
        _log(f"[WARN] Could not read master links file '{master_file}': {exc}")
        return {}


def enrich_with_master_links(rows: list, master_links: dict) -> list:
    """
    For each row in the download list that has no link yet, look it up in the
    master links dict and fill in the portal URL.

    Matching is fuzzy: strips all non-alphanumeric chars before comparing.
    Rows that already carry a link are left untouched.

    Returns the same list (mutated in-place) for convenience.
    """
    enriched  = 0
    not_found = []

    for row in rows:
        if row["link"]:
            continue  # already has a direct link

        norm = _normalize_for_match(row["rfp_id"])
        if norm in master_links:
            row["link"] = master_links[norm]["link"]
            enriched += 1
        else:
            not_found.append(row["rfp_id"])

    if enriched:
        _log(f"Master file lookup: {enriched} RFP(s) enriched with portal links.")
    if not_found:
        _log(
            f"Master file lookup: {len(not_found)} RFP(s) NOT found in master file "
            f"— will fall back to portal scraping:"
        )
        for nf in not_found:
            _log(f"  - {nf}")

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Input file loader  (CSV  or  Excel .xlsx / .xls)
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_col(headers: list[str], target: str) -> str | None:
    """Case-insensitive column lookup (spaces and underscores treated the same)."""
    for h in headers:
        if h.strip().lower().replace(" ", "_") == target.lower():
            return h
    return None


def _parse_rows(raw_rows: list[dict], source_label: str) -> list[dict]:
    """
    Validate and normalise a list of raw header→value dicts.
    Returns cleaned rows: [{'rfp_id', 'company_name', 'link'}, ...]
    'link' is an empty string when the column is absent or blank.
    """
    if not raw_rows:
        _die(f"No data rows found in {source_label}")

    headers = list(raw_rows[0].keys())
    col_id   = _resolve_col(headers, "rfp_id")
    col_co   = _resolve_col(headers, "company_name")
    col_link = _resolve_col(headers, "link")   # optional

    if not col_id:
        _die(f"Input file must have a 'RFP_ID' column. Found: {headers}")
    if not col_co:
        _die(f"Input file must have a 'Company_Name' column. Found: {headers}")

    rows: list[dict] = []
    for lineno, raw in enumerate(raw_rows, start=2):
        rfp_id  = str(raw.get(col_id)  or "").strip()
        company = str(raw.get(col_co)  or "").strip()
        link    = str(raw.get(col_link) or "").strip() if col_link else ""

        # Treat "nan" (pandas empty cell string) as blank
        if link.lower() in ("nan", "none", "n/a", "-"):
            link = ""

        if not rfp_id or not company:
            print(f"  [WARN] Row {lineno}: missing rfp_id or company_name — skipped")
            continue

        rows.append({"rfp_id": rfp_id, "company_name": company, "link": link})

    return rows


def load_input_file(file_path: str) -> list[dict]:
    """
    Load RFP rows from a CSV or Excel file.
    Returns list of dicts: [{'rfp_id', 'company_name', 'link'}, ...]
    'link' is empty string when not provided.
    """
    if not os.path.isfile(file_path):
        _die(f"Input file not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            raw_headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            raw_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw_rows.append(dict(zip(raw_headers, [str(v or "").strip() for v in row])))
            wb.close()
        except Exception as exc:
            _die(f"Could not read Excel file '{file_path}': {exc}")

    elif ext == ".csv":
        raw_rows = []
        with open(file_path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                raw_rows.append(dict(row))

    else:
        _die(f"Unsupported file type '{ext}'. Use .csv, .xlsx, or .xls")

    return _parse_rows(raw_rows, file_path)


def group_by_company(rows: list[dict]) -> dict[str, list[dict]]:
    """
    Group rows by company_name.
    Returns {company_name: [{'rfp_id', 'company_name', 'link'}, ...], ...}
    Preserves the full row dict so the link travels with each RFP.
    """
    groups: dict[str, list[dict]] = {}
    for r in rows:
        groups.setdefault(r["company_name"], []).append(r)
    return groups


# ─────────────────────────────────────────────────────────────────────────────
# Utilities
# ─────────────────────────────────────────────────────────────────────────────

def _die(msg: str):
    print(f"[ERROR] {msg}")
    sys.exit(1)


def _log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


# ─────────────────────────────────────────────────────────────────────────────
# Portal session helpers
# ─────────────────────────────────────────────────────────────────────────────

async def _wait_ready(page, timeout: int = 30000):
    """Best-effort wait for the page to settle."""
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout)
    except Exception:
        pass
    try:
        await page.wait_for_load_state("domcontentloaded", timeout=5000)
    except Exception:
        pass


async def _page_has_login_form(page) -> bool:
    """Return True if the page is showing a login form (any platform/domain)."""
    LOGIN_SELECTORS = [
        'xpath=//*[@id="_boebpb"]/div[1]/input',  # Ariba standard
        'input[name="UserName"]',
        'input[id="UserName"]',
        '#UserName',
        'input[name="username"]',
        'input[type="password"]',                  # any password field = login page
    ]
    for sel in LOGIN_SELECTORS:
        try:
            if await page.locator(sel).count() > 0:
                return True
        except Exception:
            pass
    return False


async def is_logged_in(page) -> bool:
    """Return True if the current page looks like an authenticated portal page."""
    try:
        url = page.url.lower()
        # URL-based check (covers most cases)
        if any(kw in url for kw in ("login", "signin", "logon", "sso", "auth?")):
            return False
        # Form-based check (catches login redirects with non-standard URLs)
        if await _page_has_login_form(page):
            return False
        return True
    except Exception:
        return False


async def do_login(page, username: str, password: str) -> bool:
    """
    Navigate to the portal home and log in.
    Tries multiple selectors for username, password, and submit button.
    Returns True on success, False otherwise.
    """
    # Selector lists tried in priority order
    USER_SELECTORS = [
        'xpath=//*[@id="_boebpb"]/div[1]/input',
        '#UserName',
        'input[name="UserName"]',
        'input[id="UserName"]',
        'input[type="email"]',
        'input[type="text"]',
    ]
    PASS_SELECTORS = [
        '#Password',
        'input[name="Password"]',
        'input[id="Password"]',
        'input[type="password"]',
    ]
    SUBMIT_SELECTORS = [
        'input[type="submit"]',
        'button[type="submit"]',
        '#loginButton',
        'button:has-text("Log in")',
        'button:has-text("Sign in")',
    ]

    for attempt in range(1, LOGIN_MAX_RETRIES + 1):
        try:
            _log(f"Login attempt {attempt}/{LOGIN_MAX_RETRIES} …")
            await page.goto(PORTAL_URL, wait_until="domcontentloaded", timeout=60000)
            await asyncio.sleep(3)
            await _wait_ready(page, 30000)

            # If already logged in after navigation (SSO auto-login), done
            if await is_logged_in(page):
                _log("Already authenticated after navigation.")
                return True

            # ── Find username field ───────────────────────────────────────
            user_loc = None
            for sel in USER_SELECTORS:
                loc = page.locator(sel)
                if await loc.count() > 0:
                    user_loc = loc.first
                    _log(f"  Username field found: {sel}")
                    break

            if user_loc is None:
                _log(f"  Attempt {attempt}: username field not found on page — retrying.")
                await asyncio.sleep(5)
                continue

            await user_loc.click()
            await user_loc.fill(username)

            # ── Find password field ───────────────────────────────────────
            pass_loc = None
            for sel in PASS_SELECTORS:
                loc = page.locator(sel)
                if await loc.count() > 0:
                    pass_loc = loc.first
                    break

            if pass_loc is None:
                _log(f"  Attempt {attempt}: password field not found — retrying.")
                await asyncio.sleep(5)
                continue

            await pass_loc.click()
            await pass_loc.fill(password)

            # ── Submit ───────────────────────────────────────────────────
            submitted = False
            for sel in SUBMIT_SELECTORS:
                loc = page.locator(sel)
                if await loc.count() > 0:
                    try:
                        async with page.expect_navigation(wait_until="networkidle", timeout=60000):
                            await loc.first.click()
                    except Exception:
                        await loc.first.click()
                        await _wait_ready(page, 60000)
                    submitted = True
                    break

            if not submitted:
                _log(f"  Attempt {attempt}: submit button not found — retrying.")
                await asyncio.sleep(5)
                continue

            await asyncio.sleep(3)

            if await is_logged_in(page):
                _log("Logged in successfully.")
                return True

            _log(f"  Attempt {attempt}: still on login page after submit.")
            await asyncio.sleep(5)

        except Exception as exc:
            _log(f"  Login attempt {attempt} error: {exc}")
            if attempt < LOGIN_MAX_RETRIES:
                await asyncio.sleep(5)

    _log("All login attempts failed.")
    return False


async def ensure_logged_in(page, username: str, password: str) -> bool:
    """Re-login if the session has expired; return True if authenticated."""
    if await is_logged_in(page):
        return True
    _log("Session expired or logged out — re-logging in …")
    return await do_login(page, username, password)


# ─────────────────────────────────────────────────────────────────────────────
# Company navigation + RFP scraping
# ─────────────────────────────────────────────────────────────────────────────

async def navigate_to_company(page, company_name: str):
    """
    Click 'More…' then select the target company from the portal menu.
    After this call the company's RFP listing page is open.
    """
    _log(f"Navigating to company: {company_name}")

    # Click the "More…" / "More" link
    more_link = page.get_by_role("link", name=re.compile(r"^more(\.\.\.)?$", re.IGNORECASE)).first
    await more_link.click()

    # Wait for menu to appear
    menu = page.locator('div.awmenu:not(.is-dnone)').first
    try:
        await menu.wait_for(timeout=10000)
    except Exception:
        await more_link.click()
        await menu.wait_for(timeout=8000)

    # Click the company link
    company_item = page.locator("a.w-pmi-item:visible").filter(has_text=company_name).first
    await company_item.wait_for(state="attached", timeout=10000)

    try:
        await company_item.scroll_into_view_if_needed()
    except Exception:
        pass

    try:
        async with page.expect_navigation(wait_until="domcontentloaded", timeout=60000):
            await company_item.click(timeout=5000)
    except Exception:
        try:
            handle = await company_item.element_handle()
            if handle:
                await page.evaluate("el => el.click()", handle)
        except Exception:
            pass

    await _wait_ready(page)
    _log(f"Company '{company_name}' loaded.")


async def scrape_rfps_for_company(page, company_name: str) -> list[dict]:
    """
    Scrape the portal RFP list for a given company.
    Returns a list of dicts: [{Title, Link, ID, RFP_End_Date, Status}, ...]
    """
    _log(f"Scraping RFP list for: {company_name}")

    for attempt in range(1, 4):
        try:
            # Locate SupplierFrame
            await page.wait_for_selector("iframe", timeout=20000)
            frame = None
            for f in page.frames:
                if "SupplierFrame" in (f.name or "") or "SupplierFrame" in (f.url or ""):
                    frame = f
                    break
            if not frame:
                raise RuntimeError("SupplierFrame not found")

            # Click the "Open RFP" link
            await frame.wait_for_selector('a[id*="_03mdrd"]', timeout=20000)
            await frame.click('a[id*="_03mdrd"]')

            # Wait for RFP table rows
            await frame.wait_for_selector('#_swbzed tr.tableRow1', timeout=30000)

            rows_el = await frame.query_selector_all('#_swbzed tr.tableRow1')
            rfps = []
            for row_el in rows_el:
                cells = await row_el.query_selector_all('td')
                if not cells:
                    continue
                link_el = await cells[0].query_selector('a')
                link    = (await link_el.get_attribute("href") or "").strip() if link_el else ""
                title   = (await cells[1].inner_text()).strip() if len(cells) > 1 else ""
                rfp_id  = (await cells[3].inner_text()).strip() if len(cells) > 3 else ""
                end_dt  = (await cells[5].inner_text()).strip() if len(cells) > 5 else ""
                status  = (await cells[9].inner_text()).strip() if len(cells) > 9 else ""

                rfps.append({
                    "Title":       title,
                    "Link":        link,
                    "ID":          rfp_id,
                    "RFP_End_Date": end_dt,
                    "Status":      status,
                })

            _log(f"Found {len(rfps)} RFPs in portal for '{company_name}'.")
            return rfps

        except Exception as exc:
            _log(f"Scrape attempt {attempt} failed: {exc}")
            if attempt < 3:
                await asyncio.sleep(RETRY_WAIT_SEC)
                await page.reload()
                await _wait_ready(page)

    _log(f"Could not scrape RFPs for '{company_name}' after retries.")
    return []


def match_rfp_ids(csv_ids: list[str], portal_rfps: list[dict]) -> list[dict]:
    """
    Return the subset of portal_rfps whose Title matches one of the CSV RFP IDs.
    Matching is case-insensitive and strips extra whitespace.
    Also returns unmatched IDs for reporting.
    """
    csv_map = {_clean_id(i).lower(): _clean_id(i) for i in csv_ids}

    matched  : list[dict] = []
    matched_ids: set[str] = set()

    for rfp in portal_rfps:
        title_key = _clean_id(rfp.get("Title", "")).lower()
        if title_key in csv_map:
            matched.append(rfp)
            matched_ids.add(title_key)

    unmatched = [orig for key, orig in csv_map.items() if key not in matched_ids]
    return matched, unmatched


# ─────────────────────────────────────────────────────────────────────────────
# Single-RFP download
# ─────────────────────────────────────────────────────────────────────────────

async def _click_if_visible(page, selector: str, timeout: int = 3000) -> bool:
    try:
        loc = page.locator(selector)
        if await loc.count() > 0:
            await loc.first.wait_for(state="visible", timeout=timeout)
            await loc.first.click()
            return True
    except Exception:
        pass
    return False


async def _download_single(page, rfp: dict, company_name: str, output_dir: str) -> bool:
    """
    Open the RFP link in a new tab, click through the Ariba button sequence,
    download the file, and save it locally.
    Returns True on success.
    """
    title   = _clean_id(rfp.get("Title", ""))
    link    = (rfp.get("Link") or "").strip()

    if not link:
        _log(f"  [SKIP] No link available for RFP: {title}")
        return False

    save_dir  = get_downloaded_rfp_folder(title, company_name, output_dir)
    os.makedirs(save_dir, exist_ok=True)

    new_page = await page.context.new_page()
    try:
        _log(f"  Opening RFP page …")
        await new_page.goto(link, wait_until="domcontentloaded", timeout=60000)
        await asyncio.sleep(2)

        # Check for redirect to login — both URL and form-element based
        url_lower = new_page.url.lower()
        url_is_login = any(kw in url_lower for kw in ("login", "signin", "logon", "sso", "auth?"))
        form_is_login = await _page_has_login_form(new_page)
        if url_is_login or form_is_login:
            reason = "URL" if url_is_login else "login form detected"
            _log(f"  Session expired mid-download ({reason}). Will re-login.")
            await new_page.close()
            return None  # Signal: needs re-login

        # Step 1 — Review Event details tab
        clicked = await _click_if_visible(new_page, "#_c8_tuc", timeout=5000)
        if clicked:
            _log("  Clicked #_c8_tuc (Review Event details)")
        await asyncio.sleep(10)

        # Step 2 — Content / View tab
        old_url = new_page.url
        for _ in range(20):
            clicked = await _click_if_visible(new_page, "#_iiyvqc", timeout=2000)
            if clicked and new_page.url != old_url:
                _log("  Clicked #_iiyvqc (Content tab)")
                break
            await asyncio.sleep(0.5)

        # Step 3 — Wait for Download button
        btn = new_page.locator("#_gktadc")
        await btn.wait_for(state="visible", timeout=10000)
        await new_page.wait_for_function(
            """el => el && !el.disabled && el.offsetParent !== null
                   && (() => {
                       const r = el.getBoundingClientRect();
                       const e = document.elementFromPoint(r.left + r.width/2, r.top + r.height/2);
                       return e && (e === el || el.contains(e));
                   })()""",
            arg=await btn.element_handle(),
            timeout=10000,
        )
        _log("  Download button ready — triggering download …")

        # Step 4 — Trigger download
        async with new_page.expect_download(timeout=120000) as dl_info:
            await btn.click(no_wait_after=True)

        download  = await dl_info.value
        suggested = download.suggested_filename or f"{title}.xls"
        ext       = os.path.splitext(suggested)[1] or ".xls"
        dest_file = os.path.join(save_dir, f"{title}{ext}")

        await download.save_as(dest_file)
        size_kb = os.path.getsize(dest_file) / 1024
        _log(f"  [OK] Saved → {dest_file}  ({size_kb:.1f} KB)")
        return True

    except Exception as exc:
        _log(f"  Download error: {exc}")
        return False
    finally:
        try:
            await new_page.close()
        except Exception:
            pass


async def download_with_retry(
    page,
    rfp: dict,
    company_name: str,
    output_dir: str,
    username: str,
    password: str,
) -> bool:
    """
    Wrap _download_single with:
      - auto re-login on session expiry  (session drops do NOT count as real attempts)
      - 10-second wait + retry on any network/system error
    Returns True on success.
    """
    title           = _clean_id(rfp.get("Title", ""))
    real_attempts   = 0   # counts actual download failures only
    session_retries = 0   # counts mid-download session drops
    MAX_SESSION_RETRIES = 5  # give up if session keeps dropping

    while real_attempts < MAX_DOWNLOAD_ATTEMPTS:
        _log(f"  Attempt {real_attempts + 1}/{MAX_DOWNLOAD_ATTEMPTS} for: {title}")

        # Always ensure we're logged in before each attempt
        if not await ensure_logged_in(page, username, password):
            _log("  Cannot log in — skipping this RFP.")
            return False

        result = await _download_single(page, rfp, company_name, output_dir)

        if result is True:
            return True

        if result is None:
            # Session dropped inside _download_single — re-login then retry
            # (this does NOT increment real_attempts so no attempt is wasted)
            session_retries += 1
            if session_retries > MAX_SESSION_RETRIES:
                _log(f"  Session dropped {session_retries} times in a row — giving up on this RFP.")
                return False
            _log(f"  Session drop #{session_retries} — re-logging in …")
            ok = await do_login(page, username, password)
            if not ok:
                _log("  Re-login failed — skipping.")
                return False
            await asyncio.sleep(3)
            continue  # retry WITHOUT incrementing real_attempts

        # result is False → an actual download error (network / button / timeout)
        real_attempts += 1
        session_retries = 0   # reset session-drop counter on a real attempt
        if real_attempts < MAX_DOWNLOAD_ATTEMPTS:
            _log(f"  Waiting {RETRY_WAIT_SEC}s before retry …")
            await asyncio.sleep(RETRY_WAIT_SEC)

    _log(f"  [FAIL] Gave up after {MAX_DOWNLOAD_ATTEMPTS} attempts: {title}")
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Report writer
# ─────────────────────────────────────────────────────────────────────────────

def write_report_csv(report_rows: list[dict], output_dir: str, run_started_at: str) -> str:
    """
    Write a CSV report to <output_dir>/RFP_Download_Report_<timestamp>.csv.

    Report structure
    ────────────────
    Block 1 – Run Summary  (key/value pairs with blank separator)
    Block 2 – Per-RFP detail rows

    Per-RFP columns:
        RFP_ID | Company_Name | Status | Reason | Local_File | Processed_At

    Status values:
        Downloaded            – file successfully saved locally
        Skipped - Already Exists – file was already present before the run
        Not Found on Portal   – RFP ID was not in the portal list (after retries)
        Portal Unavailable    – company page could not be reached / scraped
        Download Failed       – found on portal but download errored out

    Returns the full path of the written report file.
    """
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(output_dir, f"RFP_Download_Report_{timestamp}.csv")
    os.makedirs(output_dir, exist_ok=True)

    # ── Build summary counts ──────────────────────────────────────────────
    total          = len(report_rows)
    downloaded     = sum(1 for r in report_rows if r["Status"] == STATUS_DOWNLOADED)
    skipped        = sum(1 for r in report_rows if r["Status"] == STATUS_SKIPPED_EXISTS)
    not_on_portal  = sum(1 for r in report_rows if r["Status"] == STATUS_NOT_FOUND_PORTAL)
    portal_error   = sum(1 for r in report_rows if r["Status"] == STATUS_PORTAL_ERROR)
    failed         = sum(1 for r in report_rows if r["Status"] == STATUS_DOWNLOAD_FAILED)

    companies = sorted({r["Company_Name"] for r in report_rows})

    summary_block = [
        ["RFP Download Report"],
        ["Run Started",          run_started_at],
        ["Report Generated",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [""],
        ["── Summary ──────────────────────────"],
        ["Total RFPs in CSV",       total],
        ["Downloaded (new)",        downloaded],
        ["Skipped (already exist)", skipped],
        ["Not Found on Portal",     not_on_portal],
        ["Portal Unavailable",      portal_error],
        ["Download Failed",         failed],
        [""],
        ["── Companies processed ──────────────"],
    ]
    for co in companies:
        co_rows   = [r for r in report_rows if r["Company_Name"] == co]
        co_ok     = sum(1 for r in co_rows if r["Status"] == STATUS_DOWNLOADED)
        co_skip   = sum(1 for r in co_rows if r["Status"] == STATUS_SKIPPED_EXISTS)
        co_miss   = sum(1 for r in co_rows if r["Status"] in (STATUS_NOT_FOUND_PORTAL, STATUS_PORTAL_ERROR))
        co_fail   = sum(1 for r in co_rows if r["Status"] == STATUS_DOWNLOAD_FAILED)
        summary_block.append([
            co,
            f"total={len(co_rows)}",
            f"downloaded={co_ok}",
            f"skipped={co_skip}",
            f"not_found={co_miss}",
            f"failed={co_fail}",
        ])

    summary_block.append([""])
    summary_block.append(["── Detail ────────────────────────────"])
    summary_block.append([""])

    detail_header = ["RFP_ID", "Company_Name", "Status", "Reason", "Local_File", "Processed_At"]

    with open(report_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)

        # Summary block
        for row in summary_block:
            writer.writerow(row)

        # Detail block
        writer.writerow(detail_header)
        for r in report_rows:
            writer.writerow([
                r.get("RFP_ID",       ""),
                r.get("Company_Name", ""),
                r.get("Status",       ""),
                r.get("Reason",       ""),
                r.get("Local_File",   ""),
                r.get("Processed_At", ""),
            ])

    return report_path


def _report_row(
    rfp_id: str,
    company_name: str,
    status: str,
    reason: str = "",
    local_file: str = "",
) -> dict:
    """Build a single report-row dict."""
    return {
        "RFP_ID":       rfp_id,
        "Company_Name": company_name,
        "Status":       status,
        "Reason":       reason,
        "Local_File":   local_file,
        "Processed_At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Progress bar
# ─────────────────────────────────────────────────────────────────────────────

class ProgressTracker:
    """
    Prints a live progress bar to stdout after every RFP is processed.

    Example output:
      +-----------------------------------------------------------------------+
      | Progress  [####################--------------------]   80 / 200  40.0%|
      |           OK: 75    Skip: 3    Fail: 2    NotFound: 0    Pending: 120 |
      +-----------------------------------------------------------------------+
    """

    BAR_WIDTH = 40

    def __init__(self, total: int):
        self.total      = total
        self.done       = 0
        self.downloaded = 0
        self.skipped    = 0
        self.failed     = 0
        self.not_found  = 0

    @property
    def pending(self) -> int:
        return max(0, self.total - self.done)

    def update(self, status: str, count: int = 1):
        """Record `count` RFPs finishing with the given status and redraw."""
        self.done += count
        if   status == STATUS_DOWNLOADED:       self.downloaded += count
        elif status == STATUS_SKIPPED_EXISTS:   self.skipped    += count
        elif status == STATUS_DOWNLOAD_FAILED:  self.failed     += count
        else:                                   self.not_found  += count
        self._draw()

    def _draw(self):
        total  = self.total or 1
        filled = int(self.BAR_WIDTH * min(self.done, total) / total)
        bar    = "#" * filled + "-" * (self.BAR_WIDTH - filled)
        pct    = 100.0 * min(self.done, total) / total
        w      = 71   # inner width of the box

        line1 = f"[{bar}]  {self.done:>4} / {self.total}   ({pct:5.1f}%)"
        line2 = (
            f"OK: {self.downloaded:<5} "
            f"Skip: {self.skipped:<5} "
            f"Fail: {self.failed:<5} "
            f"NotFound: {self.not_found:<5} "
            f"Pending: {self.pending}"
        )

        # Build lines, then derive border width from the longest line
        prefix1 = "  | Progress  "
        prefix2 = "  |           "
        content_w = max(len(line1), len(line2))
        inner_w   = len("Progress  ") + content_w + 2   # "| Progress  <content> |"

        print()
        print(f"  +-{'-' * inner_w}-+")
        print(f"{prefix1}{line1:<{content_w}} |")
        print(f"{prefix2}{line2:<{content_w}} |")
        print(f"  +-{'-' * inner_w}-+")


# ─────────────────────────────────────────────────────────────────────────────
# Main runner
# ─────────────────────────────────────────────────────────────────────────────

async def run(input_file: str, output_dir: str, username: str, password: str, headless: bool):
    from playwright.async_api import async_playwright

    run_started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Step 1: Load download list (File 1) ────────────────────────────────
    rows = load_input_file(input_file)
    if not rows:
        _die("No valid rows in input file.")

    # ── Step 2: Load master links (File 2) and enrich rows ─────────────────
    print()
    _log(f"Loading master links file: {MASTER_LINKS_FILE}")
    master_links = load_master_rfp_links(MASTER_LINKS_FILE)
    if master_links:
        rows = enrich_with_master_links(rows, master_links)
    else:
        _log("[WARN] Master links file could not be loaded — all RFPs will use portal scraping.")

    os.makedirs(output_dir, exist_ok=True)
    company_groups = group_by_company(rows)

    rows_with_link    = sum(1 for r in rows if r["link"])
    rows_without_link = len(rows) - rows_with_link

    print()
    print("=" * 65)
    print("  Portal RFP Downloader  —  Two-File Mode")
    print("=" * 65)
    print(f"  Download list : {input_file}")
    print(f"  Master links  : {MASTER_LINKS_FILE}")
    print(f"  Output        : {output_dir}")
    print(f"  Total RFPs    : {len(rows)}  across  {len(company_groups)} companies")
    print(f"  Link found    : {rows_with_link}  (direct download via master file)")
    print(f"  No link       : {rows_without_link}  (will scrape portal for link)")
    print(f"  Headless      : {headless}")
    print("=" * 65)
    print()

    # ── report_rows accumulates one entry per RFP throughout the whole run ──
    report_rows: list[dict] = []

    # ── Pre-flight: which RFPs already exist locally? ──────────────────────
    to_process:  list[dict] = []
    pre_skipped: list[dict] = []
    for row in rows:
        exists, path = file_exists_locally(row['rfp_id'], row['company_name'], output_dir)
        if exists:
            _log(f"[SKIP] {row['rfp_id']} ({row['company_name']}) — exists: {path}")
            pre_skipped.append(row)
            report_rows.append(_report_row(
                rfp_id=row['rfp_id'],
                company_name=row['company_name'],
                status=STATUS_SKIPPED_EXISTS,
                reason="File already exists in local ALLRFPs folder",
                local_file=path or "",
            ))
        else:
            to_process.append(row)

    print()
    _log(f"Pre-flight: {len(pre_skipped)} already exist, {len(to_process)} to download.")
    print()

    # ── Progress tracker (covers the entire list, pre-skipped included) ────
    progress = ProgressTracker(total=len(rows))
    if pre_skipped:
        progress.update(STATUS_SKIPPED_EXISTS, len(pre_skipped))

    if not to_process:
        _log("Nothing to download — all files already exist locally.")
        report_path = write_report_csv(report_rows, output_dir, run_started_at)
        _log(f"Report saved → {report_path}")
        return

    stats = {
        "downloaded":  0,
        "failed":      0,
        "skipped_pre": len(pre_skipped),
        "no_match":    0,
        "portal_err":  0,
    }

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=headless,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"],
        )
        context = await browser.new_context(accept_downloads=True, viewport={"width": 1280, "height": 1024})
        page    = await context.new_page()

        # ── Initial login ──────────────────────────────────────────────────
        _log("Logging in to portal …")
        if not await do_login(page, username, password):
            await browser.close()
            _die("Initial login failed. Check your credentials.")

        # ── Process each company ───────────────────────────────────────────
        for company_name, company_rows in company_groups.items():

            # Only work on rows that aren't already downloaded
            pending_rows = [
                r for r in company_rows
                if not file_exists_locally(r['rfp_id'], company_name, output_dir)[0]
            ]
            if not pending_rows:
                _log(f"[{company_name}] All RFPs already downloaded — skipping company.")
                continue

            print()
            print(f"  {'─'*60}")
            _log(f"Company: {company_name}  ({len(pending_rows)} RFP(s) to download)")
            print(f"  {'─'*60}")

            # Split: rows that already carry a link vs rows that need portal scraping
            direct_rows = [r for r in pending_rows if r["link"]]
            scrape_rows = [r for r in pending_rows if not r["link"]]

            if direct_rows:
                _log(f"  {len(direct_rows)} RFP(s) have a direct link — downloading without portal scrape")
            if scrape_rows:
                _log(f"  {len(scrape_rows)} RFP(s) have no link — will scrape portal to find them")

            # ── Ensure session before any work ──────────────────────────
            if not await ensure_logged_in(page, username, password):
                _log(f"Cannot log in — skipping all {len(pending_rows)} RFP(s) for '{company_name}'.")
                for r in pending_rows:
                    report_rows.append(_report_row(
                        rfp_id=r['rfp_id'], company_name=company_name,
                        status=STATUS_PORTAL_ERROR,
                        reason="Could not log in to portal before processing this company",
                    ))
                stats["portal_err"] += len(pending_rows)
                progress.update(STATUS_PORTAL_ERROR, len(pending_rows))
                continue

            # ── DIRECT DOWNLOAD (link provided in Excel/CSV) ───────────────
            for r in direct_rows:
                rfp_id = _clean_id(r['rfp_id'])
                print()
                _log(f"[DIRECT]  {rfp_id}  |  {company_name}")
                _log(f"  Link: {r['link']}")

                exists, path = file_exists_locally(rfp_id, company_name, output_dir)
                if exists:
                    _log(f"  [SKIP] Already exists: {path}")
                    report_rows.append(_report_row(
                        rfp_id=rfp_id, company_name=company_name,
                        status=STATUS_SKIPPED_EXISTS,
                        reason="File appeared locally during run",
                        local_file=path or "",
                    ))
                    progress.update(STATUS_SKIPPED_EXISTS)
                    continue

                # Build a minimal rfp dict that _download_single understands
                rfp_dict = {"Title": rfp_id, "Link": r["link"]}
                success = await download_with_retry(
                    page, rfp_dict, company_name, output_dir, username, password
                )
                if success:
                    _, saved_path = file_exists_locally(rfp_id, company_name, output_dir)
                    report_rows.append(_report_row(
                        rfp_id=rfp_id, company_name=company_name,
                        status=STATUS_DOWNLOADED,
                        reason="Downloaded using link provided in input file",
                        local_file=saved_path or "",
                    ))
                    stats["downloaded"] += 1
                    progress.update(STATUS_DOWNLOADED)
                else:
                    report_rows.append(_report_row(
                        rfp_id=rfp_id, company_name=company_name,
                        status=STATUS_DOWNLOAD_FAILED,
                        reason=(
                            f"Download failed after {MAX_DOWNLOAD_ATTEMPTS} attempts using provided link. "
                            "Possible causes: link expired, no Excel on portal, download button unavailable, "
                            "or network timeout."
                        ),
                    ))
                    stats["failed"] += 1
                    progress.update(STATUS_DOWNLOAD_FAILED)

            # ── PORTAL SCRAPE + DOWNLOAD (no link provided) ────────────────
            if not scrape_rows:
                continue

            # Navigate to company & scrape RFP list
            nav_ok = False
            for nav_attempt in range(1, 3):   # 2 navigation attempts
                try:
                    await navigate_to_company(page, company_name)
                    nav_ok = True
                    break
                except Exception as exc:
                    _log(f"Navigation attempt {nav_attempt} failed: {exc}")
                    if nav_attempt < 2:
                        _log(f"Waiting {RETRY_WAIT_SEC}s and retrying …")
                        await asyncio.sleep(RETRY_WAIT_SEC)
                        await ensure_logged_in(page, username, password)

            if not nav_ok:
                _log(f"Could not navigate to '{company_name}' — skipping {len(scrape_rows)} RFP(s).")
                for r in scrape_rows:
                    report_rows.append(_report_row(
                        rfp_id=r['rfp_id'], company_name=company_name,
                        status=STATUS_PORTAL_ERROR,
                        reason="Company page could not be reached on portal after 2 attempts",
                    ))
                stats["portal_err"] += len(scrape_rows)
                progress.update(STATUS_PORTAL_ERROR, len(scrape_rows))
                continue

            portal_rfps = await scrape_rfps_for_company(page, company_name)

            if not portal_rfps:
                _log(f"No RFPs found on portal for '{company_name}' after retries — skipping {len(scrape_rows)} RFP(s).")
                for r in scrape_rows:
                    report_rows.append(_report_row(
                        rfp_id=r['rfp_id'], company_name=company_name,
                        status=STATUS_PORTAL_ERROR,
                        reason="Portal returned no RFP list for this company after 2 scrape attempts",
                    ))
                stats["portal_err"] += len(scrape_rows)
                progress.update(STATUS_PORTAL_ERROR, len(scrape_rows))
                continue

            # Match input IDs against scraped portal data
            scrape_ids   = [r['rfp_id'] for r in scrape_rows]
            matched_rfps, unmatched_ids = match_rfp_ids(scrape_ids, portal_rfps)

            # Record RFPs not found on portal
            if unmatched_ids:
                _log(f"[WARN] {len(unmatched_ids)} ID(s) not found in portal list for '{company_name}':")
                for uid in unmatched_ids:
                    _log(f"       - {uid}")
                    report_rows.append(_report_row(
                        rfp_id=uid, company_name=company_name,
                        status=STATUS_NOT_FOUND_PORTAL,
                        reason=(
                            "RFP ID was not found in the portal's RFP list after scraping "
                            f"({len(portal_rfps)} RFPs visible on portal). "
                            "The RFP may be closed, expired, or the ID may differ from the portal title."
                        ),
                    ))
                stats["no_match"] += len(unmatched_ids)
                progress.update(STATUS_NOT_FOUND_PORTAL, len(unmatched_ids))

            _log(f"Matched {len(matched_rfps)} RFP(s) on portal — downloading …")

            # Download each scraped+matched RFP
            for idx, rfp in enumerate(matched_rfps, start=1):
                title = _clean_id(rfp.get("Title", ""))
                print()
                _log(f"[{idx}/{len(matched_rfps)}] SCRAPED  {title}  |  {company_name}")

                exists, path = file_exists_locally(title, company_name, output_dir)
                if exists:
                    _log(f"  [SKIP] Already exists: {path}")
                    report_rows.append(_report_row(
                        rfp_id=title, company_name=company_name,
                        status=STATUS_SKIPPED_EXISTS,
                        reason="File appeared locally during run (added by another process or previous attempt)",
                        local_file=path or "",
                    ))
                    progress.update(STATUS_SKIPPED_EXISTS)
                    continue

                success = await download_with_retry(
                    page, rfp, company_name, output_dir, username, password
                )
                if success:
                    _, saved_path = file_exists_locally(title, company_name, output_dir)
                    report_rows.append(_report_row(
                        rfp_id=title, company_name=company_name,
                        status=STATUS_DOWNLOADED,
                        reason="Downloaded after scraping portal for link",
                        local_file=saved_path or "",
                    ))
                    stats["downloaded"] += 1
                    progress.update(STATUS_DOWNLOADED)
                else:
                    report_rows.append(_report_row(
                        rfp_id=title, company_name=company_name,
                        status=STATUS_DOWNLOAD_FAILED,
                        reason=(
                            f"Download failed after {MAX_DOWNLOAD_ATTEMPTS} attempts. "
                            "Possible causes: no Excel file on portal, download button unavailable, "
                            "network timeout, or session could not be recovered."
                        ),
                    ))
                    stats["failed"] += 1
                    progress.update(STATUS_DOWNLOAD_FAILED)

        await browser.close()

    # ── Final summary (console) ────────────────────────────────────────────
    print()
    print("=" * 65)
    print("  Download Run Complete")
    print("=" * 65)
    print(f"  Total RFPs in file        : {len(rows)}")
    print(f"  Skipped (already existed) : {stats['skipped_pre']}")
    print(f"  Not found on portal       : {stats['no_match']}")
    print(f"  Portal unavailable        : {stats['portal_err']}")
    print(f"  Download failed           : {stats['failed']}")
    print(f"  Downloaded (new)          : {stats['downloaded']}")
    print("=" * 65)
    print(f"  Files saved to: {output_dir}")

    # ── Write report CSV ───────────────────────────────────────────────────
    report_path = write_report_csv(report_rows, output_dir, run_started_at)
    print()
    print(f"  Report CSV → {report_path}")
    print()


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Download RFPs from Ariba portal using a CSV or Excel file (no SharePoint, no Dataverse)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Supported input formats:   .csv   .xlsx   .xls

Required columns:   RFP_ID,  Company_Name
Optional column:    Link   (direct portal URL — skips portal scraping when provided)

When Link is present   -> navigates directly to that URL and downloads.
When Link is blank     -> logs in to portal, scrapes company RFP list, finds the link, then downloads.

Examples:
  python download_from_csv.py --file my_rfps.xlsx
  python download_from_csv.py --file my_rfps.csv  --headless
  python download_from_csv.py --file my_rfps.xlsx --username me@co.com --password secret
  python download_from_csv.py --file my_rfps.xlsx --output "D:/MyRFPs"
        """,
    )
    parser.add_argument("--file",     required=True, help="Path to input file (.csv / .xlsx / .xls)")
    parser.add_argument("--output",   default=DEFAULT_OUTPUT, help=f"Local ALLRFPs folder (default: {DEFAULT_OUTPUT})")
    parser.add_argument("--username", default=None, help="Portal username (or set BAHRA_SAP_USERNAME)")
    parser.add_argument("--password", default=None, help="Portal password (or set BAHRA_SAP_PASSWORD)")
    parser.add_argument("--headless", action="store_true", default=False, help="Run browser in headless mode")

    args = parser.parse_args()

    username = args.username or os.getenv("BAHRA_SAP_USERNAME", "").strip()
    password = args.password or os.getenv("BAHRA_SAP_PASSWORD", "").strip()

    if not username or not password:
        _die(
            "Credentials required.\n"
            "  Use --username / --password  or  set environment variables:\n"
            "    BAHRA_SAP_USERNAME=...\n"
            "    BAHRA_SAP_PASSWORD=..."
        )

    asyncio.run(run(
        input_file=args.file,
        output_dir=args.output,
        username=username,
        password=password,
        headless=args.headless,
    ))


if __name__ == "__main__":
    main()

"""
Helper module to unprotect Excel files (both .xls and .xlsx formats).
This module removes password protection and worksheet/workbook protection from Excel files.
"""

import os
import zipfile
import tempfile
import shutil
import re as _re
from xml.etree import ElementTree as ET

try:
    import win32com.client as win32  # pip install pywin32
except Exception:
    win32 = None


# Register all Office OpenXML namespaces so ElementTree preserves them during rewrite.
# Without this, ET replaces prefixes with ns0/ns1/etc which can break Excel formatting.
_OOXML_NAMESPACES = {
    '': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
    'x14ac': 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac',
    'x16r2': 'http://schemas.microsoft.com/office/spreadsheetml/2017/richdata2',
    'xr': 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision',
    'xr2': 'http://schemas.microsoft.com/office/spreadsheetml/2015/revision2',
    'xr3': 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision3',
    'xr6': 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision6',
    'xr10': 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision10',
    'x15': 'http://schemas.microsoft.com/office/spreadsheetml/2010/11/main',
    'x14': 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/main',
}

for _prefix, _uri in _OOXML_NAMESPACES.items():
    ET.register_namespace(_prefix, _uri)


def detect_real_type(path: str) -> str:
    """
    Detect the actual file type by reading file headers.
    
    Args:
        path: Path to the file to check
        
    Returns:
        String indicating file type: 'xls', 'xlsx', 'zip', 'html', 'xml', 'csv', or 'unknown'
    """
    try:
        with open(path, 'rb') as f:
            header = f.read(8)
        if header[:4] == b'\xD0\xCF\x11\xE0':
            return 'xls'
        if header[:4] == b'PK\x03\x04':
            return 'xlsx'
        if header[:5] == b'<?xml':
            return 'xml'
        with open(path, 'rb') as f:
            chunk = f.read(2048).lower()
            if b'<html' in chunk or b'<table' in chunk:
                return 'html'
        # Naive CSV check
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            first = f.readline()
            if ',' in first or '\t' in first:
                return 'csv'
    except Exception:
        pass
    return 'unknown'


def _remove_protection_from_xml(xml_path: str, tag_local_name: str) -> bool:
    """
    Remove a protection element from an XML file using regex-based approach.
    This avoids ElementTree rewriting the entire XML (which mangles namespaces
    and breaks formatting/conditional formatting/etc).

    Returns True if the file was modified, False otherwise.
    """
    try:
        with open(xml_path, 'rb') as f:
            content = f.read()

        original = content
        # Match the protection tag with any namespace prefix, e.g.:
        #   <sheetProtection .../>  or  <x:sheetProtection .../>
        #   <sheetProtection ...>...</sheetProtection>
        # Handles both self-closing and paired tags
        pattern_self_close = _re.compile(
            rb'<[a-zA-Z0-9]*:?' + tag_local_name.encode() + rb'\b[^>]*/\s*>',
            _re.DOTALL
        )
        pattern_paired = _re.compile(
            rb'<([a-zA-Z0-9]*:?' + tag_local_name.encode() + rb')\b[^>]*>.*?</\1\s*>',
            _re.DOTALL
        )

        content = pattern_self_close.sub(b'', content)
        content = pattern_paired.sub(b'', content)

        if content != original:
            with open(xml_path, 'wb') as f:
                f.write(content)
            return True
        return False
    except Exception:
        return False


def unprotect_xlsx(xlsx_path: str, out_path: str) -> str:
    """
    Remove protection from .xlsx file by modifying its XML structure.
    Uses regex-based XML editing to preserve all formatting and namespaces.
    Only modifies files that actually contain protection elements.

    Args:
        xlsx_path: Path to the protected .xlsx file
        out_path: Path where the unprotected file will be saved

    Returns:
        Path to the unprotected file
    """
    tmpdir = tempfile.mkdtemp()
    try:
        # Extract the .xlsx (which is a ZIP archive)
        with zipfile.ZipFile(xlsx_path, 'r') as zf:
            zf.extractall(tmpdir)

        # Remove worksheet protections (only touch files that have protection)
        ws_dir = os.path.join(tmpdir, 'xl', 'worksheets')
        if os.path.isdir(ws_dir):
            for name in os.listdir(ws_dir):
                if name.endswith('.xml'):
                    p = os.path.join(ws_dir, name)
                    _remove_protection_from_xml(p, 'sheetProtection')

        # Remove workbook structure protection
        wb_xml = os.path.join(tmpdir, 'xl', 'workbook.xml')
        if os.path.exists(wb_xml):
            _remove_protection_from_xml(wb_xml, 'workbookProtection')

        # Repackage as .xlsx, preserving [Content_Types].xml as first entry
        with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Write [Content_Types].xml first (Excel expects it early in the archive)
            ct_path = os.path.join(tmpdir, '[Content_Types].xml')
            if os.path.exists(ct_path):
                zf.write(ct_path, '[Content_Types].xml')

            for folder, _, files in os.walk(tmpdir):
                for f in files:
                    full = os.path.join(folder, f)
                    arc = os.path.relpath(full, tmpdir).replace('\\', '/')
                    # Skip [Content_Types].xml since we already wrote it
                    if arc == '[Content_Types].xml':
                        continue
                    zf.write(full, arc)
        return out_path
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def unprotect_xls_via_com(xls_path: str, out_path: str) -> str:
    """
    Remove protection from .xls file using Excel COM automation.
    Requires Microsoft Excel to be installed.

    Args:
        xls_path: Path to the protected .xls file
        out_path: Path where the unprotected file will be saved

    Returns:
        Path to the unprotected file

    Raises:
        RuntimeError: If Excel/pywin32 is not available
    """
    if win32 is None:
        raise RuntimeError("Excel/pywin32 required to handle .xls files")

    excel = win32.DispatchEx('Excel.Application')
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(xls_path, False, False)
        # Try blank passwords; if protected with a password, this may fail silently
        try:
            wb.Unprotect()
        except Exception:
            pass
        try:
            for ws in wb.Worksheets:
                try:
                    ws.Unprotect()
                except Exception:
                    pass
        except Exception:
            pass
        wb.SaveAs(out_path, FileFormat=56)  # 56 = .xls format
        wb.Close(SaveChanges=False)
        return out_path
    finally:
        excel.Quit()


def _xlrd_colour_to_hex(colour_map, colour_index):
    """Convert xlrd colour index to hex string for openpyxl."""
    if colour_index is None or colour_index in (64, 0x7FFF):
        return None
    try:
        rgb = colour_map.get(colour_index)
        if rgb and rgb != (0, 0, 0):
            return "{:02X}{:02X}{:02X}".format(*rgb)
    except Exception:
        pass
    return None


def _copy_xls_formatting(rb, rs, ws, row_idx, col_idx, new_cell):
    """Copy formatting from xlrd cell to openpyxl cell."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
    try:
        xf_index = rs.cell_xf_index(row_idx, col_idx)
        xf = rb.xf_list[xf_index]
    except (IndexError, AttributeError):
        return

    colour_map = rb.colour_map

    # Font
    try:
        font_obj = rb.font_list[xf.font_index]
        font_kwargs = {}
        font_kwargs['name'] = font_obj.name
        font_kwargs['bold'] = font_obj.bold
        font_kwargs['italic'] = font_obj.italic
        if font_obj.height:
            font_kwargs['size'] = font_obj.height / 20.0
        if font_obj.underline_type:
            font_kwargs['underline'] = 'single'
        if font_obj.struck_out:
            font_kwargs['strike'] = True
        fg_hex = _xlrd_colour_to_hex(colour_map, font_obj.colour_index)
        if fg_hex:
            font_kwargs['color'] = fg_hex
        new_cell.font = Font(**font_kwargs)
    except Exception:
        pass

    # Background / fill
    try:
        bg = xf.background
        pattern_index = bg.pattern_colour_index
        bg_hex = _xlrd_colour_to_hex(colour_map, pattern_index)
        if bg_hex:
            new_cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
    except Exception:
        pass

    # Alignment
    try:
        align = xf.alignment
        hor_map = {0: 'general', 1: 'left', 2: 'center', 3: 'right', 4: 'fill',
                   5: 'justify', 6: 'centerContinuous', 7: 'distributed'}
        vert_map = {0: 'top', 1: 'center', 2: 'bottom', 3: 'justify', 4: 'distributed'}
        new_cell.alignment = Alignment(
            horizontal=hor_map.get(align.hor_align, 'general'),
            vertical=vert_map.get(align.vert_align, 'bottom'),
            wrap_text=bool(align.text_wrapped),
        )
    except Exception:
        pass

    # Number format
    try:
        fmt_key = xf.format_key
        fmt_str = rb.format_map.get(fmt_key)
        if fmt_str and hasattr(fmt_str, 'format_str') and fmt_str.format_str:
            new_cell.number_format = fmt_str.format_str
    except Exception:
        pass

    # Borders
    try:
        border = xf.border
        style_map = {0: None, 1: 'thin', 2: 'medium', 3: 'dashed', 4: 'dotted',
                     5: 'thick', 6: 'double', 7: 'hair'}

        def _make_side(line_style, colour_index):
            s = style_map.get(line_style)
            if not s:
                return Side()
            c = _xlrd_colour_to_hex(colour_map, colour_index)
            return Side(style=s, color=c) if c else Side(style=s)

        new_cell.border = Border(
            left=_make_side(border.left_line_style, border.left_colour_index),
            right=_make_side(border.right_line_style, border.right_colour_index),
            top=_make_side(border.top_line_style, border.top_colour_index),
            bottom=_make_side(border.bottom_line_style, border.bottom_colour_index),
        )
    except Exception:
        pass


def unprotect_xls_via_conversion(xls_path: str, out_path: str) -> str:
    """
    Fallback: Convert .xls to .xlsx without protection using xlrd + openpyxl.
    Used when pywin32/Microsoft Excel is not available.
    Preserves formatting (fonts, fills, borders, alignment, number formats,
    column widths, row heights, merged cells).

    Args:
        xls_path: Path to the protected .xls file
        out_path: Path where the unprotected file will be saved (will be .xlsx)

    Returns:
        Path to the unprotected .xlsx file
    """
    import xlrd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    # formatting_info=True enables reading fonts/fills/borders/etc from .xls
    try:
        rb = xlrd.open_workbook(xls_path, formatting_info=True)
        has_formatting = True
    except Exception:
        # Some xlrd versions or corrupt files may not support formatting_info
        rb = xlrd.open_workbook(xls_path, formatting_info=False)
        has_formatting = False

    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    for sheet_name in rb.sheet_names():
        rs = rb.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        for row_idx in range(rs.nrows):
            for col_idx in range(rs.ncols):
                cell = rs.cell(row_idx, col_idx)
                value = cell.value

                # Convert xlrd date floats to Python datetime
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, rb.datemode)
                    except Exception:
                        pass
                # Convert xlrd boolean
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(value)

                new_cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

                # Copy cell formatting if available
                if has_formatting:
                    _copy_xls_formatting(rb, rs, ws, row_idx, col_idx, new_cell)

        # Copy column widths
        if hasattr(rs, 'colinfo_map'):
            for col_idx, colinfo in rs.colinfo_map.items():
                col_letter = get_column_letter(col_idx + 1)
                # xlrd width is in 1/256th of character width; openpyxl uses character units
                ws.column_dimensions[col_letter].width = colinfo.width / 256.0

        # Copy row heights
        if hasattr(rs, 'rowinfo_map'):
            for row_idx, rowinfo in rs.rowinfo_map.items():
                if rowinfo.height:
                    # xlrd height is in twips (1/20th of a point)
                    ws.row_dimensions[row_idx + 1].height = rowinfo.height / 20.0

        # Copy merged cells
        for merged_range in rs.merged_cells:
            rlo, rhi, clo, chi = merged_range
            ws.merge_cells(
                start_row=rlo + 1, start_column=clo + 1,
                end_row=rhi, end_column=chi
            )

    # Ensure output path has .xlsx extension
    base, ext = os.path.splitext(out_path)
    if ext.lower() == '.xls':
        out_path = base + '.xlsx'

    wb.save(out_path)
    print(f"Converted .xls to .xlsx (no protection, formatting preserved): {out_path}")
    return out_path


def unprotect_xls_keep_xls(xls_path: str, out_path: str) -> str:
    """
    Remove protection from .xls file.
    Tries COM automation first, falls back to xlrd+openpyxl conversion.

    Args:
        xls_path: Path to the protected .xls file
        out_path: Path where the unprotected file will be saved

    Returns:
        Path to the unprotected file
    """
    # Try COM automation first (best quality, preserves formatting)
    if win32 is not None:
        try:
            return unprotect_xls_via_com(xls_path, out_path)
        except Exception as e:
            print(f"COM automation failed: {e}, falling back to conversion")

    # Fallback: convert .xls -> .xlsx using xlrd + openpyxl
    return unprotect_xls_via_conversion(xls_path, out_path)


def detect_real_type_from_bytes(data: bytes) -> str:
    """Detect the actual file type from raw bytes."""
    if data[:4] == b'\xD0\xCF\x11\xE0':
        return 'xls'
    if data[:4] == b'PK\x03\x04':
        return 'xlsx'
    if data[:5] == b'<?xml':
        return 'xml'
    chunk = data[:2048].lower()
    if b'<html' in chunk or b'<table' in chunk:
        return 'html'
    return 'unknown'


def unprotect_xlsx_bytes(data: bytes) -> bytes:
    """Remove protection from .xlsx bytes in memory. Returns unprotected .xlsx bytes.
    Uses regex-based XML editing to preserve all formatting and namespaces."""
    from io import BytesIO
    tmpdir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(BytesIO(data), 'r') as zf:
            zf.extractall(tmpdir)

        # Remove worksheet protections using regex (preserves formatting)
        ws_dir = os.path.join(tmpdir, 'xl', 'worksheets')
        if os.path.isdir(ws_dir):
            for name in os.listdir(ws_dir):
                if name.endswith('.xml'):
                    p = os.path.join(ws_dir, name)
                    _remove_protection_from_xml(p, 'sheetProtection')

        # Remove workbook structure protection
        wb_xml = os.path.join(tmpdir, 'xl', 'workbook.xml')
        if os.path.exists(wb_xml):
            _remove_protection_from_xml(wb_xml, 'workbookProtection')

        out_buf = BytesIO()
        with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Write [Content_Types].xml first
            ct_path = os.path.join(tmpdir, '[Content_Types].xml')
            if os.path.exists(ct_path):
                zf.write(ct_path, '[Content_Types].xml')

            for folder, _, files in os.walk(tmpdir):
                for f in files:
                    full = os.path.join(folder, f)
                    arc = os.path.relpath(full, tmpdir).replace('\\', '/')
                    if arc == '[Content_Types].xml':
                        continue
                    zf.write(full, arc)
        return out_buf.getvalue()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def unprotect_xls_bytes_to_xlsx(data: bytes) -> bytes:
    """Convert .xls bytes to unprotected .xlsx bytes using xlrd + openpyxl (in memory).
    Preserves formatting (fonts, fills, borders, alignment, number formats,
    column widths, row heights, merged cells)."""
    import xlrd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # Try with formatting_info=True first
    try:
        rb = xlrd.open_workbook(file_contents=data, formatting_info=True)
        has_formatting = True
    except Exception:
        rb = xlrd.open_workbook(file_contents=data, formatting_info=False)
        has_formatting = False

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in rb.sheet_names():
        rs = rb.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        for row_idx in range(rs.nrows):
            for col_idx in range(rs.ncols):
                cell = rs.cell(row_idx, col_idx)
                value = cell.value

                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, rb.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(value)

                new_cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

                # Copy cell formatting if available
                if has_formatting:
                    _copy_xls_formatting(rb, rs, ws, row_idx, col_idx, new_cell)

        # Copy column widths
        if hasattr(rs, 'colinfo_map'):
            for col_idx, colinfo in rs.colinfo_map.items():
                col_letter = get_column_letter(col_idx + 1)
                ws.column_dimensions[col_letter].width = colinfo.width / 256.0

        # Copy row heights
        if hasattr(rs, 'rowinfo_map'):
            for row_idx, rowinfo in rs.rowinfo_map.items():
                if rowinfo.height:
                    ws.row_dimensions[row_idx + 1].height = rowinfo.height / 20.0

        for merged_range in rs.merged_cells:
            rlo, rhi, clo, chi = merged_range
            ws.merge_cells(
                start_row=rlo + 1, start_column=clo + 1,
                end_row=rhi, end_column=chi
            )

    out_buf = BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue()


def unprotect_excel_bytes(data: bytes, filename: str = "") -> tuple[bytes, str]:
    """
    Unprotect Excel file from raw bytes (in memory, no disk I/O).

    Args:
        data: Raw Excel file bytes
        filename: Original filename (used for extension hint)

    Returns:
        Tuple of (unprotected_bytes, output_filename)
    """
    real_type = detect_real_type_from_bytes(data)
    base, ext = os.path.splitext(filename)

    if real_type == 'xlsx':
        out_bytes = unprotect_xlsx_bytes(data)
        return out_bytes, f"{base}_unprotected.xlsx"
    elif real_type == 'xls':
        out_bytes = unprotect_xls_bytes_to_xlsx(data)
        return out_bytes, f"{base}_unprotected.xlsx"
    else:
        # Return as-is if not a recognized Excel format
        return data, filename


def unprotect_excel_file(file_path: str, output_path: str = None) -> str:
    """
    Unprotect an Excel file (auto-detects format and applies appropriate method).
    
    Args:
        file_path: Path to the Excel file to unprotect
        output_path: Optional path for the output file. If None, creates a file with 
                    '_unprotected' suffix in the same directory
        
    Returns:
        Path to the unprotected file, or None if file couldn't be processed
        
    Raises:
        FileNotFoundError: If the input file doesn't exist
        RuntimeError: If Excel is required but not available
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    base, ext = os.path.splitext(os.path.basename(file_path))
    
    # Check if already unprotected
    if base.endswith('_unprotected'):
        print(f"File already unprotected: {file_path}")
        return file_path
    
    # Determine output path
    if output_path is None:
        output_path = os.path.join(os.path.dirname(file_path), f"{base}_unprotected{ext}")
    
    # Detect actual file type
    real_type = detect_real_type(file_path)
    
    if real_type == 'xlsx':
        return unprotect_xlsx(file_path, output_path)
    elif real_type == 'xls':
        return unprotect_xls_keep_xls(file_path, output_path)
    else:
        print(f"Skip non-Excel or unsupported file ({real_type}): {file_path}")
        return None


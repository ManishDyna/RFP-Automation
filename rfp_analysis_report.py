"""
RFP Analysis Report Generator

Cross-references Portal RFPs with Dataverse database and local file system.
Generates an Excel report with:
  - RFP Company Name
  - RFP Name / ID
  - End Date
  - Publish Date
  - Participant Status in Portal
  - Participant Status in Dataverse
  - File Exist in Local (Yes/No)
  - Local Files Path

Usage:
    python rfp_analysis_report.py
"""

import sys
import os
import io
import re
import logging
from datetime import datetime

# Fix Windows console encoding
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from bs4 import BeautifulSoup

from config.config import OUTPUT_DIR
from helpers.core_helper import (
    get_rfp_activity_data_from_db,
    clean_rfp_title,
    normalize_filename,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

PORTAL_RFPS_DIR = os.path.join(OUTPUT_DIR, "Portal-Rfps")

# Company name mapping: portal file name -> local folder name / Dataverse name
COMPANY_MAP = {
    "Saudi Electricity Company": {
        "local_folder": "Saudi Electricity Company",
        "dataverse_name": "Saudi Electricity Company",
    },
    "Aramco E-Marketplace": {
        "local_folder": "Aramco e-Marketplace",
        "dataverse_name": "Aramco e-Marketplace",
    },
}


def normalize_rfp_id(text: str) -> str:
    """Normalize RFP title/ID for matching by removing extra spaces and lowercasing."""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', text).strip().lower()


def extract_contract_number(title: str) -> str:
    """Extract the contract number (e.g., C001738316) from an RFP title."""
    if not title:
        return ""
    match = re.search(r'[CR]\d{7,10}', title)
    if match:
        return match.group(0)
    # Try pure numeric IDs (Aramco style)
    match = re.search(r'\d{10}', title)
    if match:
        return match.group(0)
    return ""


def parse_sec_portal_file(filepath: str) -> list[dict]:
    """
    Parse Saudi Electricity Company portal .xls file (HTML format).
    Returns list of dicts with: company, title, id, end_time, event_type, participated, portal_status
    """
    logger.info(f"Parsing SEC portal file: {filepath}")
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    soup = BeautifulSoup(content, 'lxml')
    tables = soup.find_all('table')

    if not tables:
        logger.warning("No tables found in SEC portal file")
        return []

    table = tables[0]
    all_cells = []
    for row in table.find_all('tr'):
        for cell in row.find_all(['td', 'th']):
            text = cell.get_text(strip=True)
            colspan = int(cell.get('colspan', 1))
            all_cells.append({'text': text, 'colspan': colspan})

    current_status = 'Unknown'
    rfps = []
    data_cells = []
    # Texts that belong to the header row and should be skipped
    header_texts = {'Title', 'ID', 'End Time', 'Event Type', 'Participated'}

    for cell in all_cells:
        # Status header: colspan >= 5 and matches pattern
        if cell['colspan'] >= 5 and re.match(r'Status:\s*\w+', cell['text']):
            current_status = cell['text']
            data_cells = []
            continue

        # Skip the combined header cell
        if cell['text'].startswith('TitleID'):
            continue

        # Skip individual header texts when they appear as standalone cells
        if cell['colspan'] == 1 and cell['text'] in header_texts:
            continue

        if cell['colspan'] == 1 and cell['text']:
            data_cells.append(cell['text'])
            if len(data_cells) == 5:
                rfps.append({
                    'company': 'Saudi Electricity Company',
                    'title': data_cells[0],
                    'doc_id': data_cells[1],
                    'end_time': data_cells[2],
                    'event_type': data_cells[3],
                    'portal_participated': data_cells[4],
                    'portal_status': current_status,
                })
                data_cells = []

    logger.info(f"Parsed {len(rfps)} RFPs from SEC portal file")
    return rfps


def parse_aramco_portal_file(filepath: str) -> list[dict]:
    """
    Parse Aramco E-Marketplace portal .xls file.
    This file uses frames referencing external files which may not exist.
    Falls back to trying different parsing strategies.
    """
    logger.info(f"Parsing Aramco portal file: {filepath}")

    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    # Check if it's a frameset (references external files)
    if 'frameset' in content.lower() and 'frame src=' in content.lower():
        # Try to find the referenced data file
        soup = BeautifulSoup(content, 'lxml')
        frames = soup.find_all('frame')
        for frame in frames:
            src = frame.get('src', '')
            if src and 'sheet' in src.lower():
                # Try to resolve the path relative to the portal file
                base_dir = os.path.dirname(filepath)
                # URL-decode the src
                from urllib.parse import unquote
                decoded_src = unquote(src)
                ref_path = os.path.join(base_dir, decoded_src)

                if os.path.exists(ref_path):
                    logger.info(f"Found referenced sheet file: {ref_path}")
                    return _parse_aramco_html_table(ref_path)
                else:
                    logger.warning(
                        f"Aramco portal file references '{decoded_src}' but file not found at '{ref_path}'. "
                        f"Please place the data files folder next to the .xls file."
                    )

        logger.warning("Aramco portal file is a frameset with missing data files. Skipping.")
        return []

    # If not a frameset, try direct HTML table parsing
    return _parse_aramco_html_table(filepath)


def _parse_aramco_html_table(filepath: str) -> list[dict]:
    """Parse Aramco data from an HTML table file."""
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    soup = BeautifulSoup(content, 'lxml')
    tables = soup.find_all('table')

    if not tables:
        logger.warning("No tables found in Aramco file")
        return []

    rfps = []
    for table in tables:
        rows = table.find_all('tr')
        if len(rows) < 2:
            continue

        # Try to find header row
        header_row = rows[0]
        headers = [cell.get_text(strip=True).lower() for cell in header_row.find_all(['th', 'td'])]

        if 'title' not in headers and 'name' not in headers:
            continue

        for row in rows[1:]:
            cells = [cell.get_text(strip=True) for cell in row.find_all(['td'])]
            if len(cells) >= 4:
                rfps.append({
                    'company': 'Aramco e-Marketplace',
                    'title': cells[0] if len(cells) > 0 else '',
                    'doc_id': cells[1] if len(cells) > 1 else '',
                    'end_time': cells[2] if len(cells) > 2 else '',
                    'event_type': cells[3] if len(cells) > 3 else 'RFP',
                    'portal_participated': cells[4] if len(cells) > 4 else '',
                    'portal_status': '',
                })

    logger.info(f"Parsed {len(rfps)} RFPs from Aramco portal file")
    return rfps


def load_portal_rfps() -> list[dict]:
    """Load all portal RFPs from XLS files in Portal-Rfps folder."""
    all_rfps = []

    if not os.path.exists(PORTAL_RFPS_DIR):
        logger.error(f"Portal-Rfps directory not found: {PORTAL_RFPS_DIR}")
        return all_rfps

    for filename in os.listdir(PORTAL_RFPS_DIR):
        if not filename.lower().endswith('.xls'):
            continue

        filepath = os.path.join(PORTAL_RFPS_DIR, filename)
        base_name = os.path.splitext(filename)[0]

        if 'saudi' in base_name.lower() or 'sec' in base_name.lower():
            rfps = parse_sec_portal_file(filepath)
        elif 'aramco' in base_name.lower():
            rfps = parse_aramco_portal_file(filepath)
        else:
            logger.warning(f"Unknown portal file: {filename}, attempting generic parse")
            rfps = parse_sec_portal_file(filepath)  # try SEC-style parse

        all_rfps.extend(rfps)

    logger.info(f"Total portal RFPs loaded: {len(all_rfps)}")
    return all_rfps


def load_dataverse_rfps() -> dict:
    """
    Load RFP data from Dataverse and index by normalized RFP_ID.
    Returns dict: normalized_rfp_id -> row dict
    """
    logger.info("Fetching RFP data from Dataverse...")
    rows = get_rfp_activity_data_from_db()
    logger.info(f"Fetched {len(rows)} RFP records from Dataverse")

    indexed = {}
    for row in rows:
        rfp_id = row.get('RFP_ID', '')
        if rfp_id:
            key = normalize_rfp_id(rfp_id)
            indexed[key] = row

    return indexed


def check_local_files(rfp_title: str, company_name: str) -> tuple[bool, str]:
    """
    Check if RFP files exist locally for the given RFP title and company.
    Returns (exists: bool, path: str)
    """
    local_folder_name = COMPANY_MAP.get(company_name, {}).get('local_folder', company_name)
    clean_title = clean_rfp_title(rfp_title)

    # Check the expected folder path
    rfp_folder = os.path.join(OUTPUT_DIR, local_folder_name, clean_title)
    downloaded_rfp_dir = os.path.join(rfp_folder, "downloaded-rfp")

    if os.path.exists(downloaded_rfp_dir):
        # Check for Excel files in downloaded-rfp
        for f in os.listdir(downloaded_rfp_dir):
            if f.lower().endswith(('.xls', '.xlsx')):
                return True, os.path.join(downloaded_rfp_dir, f)

    if os.path.exists(rfp_folder):
        # Check for Excel files directly in RFP folder
        for f in os.listdir(rfp_folder):
            if f.lower().endswith(('.xls', '.xlsx')):
                return True, os.path.join(rfp_folder, f)
        # Folder exists but no Excel files
        return False, rfp_folder

    # Try fuzzy match on local folders - the titles may have different spacing
    company_dir = os.path.join(OUTPUT_DIR, local_folder_name)
    if os.path.exists(company_dir):
        normalized_title = normalize_filename(clean_title)
        for folder_name in os.listdir(company_dir):
            if normalize_filename(folder_name) == normalized_title:
                folder_path = os.path.join(company_dir, folder_name)
                dl_path = os.path.join(folder_path, "downloaded-rfp")
                if os.path.exists(dl_path):
                    for f in os.listdir(dl_path):
                        if f.lower().endswith(('.xls', '.xlsx')):
                            return True, os.path.join(dl_path, f)
                return False, folder_path

    return False, ""


def build_report() -> pd.DataFrame:
    """Build the cross-reference analysis report."""
    # Step 1: Load portal RFPs
    portal_rfps = load_portal_rfps()
    if not portal_rfps:
        logger.error("No portal RFPs loaded. Check the Portal-Rfps folder.")
        return pd.DataFrame()

    # Step 2: Load Dataverse data
    dataverse_index = load_dataverse_rfps()

    # Step 3: Cross-reference and build report rows
    report_rows = []

    for rfp in portal_rfps:
        company = rfp['company']
        title = rfp['title']
        portal_participated = rfp['portal_participated']
        end_time = rfp['end_time']
        portal_status = rfp.get('portal_status', '')

        # Look up in Dataverse by normalized title
        norm_title = normalize_rfp_id(title)
        dv_row = dataverse_index.get(norm_title)

        # If not found by exact normalized match, try fuzzy matching
        if not dv_row:
            contract_num = extract_contract_number(title)
            if contract_num:
                for dv_key, dv_val in dataverse_index.items():
                    dv_rfp_id = dv_val.get('RFP_ID', '')
                    if contract_num in dv_rfp_id:
                        dv_row = dv_val
                        break

        # Extract Dataverse fields
        dv_participated = ''
        publish_date = ''
        if dv_row:
            dv_participated = dv_row.get('participated', '') or ''
            publish_date = dv_row.get('publish_time', '') or ''
            # Format publish_time if it's ISO format
            if publish_date and 'T' in str(publish_date):
                try:
                    dt = datetime.fromisoformat(publish_date.replace('Z', '+00:00'))
                    publish_date = dt.strftime('%m/%d/%Y %I:%M %p')
                except (ValueError, AttributeError):
                    pass

        # Check local files
        file_exists, local_path = check_local_files(title, company)

        report_rows.append({
            'Company Name': company,
            'RFP Name / Title': title,
            'End Date': end_time,
            'Publish Date': publish_date,
            'Portal Status Group': portal_status,
            'Participant Status (Portal)': portal_participated,
            'Participant Status (Dataverse)': dv_participated,
            'Found in Dataverse': 'Yes' if dv_row else 'No',
            'File Exist in Local': 'Yes' if file_exists else 'No',
            'Local File Path': local_path,
        })

    df = pd.DataFrame(report_rows)
    return df


def print_summary(df: pd.DataFrame):
    """Print summary statistics of the report."""
    if df.empty:
        print("\nNo data to report.")
        return

    total = len(df)
    in_dataverse = len(df[df['Found in Dataverse'] == 'Yes'])
    with_local_files = len(df[df['File Exist in Local'] == 'Yes'])

    print(f"\n{'=' * 70}")
    print(f"  RFP Cross-Reference Analysis Report")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'=' * 70}")
    print(f"\n  Total Portal RFPs           : {total}")
    print(f"  Found in Dataverse          : {in_dataverse} ({in_dataverse/total*100:.1f}%)")
    print(f"  Not in Dataverse            : {total - in_dataverse} ({(total-in_dataverse)/total*100:.1f}%)")
    print(f"  With Local Files            : {with_local_files} ({with_local_files/total*100:.1f}%)")
    print(f"  Without Local Files         : {total - with_local_files} ({(total-with_local_files)/total*100:.1f}%)")

    # Per company
    print(f"\n  By Company:")
    for company in df['Company Name'].unique():
        c_df = df[df['Company Name'] == company]
        c_total = len(c_df)
        c_dv = len(c_df[c_df['Found in Dataverse'] == 'Yes'])
        c_local = len(c_df[c_df['File Exist in Local'] == 'Yes'])
        print(f"    {company}: {c_total} total, {c_dv} in Dataverse, {c_local} with local files")

    # Portal participation breakdown
    print(f"\n  Portal Participation Status:")
    for status, count in df['Participant Status (Portal)'].value_counts().items():
        print(f"    {status}: {count}")

    # Portal status group breakdown
    print(f"\n  Portal Status Groups:")
    for status, count in df['Portal Status Group'].value_counts().items():
        print(f"    {status}: {count}")

    # Dataverse participation breakdown
    print(f"\n  Dataverse Participation Status:")
    dv_found = df[df['Found in Dataverse'] == 'Yes']
    if not dv_found.empty:
        for status, count in dv_found['Participant Status (Dataverse)'].value_counts().items():
            label = status if status else '(empty)'
            print(f"    {label}: {count}")
    else:
        print(f"    (no records found in Dataverse)")

    # Mismatches between portal and dataverse participation
    both = df[(df['Found in Dataverse'] == 'Yes')]
    if not both.empty:
        mismatches = both[
            both['Participant Status (Portal)'].str.lower() != both['Participant Status (Dataverse)'].str.lower()
        ]
        print(f"\n  Status Mismatches (Portal vs Dataverse): {len(mismatches)}")

    print(f"{'=' * 70}\n")


def main():
    output_filename = f"RFP_Analysis_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(os.getcwd(), output_filename)

    logger.info("Starting RFP Analysis Report generation...")

    df = build_report()

    if df.empty:
        logger.error("Report is empty. No data to export.")
        return

    # Print summary
    print_summary(df)

    # Export to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='RFP Analysis')

        # Auto-adjust column widths
        worksheet = writer.sheets['RFP Analysis']
        for col_idx, column in enumerate(df.columns, 1):
            max_length = max(
                len(str(column)),
                df[column].astype(str).map(len).max() if len(df) > 0 else 0
            )
            # Cap at 60 chars width
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = adjusted_width

        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Freeze top row
        worksheet.freeze_panes = 'A2'

        # Add conditional formatting for File Exist column
        from openpyxl.styles import Font as OpFont
        green_font = OpFont(color='006100')
        red_font = OpFont(color='9C0006')

        file_exist_col_idx = list(df.columns).index('File Exist in Local') + 1
        dv_found_col_idx = list(df.columns).index('Found in Dataverse') + 1

        for row_idx in range(2, len(df) + 2):
            # Color File Exist column
            cell = worksheet.cell(row=row_idx, column=file_exist_col_idx)
            if cell.value == 'Yes':
                cell.font = green_font
            elif cell.value == 'No':
                cell.font = red_font

            # Color Found in Dataverse column
            cell = worksheet.cell(row=row_idx, column=dv_found_col_idx)
            if cell.value == 'Yes':
                cell.font = green_font
            elif cell.value == 'No':
                cell.font = red_font

    logger.info(f"Report saved to: {output_path}")
    print(f"\nReport saved: {output_path}")


if __name__ == "__main__":
    main()

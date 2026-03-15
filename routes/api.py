"""
JSON API endpoints for React frontend.
All routes are prefixed with /api
"""

from fastapi import APIRouter, Request, HTTPException, Query, Depends
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from services.user_service import (
    authenticate_user, list_users, get_user, create_user, update_user, delete_user, get_user_by_email
)
from services.dashboard_service import (
    get_dashboard_data_cached, get_all_rfp_data_cached, get_logs_data_cached,
    get_material_insights_cached, get_material_insights_grouped_cached
)
from services.sap_service import create_sap_password_record, list_sap_password_records_cached
from services.dynamic_role_service import get_user_permissions
from services.audit_service import log_event, AuditAction, AuditCategory
from services.user_lifecycle_service import (
    is_account_locked, is_user_active, record_failed_login, clear_failed_attempts,
    update_last_login, get_or_create_user_status, activate_user, deactivate_user,
    unlock_user, validate_password_strength, update_password_changed, get_user_status,
)
from middleware.auth import get_current_user, require_permission, get_request_ip
from services.system_settings_service import get_setting
import time
import hmac
import hashlib
import base64
import json
import os
import glob
from collections import defaultdict
import threading

router = APIRouter(prefix="/api", tags=["API"])


# ==================== RATE LIMITING ====================
# Simple in-memory rate limiter for login attempts
_FAILED_ATTEMPTS = defaultdict(list)
_RATE_LIMIT_LOCK = threading.Lock()
LOCKOUT_THRESHOLD = 5  # Max failed attempts before lockout
LOCKOUT_DURATION = 300  # Lockout duration in seconds (5 minutes)
ATTEMPT_WINDOW = 300  # Time window to track attempts (5 minutes)


def _get_client_ip(request: Request) -> str:
    """Get client IP from request, handling proxies."""
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _check_rate_limit(identifier: str) -> tuple[bool, int]:
    """
    Check if identifier (email or IP) is rate limited.
    Returns (is_locked, seconds_remaining).
    """
    now = time.time()
    with _RATE_LIMIT_LOCK:
        # Clean old attempts outside the window
        _FAILED_ATTEMPTS[identifier] = [
            t for t in _FAILED_ATTEMPTS[identifier]
            if now - t < ATTEMPT_WINDOW
        ]
        attempts = _FAILED_ATTEMPTS[identifier]

        if len(attempts) >= LOCKOUT_THRESHOLD:
            # Check if still in lockout period
            oldest_in_window = min(attempts) if attempts else now
            lockout_end = oldest_in_window + LOCKOUT_DURATION
            if now < lockout_end:
                return True, int(lockout_end - now)

        return False, 0


def _record_failed_attempt(identifier: str):
    """Record a failed login attempt."""
    with _RATE_LIMIT_LOCK:
        _FAILED_ATTEMPTS[identifier].append(time.time())


def _clear_failed_attempts(identifier: str):
    """Clear failed attempts after successful login."""
    with _RATE_LIMIT_LOCK:
        _FAILED_ATTEMPTS.pop(identifier, None)


# ==================== AUTH ENDPOINTS ====================

@router.post("/login")
async def api_login(request: Request):
    """Login endpoint for React frontend with rate limiting protection."""
    data = await request.json()
    email = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "").strip()
    if not email or not password:
        raise HTTPException(status_code=400, detail="Email and password are required")

    # Check rate limiting by email
    is_locked, seconds_remaining = _check_rate_limit(email)
    if is_locked:
        raise HTTPException(
            status_code=429,
            detail=f"Too many failed attempts. Try again in {seconds_remaining} seconds."
        )

    # Also check by IP to prevent distributed attacks
    client_ip = _get_client_ip(request)
    ip_locked, ip_seconds = _check_rate_limit(f"ip:{client_ip}")
    if ip_locked:
        raise HTTPException(
            status_code=429,
            detail=f"Too many failed attempts from this IP. Try again in {ip_seconds} seconds."
        )

    user = authenticate_user(email=email, password=password)
    if not user:
        # Record failed attempt for both email and IP
        _record_failed_attempt(email)
        _record_failed_attempt(f"ip:{client_ip}")

        # Audit: failed login
        log_event(
            action=AuditAction.LOGIN_FAILED,
            category=AuditCategory.AUTH,
            actor_email=email,
            target_type="Session",
            details=json.dumps({"reason": "Invalid credentials"}),
            ip_address=client_ip,
        )
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Check if account is locked (persistent lockout in Dataverse)
    user_id = user.get("record_id", "")
    locked, minutes_remaining = is_account_locked(user_id)
    if locked:
        log_event(
            action=AuditAction.LOGIN_FAILED,
            category=AuditCategory.AUTH,
            actor_email=email,
            target_type="Session",
            details=json.dumps({"reason": "Account locked", "minutes_remaining": minutes_remaining}),
            ip_address=client_ip,
        )
        raise HTTPException(
            status_code=423,
            detail=f"Account is locked. Try again in {minutes_remaining} minutes."
        )

    # Check if account is active
    if not is_user_active(user_id):
        log_event(
            action=AuditAction.LOGIN_FAILED,
            category=AuditCategory.AUTH,
            actor_email=email,
            target_type="Session",
            details=json.dumps({"reason": "Account deactivated"}),
            ip_address=client_ip,
        )
        raise HTTPException(status_code=403, detail="Your account has been deactivated. Please contact an administrator.")

    # Clear failed attempts on successful login
    _clear_failed_attempts(email)
    _clear_failed_attempts(f"ip:{client_ip}")
    clear_failed_attempts(user_id)
    update_last_login(user_id)

    # Load user permissions from dynamic RBAC
    user["permissions"] = get_user_permissions(user)

    request.session["user"] = user
    request.session["last_activity"] = int(time.time())

    # Audit: successful login
    log_event(
        action=AuditAction.LOGIN,
        category=AuditCategory.AUTH,
        actor_email=email,
        actor_name=user.get("name", ""),
        target_type="Session",
        ip_address=client_ip,
    )

    return JSONResponse({"ok": True, "redirect": "/dashboard"})


@router.post("/logout")
async def api_logout(request: Request):
    """Logout endpoint"""
    user = request.session.get("user")
    if user:
        log_event(
            action=AuditAction.LOGOUT,
            category=AuditCategory.AUTH,
            actor_email=user.get("email", ""),
            actor_name=user.get("name", ""),
            target_type="Session",
            ip_address=get_request_ip(request),
        )
    request.session.clear()
    return JSONResponse({"ok": True, "redirect": "/login"})


@router.get("/session/status")
async def api_session_status(request: Request):
    """Check if session is valid - returns 'valid' field for React frontend"""
    if not request.session.get("user"):
        return JSONResponse(status_code=401, content={"valid": False, "message": "No active session"})

    user = request.session.get("user")

    # Ensure permissions are always present in session
    if "permissions" not in user:
        try:
            user["permissions"] = get_user_permissions(user)
            request.session["user"] = user
        except Exception as e:
            print(f"[SESSION] Failed to load permissions: {e}")
            user["permissions"] = []

    return JSONResponse({
        "valid": True,
        "user": user,
        "last_activity": request.session.get("last_activity", int(time.time()))
    })


@router.post("/session/refresh")
async def api_refresh_session(request: Request):
    """Refresh the current session"""
    if not request.session.get("user"):
        return JSONResponse(status_code=401, content={"ok": False, "message": "No active session"})

    user = request.session.get("user")
    request.session["user"] = user
    request.session["last_activity"] = int(time.time())
    return JSONResponse({"ok": True, "message": "Session refreshed"})


def _sign_token(secret: str, payload: dict, ttl_seconds: int = 1800) -> str:
    data = dict(payload or {})
    data["exp"] = int(time.time()) + ttl_seconds
    raw = json.dumps(data, separators=(",", ":"))
    sig = hmac.new(secret.encode(), raw.encode(), hashlib.sha256).digest()
    return base64.urlsafe_b64encode(raw.encode()).decode().rstrip("=") + "." + base64.urlsafe_b64encode(sig).decode().rstrip("=")


@router.post("/forgot")
async def api_forgot(request: Request):
    """Forgot password endpoint"""
    body = await request.json()
    email = (body.get("email") or "").strip()
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")

    users = get_user_by_email(email) or []
    if not users:
        raise HTTPException(status_code=404, detail="Email not found")

    secret = request.app.state.__dict__.get("secret_key", "change-me-please")
    token = _sign_token(secret, {"email": email}, ttl_seconds=1800)
    base_url = str(request.base_url).rstrip("/")
    reset_link = f"{base_url}/reset-password?token={token}"

    payload = {
        "to": email,
        "subject": "Reset your password",
        "isHtml": True,
        "body": f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin:0;padding:0;background-color:#f4f6f9;font-family:'Segoe UI',Arial,Helvetica,sans-serif;">
    <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%" style="background-color:#f4f6f9;">
        <tr>
            <td align="center" style="padding:40px 20px;">
                <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="600" style="max-width:600px;background-color:#ffffff;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,0.08);">
                    <!-- Header -->
                    <tr>
                        <td style="background-color:#4f46e5;padding:30px 40px;border-radius:12px 12px 0 0;text-align:center;">
                            <h1 style="margin:0;font-size:24px;color:#ffffff;font-weight:600;">Bahra E-Bidding</h1>
                        </td>
                    </tr>
                    <!-- Body -->
                    <tr>
                        <td style="padding:40px;">
                            <h2 style="margin:0 0 16px 0;font-size:22px;color:#1a1a2e;font-weight:600;">Reset Your Password</h2>
                            <p style="margin:0 0 24px 0;font-size:15px;color:#555555;line-height:1.6;">
                                We received a request to reset your password for your Bahra E-Bidding account. Click the button below to set a new password.
                            </p>
                            <!-- Button -->
                            <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%">
                                <tr>
                                    <td align="center" style="padding:8px 0 32px 0;">
                                        <a href="{reset_link}" target="_blank"
                                           style="background-color:#4f46e5;color:#ffffff;text-decoration:none;padding:14px 36px;border-radius:8px;font-size:16px;font-weight:600;display:inline-block;letter-spacing:0.5px;">
                                            Reset Password
                                        </a>
                                    </td>
                                </tr>
                            </table>
                            <!-- Link fallback -->
                            <p style="margin:0 0 8px 0;font-size:13px;color:#888888;">If the button above doesn't work, copy and paste the link below into your browser:</p>
                            <p style="margin:0 0 24px 0;word-break:break-all;font-size:13px;">
                                <a href="{reset_link}" target="_blank" style="color:#4f46e5;text-decoration:underline;">{reset_link}</a>
                            </p>
                            <!-- Divider -->
                            <hr style="border:none;border-top:1px solid #eeeeee;margin:24px 0;">
                            <p style="margin:0 0 8px 0;font-size:13px;color:#999999;line-height:1.5;">
                                This link will expire in <strong>30 minutes</strong>. If you didn't request a password reset, you can safely ignore this email.
                            </p>
                        </td>
                    </tr>
                    <!-- Footer -->
                    <tr>
                        <td style="background-color:#f9fafb;padding:24px 40px;border-radius:0 0 12px 12px;text-align:center;">
                            <p style="margin:0;font-size:13px;color:#888888;">
                                Thanks,<br><strong>Bahra E-Bidding Automation Team</strong>
                            </p>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>"""
    }
    import requests
    resp = requests.post(get_setting("FORGOT_PASSWORD_FLOW_URL", ""), json=payload)
    if not (200 <= resp.status_code < 300):
        raise HTTPException(status_code=502, detail=f"Flow error: {resp.status_code}")
    return JSONResponse({"ok": True})


def _verify_token(secret: str, token: str) -> dict:
    try:
        raw_b64, sig_b64 = token.split(".", 1)
        raw = base64.urlsafe_b64decode(raw_b64 + "==")
        sig = base64.urlsafe_b64decode(sig_b64 + "==")
        good = hmac.new(secret.encode(), raw, hashlib.sha256).digest()
        if not hmac.compare_digest(sig, good):
            raise ValueError("bad signature")
        data = json.loads(raw.decode())
        if int(data.get("exp", 0)) < int(time.time()):
            raise ValueError("expired")
        return data
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or expired token")


@router.post("/reset-password")
async def api_reset_password(request: Request):
    """Reset password endpoint"""
    body = await request.json()
    token = (body.get("token") or "").strip()
    new_password = (body.get("password") or "")
    if not token:
        raise HTTPException(status_code=400, detail="Token is required")
    if not new_password:
        raise HTTPException(status_code=400, detail="Password is required")

    secret = request.app.state.__dict__.get("secret_key", "change-me-please")
    data = _verify_token(secret, token)
    email = (data.get("email") or "").strip()
    users = get_user_by_email(email) or []
    if not users:
        raise HTTPException(status_code=404, detail="User not found")
    record_id = users[0].get("record_id")
    ok = update_user(record_id, {"password": str(new_password)})
    if not ok:
        raise HTTPException(status_code=500, detail="Failed to update password")

    # Track password change timestamp
    update_password_changed(record_id)

    # Audit log
    log_event(
        action=AuditAction.PASSWORD_RESET,
        category=AuditCategory.AUTH,
        actor_email=email,
        target_type="User",
        target_id=record_id,
        ip_address=get_request_ip(request),
    )

    return JSONResponse({"ok": True})


# ==================== DASHBOARD ENDPOINTS ====================

@router.get("/dashboard/data")
async def api_dashboard_data(request: Request, refresh: int = Query(0), user: dict = Depends(require_permission("dashboard.view"))):
    """Get dashboard data as JSON"""
    data = get_dashboard_data_cached(force_refresh=bool(refresh))
    return JSONResponse(data)


def _get_filtered_rfp_rows(status="downloaded", search="", start_date="", end_date="",
                           company="", material_match="", keyword_match="",
                           participation="", refresh=0):
    """Shared helper: load RFP data, normalize statuses, apply filters.
    Returns (detailed_rows, filtered_rows)."""
    from datetime import datetime

    def _parse_date(s):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    def _parse_end_datetime(date_str):
        if not date_str:
            return None
        s = str(date_str).strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S",
                     "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s[:len(fmt) + 5].split("+")[0].split("Z")[0].strip(), fmt)
            except (ValueError, TypeError):
                continue
        try:
            from dateutil import parser as du_parser
            dt = du_parser.parse(s)
            return dt.replace(tzinfo=None)
        except Exception:
            pass
        return None

    def _normalize_participation(raw_status, rfp_end_date_str):
        value = (raw_status or "").strip().lower()
        if value in ("submitted", "yes"):
            return "submitted"
        if value == "declined":
            return "declined"
        if value == "saved_draft":
            return "saved_draft"
        if value in ("", "no", "open", "not participated"):
            end_dt = _parse_end_datetime(rfp_end_date_str)
            if end_dt and end_dt < datetime.now():
                return "not_participant"
            return "open"
        return "other"

    all_rfp_data = get_all_rfp_data_cached(force_refresh=bool(refresh))
    downloaded_rows = all_rfp_data.get("downloaded_rfps") or []

    detailed_rows = []
    for row in downloaded_rows:
        status_key = _normalize_participation(row.get("participated"), row.get("RFP_End_Date", ""))
        detailed_rows.append({**row, "status_key": status_key})

    # Filter by status
    selected_filter = (status or "downloaded").lower()
    filtered_rows = detailed_rows
    if selected_filter != "downloaded":
        filtered_rows = [r for r in filtered_rows if r["status_key"] == selected_filter]

    # Filter by company
    if company:
        filtered_rows = [r for r in filtered_rows if (r.get("Company_Name") or "") == company]

    # Filter by search
    if search:
        query = search.lower()
        filtered_rows = [
            r for r in filtered_rows
            if query in (r.get("RFP_ID") or "").lower()
            or query in (r.get("Company_Name") or "").lower()
            or query in (r.get("Owner_Name") or "").lower()
        ]

    # Filter by date range
    start_dt = _parse_date(start_date) if start_date else None
    end_dt = _parse_date(end_date) if end_date else None
    if start_dt or end_dt:
        def within_range(row):
            end_str = row.get("RFP_End_Date")
            if not end_str:
                return False
            parsed = _parse_date(end_str[:10]) if len(end_str) >= 10 else None
            if not parsed:
                return False
            if start_dt and parsed.date() < start_dt.date():
                return False
            if end_dt and parsed.date() > end_dt.date():
                return False
            return True
        filtered_rows = [r for r in filtered_rows if within_range(r)]

    # Filter by material matching
    if material_match:
        if material_match.lower() == "matched":
            filtered_rows = [r for r in filtered_rows if (r.get("Material_Matched") or "").lower() == "yes"]
        elif material_match.lower() == "not_matched":
            filtered_rows = [r for r in filtered_rows if (r.get("Material_Matched") or "").lower() in ("no", "")]

    # Filter by keyword matching
    if keyword_match:
        if keyword_match.lower() == "matched":
            filtered_rows = [r for r in filtered_rows if (r.get("Keyword_Matched") or "").lower() == "yes"]
        elif keyword_match.lower() == "not_matched":
            filtered_rows = [r for r in filtered_rows if (r.get("Keyword_Matched") or "").lower() in ("no", "")]

    # Filter by participation status
    if participation:
        participation_lower = participation.lower()
        if participation_lower == "participated":
            filtered_rows = [r for r in filtered_rows if r["status_key"] == "submitted"]
        elif participation_lower == "not_participated":
            filtered_rows = [r for r in filtered_rows if r["status_key"] in ("open", "not_participant")]
        elif participation_lower == "declined":
            filtered_rows = [r for r in filtered_rows if r["status_key"] == "declined"]

    return detailed_rows, filtered_rows


@router.get("/dashboard/rfp-details")
async def api_rfp_details(
    request: Request,
    status: str = Query("downloaded"),
    search: str = Query(""),
    start_date: str = Query(""),
    end_date: str = Query(""),
    company: str = Query(""),
    material_match: str = Query(""),  # New: "matched" or "not_matched"
    keyword_match: str = Query(""),   # New: "matched" or "not_matched"
    participation: str = Query(""),   # New: "participated", "not_participated", "declined"
    limit: int = Query(50),           # New: Number of records per page
    offset: int = Query(0),           # New: Starting position
    refresh: int = Query(0),
    user: dict = Depends(require_permission("rfp.view")),
):
    """Get RFP details as JSON with pagination and new filters"""
    detailed_rows, filtered_rows = _get_filtered_rfp_rows(
        status=status, search=search, start_date=start_date, end_date=end_date,
        company=company, material_match=material_match, keyword_match=keyword_match,
        participation=participation, refresh=refresh,
    )

    # Apply pagination
    total_filtered = len(filtered_rows)
    paginated_rows = filtered_rows[offset:offset + limit]

    # Count statuses from FILTERED rows (not all detailed_rows)
    from collections import Counter
    status_counts = Counter(r["status_key"] for r in filtered_rows)

    # Also provide total counts (unfiltered) for reference
    total_status_counts = Counter(r["status_key"] for r in detailed_rows)
    total_status_counts["downloaded"] = len(detailed_rows)

    # Get unique companies for filter dropdown
    unique_companies = list(set(r.get("Company_Name") or "" for r in detailed_rows if r.get("Company_Name")))
    unique_companies.sort()

    return JSONResponse({
        "rfps": paginated_rows,
        "status_counts": dict(status_counts),  # Filtered counts
        "total_status_counts": dict(total_status_counts),  # Total counts (unfiltered)
        "total": len(detailed_rows),
        "total_rows": len(detailed_rows),
        "total_filtered": total_filtered,
        "shown_rows": len(paginated_rows),
        "offset": offset,
        "limit": limit,
        "has_more": offset + limit < total_filtered,
        "unique_companies": unique_companies
    })


@router.get("/dashboard/rfp-details/export")
async def api_rfp_details_export(
    request: Request,
    format: str = Query("csv"),
    status: str = Query("downloaded"),
    search: str = Query(""),
    start_date: str = Query(""),
    end_date: str = Query(""),
    company: str = Query(""),
    material_match: str = Query(""),
    keyword_match: str = Query(""),
    participation: str = Query(""),
    refresh: int = Query(0),
    user: dict = Depends(require_permission("rfp.view")),
):
    """Export filtered RFP data as CSV or Excel"""
    import csv
    import io

    _, filtered_rows = _get_filtered_rfp_rows(
        status=status, search=search, start_date=start_date, end_date=end_date,
        company=company, material_match=material_match, keyword_match=keyword_match,
        participation=participation, refresh=refresh,
    )

    # Define export columns
    headers = ["RFP ID", "Company", "Owner", "Published", "Deadline",
               "Status", "Participation", "Material Match", "Keyword Match", "Portal Link"]

    def _participation_label(status_key):
        return {"submitted": "Participated", "declined": "Declined",
                "not_participant": "Not Participant", "open": "Open"}.get(status_key, status_key or "")

    def _match_label(val):
        return "Yes" if (val or "").lower() == "yes" else "No"

    def _row_to_export(r):
        return [
            r.get("RFP_ID", ""),
            r.get("Company_Name", ""),
            r.get("Owner_Name", ""),
            r.get("Publish_Time", ""),
            r.get("RFP_End_Date", ""),
            (r.get("status_key", "") or "").replace("_", " ").title(),
            _participation_label(r.get("status_key", "")),
            _match_label(r.get("Material_Matched", "")),
            _match_label(r.get("Keyword_Matched", "")),
            r.get("Link", ""),
        ]

    if format.lower() == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFP Data"

        # Style header row
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, row_data in enumerate(filtered_rows, 2):
            for col_idx, value in enumerate(_row_to_export(row_data), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row_idx in range(2, len(filtered_rows) + 2):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=RFP_Data_Export.xlsx"},
        )

    # Default: CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row_data in filtered_rows:
        writer.writerow(_row_to_export(row_data))

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=RFP_Data_Export.csv"},
    )


@router.get("/dashboard/material-insights")
async def api_material_insights(
    request: Request,
    rfp_id: str = Query(""),
    company: str = Query(""),
    material_match: str = Query(""),
    keyword_match: str = Query(""),
    participated: str = Query(""),
    search: str = Query(""),
    limit: int = Query(50),
    offset: int = Query(0),
    refresh: int = Query(0),
    user: dict = Depends(require_permission("material_insights.view")),
):
    """Get material insights data from bahra_rfps table with filtering and pagination."""
    data = get_material_insights_cached(force_refresh=bool(refresh))
    materials = data.get("materials", [])

    # Apply filters
    filtered = materials

    if company:
        filtered = [m for m in filtered if m["company"] == company]

    if rfp_id:
        filtered = [m for m in filtered if m["rfp_id"] == rfp_id]

    if material_match:
        if material_match.lower() == "yes":
            filtered = [m for m in filtered if m["material_matched"].lower() == "yes"]
        elif material_match.lower() == "no":
            filtered = [m for m in filtered if m["material_matched"].lower() != "yes"]

    if keyword_match:
        if keyword_match.lower() == "yes":
            filtered = [m for m in filtered if m["keyword_matched"].lower() == "yes"]
        elif keyword_match.lower() == "no":
            filtered = [m for m in filtered if m["keyword_matched"].lower() != "yes"]

    if participated:
        p = participated.lower()
        if p == "submitted":
            filtered = [m for m in filtered if m.get("participated", "").lower() in ("submitted", "yes")]
        elif p == "declined":
            filtered = [m for m in filtered if m.get("participated", "").lower() == "declined"]
        elif p == "open":
            filtered = [m for m in filtered if m.get("participated", "").lower() not in ("submitted", "yes", "declined")]

    if search:
        q = search.lower()
        filtered = [
            m for m in filtered
            if q in (m.get("rfp_id") or "").lower()
            or q in (m.get("company") or "").lower()
        ]

    total_filtered = len(filtered)
    paginated = filtered[offset:offset + limit]

    return JSONResponse({
        "materials": paginated,
        "stats": data.get("stats", {}),
        "item_stats": data.get("item_stats", {}),
        "unique_rfps": data.get("unique_rfps", {}),
        "total_filtered": total_filtered,
        "total": len(materials),
        "offset": offset,
        "limit": limit,
        "has_more": offset + limit < total_filtered,
    })


@router.get("/dashboard/material-insights-grouped")
async def api_material_insights_grouped(
    request: Request,
    tab: str = Query("materials"),
    company: str = Query(""),
    search: str = Query(""),
    participated: str = Query(""),
    limit: int = Query(50),
    offset: int = Query(0),
    refresh: int = Query(0),
    user: dict = Depends(require_permission("material_insights.view")),
):
    """Get material insights grouped by material code or keyword."""
    data = get_material_insights_grouped_cached(force_refresh=bool(refresh))

    if tab == "keywords":
        items = data.get("keywords", [])
    else:
        items = data.get("materials", [])

    filtered = items

    if company:
        filtered = [item for item in filtered if company in item.get("companies", [])]

    if participated:
        p = participated.lower()
        if p == "submitted":
            filtered = [
                item for item in filtered
                if any(r.get("participated", "") in ("submitted", "yes") for r in item.get("rfps", []))
            ]
        elif p == "declined":
            filtered = [
                item for item in filtered
                if any(r.get("participated", "") == "declined" for r in item.get("rfps", []))
            ]
        elif p == "open":
            filtered = [
                item for item in filtered
                if any(r.get("participated", "") not in ("submitted", "yes", "declined") for r in item.get("rfps", []))
            ]

    if search:
        q = search.lower()
        if tab == "keywords":
            filtered = [item for item in filtered if q in item.get("keyword", "").lower()]
        else:
            filtered = [
                item for item in filtered
                if q in item.get("material_code", "").lower()
                or q in item.get("material_description", "").lower()
            ]

    has_filters = bool(company or participated or search)
    total_filtered = len(filtered)
    paginated = filtered[offset:offset + limit]

    all_companies = set()
    for item in items:
        all_companies.update(item.get("companies", []))

    # Recompute stats and charts from filtered data when filters are active
    if has_filters:
        all_rfp_ids = set()
        submitted_count = 0
        for item in filtered:
            for r in item.get("rfps", []):
                all_rfp_ids.add(r.get("rfp_id", ""))
                if r.get("participated", "") in ("submitted", "yes"):
                    submitted_count += 1

        # Get the other tab's data for cross-tab stats
        other_items = data.get("keywords", []) if tab != "keywords" else data.get("materials", [])

        filtered_stats = {
            "total_unique_materials": total_filtered if tab == "materials" else len(other_items),
            "total_unique_keywords": total_filtered if tab == "keywords" else len(other_items),
            "total_rfps_with_matches": len(all_rfp_ids),
            "total_material_rfp_links": sum(item.get("rfp_count", 0) for item in (filtered if tab == "materials" else other_items)),
            "total_keyword_rfp_links": sum(item.get("rfp_count", 0) for item in (filtered if tab == "keywords" else other_items)),
            "submitted_rfp_count": submitted_count,
        }

        # Recompute chart data from filtered items
        if tab == "materials":
            filtered_top_chart = [
                {"material": m.get("material_code", ""), "description": (m.get("material_description", "") or "")[:40], "rfp_count": m.get("rfp_count", 0)}
                for m in filtered[:10]
            ]
            filtered_keyword_chart = data.get("keyword_chart", [])
        else:
            filtered_top_chart = data.get("top_materials_chart", [])
            filtered_keyword_chart = [
                {"keyword": k.get("keyword", ""), "rfp_count": k.get("rfp_count", 0)}
                for k in filtered
            ]

        response_stats = filtered_stats
        response_top_chart = filtered_top_chart
        response_keyword_chart = filtered_keyword_chart
    else:
        response_stats = data.get("stats", {})
        response_top_chart = data.get("top_materials_chart", [])
        response_keyword_chart = data.get("keyword_chart", [])

    return JSONResponse({
        "items": paginated,
        "stats": response_stats,
        "top_materials_chart": response_top_chart,
        "keyword_chart": response_keyword_chart,
        "total_filtered": total_filtered,
        "total": len(items),
        "offset": offset,
        "limit": limit,
        "has_more": offset + limit < total_filtered,
        "unique_companies": sorted(list(all_companies)),
    })


@router.get("/dashboard/view-logs")
async def api_view_logs(request: Request, page: int = Query(1), page_size: int = Query(50), force_refresh: bool = Query(False), user: dict = Depends(require_permission("logs.view"))):
    """Get automation logs as JSON, grouped by run_id and paginated by runs."""

    # get_logs_data_cached returns a list directly, not a dict
    logs_list = get_logs_data_cached(force_refresh=force_refresh)
    if not isinstance(logs_list, list):
        logs_list = []

    # Map field names to what frontend expects
    mapped_logs = []
    for log in logs_list:
        mapped_logs.append({
            "id": log.get("RunID") or log.get("id"),
            "run_id": log.get("RunID", "-"),
            "event_time": log.get("formatted_timestamp") or log.get("Timestamp", "-"),
            "event_type": log.get("Category", "-"),
            "rfp_id": log.get("RFP_ID", "-"),
            "action": log.get("Action", "-"),
            "status": log.get("automation_status", "-"),
            "details": log.get("Message", "-"),
        })

    # Group logs by run_id so pagination works on runs, not individual entries
    from collections import OrderedDict
    run_groups: dict[str, list] = OrderedDict()
    for log in mapped_logs:
        key = log.get("run_id", "-")
        if key not in run_groups:
            run_groups[key] = []
        run_groups[key].append(log)

    total_runs = len(run_groups)
    all_run_ids = list(run_groups.keys())

    # Paginate by runs
    start = (page - 1) * page_size
    end = start + page_size
    paginated_run_ids = all_run_ids[start:end]

    # Collect all log entries for the paginated runs
    paginated_logs = []
    for run_id in paginated_run_ids:
        paginated_logs.extend(run_groups[run_id])

    return JSONResponse({
        "logs": paginated_logs,
        "total": len(mapped_logs),
        "total_runs": total_runs,
        "page": page,
        "page_size": page_size,
    })


# ==================== ERROR FILE ENDPOINTS ====================

@router.get("/error-files/list")
async def api_list_error_files(request: Request, rfp_id: str = Query(None)):
    """List error log files from the LOGS directory, optionally filtered by RFP ID.
    Scans both top-level files and subdirectories (which may contain screenshot.png)."""
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    failure_logs_dir = get_setting("FAILURE_LOGS_DIR", "")
    if not os.path.isdir(failure_logs_dir):
        return JSONResponse({"files": []})

    def _matches_rfp(name: str) -> bool:
        if not rfp_id:
            return True
        safe_rfp = rfp_id.replace(" ", "_").replace("/", "_").replace("\\", "_")
        return safe_rfp.lower() in name.lower() or rfp_id.lower() in name.lower()

    files = []

    for entry in os.listdir(failure_logs_dir):
        entry_path = os.path.join(failure_logs_dir, entry)

        if os.path.isfile(entry_path):
            # Top-level files (json, txt, png)
            if not entry.endswith((".json", ".txt", ".png")):
                continue
            if not _matches_rfp(entry):
                continue
            stat = os.stat(entry_path)
            files.append({
                "filename": entry,
                "size": stat.st_size,
                "modified": stat.st_mtime,
                "type": "screenshot" if entry.endswith(".png") else
                        "report" if entry.endswith(".txt") else "json",
            })

        elif os.path.isdir(entry_path):
            # Subdirectories (contain json, txt, screenshot.png)
            if not _matches_rfp(entry):
                continue
            for sub_file in os.listdir(entry_path):
                sub_path = os.path.join(entry_path, sub_file)
                if not os.path.isfile(sub_path):
                    continue
                if not sub_file.endswith((".json", ".txt", ".png")):
                    continue
                stat = os.stat(sub_path)
                # Use folder/file as the path so we can serve it
                relative_name = f"{entry}/{sub_file}"
                files.append({
                    "filename": relative_name,
                    "size": stat.st_size,
                    "modified": stat.st_mtime,
                    "type": "screenshot" if sub_file.endswith(".png") else
                            "report" if sub_file.endswith(".txt") else "json",
                })

    # Sort by modified time descending
    files.sort(key=lambda f: f["modified"], reverse=True)
    return JSONResponse({"files": files})


def _resolve_log_file_path(filename: str) -> str | None:
    """Resolve a filename (may include one subfolder) to an absolute path inside FAILURE_LOGS_DIR.
    Returns None if the resolved path is outside FAILURE_LOGS_DIR (path traversal)."""
    failure_logs_dir = get_setting("FAILURE_LOGS_DIR", "")
    # Allow at most one subfolder: "subfolder/file.ext" or just "file.ext"
    parts = filename.replace("\\", "/").split("/")
    if len(parts) > 2:
        return None
    # Rebuild safely
    safe_parts = [os.path.basename(p) for p in parts]
    fpath = os.path.join(failure_logs_dir, *safe_parts)
    # Verify it's inside FAILURE_LOGS_DIR
    if not os.path.normpath(fpath).startswith(os.path.normpath(failure_logs_dir)):
        return None
    return fpath


@router.get("/error-files/content/{filename:path}")
async def api_get_error_file_content(request: Request, filename: str):
    """Get content of a text/json error file from LOGS directory (supports subfolder/file)"""
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    fpath = _resolve_log_file_path(filename)
    if not fpath or not os.path.isfile(fpath):
        raise HTTPException(status_code=404, detail="File not found")

    if fpath.endswith(".json"):
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)
        return JSONResponse({"filename": filename, "type": "json", "content": data})
    elif fpath.endswith(".txt"):
        with open(fpath, "r", encoding="utf-8") as f:
            text = f.read()
        return JSONResponse({"filename": filename, "type": "text", "content": text})
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")


@router.get("/error-files/screenshot/{filename:path}")
async def api_get_screenshot(request: Request, filename: str):
    """Serve a screenshot image from LOGS directory (supports subfolder/screenshot.png)"""
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    fpath = _resolve_log_file_path(filename)
    if not fpath or not os.path.isfile(fpath) or not fpath.endswith(".png"):
        raise HTTPException(status_code=404, detail="Screenshot not found")

    return FileResponse(fpath, media_type="image/png")


# ==================== PROFILE ENDPOINTS ====================

@router.get("/profile")
async def api_get_profile(request: Request):
    """Get current user profile"""
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    user = request.session.get("user")
    return JSONResponse({
        "name": user.get("name"),
        "email": user.get("email"),
        "mobile": user.get("mobile") or user.get("mobile_number"),
        "role": user.get("role"),
        "record_id": user.get("record_id")
    })


@router.post("/profile/update")
async def api_update_profile(request: Request):
    """Update user profile"""
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    user = request.session.get("user")
    body = await request.json()

    updates = {}
    session_updates = {}
    if "name" in body:
        updates["name"] = body["name"]
        session_updates["name"] = body["name"]
    if "mobile" in body:
        # Dataverse uses mobile_number, but frontend uses mobile
        updates["mobile_number"] = body["mobile"]
        session_updates["mobile"] = body["mobile"]

    if updates:
        ok = update_user(user.get("record_id"), updates)
        if ok:
            # Update session
            for k, v in session_updates.items():
                user[k] = v
            request.session["user"] = user
            return JSONResponse({"ok": True})
        raise HTTPException(status_code=500, detail="Failed to update profile")

    return JSONResponse({"ok": True})


@router.post("/profile/change-password")
async def api_change_password(request: Request):
    """Change user password"""
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    user = request.session.get("user")
    body = await request.json()
    current_password = body.get("current_password")
    new_password = body.get("new_password")

    if not current_password or not new_password:
        raise HTTPException(status_code=400, detail="Current and new password required")

    # Validate password strength
    is_valid, error_msg = validate_password_strength(new_password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)

    # Verify current password
    auth_user = authenticate_user(user.get("email"), current_password)
    if not auth_user:
        raise HTTPException(status_code=401, detail="Current password is incorrect")

    ok = update_user(user.get("record_id"), {"password": new_password})
    if not ok:
        raise HTTPException(status_code=500, detail="Failed to change password")

    # Track password change timestamp
    update_password_changed(user.get("record_id", ""))

    # Audit log
    log_event(
        action=AuditAction.PASSWORD_CHANGED,
        category=AuditCategory.AUTH,
        actor_email=user.get("email", ""),
        actor_name=user.get("name", ""),
        target_type="User",
        target_id=user.get("record_id", ""),
        ip_address=get_request_ip(request),
    )

    return JSONResponse({"ok": True})


# ==================== SAP PASSWORD ENDPOINTS ====================

@router.post("/sap/change-password")
async def api_sap_change_password(request: Request, user: dict = Depends(require_permission("sap_password.change"))):
    """Change SAP password"""
    body = await request.json()
    username = body.get("username")
    password = body.get("password")

    if not username or not password:
        raise HTTPException(status_code=400, detail="Username and password required")
    ok = create_sap_password_record(
        password=password,
        user_email=user.get("email"),
        username=username
    )
    if not ok:
        raise HTTPException(status_code=500, detail="Failed to save SAP password")

    return JSONResponse({"ok": True})


@router.get("/dashboard/sap-password-logs")
async def api_sap_password_logs(request: Request, user: dict = Depends(require_permission("sap_password.view"))):
    """Get SAP password change logs"""

    logs = list_sap_password_records_cached(force_refresh=False)
    return JSONResponse({"logs": logs})


# ==================== USER MANAGEMENT ENDPOINTS ====================

@router.get("/users/user-list")
async def api_user_list(request: Request, refresh: int = Query(0), user: dict = Depends(require_permission("user_management.view"))):
    """Get list of users as JSON"""
    users = list_users(top=200)
    return JSONResponse({"users": users})


@router.post("/users/create")
async def api_create_user(request: Request, user: dict = Depends(require_permission("user_management.create"))):
    """Create a new user"""

    body = await request.json()
    ok = create_user(body)
    if not ok:
        raise HTTPException(status_code=400, detail="Failed to create user")

    # Audit log
    log_event(
        action=AuditAction.USER_CREATED,
        category=AuditCategory.USER,
        actor_email=user.get("email", ""),
        actor_name=user.get("name", ""),
        target_type="User",
        target_id=body.get("email", ""),
        details=json.dumps({"name": body.get("name"), "email": body.get("email"), "role": body.get("role")}),
        ip_address=get_request_ip(request),
    )

    return JSONResponse({"ok": True})


@router.put("/users/update/{record_id}")
async def api_update_user(request: Request, record_id: str, user: dict = Depends(require_permission("user_management.edit"))):
    """Update a user"""

    body = await request.json()
    ok = update_user(record_id, body)
    if not ok:
        raise HTTPException(status_code=400, detail="Failed to update user")

    # Audit log
    log_event(
        action=AuditAction.USER_UPDATED,
        category=AuditCategory.USER,
        actor_email=user.get("email", ""),
        actor_name=user.get("name", ""),
        target_type="User",
        target_id=record_id,
        details=json.dumps({"updated_fields": list(body.keys())}),
        ip_address=get_request_ip(request),
    )

    return JSONResponse({"ok": True})


@router.delete("/users/delete/{record_id}")
async def api_delete_user(request: Request, record_id: str, user: dict = Depends(require_permission("user_management.delete"))):
    """Delete a user"""

    ok = delete_user(record_id)
    if not ok:
        raise HTTPException(status_code=400, detail="Failed to delete user")

    # Audit log
    log_event(
        action=AuditAction.USER_DELETED,
        category=AuditCategory.USER,
        actor_email=user.get("email", ""),
        actor_name=user.get("name", ""),
        target_type="User",
        target_id=record_id,
        ip_address=get_request_ip(request),
    )

    return JSONResponse({"ok": True})


# ==================== RFP VALIDATION ENDPOINTS ====================

@router.get("/validate-rfp")
async def api_validate_rfp(request: Request, rfp_id: str = Query(...)):
    """Validate if an RFP exists in the database and return its company name."""
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    rfp_id = (rfp_id or "").strip()
    if not rfp_id:
        raise HTTPException(status_code=400, detail="RFP ID is required")

    from services.dashboard_service import get_all_rfp_data_cached

    all_rfp_data = get_all_rfp_data_cached(force_refresh=False)
    downloaded_rfps = all_rfp_data.get("downloaded_rfps") or []

    # Search for the RFP by ID (case-insensitive)
    rfp_id_lower = rfp_id.lower()
    matched_rfp = None
    for rfp in downloaded_rfps:
        if (rfp.get("RFP_ID") or "").lower() == rfp_id_lower:
            matched_rfp = rfp
            break

    if not matched_rfp:
        raise HTTPException(
            status_code=404,
            detail="This RFP cannot be submitted because it was not found in the database. Please download it first and then try to submit."
        )

    return JSONResponse({
        "ok": True,
        "rfp_id": matched_rfp.get("RFP_ID", ""),
        "company": matched_rfp.get("Company_Name", ""),
        "status": matched_rfp.get("participated", ""),
    })


# ==================== SCHEDULE ENDPOINTS ====================

@router.post("/schedule/save")
async def api_save_schedule(request: Request, user: dict = Depends(require_permission("schedule_automation.manage"))):
    """Save automation schedule"""

    from helpers.core_helper import DATAVERSE
    from services.system_settings_service import get_setting as _get_setting

    AUTOMATION_SCHEDULE_TABLE_API = _get_setting("AUTOMATION_SCHEDULE_TABLE_API", "")
    AUTOMATION_SCHEDULE_TABLE_LOGICAL = _get_setting("AUTOMATION_SCHEDULE_TABLE_LOGICAL", "")

    body = await request.json()

    # Get frequency maps for conversion
    def _get_frequency_maps():
        try:
            maps = DATAVERSE.get_choice_options(
                AUTOMATION_SCHEDULE_TABLE_LOGICAL,
                "cr673_frequency",
            )
            return maps
        except Exception:
            label_to_value = {
                "Month": 415300000,
                "Week": 415300001,
                "Day": 415300002,
                "Hour": 415300003,
                "Minute": 415300004,
                "Year": 415300005,
            }
            return {
                "label_to_value": label_to_value,
                "value_to_label": {v: k for k, v in label_to_value.items()}
            }

    try:
        freq_label = body.get("frequency")
        maps = _get_frequency_maps()
        freq_value = maps["label_to_value"].get(freq_label) if isinstance(freq_label, str) else freq_label
        if freq_value is None:
            raise HTTPException(status_code=400, detail="Unsupported frequency. Use Month, Week, Day, Hour, Minute, or Year.")

        # Get start_time and validate - empty string should be treated as None
        start_time = body.get("start_time")
        if isinstance(start_time, str):
            start_time = start_time.strip() or None

        data = {
            "job_name": body.get("job_name") or "RFP Automation",
            "interval": int(body.get("interval")) if body.get("interval") else 1,
            "frequency": int(freq_value),
            "timezone": body.get("timezone"),
            "start_time": start_time,
            "max_concurrency": int(body.get("max_concurrency") or 1),
            "notes": body.get("notes"),
            "is_active": bool(body.get("is_active", True)),
            "created_by": user.get("email"),
        }

        # Filter out None and empty string values
        data = {k: v for k, v in data.items() if v is not None and v != ""}

        ok = DATAVERSE.insert_row(
            table_api_name=AUTOMATION_SCHEDULE_TABLE_API,
            data=data,
            table_logical_name=AUTOMATION_SCHEDULE_TABLE_LOGICAL,
        )
        if not ok:
            raise HTTPException(status_code=500, detail="Failed to save schedule")

        return JSONResponse({"ok": True})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== USER LIFECYCLE ENDPOINTS ====================

@router.post("/users/{record_id}/activate")
async def api_activate_user(request: Request, record_id: str, user: dict = Depends(require_permission("user_management.activate"))):
    """Activate a user account."""

    ok = activate_user(record_id, admin_email=user.get("email", ""))
    if not ok:
        raise HTTPException(status_code=500, detail="Failed to activate user")

    log_event(
        action=AuditAction.USER_ACTIVATED,
        category=AuditCategory.USER,
        actor_email=user.get("email", ""),
        actor_name=user.get("name", ""),
        target_type="User",
        target_id=record_id,
        ip_address=get_request_ip(request),
    )
    return JSONResponse({"ok": True, "message": "User activated"})


@router.post("/users/{record_id}/deactivate")
async def api_deactivate_user(request: Request, record_id: str, user: dict = Depends(require_permission("user_management.activate"))):
    """Deactivate a user account."""

    ok = deactivate_user(record_id, admin_email=user.get("email", ""))
    if not ok:
        raise HTTPException(status_code=500, detail="Failed to deactivate user")

    log_event(
        action=AuditAction.USER_DEACTIVATED,
        category=AuditCategory.USER,
        actor_email=user.get("email", ""),
        actor_name=user.get("name", ""),
        target_type="User",
        target_id=record_id,
        ip_address=get_request_ip(request),
    )
    return JSONResponse({"ok": True, "message": "User deactivated"})


@router.post("/users/{record_id}/unlock")
async def api_unlock_user(request: Request, record_id: str, user: dict = Depends(require_permission("user_management.activate"))):
    """Unlock a locked user account."""

    ok = unlock_user(record_id)
    if not ok:
        raise HTTPException(status_code=500, detail="Failed to unlock user")

    log_event(
        action=AuditAction.USER_UNLOCKED,
        category=AuditCategory.USER,
        actor_email=user.get("email", ""),
        actor_name=user.get("name", ""),
        target_type="User",
        target_id=record_id,
        ip_address=get_request_ip(request),
    )
    return JSONResponse({"ok": True, "message": "User unlocked"})


@router.get("/users/{record_id}/status")
async def api_user_status(request: Request, record_id: str, user: dict = Depends(require_permission("user_management.view"))):
    """Get user lifecycle status (active, locked, last login, etc.)."""

    status = get_user_status(record_id)
    if not status:
        return JSONResponse({"ok": True, "status": {"is_active": True, "failed_attempts": 0}})

    return JSONResponse({"ok": True, "status": status})


# ==================== AUDIT LOG ENDPOINTS ====================

@router.get("/audit-logs")
async def api_list_audit_logs(
    request: Request,
    page: int = Query(1),
    page_size: int = Query(50),
    category: str = Query(""),
    action: str = Query(""),
    actor_email: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    user: dict = Depends(require_permission("audit_logs.view")),
):
    """Query audit logs with pagination and filters."""

    from services.audit_service import list_audit_logs, count_audit_logs

    filters = {}
    if category:
        filters["category"] = category
    if action:
        filters["action"] = action
    if actor_email:
        filters["actor_email"] = actor_email

    skip = (page - 1) * page_size
    logs = list_audit_logs(
        top=page_size,
        skip=skip,
        filters=filters if filters else None,
        date_from=date_from or None,
        date_to=date_to or None,
    )
    total = count_audit_logs(
        filters=filters if filters else None,
        date_from=date_from or None,
        date_to=date_to or None,
    )

    return JSONResponse({
        "ok": True,
        "logs": logs,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0,
    })

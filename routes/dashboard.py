
from fastapi import APIRouter, Request, HTTPException, Query, Body, UploadFile, File, Depends
from middleware.auth import get_current_user, require_permission, require_admin
from services.dashboard_service import (
    get_dashboard_data,
    get_logs_data,
    get_logs_data_cached,
    get_dashboard_data_cached,
    get_all_rfp_data_cached,
    invalidate_dashboard_caches,
)
from services.user_service import get_user, update_user, authenticate_user, get_user_by_email
from services.sap_service import create_sap_password_record, list_sap_password_records, list_sap_password_records_cached, invalidate_sap_password_cache
from fastapi.responses import JSONResponse, FileResponse, Response
import os
import re
import math
from hashlib import md5
import json
from functools import lru_cache
from datetime import datetime, timedelta
import asyncio
from automation_logic import run_automation_submit, run_automation_download_all_rfps
from collections import Counter
from typing import Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

# Import run state management from automation routes
from routes.automation import _RUN_STATE, _set_state, _add_submitting_rfp, _remove_submitting_rfp, _is_rfp_submitting, _run_async_in_thread
from services.system_settings_service import get_setting
from helpers.core_helper import (
    DATAVERSE,
    update_rfp_participation_status,
    get_rfp_excel_file_path,
    get_rfp_material_file_path,
    get_rfp_tds_folder_path,
    get_rfp_saved_excel_file_path,
    get_sharepoint_rfp_savedrfp_path,
    get_sharepoint_rfp_tds_path,
    get_sharepoint_rfp_material_path,
    clean_rfp_title,
    find_column_name,
    extract_materials_from_excel,
    extract_keywords_from_text,
    _find_other_content_sheet_name,
    find_rfp_file_across_companies,
)
from helpers.sharepoint_helper import GraphClient
import tempfile
from io import BytesIO
import pandas as pd
from helpers.unprotect_xls import unprotect_excel_file, unprotect_excel_bytes
import glob
import base64

router = APIRouter(prefix="/dashboard", tags=["Dashboard"])


@router.get("/clear-cache")
async def clear_cache_and_refresh(user: dict = Depends(require_admin)):
    """Clear all dashboard caches and redirect to dashboard with fresh data."""
    from fastapi.responses import RedirectResponse
    invalidate_dashboard_caches()
    return RedirectResponse(url="/dashboard?refresh=1", status_code=303)


# ===== RFP details helpers =====
_RFP_STATUS_META = {
    "open": {"label": "Open", "badge": "bg-warning"},
    "submitted": {"label": "Submitted", "badge": "bg-success"},
    "saved_draft": {"label": "Saved Draft", "badge": "bg-secondary"},
    "declined": {"label": "Declined", "badge": "bg-danger"},
    "other": {"label": "Other", "badge": "bg-info"},
}

_RFP_FILTER_OPTIONS = [
    {"value": "downloaded", "label": "Downloaded (All)"},
    {"value": "open", "label": "Open"},
    {"value": "submitted", "label": "Submitted"},
    {"value": "saved_draft", "label": "Saved Draft"},
    {"value": "declined", "label": "Declined"},
]

def _normalize_participation(raw_status: str) -> str:
    value = (raw_status or "").strip().lower()
    if value in ("submitted", "yes"):
        return "submitted"
    if value == "declined":
        return "declined"
    if value == "saved_draft":
        return "saved_draft"
    if value in ("", "no", "open", "not participated"):
        return "open"
    return "other"

def _parse_date_filter(date_str: str) -> Optional[datetime]:
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return None

try:
    from dateutil import parser as _dateutil_parser  # type: ignore
except Exception:
    _dateutil_parser = None

def _parse_end_datetime(value) -> Optional[datetime]:
    if not value:
        return None
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime()
        except Exception:
            pass
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            # assume epoch seconds
            return datetime.fromtimestamp(value)
        except Exception:
            pass
    if isinstance(value, str):
        cleaned = value.strip()
        try:
            return datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
        except Exception:
            pass
        if _dateutil_parser:
            try:
                return _dateutil_parser.parse(cleaned)
            except Exception:
                pass
        value = cleaned
    fallback_formats = [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %I:%M %p",
        "%d-%m-%Y",
        "%m-%d-%Y %H:%M",
        "%m-%d-%Y %H:%M:%S",
        "%m-%d-%Y %I:%M %p",
        "%m-%d-%Y",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %I:%M %p",
        "%d/%m/%Y",
    ]
    for fmt in fallback_formats:
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return None

# 4. Routes
def _make_etag(payload: dict) -> str:
    return md5(json.dumps(payload, sort_keys=True, default=str).encode("utf-8")).hexdigest()

_FREQ_CACHE = {"label_to_value": None, "value_to_label": None}

def _get_frequency_maps():
    # Load once and cache; column logical name assumed to be 'cr673_frequency'
    if not _FREQ_CACHE["label_to_value"] or not _FREQ_CACHE["value_to_label"]:
        try:
            maps = DATAVERSE.get_choice_options(
                get_setting("AUTOMATION_SCHEDULE_TABLE_LOGICAL", ""),
                "cr673_frequency",
            )
            _FREQ_CACHE.update(maps)
        except Exception:
            # Fallback defaults (update to your org values if needed)
            _FREQ_CACHE["label_to_value"] = {
                "Month": 415300000,
                "Week": 415300001,
                "Day": 415300002,
                "Hour": 415300003,
                "Minute": 415300004,
                "Year": 415300005,
            }
            _FREQ_CACHE["value_to_label"] = {v: k for k, v in _FREQ_CACHE["label_to_value"].items()}
    return _FREQ_CACHE

# ===== RFP status update =====
@router.post("/rfp/status")
async def update_rfp_status(request: Request, payload: dict = Body(...), user: dict = Depends(get_current_user)):
    try:
        rfp_id = (payload.get("rfp_id") or "").strip()
        status = (payload.get("status") or "").strip()
        if not rfp_id or not status:
            raise HTTPException(status_code=400, detail="rfp_id and status are required")

        # Normalize status to lowercase for consistent database storage
        # Frontend may send "Submitted" but database expects "submitted"
        status_normalized = status.lower()

        # Validate status against allowed values
        VALID_RFP_STATUSES = get_setting("VALID_RFP_STATUSES", ["no", "saved_draft", "submitted", "declined"])
        if status_normalized not in [s.lower() for s in VALID_RFP_STATUSES]:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid status: '{status}'. Valid values are: {', '.join(VALID_RFP_STATUSES)}"
            )

        ok = update_rfp_participation_status(rfp_id, status_normalized)
        if not ok:
            raise HTTPException(status_code=404, detail="RFP not found to update")
        # Invalidate caches so the next dashboard fetch gets fresh data
        invalidate_dashboard_caches()
        return JSONResponse({"ok": True, "message": f"Status updated to {status_normalized}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating status: {str(e)}")

@router.get("/rfp-status/{rfp_id}")
async def rfp_submit_status(rfp_id: str, user: dict = Depends(get_current_user)):
    """Check if specific RFP is currently being submitted"""
    return JSONResponse({
        "ok": True,
        "rfp_id": rfp_id,
        "is_submitting": _is_rfp_submitting(rfp_id)
    })

@router.post("/download-all-rfps")
async def download_all_rfps(request: Request, user: dict = Depends(require_permission("rfp.download"))):
    """API endpoint to trigger download all RFPs automation (non-blocking background task)"""

    # Check if download is already running
    if _RUN_STATE.get("download"):
        return JSONResponse({"ok": False, "message": "Download already running"}, status_code=409)

    try:
        # Get request body
        body = await request.json()
        selected_company = body.get("company", "").strip()

        # Define the background task
        async def _download_task():
            try:
                _set_state("download", True)
                await run_automation_download_all_rfps(selected_company=selected_company)
            finally:
                _set_state("download", False)

        # Run in separate thread with ProactorEventLoop for Windows Playwright compatibility
        # This returns immediately, allowing concurrent operations
        _run_async_in_thread(_download_task)

        return JSONResponse({
            "ok": True,
            "started": True,
            "message": "Download all RFPs started in background"
        }, status_code=202)
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error starting download all RFPs: {str(e)}")

@router.post("/profile")
async def update_profile(request: Request):
    # Simple session check
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    try:
        data = await request.json()
        current_user = request.session.get("user")
        
        # Prepare updates
        updates = {}
        
        # Update display name if provided
        if data.get("name") and data.get("name").strip():
            updates["name"] = data.get("name").strip()
        
        # Update mobile number if provided
        if data.get("mobile_number") is not None:
            updates["mobile_number"] = data.get("mobile_number").strip() if data.get("mobile_number") else ""
        
        # Handle password change if provided
        if data.get("current_password") and data.get("new_password"):
            # Verify current password
            if not authenticate_user(current_user["email"], data.get("current_password")):
                raise HTTPException(status_code=400, detail="Current password is incorrect")
            
            # Validate new password
            new_password = data.get("new_password")
            confirm_password = data.get("confirm_password")
            
            if new_password != confirm_password:
                raise HTTPException(status_code=400, detail="New passwords do not match")
            
            if len(new_password) < 6:
                raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
            
            updates["password"] = new_password
        
        # If no updates, return success
        if not updates:
            return JSONResponse({"ok": True, "message": "No changes to update"})
        
        # Get user record ID from current session (we need to fetch it)
        # For now, we'll use email to find the user record
        # In a real implementation, you'd store the record_id in the session
        user_records = get_user_by_email(current_user["email"])
        if not user_records:
            raise HTTPException(status_code=404, detail="User not found")
        
        user_record_id = user_records[0]["record_id"]
        
        # Update user in database
        success = update_user(user_record_id, updates)
        
        if not success:
            raise HTTPException(status_code=500, detail="Failed to update profile")
        
        # Update session with new data
        updated_user = {**current_user}
        updated_user.update(updates)
        request.session["user"] = updated_user
        
        return JSONResponse({
            "ok": True, 
            "message": "Profile updated successfully",
            "user": updated_user
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating profile: {str(e)}")

@router.post("/sap-password")
async def update_sap_password(request: Request, user: dict = Depends(require_permission("sap_password.change"))):
    try:
        data = await request.json()
        # allow any password (no validation)
        password = (data.get("password") or "")
        username = (data.get("username") or "").strip()
        user_email = user.get("email") if isinstance(user, dict) else None
        if not username and isinstance(user, dict):
            username = user.get("name") or ""

        ok = create_sap_password_record(password=password, user_email=user_email, username=username)
        if not ok:
            raise HTTPException(status_code=500, detail="Failed to save SAP password")
        # New record added; make sure logs reflect it immediately
        invalidate_sap_password_cache()

        return JSONResponse({"ok": True, "message": "SAP password updated"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ================= Automation Schedule APIs =================

def _safe_use_display_names() -> bool:
    return bool(get_setting("AUTOMATION_SCHEDULE_TABLE_LOGICAL", ""))

@router.get("/schedule-automation/latest")
async def get_latest_schedule(request: Request, user: dict = Depends(require_permission("schedule_automation.manage"))):
    try:
        rows = DATAVERSE.get_rows_from_dataverse(
            table_api_name=get_setting("AUTOMATION_SCHEDULE_TABLE_API", ""),
            select_columns=[
                "job_name",
                "interval",
                "frequency",
                "timezone",
                "start_time",
                "max_concurrency",
                "notes",
                "is_active",
                "created_by",
                "created_at",
            ],
            top=1,
            order_by="id desc",
            table_logical_name=get_setting("AUTOMATION_SCHEDULE_TABLE_LOGICAL", ""),
            use_display_names=_safe_use_display_names(),
        )
        latest = rows[0] if rows else {}
        # Map numeric frequency back to label for UI using dynamic map
        maps = _get_frequency_maps()
        if isinstance(latest.get("frequency"), int):
            latest["frequency"] = maps["value_to_label"].get(latest["frequency"], latest["frequency"])
        return JSONResponse({"ok": True, "data": latest})
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "message": str(e)})




@router.post("/schedule-automation")
async def save_schedule(request: Request, payload: dict = Body(...), user: dict = Depends(require_permission("schedule_automation.manage"))):
    try:
        freq_label = payload.get("frequency")
        maps = _get_frequency_maps()
        freq_value = maps["label_to_value"].get(freq_label) if isinstance(freq_label, str) else freq_label
        if freq_value is None:
            raise RuntimeError("Unsupported frequency. Use Month, Week, Day, Hour, Minute, or Year.")

        # Get start_time and validate - empty string should be treated as None
        start_time = payload.get("start_time")
        if isinstance(start_time, str):
            start_time = start_time.strip() or None

        data = {
            "job_name": payload.get("job_name") or "RFP Automation",
            "interval": int(payload.get("interval")),
            "frequency": int(freq_value),
            "timezone": payload.get("timezone"),
            "start_time": start_time,
            "max_concurrency": int(payload.get("max_concurrency") or 1),
            "notes": payload.get("notes"),
            "is_active": bool(payload.get("is_active", True)),
            "created_by": payload.get("created_by"),
            "created_at": payload.get("created_at"),
        }

        # Filter out None and empty string values
        data = {k: v for k, v in data.items() if v is not None and v != ""}

        ok = DATAVERSE.insert_row(
            table_api_name=get_setting("AUTOMATION_SCHEDULE_TABLE_API", ""),
            data=data,
            table_logical_name=get_setting("AUTOMATION_SCHEDULE_TABLE_LOGICAL", ""),
            use_display_names=_safe_use_display_names(),
        )
        if not ok:
            raise RuntimeError("Insert failed")
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "message": str(e)})


# ================= RFP Excel Viewing =================

@router.get("/view-excel/{rfp_id}")
async def view_rfp_excel(request: Request, rfp_id: str, company: str = None):
    """
    View and edit RFP Excel file (unprotected version).
    Downloads directly from SharePoint, unprotects in memory, and streams to browser.
    No local file storage.
    """
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        # Resolve company
        selected_company = (company or "").strip()
        if not selected_company:
            found_path, found_company = find_rfp_file_across_companies(rfp_id)
            if found_company:
                selected_company = found_company
            else:
                selected_company = get_setting("COMPANY_NAME", "")

        # Connect to SharePoint and find the Excel file
        graph_client = GraphClient(get_setting("CLIENT_ID", ""), get_setting("CLIENT_SECRET", ""), get_setting("TENANT_ID", ""), get_setting("SHAREPOINT_HOSTNAME", ""), get_setting("SITE_PATH", ""), get_setting("DRIVE_NAME", ""))
        graph_client.auth()
        graph_client.resolve_site_and_drive()

        sp_file_info = None
        # Strategy 1: Look in exact RFP folder path
        sp_material_path = get_sharepoint_rfp_material_path(rfp_id, selected_company)
        sp_files = graph_client.list_files_in_directory(sp_material_path, ['.xls', '.xlsx'])
        if sp_files:
            sp_file_info = sp_files[0]

        # Strategy 2: Search company folder for any file/folder matching the RFP ID
        if not sp_file_info:
            import re as _re
            safe_company = _re.sub(r'[<>:"/\\|?*]', '_', selected_company).strip().rstrip('.')
            sp_company_path = f"{get_setting('SP_BASE_FOLDER', '')}/ALLRFPs/{safe_company}"
            sp_all_files = graph_client.list_files_in_directory(sp_company_path, ['.xls', '.xlsx'])
            matching = [f for f in sp_all_files if rfp_id in f.get('path', '')]
            if matching:
                sp_file_info = matching[0]

        if not sp_file_info:
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found on SharePoint for RFP: {rfp_id}"
            )

        # Download file content from SharePoint into memory
        file_content = graph_client.get_file_content_from_sharepoint(sp_file_info['path'])
        raw_bytes = file_content.read()
        original_filename = sp_file_info.get('name', f'{rfp_id}.xlsx')

        # Unprotect in memory
        unprotected_bytes, out_filename = unprotect_excel_bytes(raw_bytes, original_filename)

        # Determine media type
        media_type = "application/vnd.ms-excel" if out_filename.endswith('.xls') else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return Response(
            content=unprotected_bytes,
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{out_filename}"',
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error viewing Excel file: {str(e)}")


@router.post("/save-excel/{rfp_id}")
async def save_rfp_excel(request: Request, rfp_id: str, file: UploadFile = File(...), company: str = None):
    """
    Save edited Excel file back to disk.

    This endpoint:
    1. Receives the edited Excel file from the frontend
    2. Saves it back to the unprotected version in ALLRFPs folder
    3. Optionally updates the original file as well
    """
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        # Resolve company - either from query param or by searching
        selected_company = (company or "").strip()
        if not selected_company:
            # Try to find company from database or file system
            found_path, found_company = find_rfp_file_across_companies(rfp_id)
            if found_company:
                selected_company = found_company
            else:
                selected_company = get_setting("COMPANY_NAME", "")

        # Use new folder structure: ALLRFPs/Company/RFP_title/downloaded-rfp/RFP_title.xls
        found_file = get_rfp_excel_file_path(rfp_id, selected_company)

        # If not found in new structure, try searching recursively
        if not os.path.exists(found_file):
            found_file = None
            _output_dir = get_setting("OUTPUT_DIR", "")
            company_folder = os.path.join(_output_dir, selected_company)
            search_dirs = [company_folder, _output_dir] if os.path.isdir(company_folder) else [_output_dir]

            for search_dir in search_dirs:
                if not os.path.isdir(search_dir):
                    continue
                for pattern in [f"**/*{rfp_id}*.xls", f"**/*{rfp_id}*.xlsx"]:
                    matches = glob.glob(os.path.join(search_dir, pattern), recursive=True)
                    matches = [f for f in matches if '_unprotected' not in os.path.basename(f)]
                    if matches:
                        found_file = matches[0]
                        break
                if found_file:
                    break

        if not found_file or not os.path.exists(found_file):
            raise HTTPException(
                status_code=404,
                detail=f"Original Excel file not found for RFP ID: {rfp_id}"
            )

        # Determine the unprotected file path (in same folder as original)
        base_name = os.path.basename(found_file)
        name_without_ext, ext = os.path.splitext(base_name)
        unprotected_filename = f"{name_without_ext}_unprotected{ext}"
        unprotected_path = os.path.join(os.path.dirname(found_file), unprotected_filename)
        
        # Read the uploaded file content
        file_content = await file.read()
        
        # Save to the unprotected file
        with open(unprotected_path, 'wb') as f:
            f.write(file_content)
        
        # Get user info for logging
        user = request.session.get("user", {})
        user_email = user.get("email", "unknown")
        
        return JSONResponse({
            "ok": True,
            "message": f"Excel file saved successfully to {unprotected_filename}",
            "file_path": unprotected_path,
            "saved_by": user_email
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving Excel file: {str(e)}")

# ================= RFP Materials Extraction and Matching =================

# In-memory cache for master files (shared across requests)
_MASTER_CACHE = {"data": None, "timestamp": None, "path": None, "file_mtime": None}
_KEYWORDS_CACHE = {"data": None, "timestamp": None, "path": None, "file_mtime": None}
_MATCH_PERCENTAGE_CACHE = {}  # {rfp_id: {"percentage": float, "file_mtime": float, "total_materials": int, "matched_count": int}}
# Bump this version whenever match logic changes — forces cache invalidation
_MATCH_CACHE_VERSION = 3  # v3: all items + exact code + keyword matching (same as download time)

# Cache TTL (5 minutes)
CACHE_TTL_SECONDS = 300

def get_cached_master_data(graph_client, master_csv_local):
    """Get master CSV data with caching"""
    current_time = datetime.now()
    
    # Check if cache is valid
    if (_MASTER_CACHE["data"] is not None and 
        _MASTER_CACHE["timestamp"] and 
        (current_time - _MASTER_CACHE["timestamp"]).total_seconds() < CACHE_TTL_SECONDS and
        _MASTER_CACHE["path"] and
        os.path.exists(_MASTER_CACHE["path"]) and
        os.path.getmtime(_MASTER_CACHE["path"]) == _MASTER_CACHE.get("file_mtime")):
        return _MASTER_CACHE["data"]
    
    # Download and cache
    master_csv_path = graph_client.download_file_from_sharepoint(
        sp_path=f"{get_setting('SP_BASE_FOLDER', '')}/master-files/material.csv",
        local_path=master_csv_local
    )
    master = pd.read_csv(master_csv_path)
    _MASTER_CACHE["data"] = master
    _MASTER_CACHE["timestamp"] = current_time
    _MASTER_CACHE["path"] = master_csv_path
    _MASTER_CACHE["file_mtime"] = os.path.getmtime(master_csv_path)
    
    return master

def get_cached_keywords(graph_client, keywords_csv_local):
    """Get keywords list with caching"""
    current_time = datetime.now()
    
    # Check if cache is valid
    if (_KEYWORDS_CACHE["data"] is not None and 
        _KEYWORDS_CACHE["timestamp"] and 
        (current_time - _KEYWORDS_CACHE["timestamp"]).total_seconds() < CACHE_TTL_SECONDS and
        _KEYWORDS_CACHE["path"] and
        os.path.exists(_KEYWORDS_CACHE["path"]) and
        os.path.getmtime(_KEYWORDS_CACHE["path"]) == _KEYWORDS_CACHE.get("file_mtime")):
        return _KEYWORDS_CACHE["data"]
    
    # Download and cache
    keywords_list = []
    try:
        keywords_csv_path = graph_client.download_file_from_sharepoint(
            sp_path=f"{get_setting('SP_BASE_FOLDER', '')}/master-files/unique_keywords.csv",
            local_path=keywords_csv_local
        )
        keywords_df = pd.read_csv(keywords_csv_path)
        keywords_col = find_column_name(keywords_df.columns, "keywords") or find_column_name(keywords_df.columns, "keyword")
        if keywords_col:
            # Use to_dict('records') for better performance (faster than iterrows)
            for row in keywords_df.to_dict('records'):
                keyword_value = str(row.get(keywords_col, "")).strip()
                if keyword_value and keyword_value.lower() != 'nan':
                    for kw in keyword_value.split(','):
                        kw_clean = kw.strip().upper()
                        if kw_clean:
                            keywords_list.append(kw_clean)
    except Exception as e:
        print(f"[WARN] Could not load keywords: {e}")
        keywords_list = []
    
    _KEYWORDS_CACHE["data"] = keywords_list
    _KEYWORDS_CACHE["timestamp"] = current_time
    _KEYWORDS_CACHE["path"] = keywords_csv_local
    if os.path.exists(keywords_csv_local):
        _KEYWORDS_CACHE["file_mtime"] = os.path.getmtime(keywords_csv_local)
    
    return keywords_list

def ensure_rfp_excel_from_sharepoint(rfp_id, company, graph_client):
    """
    Ensure RFP Excel file exists locally by downloading from SharePoint if needed.
    Returns local excel_path or None if not found anywhere.
    """
    if not company:
        # Try local search first
        excel_path, found_company = find_rfp_file_across_companies(rfp_id)
        if excel_path and os.path.exists(excel_path):
            return excel_path
        company = found_company or get_setting("COMPANY_NAME", "")

    excel_path = get_rfp_excel_file_path(rfp_id, company)
    if os.path.exists(excel_path):
        return excel_path

    # File not local — try downloading from SharePoint
    if not graph_client:
        return None

    clean_title = clean_rfp_title(rfp_id)
    for ext in ['.xls', '.xlsx']:
        filename = f"{clean_title}{ext}"
        sp_path = get_sharepoint_rfp_material_path(rfp_id, company, filename)
        try:
            local_path = os.path.join(
                get_rfp_material_file_path(rfp_id, company),
                filename
            )
            graph_client.download_file_from_sharepoint(sp_path, local_path)
            print(f"[OK] Downloaded RFP Excel from SharePoint: {sp_path}")
            return local_path
        except Exception as e:
            print(f"[WARN] SharePoint download attempt ({ext}): {e}")
            continue

    print(f"[ERROR] RFP Excel not found on SharePoint for {rfp_id} (company: {company})")
    return None


def _parse_matched_data_json(matched_data_str: str):
    """Parse Matched_Data JSON string into match percentage result, or return None."""
    matched_data_str = (matched_data_str or "").strip()
    if not matched_data_str:
        return None
    try:
        data = json.loads(matched_data_str)
    except Exception:
        return None

    # New categorized format (dict with summary)
    if isinstance(data, dict) and "summary" in data:
        s = data["summary"]
        return {
            "match_percentage": s.get("match_percentage", 0),
            "total_materials": data.get("total_items", 0),
            "matched_count": s.get("exact_match_count", 0) + s.get("keyword_match_count", 0),
        }

    # Old flat format (list of items) — backward compatibility
    if isinstance(data, list) and data:
        if not any("is_matched" in item for item in data):
            return None
        total_materials = len(data)
        matched_count = sum(1 for item in data if bool(item.get("is_matched", True)))
        match_percentage = round((matched_count / total_materials * 100) if total_materials > 0 else 0, 1)
        return {
            "match_percentage": match_percentage,
            "total_materials": total_materials,
            "matched_count": matched_count,
        }

    return None


def _batch_get_match_percentages_from_dataverse(rfp_ids: list):
    """
    Batch query Dataverse for Matched_Data of multiple RFPs at once.
    Returns dict mapping rfp_id -> result dict or None (if not found/invalid).
    Uses OData 'or' filter in chunks to stay within URL length limits.
    """
    results = {rid: None for rid in rfp_ids}
    if not rfp_ids:
        return results

    # Chunk into groups of 15 to avoid URL length limits
    chunk_size = 15
    for i in range(0, len(rfp_ids), chunk_size):
        chunk = rfp_ids[i:i + chunk_size]
        filter_parts = [f"RFP_ID eq '{rid.replace(chr(39), chr(39)*2)}'" for rid in chunk]
        filter_expr = " or ".join(filter_parts)
        try:
            result = DATAVERSE.query_rows(
                get_setting("RFP_ACTIVITY_LOG_TABLE_API", "cr673_bahra_rfps_v2s"),
                filter_expr=filter_expr,
                select="RFP_ID,Matched_Data",
                top=chunk_size,
                table_logical_name=get_setting("RFP_ACTIVITY_LOG_TABLE_LOGICAL", "cr673_bahra_rfps_v2"),
                use_display_names=True
            )
            rows = result.get("value", []) if isinstance(result, dict) else []
            for row in rows:
                rfp_id = row.get("RFP_ID", "")
                if rfp_id in results:
                    parsed = _parse_matched_data_json(row.get("Matched_Data"))
                    results[rfp_id] = parsed
        except Exception as e:
            print(f"[BatchMatch] Batch Dataverse query failed for chunk: {e}")

    return results


def _get_match_percentage_from_dataverse(rfp_id: str):
    """
    Try to get match percentage summary from stored Matched_Data in Dataverse.
    Returns dict with match_percentage/total_materials/matched_count, or None if not available.
    """
    try:
        filter_val = rfp_id.replace("'", "''")
        result = DATAVERSE.query_rows(
            get_setting("RFP_ACTIVITY_LOG_TABLE_API", "cr673_bahra_rfps_v2s"),
            filter_expr=f"RFP_ID eq '{filter_val}'",
            select="RFP_ID,Matched_Data",
            top=1,
            table_logical_name=get_setting("RFP_ACTIVITY_LOG_TABLE_LOGICAL", "cr673_bahra_rfps_v2"),
            use_display_names=True
        )
        rows = result.get("value", []) if isinstance(result, dict) else []
        if not rows:
            return None
        return _parse_matched_data_json(rows[0].get("Matched_Data"))
    except Exception as e:
        print(f"[BatchMatch] Dataverse lookup failed for {rfp_id}: {e}")
        return None


def calculate_match_percentage_optimized(rfp_id, master, master_col, keywords_list, company: str = None, graph_client=None, master_code_set: set = None, skip_dataverse: bool = False):
    """
    Calculate match percentage for a single RFP (optimized with caching).
    Uses ALL items from Excel (no intent filter).
    Matching uses exact code + keyword matching (same logic as download time).
    Primary: checks Dataverse for stored Matched_Data.
    Fallback: uses locally available Excel files.
    """
    # Check cache first
    if rfp_id in _MATCH_PERCENTAGE_CACHE:
        cached = _MATCH_PERCENTAGE_CACHE[rfp_id]
        if cached.get("cache_version") == _MATCH_CACHE_VERSION:
            # For Dataverse-sourced cache, no file_mtime to check
            if cached.get("source") == "dataverse":
                return cached
            # For file-sourced cache, validate file_mtime
            file_mtime = cached.get("file_mtime")
            if file_mtime:
                excel_path = ensure_rfp_excel_from_sharepoint(rfp_id, company, None)
                if excel_path and os.path.exists(excel_path) and os.path.getmtime(excel_path) == file_mtime:
                    return cached

    # Primary: try Dataverse stored data (skip if batch already checked)
    if not skip_dataverse:
        dv_result = _get_match_percentage_from_dataverse(rfp_id)
        if dv_result is not None:
            dv_result["source"] = "dataverse"
            dv_result["cache_version"] = _MATCH_CACHE_VERSION
            _MATCH_PERCENTAGE_CACHE[rfp_id] = dv_result
            return dv_result

    # Fallback: Excel file matching (download from SharePoint if not local)
    excel_path = ensure_rfp_excel_from_sharepoint(rfp_id, company, graph_client)

    if not excel_path or not os.path.exists(excel_path):
        return {"match_percentage": 0, "total_materials": 0, "matched_count": 0, "file_mtime": None}

    file_mtime = os.path.getmtime(excel_path)

    # Extract ALL materials from Excel (no intent filter)
    materials_data = extract_materials_from_excel(excel_path, include_details=True, filter_by_intent=False)

    if not materials_data:
        result = {"match_percentage": 0, "total_materials": 0, "matched_count": 0, "file_mtime": file_mtime}
        _MATCH_PERCENTAGE_CACHE[rfp_id] = result
        return result

    # Build master_code_set if not provided (backward compat)
    if master_code_set is None:
        master_code_set = set(master[master_col].astype(str))

    # Match using exact code + keyword (same logic as download time)
    matched_count = 0
    for mat_data in materials_data:
        mat_code = mat_data["material_code"]
        name_text = mat_data.get("name", "")
        description_text = mat_data.get("description", "")

        # Method 1: Exact Material Code Match (O(1) set lookup)
        is_matched = mat_code in master_code_set

        # Method 2: Keyword Matching (only if exact match failed)
        if not is_matched and keywords_list:
            name_keywords = extract_keywords_from_text(name_text)
            desc_keywords = extract_keywords_from_text(description_text)
            all_material_keywords = set(name_keywords + desc_keywords)

            for csv_keyword in keywords_list:
                for mat_keyword in all_material_keywords:
                    if csv_keyword in mat_keyword or mat_keyword in csv_keyword:
                        is_matched = True
                        break
                if is_matched:
                    break

        if is_matched:
            matched_count += 1

    total_materials = len(materials_data)
    match_percentage = round((matched_count / total_materials * 100) if total_materials > 0 else 0, 1)

    result = {
        "match_percentage": match_percentage,
        "total_materials": total_materials,
        "matched_count": matched_count,
        "file_mtime": file_mtime,
        "cache_version": _MATCH_CACHE_VERSION
    }
    _MATCH_PERCENTAGE_CACHE[rfp_id] = result
    return result


# ==================== DYNAMIC FORM GENERATION HELPERS ====================

def rgb_to_tuple(rgb_str):
    """Convert hex or ARGB (e.g. 'FFFFC000') to RGB tuple."""
    if not rgb_str:
        return None
    rgb_str = rgb_str.replace("0x", "").replace("#", "").upper()
    if len(rgb_str) == 8:  # remove alpha if present
        rgb_str = rgb_str[2:]
    try:
        r = int(rgb_str[0:2], 16)
        g = int(rgb_str[2:4], 16)
        b = int(rgb_str[4:6], 16)
        return (r, g, b)
    except Exception:
        return None


def is_yellow_rgb(r, g, b, strict=True):
    """
    Detect yellow color.
    
    Args:
        r, g, b: RGB color values (0-255)
        strict: If True, only detect exact RGB(255, 255, 153). 
                If False, detect any yellow-ish color.
    
    Returns:
        bool: True if the color is yellow
    """
    if strict:
        # Exact match for RGB(255, 255, 153) - the light yellow shade
        return (r == 255 and g == 255 and b == 153)
    else:
        # Flexible yellow detection: any color that looks yellow
        # Yellow has high R and G, but low B
        # Criteria:
        # 1. R >= 180 (red component high)
        # 2. G >= 180 (green component high)
        # 3. B < 200 (blue component lower than red/green)
        # 4. R and G are similar (within 75 of each other)
        # 5. Not too dark (R+G+B > 400)
        # 6. Not white (B must be significantly lower: B < min(R,G) - 20)
        
        if r < 180 or g < 180:  # Too dark or not yellow
            return False
        
        if b >= 200:  # Too much blue, probably white or light blue
            return False
        
        if abs(r - g) > 75:  # Red and green should be similar for yellow
            return False
        
        if (r + g + b) < 400:  # Too dark overall
            return False
        
        if b >= min(r, g) - 20:  # Blue too close to red/green, probably white
            return False
        
        return True


def is_yellow_cell_xls(sheet, row_idx, col_idx, workbook, debug=False, strict=True):
    """
    Check if a cell has yellow background color in .xls file using RGB detection.
    
    Args:
        sheet: xlrd sheet object
        row_idx: Row index
        col_idx: Column index
        workbook: xlrd workbook object
        debug: Print debug info
        strict: If False, use flexible yellow detection
    
    Returns tuple: (is_yellow: bool, color_rgb: tuple or None)
    """
    try:
        cell_xf_index = sheet.cell_xf_index(row_idx, col_idx)
        xf = workbook.xf_list[cell_xf_index]
        bg_color_index = xf.background.pattern_colour_index
        
        color = workbook.colour_map.get(bg_color_index)
        
        if color:
            r, g, b = color[0], color[1], color[2]
            is_yellow = is_yellow_rgb(r, g, b, strict=strict)
            if debug and (is_yellow or bg_color_index not in [64, 65]):
                print(f"         Cell ({row_idx}, {col_idx}): RGB({r}, {g}, {b}) - Color Index: {bg_color_index} - Yellow: {is_yellow}")
            return is_yellow, color
        
        return False, None
    except Exception as e:
        if debug:
            print(f"         Error checking cell ({row_idx}, {col_idx}): {e}")
        return False, None


def is_yellow_cell_xlsx(cell, debug=False, strict=True):
    """
    Check if a cell has yellow background color in .xlsx file using RGB detection.
    Handles theme/indexed colors and RGB values from openpyxl.
    
    Args:
        cell: openpyxl cell object
        debug: Print debug info
        strict: If False, use flexible yellow detection
    
    Returns tuple: (is_yellow: bool, rgb_tuple: tuple or None)
    """
    try:
        fill = cell.fill
        if not fill or not fill.start_color:
            return False, None
        
        color = fill.start_color
        rgb_tuple = None
        color_type = color.type
        
        if color.type == "rgb" and color.rgb:
            rgb_tuple = rgb_to_tuple(color.rgb)
        elif color.type == "theme":
            # Theme colors can't be read directly; skip or assume yellow theme = False
            if debug:
                print(f"         Cell {cell.coordinate}: Theme color (skipped)")
            return False, None
        elif color.type == "indexed" and color.indexed is not None:
            # Map indexed colors (Excel palette)
            try:
                from openpyxl.styles.colors import COLOR_INDEX
                if color.indexed < len(COLOR_INDEX):
                    rgb_tuple = rgb_to_tuple(COLOR_INDEX[color.indexed])
            except:
                return False, None
        
        if rgb_tuple:
            is_yellow = is_yellow_rgb(*rgb_tuple, strict=strict)
            if debug and (is_yellow or rgb_tuple != (255, 255, 255)):
                print(f"         Cell {cell.coordinate}: RGB{rgb_tuple} - Type: {color_type} - Yellow: {is_yellow}")
            return is_yellow, rgb_tuple
        
        return False, None
    except Exception as e:
        if debug:
            print(f"         Error checking cell {cell.coordinate}: {e}")
        return False, None


def get_cell_label(sheet, row, col, is_xlsx=False):
    """
    Get label for a yellow cell by checking adjacent cells.
    Priority: left cell > cell above > further left > row header.
    """
    def get_cell_val(r, c):
        try:
            if is_xlsx:
                cell = sheet.cell(row=r+1, column=c+1)  # openpyxl is 1-indexed
                return str(cell.value).strip() if cell.value else ""
            else:
                val = sheet.cell_value(r, c)
                return str(val).strip() if val else ""
        except:
            return ""
    
    # Try immediate left cell
    if col > 0:
        left_val = get_cell_val(row, col - 1)
        if left_val:
            return left_val
    
    # Try cell above
    if row > 0:
        above_val = get_cell_val(row - 1, col)
        if above_val:
            return above_val
    
    # Try further left (up to 4 columns)
    for c in range(col - 2, max(-1, col - 5), -1):
        if c >= 0:
            left_val = get_cell_val(row, c)
            if left_val:
                return left_val
    
    # Try first column of the row (row header)
    if col > 0:
        first_col = get_cell_val(row, 0)
        if first_col:
            return first_col
    
    # Fallback: use position
    return f"Field at Row {row + 1}, Column {col + 1}"


def get_dropdown_options_xlsx(sheet, cell):
    """
    Extract dropdown options from Excel data validation for XLSX files.
    
    Args:
        sheet: openpyxl worksheet object
        cell: openpyxl cell object
    
    Returns:
        list: List of dropdown options, or None if no dropdown
    """
    try:
        # Check if the cell has data validation
        for dv in sheet.data_validations.dataValidation:
            if cell.coordinate in dv.cells:
                # Check if it's a list validation
                if dv.type == "list":
                    # Formula-based list (e.g., =Sheet1!$A$1:$A$10)
                    if dv.formula1:
                        formula = str(dv.formula1)
                        # Remove quotes if present
                        formula = formula.strip('"')
                        
                        # Check if it's a simple comma-separated list
                        if ',' in formula and '!' not in formula:
                            options = [opt.strip() for opt in formula.split(',')]
                            return options
                        
                        # Check if it's a range reference (e.g., $A$1:$A$10)
                        if '!' in formula or ':' in formula:
                            # Try to resolve the range
                            try:
                                # Parse reference like Sheet1!$A$1:$A$10
                                if '!' in formula:
                                    sheet_name, cell_range = formula.split('!')
                                    ref_sheet = sheet.parent[sheet_name.strip("'")]
                                else:
                                    ref_sheet = sheet
                                    cell_range = formula
                                
                                # Extract values from the range
                                options = []
                                for row in ref_sheet[cell_range]:
                                    for cell in row if isinstance(row, tuple) else [row]:
                                        if cell.value:
                                            options.append(str(cell.value))
                                return options if options else None
                            except:
                                pass
                
                return None
        
        return None
    except Exception as e:
        return None


def get_dropdown_options_xls(workbook, sheet, row_idx, col_idx, label):
    """
    Extract dropdown options from Excel data validation for XLS files.
    Since xlrd doesn't support data validation directly, we look for:
    1. DV_Sheet or validation sheet with dropdown values
    2. Adjacent columns with dropdown values (common pattern)
    3. Hidden sheets with validation lists
    
    Args:
        workbook: xlrd workbook object
        sheet: xlrd sheet object
        row_idx: row index of the cell
        col_idx: column index of the cell
        label: label/header of the field
    
    Returns:
        list: List of dropdown options, or None if no dropdown found
    """
    try:
        # Method 1: Look for DV_Sheet or validation sheet
        validation_sheet_names = ['dv_sheet', 'validation', 'dropdowns', 'data_validation', 'dv', 'lists']
        
        for sheet_name in workbook.sheet_names():
            if any(vs in sheet_name.lower() for vs in validation_sheet_names):
                dv_sheet = workbook.sheet_by_name(sheet_name)
                print(f"          Found validation sheet: '{sheet_name}' with {dv_sheet.ncols} columns")
                
                # Show all column headers in DV_sheet for debugging
                dv_headers = []
                for c in range(dv_sheet.ncols):
                    try:
                        h = str(dv_sheet.cell_value(0, c)).strip()
                        dv_headers.append(f"{chr(65+c)}:'{h}'")
                    except:
                        dv_headers.append(f"{chr(65+c)}:''")
                print(f"          DV_sheet columns: {', '.join(dv_headers)}")
                print(f"          Field '{label}' is at column {chr(65+col_idx)} (index {col_idx}) in Material sheet")
                
                # PRIORITY 1: Match by COLUMN POSITION (only if DV_sheet has enough columns)
                # The dropdown for column C in Material sheet should be in column C of DV_sheet
                if col_idx < dv_sheet.ncols:
                    options = []
                    dv_header = ""
                    try:
                        dv_header = str(dv_sheet.cell_value(0, col_idx)).strip()
                        print(f"          Column {chr(65+col_idx)} in DV_sheet has header: '{dv_header}'")
                    except:
                        pass
                    
                    # Extract all non-empty values from this column
                    for row in range(1, dv_sheet.nrows):
                        try:
                            val = dv_sheet.cell_value(row, col_idx)
                            if val and str(val).strip():
                                val_str = str(val).strip()
                                # Skip obvious non-dropdown values
                                if val_str.lower() not in ['', 'n/a', 'none', '-', 'null']:
                                    options.append(val_str)
                        except:
                            continue
                    
                    # Remove duplicates while preserving order
                    unique_options = list(dict.fromkeys(options))
                    
                    # Validate: Check if this looks like a real dropdown list
                    if 2 <= len(unique_options) <= 500:  # Reasonable dropdown size
                        dv_info = f" (DV_sheet column {chr(65+col_idx)}: '{dv_header}')" if dv_header else f" (DV_sheet column {chr(65+col_idx)})"
                        print(f"         [OK] COLUMN POSITION MATCH{dv_info}: Found {len(unique_options)} values: {unique_options[:5]}{'...' if len(unique_options) > 5 else ''}")
                        return unique_options
                    else:
                        print(f"         [WARN] Column {chr(65+col_idx)} in DV_sheet has {len(unique_options)} unique values - doesn't look like a dropdown")
                else:
                    print(f"         [WARN] Column {chr(65+col_idx)} doesn't exist in DV_sheet (only has {dv_sheet.ncols} columns)")
                
                # PRIORITY 2: Match by DATA PATTERN (check what type of data each column contains)
                print(f"          Trying to match by analyzing DV_sheet data patterns...")
                for col in range(dv_sheet.ncols):
                    try:
                        header_val = str(dv_sheet.cell_value(0, col)).strip()
                        
                        # Get sample data from this column
                        sample_data = []
                        for row in range(1, min(20, dv_sheet.nrows)):
                            val = dv_sheet.cell_value(row, col)
                            if val and str(val).strip():
                                sample_data.append(str(val).strip())
                        
                        if len(sample_data) < 2:
                            continue
                        
                        # Analyze data pattern
                        col_letter = chr(65+col)
                        print(f"          Column {col_letter} '{header_val}': {len(sample_data)} values, sample: {sample_data[:3]}")
                        
                        # Check if this column matches the field we're looking for
                        is_match = False
                        match_reason = ""
                        
                        # For country fields, look for country code pattern (e.g., "AD Andorra")
                        if 'country' in label.lower() or 'origin' in label.lower():
                            country_like = sum(1 for s in sample_data[:20] if (
                                re.match(r'^[A-Z]{2}\s+', s) or  # "AD Andorra"
                                any(c in s for c in ['Arabia', 'Afghanistan', 'Albania', 'Argentina', 'Armenia', 'Australia'])
                            ))
                            if country_like >= 3:
                                is_match = True
                                match_reason = f"country data pattern ({country_like}/20 samples)"
                        
                        # For delivery time fields
                        elif 'delivery' in label.lower() or 'time' in label.lower():
                            time_like = sum(1 for s in sample_data[:10] if (
                                re.search(r'\d+.*day', s.lower()) or
                                re.search(r'\d+[-~]\d+', s) or  # "7-14", "1~3"
                                any(t in s.lower() for t in ['week', 'month', 'immediate'])
                            ))
                            if time_like >= 2:
                                is_match = True
                                match_reason = f"time/delivery pattern ({time_like}/10 samples)"
                        
                        # For factory/yes-no fields
                        elif 'factory' in label.lower() or 'own' in label.lower():
                            yn_like = sum(1 for s in sample_data[:5] if s.lower() in ['yes', 'no', 'y', 'n'])
                            if yn_like >= 2:
                                is_match = True
                                match_reason = f"yes/no pattern ({yn_like}/5 samples)"
                        
                        # Header name matching as secondary check
                        if not is_match and header_val:
                            header_clean = header_val.lower().replace(' ', '').replace('_', '').replace('*', '')
                            label_clean = label.lower().replace(' ', '').replace('_', '').replace('*', '')
                            if header_clean in label_clean or label_clean in header_clean:
                                is_match = True
                                match_reason = f"header match '{header_val}'"
                        
                        if is_match:
                            # Extract all values from this column
                            options = []
                            for row in range(1, dv_sheet.nrows):
                                val = dv_sheet.cell_value(row, col)
                                if val and str(val).strip():
                                    val_str = str(val).strip()
                                    if val_str.lower() not in ['', 'n/a', 'none', '-', 'null']:
                                        options.append(val_str)
                            
                            unique_options = list(dict.fromkeys(options))
                            if 2 <= len(unique_options) <= 500:
                                print(f"         [OK] DATA PATTERN MATCH (column {col_letter} '{header_val}' via {match_reason}): Found {len(unique_options)} values: {unique_options[:5]}{'...' if len(unique_options) > 5 else ''}")
                                return unique_options
                    except Exception as e:
                        print(f"         [WARN] Error analyzing column {chr(65+col)}: {e}")
                        continue
                
                print(f"         [ERROR] No matching dropdown column found in DV_sheet for field '{label}'")
        
        # Method 2: Check if there's a note/comment on the cell (some Excel files store validation in notes)
        # xlrd doesn't support reading notes directly, so we skip this
        
        # Method 3: Look in the same sheet for dropdown values in adjacent areas
        # Some Excel files have validation lists to the right of the data
        # Check columns far to the right for potential dropdown lists
        if sheet.ncols > 20:  # If sheet has many columns, check rightmost columns
            for check_col in range(sheet.ncols - 10, sheet.ncols):
                try:
                    header = str(sheet.cell_value(0, check_col)).strip().lower()
                    if header and (label.lower() in header or header in label.lower()):
                        options = []
                        for row in range(1, min(100, sheet.nrows)):  # Check first 100 rows
                            val = sheet.cell_value(row, check_col)
                            if val and str(val).strip() and str(val).strip() not in options:
                                options.append(str(val).strip())
                        if options and len(options) >= 2:
                            print(f"         [OK] Found dropdown values in same sheet column {chr(65 + check_col)}: {options[:5]}{'...' if len(options) > 5 else ''}")
                            return options
                except:
                    continue
        
        return None
        
    except Exception as e:
        print(f"         [WARN] Error extracting dropdown options from XLS: {e}")
        return None


def infer_field_type(label):
    """
    Infer HTML input type from label text.
    Returns: text, number, date, email, tel, dropdown, or file
    """
    label_lower = label.lower()
    
    # File/Attachment fields (check first as they're specific)
    # TDS files are the primary attachment type for materials
    if any(kw in label_lower for kw in ['tds', 'technical data sheet', 'attach', 'upload', 'file', 'document', 'supporting file']):
        return 'file'
    
    # Date fields
    if any(kw in label_lower for kw in ['date', 'validity', 'deadline', 'expiry', 'dd-mm-yyyy', 'mm/dd/yyyy']):
        return 'date'
    
    # Phone/Mobile fields
    if any(kw in label_lower for kw in ['phone', 'mobile', 'telephone', 'contact number']):
        return 'tel'
    
    # Email fields
    if 'email' in label_lower or 'e-mail' in label_lower:
        return 'email'
    
    # Dropdown fields (CHECK BEFORE NUMBER - to avoid "country" matching "count")
    # Yes/No questions, Country, Delivery, etc.
    if any(kw in label_lower for kw in ['yes/no', 'vendor', 'local', 'are you', 'country', 'origin', 'factory', 'status', 'own factory', 'product own', 'delivery']):
        return 'dropdown'
    
    # Special case: "time" + "day" combination is usually a dropdown (delivery time options)
    if ('time' in label_lower and 'day' in label_lower) or ('delivery' in label_lower and 'day' in label_lower):
        return 'dropdown'
    
    # Number fields (check AFTER dropdown to avoid false matches)
    # But exclude "days" if it's part of "delivery" or "time" context
    if any(kw in label_lower for kw in ['price', 'amount', 'quantity', 'qty', 'number', 'count']):
        return 'number'
    
    # Check for standalone "days" (not in time/delivery context)
    if 'days' in label_lower and 'time' not in label_lower and 'delivery' not in label_lower:
        return 'number'
    
    # Default: text
    return 'text'


def parse_material_sheet_xls(sheet, sheet_idx, sheet_name, workbook):
    """
    Parse material sheet (like 'Other Content') grouping by materials.
    Uses column headers as labels instead of adjacent cells.
    """
    # Find header row (usually row 0 or 1)
    header_row_idx = 0
    for row_idx in range(min(3, sheet.nrows)):
        # Check if this row has typical header keywords
        row_text = " ".join([str(sheet.cell_value(row_idx, c)) for c in range(min(10, sheet.ncols))]).lower()
        if any(kw in row_text for kw in ['price', 'quantity', 'manufacturer', 'name', 'number']):
            header_row_idx = row_idx
            break
    
    print(f"          Using row {header_row_idx + 1} as header row")
    
    # Read column headers
    headers = {}
    name_col_idx = None
    for col_idx in range(sheet.ncols):
        try:
            header_val = str(sheet.cell_value(header_row_idx, col_idx)).strip()
            if header_val:
                headers[col_idx] = header_val
                print(f"            Col {chr(65 + col_idx) if col_idx < 26 else col_idx}: '{header_val}'")
                # Find Name column for material detection
                if 'name' in header_val.lower().replace(" ", "").replace("_", ""):
                    name_col_idx = col_idx
        except:
            continue
    
    # Find material rows and their blank yellow cells
    materials_data = {}  # {material_code: {fields: [], row_idx: }}
    
    for row_idx in range(header_row_idx + 1, sheet.nrows):
        # Extract material code from Name column using 9-digit regex (same as existing logic)
        material_code = None
        if name_col_idx is not None:
            name_value = str(sheet.cell_value(row_idx, name_col_idx))
            material_codes = re.findall(r"\d{9}", name_value)
            if material_codes:
                material_code = material_codes[0]  # Use first match
        else:
            # Fallback: Check first few columns for 8-10 digit codes
            for col_idx in range(min(5, sheet.ncols)):
                cell_val = str(sheet.cell_value(row_idx, col_idx))
                if cell_val and cell_val.isdigit() and len(cell_val) >= 8:
                    material_code = cell_val
                    break
        
        if not material_code:
            continue
        
        print(f"          Material: {material_code} (Row {row_idx + 1})")
        
        # Find ALL yellow cells in this material row (blank or filled)
        material_fields = []
        yellow_cells_found = 0
        blank_cells_included = 0
        
        for col_idx in range(sheet.ncols):
            is_yellow, color_rgb = is_yellow_cell_xls(sheet, row_idx, col_idx, workbook, debug=False)
            
            if is_yellow:
                yellow_cells_found += 1
                cell_value = sheet.cell_value(row_idx, col_idx)
                is_empty = (cell_value == "" or cell_value is None or 
                           (isinstance(cell_value, str) and cell_value.strip() == ""))
                
                col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                rgb_str = f"RGB({color_rgb[0]}, {color_rgb[1]}, {color_rgb[2]})" if color_rgb else "N/A"
                value_str = f"'{cell_value}'" if cell_value else "(empty)"
                
                # Show ALL yellow cells in debug output
                print(f"             {col_letter}{row_idx+1}: {rgb_str} - Value: {value_str} - Blank: {is_empty}")
                
                # Only include BLANK yellow cells in form
                if is_empty:
                    blank_cells_included += 1
                    # Use column header as label
                    label = headers.get(col_idx, f"Field at Column {col_letter}")
                    field_type = infer_field_type(label)
                    
                    # Check if field is required (has * in header)
                    is_required = '*' in label or 'required' in label.lower()
                    
                    # Check for dropdown options - try to extract from Excel first
                    dropdown_options = None
                    if field_type == 'dropdown':
                        # Try to extract dropdown options from Excel data validation or DV sheet
                        dropdown_options = get_dropdown_options_xls(workbook, sheet, row_idx, col_idx, label)
                        
                        # If not found in Excel, use smart fallback ONLY for specific known fields
                        if not dropdown_options:
                            label_check = label.lower()
                            if 'country' in label_check or 'origin' in label_check:
                                dropdown_options = ['Saudi Arabia', 'UAE', 'USA', 'China', 'Germany', 'Italy', 'France', 'UK', 'Japan', 'South Korea', 'India', 'Turkey', 'Spain', 'Canada', 'Brazil', 'Mexico', 'Other']
                                print(f"             Using fallback country list")
                            elif 'factory' in label_check or 'own factory' in label_check or 'product own' in label_check:
                                dropdown_options = ['Yes', 'No']
                                print(f"             Using fallback Yes/No")
                            elif 'vendor' in label_check or 'local' in label_check:
                                dropdown_options = ['Yes', 'No']
                                print(f"             Using fallback Yes/No")
                            else:
                                # If no dropdown in Excel, change field type to 'text' instead of forcing dropdown
                                field_type = 'text'
                                print(f"            [WARN] No dropdown values found in Excel for '{label}' - changing to TEXT input")
                    
                    field = {
                        "id": f"material_{sheet_idx}_{row_idx}_{col_idx}",
                        "label": label.replace('*', '').strip(),
                        "type": field_type,
                        "required": is_required,
                        "row": row_idx,
                        "col": col_idx,
                        "sheet_index": sheet_idx,
                        "sheet_name": sheet_name,
                        "material_code": material_code
                    }
                    
                    # Add dropdown options if available
                    if dropdown_options:
                        field["options"] = dropdown_options
                    
                    material_fields.append(field)
                    options_str = f", Options: {dropdown_options}" if dropdown_options else ""
                    print(f"INCLUDED - Label: '{field['label']}' - Type: {field_type} - Required: {is_required}{options_str}")
                else:
                    print(f"SKIPPED (already has value)")
        
        print(f"Material {material_code}: {yellow_cells_found} yellow cells found, {blank_cells_included} blank (included in form)")
        
        if material_fields:
            materials_data[material_code] = {
                "fields": material_fields,
                "row_idx": row_idx
            }
    
    # Build section structure
    section = {
        "sheet_name": sheet_name,
        "sheet_index": sheet_idx,
        "is_material_sheet": True,
        "materials": [],
        "fields": []  # Flattened list for counting
    }
    
    for material_code, mat_data in materials_data.items():
        section["materials"].append({
            "material_code": material_code,
            "fields": mat_data["fields"]
        })
        section["fields"].extend(mat_data["fields"])
    
    print(f"       Material sheet '{sheet_name}': Found {len(materials_data)} materials with {len(section['fields'])} total fields")
    
    return section


def extract_material_listing_fields(excel_path):
    """
    Extract yellow cell fields from the material listing sheet (e.g., 'Other Content').
    These are fields that apply to each material (like Unit Price, Quantity, etc.).
    Returns list of field definitions.
    """
    sheet_name = _find_other_content_sheet_name(excel_path) or "Other Content"
    file_ext = os.path.splitext(excel_path)[1].lower()
    material_fields = []
    
    try:
        print(f" Extracting material listing fields from '{sheet_name}'...")
        
        if file_ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(excel_path, formatting_info=True)
            
            try:
                sheet = wb.sheet_by_name(sheet_name)
            except:
                print(f"[WARN] Sheet '{sheet_name}' not found, skipping material fields")
                return []
            
            # Find header row (search first 10 rows or all rows if less)
            header_row_idx = 0
            for row_idx in range(min(10, sheet.nrows)):
                row_text = " ".join([str(sheet.cell_value(row_idx, c)) for c in range(sheet.ncols)]).lower()
                if any(kw in row_text for kw in ['price', 'quantity', 'manufacturer', 'material', 'code', 'name', 'description']):
                    header_row_idx = row_idx
                    break
            
            print(f"          Using row {header_row_idx + 1} as header row")
            
            # Find Name column (use same logic as extract_materials_from_excel)
            name_col_idx = None
            for col_idx in range(sheet.ncols):
                header_val = str(sheet.cell_value(header_row_idx, col_idx)).strip().lower()
                if 'name' in header_val.replace(" ", "").replace("_", ""):
                    name_col_idx = col_idx
                    col_letter = chr(65 + col_idx) if col_idx < 26 else f'Col{col_idx}'
                    print(f"Found Name column at: {col_letter}")
                    break
            
            if name_col_idx is None:
                print(f"No 'Name' column found in sheet")
                return []
            
            # Find first material row by searching Name column for 9-digit material codes (same as existing logic)
            material_row_idx = None
            for row_idx in range(header_row_idx + 1, sheet.nrows):
                name_value = str(sheet.cell_value(row_idx, name_col_idx))
                # Extract 9-digit material codes using regex (same as extract_materials_from_excel)
                material_codes = re.findall(r"\d{9}", name_value)
                if material_codes:
                    material_row_idx = row_idx
                    print(f"Found material row at: {row_idx + 1} with material code(s): {material_codes}")
                    break
            
            if not material_row_idx:
                print(f"[WARN] No material rows found in sheet")
                return []
            
            print(f"Checking row {material_row_idx + 1} for yellow cells (using flexible detection)")
            print(f"Total columns to scan: {sheet.ncols}")
            
            # Scan ALL columns for yellow cells (use flexible detection for material listing)
            for col_idx in range(sheet.ncols):
                # Get header for this column
                header_val = str(sheet.cell_value(header_row_idx, col_idx)).strip() if col_idx < sheet.ncols else ""
                col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                
                # Use the same function that works in other sheets, but with flexible detection
                is_yellow, color_rgb = is_yellow_cell_xls(sheet, material_row_idx, col_idx, wb, debug=True, strict=False)
                
                # Check if this is a TDS/file field (always include even if not yellow)
                # TDS files are required for all materials and must be included
                is_file_field = False
                if header_val:
                    inferred_type = infer_field_type(header_val)
                    is_file_field = (inferred_type == 'file')
                    # Also check explicitly for common TDS field names
                    if any(kw in header_val.lower() for kw in ['tds', 'technical data sheet']):
                        is_file_field = True
                
                if is_yellow or is_file_field:
                    cell_value = sheet.cell_value(material_row_idx, col_idx)
                    is_empty = (cell_value == "" or cell_value is None or 
                               (isinstance(cell_value, str) and cell_value.strip() == ""))
                    
                    if not header_val:
                        print(f"{col_letter}: Cell detected but NO HEADER - skipping")
                        continue  # Skip columns with no header
                    
                    # Log why this field was included
                    if is_file_field and not is_yellow:
                        print(f"{col_letter}: '{header_val}' - INCLUDED (TDS/file field - REQUIRED for all materials, not yellow)")
                    
                    # Skip common read-only identifier columns
                    skip_columns = ['number', 'name', 'alternative', 'bundle', 'tier', 'answer', 
                                  'description', 'material code', 'item text', 'material po text', 
                                  'comment', 'intend to respond', 'reason for not bidding']
                    if header_val.lower() in skip_columns:
                        print(f"           {col_letter}: '{header_val}' - SKIPPED (identifier column)")
                        continue  # Skip identifier/info columns
                    
                    label = header_val
                    
                    # Check if required (has * in header OR is a file field - files are always required)
                    is_required = '*' in label or 'required' in label.lower()
                    
                    # Infer field type
                    field_type = infer_field_type(label)
                    print(f"          Field '{label}' detected as type: '{field_type}'")
                    
                    # TDS/File fields are ALWAYS required for materials
                    if field_type == 'file':
                        is_required = True
                        print(f"          {col_letter}: TDS/File field marked as REQUIRED")
                    
                    # Check for dropdown options - try to extract from Excel first
                    dropdown_options = None
                    if field_type == 'dropdown':
                        print(f"          Extracting dropdown options for '{label}' at column {col_letter} (idx={col_idx})...")
                        # Try to extract dropdown options from Excel data validation or DV sheet
                        dropdown_options = get_dropdown_options_xls(wb, sheet, material_row_idx, col_idx, label)
                        print(f"          Dropdown extraction result: {len(dropdown_options) if dropdown_options else 0} options found")
                        
                        # If not found in Excel, use smart fallback ONLY for specific known fields
                        if not dropdown_options:
                            label_check = label.lower()
                            if 'country' in label_check or 'origin' in label_check:
                                # Country fields should have dropdown even if not in Excel
                                dropdown_options = ['Saudi Arabia', 'UAE', 'USA', 'China', 'Germany', 'Italy', 'France', 'UK', 'Japan', 'South Korea', 'India', 'Turkey', 'Spain', 'Canada', 'Brazil', 'Mexico', 'Other']
                                print(f"          {col_letter}: Detected COUNTRY field - using fallback country list")
                            elif 'factory' in label_check or 'own factory' in label_check or 'product own' in label_check:
                                # Factory fields are typically Yes/No
                                dropdown_options = ['Yes', 'No']
                                print(f"          {col_letter}: Detected FACTORY field - using fallback Yes/No")
                            elif 'vendor' in label_check or 'local' in label_check:
                                # Vendor fields are typically Yes/No
                                dropdown_options = ['Yes', 'No']
                                print(f"          {col_letter}: Detected VENDOR field - using fallback Yes/No")
                            else:
                                # For other fields (like delivery time, etc.), if no dropdown in Excel, 
                                # change field type to 'text' instead of forcing a dropdown
                                field_type = 'text'
                                print(f"         [WARN] {col_letter}: No dropdown values found in Excel for '{label}' - changing to TEXT input")
                    
                    print(f"          {col_letter}: '{label}' - Type: {field_type} - Blank: {is_empty} - Required: {is_required}")
                    
                    field = {
                        "id": f"material_field_{col_idx}",
                        "label": label.replace('*', '').strip(),
                        "type": field_type,
                        "required": is_required,
                        "col": col_idx,
                        "col_letter": col_letter,
                        "default_value": "" if is_empty else str(cell_value)
                    }
                    
                    # Add dropdown options if available
                    if dropdown_options:
                        field["options"] = dropdown_options
                        options_summary = f" - {len(dropdown_options)} dropdown options: {dropdown_options[:3]}{'...' if len(dropdown_options) > 3 else ''}"
                    else:
                        options_summary = " - NO DROPDOWN OPTIONS" if field_type == 'dropdown' else ""
                    
                    material_fields.append(field)
                    rgb_str = f"RGB({color_rgb[0]}, {color_rgb[1]}, {color_rgb[2]})" if color_rgb else "N/A"
                    print(f"         [OK] {col_letter}: '{label}' - Type: {field_type} - Required: {is_required} - {rgb_str}{options_summary}")
            
            # Summary of fields found
            print(f"       Material listing fields summary:")
            print(f"         Total fields found: {len(material_fields)}")
            file_fields = [f for f in material_fields if f['type'] == 'file']
            if file_fields:
                print(f"         [Attachment] TDS/File fields (REQUIRED for all materials): {len(file_fields)}")
                for ff in file_fields:
                    print(f"            - {ff['label']}")
            else:
                print(f"         [WARN]  No TDS/file fields found - materials will not have file upload option!")
            
            # Dropdown fields summary
            dropdown_fields = [f for f in material_fields if f['type'] == 'dropdown']
            if dropdown_fields:
                print(f"          Dropdown fields: {len(dropdown_fields)}")
                for df in dropdown_fields:
                    options_preview = df.get('options', [])[:3]
                    options_str = ', '.join(options_preview) + ('...' if len(df.get('options', [])) > 3 else '')
                    print(f"            - {df['label']}: [{options_str}]")
        
        else:  # .xlsx
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=False)
            
            if sheet_name not in wb.sheetnames:
                print(f"      [WARN] Sheet '{sheet_name}' not found, skipping material fields")
                return []
            
            sheet = wb[sheet_name]
            
            # Find header row (search first 10 rows or all rows if less)
            header_row_idx = 1
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                row_text = " ".join([str(sheet.cell(row_idx, c).value or "") for c in range(1, sheet.max_column + 1)]).lower()
                if any(kw in row_text for kw in ['price', 'quantity', 'manufacturer', 'material', 'code', 'name', 'description']):
                    header_row_idx = row_idx
                    break
            
            print(f"          Using row {header_row_idx} as header row")
            
            # Find Name column (use same logic as extract_materials_from_excel)
            name_col_idx = None
            for col_idx in range(1, sheet.max_column + 1):
                header_val = str(sheet.cell(header_row_idx, col_idx).value or "").strip().lower()
                if 'name' in header_val.replace(" ", "").replace("_", ""):
                    name_col_idx = col_idx
                    col_letter = sheet.cell(header_row_idx, col_idx).column_letter
                    print(f"          Found Name column at: {col_letter}")
                    break
            
            if name_col_idx is None:
                print(f"          No 'Name' column found in sheet")
                return []
            
            # Find first material row by searching Name column for 9-digit material codes (same as existing logic)
            material_row_idx = None
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                name_value = str(sheet.cell(row_idx, name_col_idx).value or "")
                # Extract 9-digit material codes using regex (same as extract_materials_from_excel)
                material_codes = re.findall(r"\d{9}", name_value)
                if material_codes:
                    material_row_idx = row_idx
                    print(f"          Found material row at: {row_idx} with material code(s): {material_codes}")
                    break
            
            if not material_row_idx:
                print(f"      [WARN] No material rows found in sheet")
                return []
            
            print(f"          Checking row {material_row_idx} for yellow cells (using flexible detection)")
            print(f"          Total columns to scan: {sheet.max_column}")
            
            # Scan ALL columns for yellow cells (use flexible detection for material listing)
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(material_row_idx, col_idx)
                
                # Get header for this column
                header_cell = sheet.cell(header_row_idx, col_idx)
                header_val = str(header_cell.value or "").strip()
                
                # Use the same function that works in other sheets, but with flexible detection
                is_yellow, rgb_tuple = is_yellow_cell_xlsx(cell, debug=True, strict=False)
                
                # Check if this is a TDS/file field (always include even if not yellow)
                # TDS files are required for all materials and must be included
                is_file_field = False
                if header_val:
                    inferred_type = infer_field_type(header_val)
                    is_file_field = (inferred_type == 'file')
                    # Also check explicitly for common TDS field names
                    if any(kw in header_val.lower() for kw in ['tds', 'technical data sheet']):
                        is_file_field = True
                
                if is_yellow or is_file_field:
                    cell_value = cell.value
                    is_empty = (cell_value is None or str(cell_value).strip() == "")
                    
                    if not header_val:
                        print(f"         [WARN]  {cell.column_letter}: Cell detected but NO HEADER - skipping")
                        continue  # Skip columns with no header
                    
                    # Log why this field was included
                    if is_file_field and not is_yellow:
                        print(f"         [Attachment] {cell.column_letter}: '{header_val}' - INCLUDED (TDS/file field - REQUIRED for all materials, not yellow)")
                    
                    # Skip common read-only identifier columns
                    skip_columns = ['number', 'name', 'alternative', 'bundle', 'tier', 'answer', 
                                  'description', 'material code', 'item text', 'material po text', 
                                  'comment', 'intend to respond', 'reason for not bidding']
                    if header_val.lower() in skip_columns:
                        print(f"           {cell.column_letter}: '{header_val}' - SKIPPED (identifier column)")
                        continue  # Skip identifier/info columns
                    
                    label = header_val
                    
                    # Check if required (has * in header OR is a file field - files are always required)
                    is_required = '*' in label or 'required' in label.lower()
                    
                    # Infer field type
                    field_type = infer_field_type(label)
                    
                    # TDS/File fields are ALWAYS required for materials
                    if field_type == 'file':
                        is_required = True
                        print(f"          {cell.column_letter}: TDS/File field marked as REQUIRED")
                    
                    # Check for dropdown options from Excel data validation
                    dropdown_options = None
                    if field_type == 'dropdown':
                        dropdown_options = get_dropdown_options_xlsx(sheet, cell)
                        if dropdown_options:
                            print(f"          {cell.column_letter}: Dropdown options found from Excel: {dropdown_options}")
                        else:
                            # Fallback: provide common options based on field name
                            label_check = label.lower()
                            if 'country' in label_check or 'origin' in label_check:
                                dropdown_options = ['Saudi Arabia', 'UAE', 'USA', 'China', 'Germany', 'Italy', 'France', 'UK', 'Japan', 'South Korea', 'India', 'Turkey', 'Spain', 'Canada', 'Brazil', 'Mexico', 'Other']
                                print(f"          {cell.column_letter}: Detected COUNTRY field - using country list")
                            elif 'factory' in label_check or 'own factory' in label_check or 'product own' in label_check:
                                dropdown_options = ['Yes', 'No']
                                print(f"          {cell.column_letter}: Detected FACTORY field - using Yes/No")
                            elif 'vendor' in label_check or 'local' in label_check:
                                dropdown_options = ['Yes', 'No']
                                print(f"          {cell.column_letter}: Detected VENDOR field - using Yes/No")
                            else:
                                dropdown_options = ['Yes', 'No', 'N/A']
                                print(f"          {cell.column_letter}: Generic dropdown - using Yes/No/N/A")
                            
                            print(f"          {cell.column_letter}: Fallback dropdown options: {dropdown_options}")
                    
                    print(f"          {cell.column_letter}: '{label}' - Type: {field_type} - Blank: {is_empty} - Required: {is_required}")
                    
                    field = {
                        "id": f"material_field_{col_idx}",
                        "label": label.replace('*', '').strip(),
                        "type": field_type,
                        "required": is_required,
                        "col": col_idx,
                        "col_letter": cell.column_letter,
                        "default_value": "" if is_empty else str(cell_value)
                    }
                    
                    # Add dropdown options if available
                    if dropdown_options:
                        field["options"] = dropdown_options
                    
                    material_fields.append(field)
                    rgb_str = f"RGB{rgb_tuple}" if rgb_tuple else "N/A"
                    print(f"         [OK] {cell.column_letter}: '{label}' - Type: {field_type} - Required: {is_required} - {rgb_str}")
            
            # Summary of fields found
            print(f"       Material listing fields summary:")
            print(f"         Total fields found: {len(material_fields)}")
            
            # File fields
            file_fields = [f for f in material_fields if f['type'] == 'file']
            if file_fields:
                print(f"         [Attachment] TDS/File fields (REQUIRED for all materials): {len(file_fields)}")
                for ff in file_fields:
                    print(f"            - {ff['label']}")
            else:
                print(f"         [WARN]  No TDS/file fields found - materials will not have file upload option!")
            
            # Dropdown fields
            dropdown_fields = [f for f in material_fields if f['type'] == 'dropdown']
            if dropdown_fields:
                print(f"          Dropdown fields: {len(dropdown_fields)}")
                for df in dropdown_fields:
                    options_preview = df.get('options', [])[:3]
                    options_str = ', '.join(options_preview) + ('...' if len(df.get('options', [])) > 3 else '')
                    print(f"            - {df['label']}: [{options_str}]")
        
        print(f"      [OK] Found {len(material_fields)} material listing fields from yellow cells")
        return material_fields
    
    except Exception as e:
        print(f"      [ERROR] Error extracting material listing fields: {e}")
        import traceback
        traceback.print_exc()
        return []


def parse_excel_for_dynamic_form(excel_path):
    """
    Parse Excel file and generate dynamic form structure based on yellow cells.
    For material sheets (like 'Other Content'), groups fields by material and uses column headers.
    Returns dict with sections (sheets) and fields (yellow cells).
    """
    form_structure = {
        "sections": [],
        "total_fields": 0,
        "material_listing_fields": []  # Fields that apply to each material
    }
    
    # Extract material listing fields (for the Materials Required section)
    material_listing_fields = extract_material_listing_fields(excel_path)
    form_structure["material_listing_fields"] = material_listing_fields
    
    file_ext = os.path.splitext(excel_path)[1].lower()
    
    try:
        if file_ext == '.xls':
            # Parse .xls file with xlrd
            import xlrd
            
            wb = xlrd.open_workbook(excel_path, formatting_info=True)
            
            for sheet_idx, sheet_name in enumerate(wb.sheet_names()):
                # Skip certain sheets (instructions, attachments)
                if any(skip in sheet_name.lower() for skip in ['instruction', 'attachment', 'dv_sheet']):
                    continue
                
                sheet = wb.sheet_by_index(sheet_idx)
                
                # Check if this is a material sheet (has "content" or "material" in name)
                is_material_sheet = any(kw in sheet_name.lower() for kw in ['content', 'material'])
                
                if is_material_sheet:
                    # For material sheets, group by material and use column headers
                    print(f"       Scanning MATERIAL sheet '{sheet_name}' for blank yellow cells...")
                    section = parse_material_sheet_xls(sheet, sheet_idx, sheet_name, wb)
                    if section and section.get("materials"):
                        form_structure["sections"].append(section)
                        form_structure["total_fields"] += len(section.get("fields", []))
                else:
                    # For non-material sheets, use adjacent cell labels
                    section = {
                        "sheet_name": sheet_name,
                        "sheet_index": sheet_idx,
                        "fields": []
                    }
                    
                    print(f"       Scanning sheet '{sheet_name}' for blank yellow cells...")
                    yellow_count = 0
                    blank_yellow_count = 0
                    
                    for row_idx in range(sheet.nrows):
                        for col_idx in range(sheet.ncols):
                            is_yellow, color_rgb = is_yellow_cell_xls(sheet, row_idx, col_idx, wb, debug=False)
                            
                            if is_yellow:
                                yellow_count += 1
                                cell_value = sheet.cell_value(row_idx, col_idx)
                                is_empty = (cell_value == "" or cell_value is None or 
                                           (isinstance(cell_value, str) and cell_value.strip() == ""))
                                
                                col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                                rgb_str = f"RGB({color_rgb[0]}, {color_rgb[1]}, {color_rgb[2]})" if color_rgb else "N/A"
                                value_str = f"'{cell_value}'" if cell_value else "(empty)"
                                print(f" {col_letter}{row_idx+1}: {rgb_str} - Value: {value_str} - Blank: {is_empty}")
                                
                                if is_empty:
                                    blank_yellow_count += 1
                                    label = get_cell_label(sheet, row_idx, col_idx, is_xlsx=False)
                                    field_type = infer_field_type(label)
                                    
                                    # Check for dropdown options - try to extract from Excel first
                                    dropdown_options = None
                                    if field_type == 'dropdown':
                                        # Try to extract dropdown options from Excel data validation or DV sheet
                                        dropdown_options = get_dropdown_options_xls(wb, sheet, row_idx, col_idx, label)
                                        
                                        # If not found in Excel, use smart fallback ONLY for specific known fields
                                        if not dropdown_options:
                                            label_check = label.lower()
                                            if 'country' in label_check or 'origin' in label_check:
                                                dropdown_options = ['Saudi Arabia', 'UAE', 'USA', 'China', 'Germany', 'Italy', 'France', 'UK', 'Japan', 'South Korea', 'India', 'Turkey', 'Spain', 'Canada', 'Brazil', 'Mexico', 'Other']
                                            elif 'factory' in label_check or 'own factory' in label_check or 'product own' in label_check:
                                                dropdown_options = ['Yes', 'No']
                                            elif 'vendor' in label_check or 'local' in label_check:
                                                dropdown_options = ['Yes', 'No']
                                            else:
                                                # If no dropdown in Excel, change field type to 'text' instead of forcing dropdown
                                                field_type = 'text'
                                    
                                    field = {
                                        "id": f"field_{sheet_idx}_{row_idx}_{col_idx}",
                                        "label": label,
                                        "type": field_type,
                                        "required": True,
                                        "row": row_idx,
                                        "col": col_idx,
                                        "sheet_index": sheet_idx,
                                        "sheet_name": sheet_name
                                    }
                                    
                                    # Add dropdown options if available
                                    if dropdown_options:
                                        field["options"] = dropdown_options
                                    
                                    section["fields"].append(field)
                                    form_structure["total_fields"] += 1
                                    print(f"            [OK] INCLUDED in form - Label: '{label}' - Type: {field_type}")
                    
                    print(f"       Sheet '{sheet_name}': {yellow_count} yellow cells, {blank_yellow_count} blank (included in form)")
                    
                    if section["fields"]:
                        form_structure["sections"].append(section)
        
        else:
            # Parse .xlsx file with openpyxl
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_path, data_only=False)
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                # Skip certain sheets
                if any(skip in sheet_name.lower() for skip in ['instruction', 'attachment', 'dv_sheet']):
                    continue
                
                sheet = wb[sheet_name]
                section = {
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_idx,
                    "fields": []
                }
                
                # Find all BLANK yellow cells in this sheet (cells to fill)
                print(f"       Scanning sheet '{sheet_name}' for blank yellow cells...")
                yellow_count = 0
                blank_yellow_count = 0
                
                for row in sheet.iter_rows():
                    for cell in row:
                        is_yellow, rgb_tuple = is_yellow_cell_xlsx(cell, debug=False)
                        
                        if is_yellow:
                            yellow_count += 1
                            # Check if cell is blank/empty
                            cell_value = cell.value
                            is_empty = (cell_value is None or cell_value == "" or 
                                       (isinstance(cell_value, str) and cell_value.strip() == ""))
                            
                            # Debug output for all yellow cells
                            rgb_str = f"RGB{rgb_tuple}" if rgb_tuple else "N/A"
                            value_str = f"'{cell_value}'" if cell_value else "(empty)"
                            print(f"          {cell.coordinate}: {rgb_str} - Value: {value_str} - Blank: {is_empty}")
                            
                            # Only include BLANK yellow cells (cells user needs to fill)
                            if is_empty:
                                blank_yellow_count += 1
                                row_idx = cell.row - 1  # Convert to 0-indexed
                                col_idx = cell.column - 1  # Convert to 0-indexed
                                
                                # Get label from adjacent cells
                                label = get_cell_label(sheet, row_idx, col_idx, is_xlsx=True)
                                field_type = infer_field_type(label)
                                
                                # Check for dropdown options from Excel data validation
                                dropdown_options = None
                                if field_type == 'dropdown':
                                    dropdown_options = get_dropdown_options_xlsx(sheet, cell)
                                    if not dropdown_options:
                                        # Fallback: provide common options based on field name
                                        label_check = label.lower()
                                        if 'country' in label_check or 'origin' in label_check:
                                            dropdown_options = ['Saudi Arabia', 'UAE', 'USA', 'China', 'Germany', 'Italy', 'France', 'UK', 'Japan', 'South Korea', 'India', 'Turkey', 'Spain', 'Canada', 'Brazil', 'Mexico', 'Other']
                                        elif 'factory' in label_check or 'own factory' in label_check or 'product own' in label_check:
                                            dropdown_options = ['Yes', 'No']
                                        elif 'vendor' in label_check or 'local' in label_check:
                                            dropdown_options = ['Yes', 'No']
                                        else:
                                            dropdown_options = ['Yes', 'No', 'N/A']
                                
                                field = {
                                    "id": f"field_{sheet_idx}_{row_idx}_{col_idx}",
                                    "label": label,
                                    "type": field_type,
                                    "required": True,
                                    "row": row_idx,
                                    "col": col_idx,
                                    "sheet_index": sheet_idx,
                                    "sheet_name": sheet_name
                                }
                                
                                # Add dropdown options if available
                                if dropdown_options:
                                    field["options"] = dropdown_options
                                
                                section["fields"].append(field)
                                form_structure["total_fields"] += 1
                                print(f"            [OK] INCLUDED in form - Label: '{label}' - Type: {field_type}")
                
                print(f"       Sheet '{sheet_name}': {yellow_count} yellow cells, {blank_yellow_count} blank (included in form)")
                
                if section["fields"]:
                    form_structure["sections"].append(section)
    
    except Exception as e:
        print(f"[ERROR] Error parsing Excel for dynamic form: {e}")
        import traceback
        traceback.print_exc()
    
    return form_structure

def _try_materials_from_dataverse(rfp_id: str):
    """
    Try to build dialog response from stored Matched_Data JSON in Dataverse.
    Returns dict (response body) if successful, None if fallback needed.
    """
    try:
        filter_val = rfp_id.replace("'", "''")
        result = DATAVERSE.query_rows(
            get_setting("RFP_ACTIVITY_LOG_TABLE_API", "cr673_bahra_rfps_v2s"),
            filter_expr=f"RFP_ID eq '{filter_val}'",
            select="RFP_ID,Matched_Data",
            top=1,
            table_logical_name=get_setting("RFP_ACTIVITY_LOG_TABLE_LOGICAL", "cr673_bahra_rfps_v2"),
            use_display_names=True
        )
        rows = result.get("value", []) if isinstance(result, dict) else []
        if not rows:
            return None

        matched_data_str = (rows[0].get("Matched_Data") or "").strip()
        if not matched_data_str:
            return None

        data = json.loads(matched_data_str)

        materials_list = []

        # New categorized format (dict with summary)
        if isinstance(data, dict) and "summary" in data:
            for item in data.get("exact_matches", []):
                materials_list.append({
                    "material_code": item.get("material_code", ""),
                    "name": item.get("excel_name", ""),
                    "description": item.get("excel_description", ""),
                    "is_matched": True,
                    "match_method": "exact_code",
                    "master_description": item.get("material_description", ""),
                    "master_data": {},
                    "selected": False,
                    "reason": "",
                })
            for item in data.get("keyword_matches", []):
                materials_list.append({
                    "material_code": item.get("material_code", ""),
                    "name": item.get("excel_name", ""),
                    "description": item.get("excel_description", ""),
                    "is_matched": True,
                    "match_method": "keyword",
                    "matched_keyword": item.get("matched_keyword", ""),
                    "master_description": item.get("material_description", ""),
                    "master_data": {},
                    "selected": False,
                    "reason": "",
                })
            for item in data.get("not_matched", []):
                materials_list.append({
                    "material_code": item.get("material_code", ""),
                    "name": item.get("excel_name", ""),
                    "description": item.get("excel_description", ""),
                    "is_matched": False,
                    "match_method": None,
                    "master_description": "",
                    "master_data": {},
                    "selected": False,
                    "reason": "",
                })

            s = data["summary"]
            return {
                "ok": True,
                "rfp_id": rfp_id,
                "total_materials": data.get("total_items", len(materials_list)),
                "matched_count": s.get("exact_match_count", 0) + s.get("keyword_match_count", 0),
                "unmatched_count": s.get("not_matched_count", 0),
                "exact_code_matches": s.get("exact_match_count", 0),
                "keyword_matches": s.get("keyword_match_count", 0),
                "match_percentage": s.get("match_percentage", 0),
                "materials": materials_list,
            }

        # Old flat format (list) — backward compatibility
        if isinstance(data, list) and data:
            if not any("is_matched" in item for item in data):
                return None

            for item in data:
                is_matched = bool(item.get("is_matched", True))
                raw_method = item.get("MatchMethod")
                if not is_matched:
                    match_method = None
                elif raw_method and str(raw_method).lower() == "keyword":
                    match_method = "keyword"
                else:
                    match_method = "exact_code"

                materials_list.append({
                    "material_code": str(item.get("ExtractedMaterial") or item.get("Material") or "").strip(),
                    "name": str(item.get("ExcelName") or item.get("ColumnName") or "").strip(),
                    "description": str(item.get("ExcelDescription") or "").strip(),
                    "is_matched": is_matched,
                    "match_method": match_method,
                    "master_description": str(item.get("Material Description") or "").strip() if is_matched else "",
                    "master_data": {},
                    "selected": False,
                    "reason": "",
                })

            total_materials = len(materials_list)
            matched_count = sum(1 for m in materials_list if m["is_matched"])
            exact_code_matches = sum(1 for m in materials_list if m.get("match_method") == "exact_code")
            keyword_matches_count = sum(1 for m in materials_list if m.get("match_method") == "keyword")
            match_percentage = round((matched_count / total_materials * 100) if total_materials > 0 else 0, 1)

            return {
                "ok": True,
                "rfp_id": rfp_id,
                "total_materials": total_materials,
                "matched_count": matched_count,
                "unmatched_count": total_materials - matched_count,
                "exact_code_matches": exact_code_matches,
                "keyword_matches": keyword_matches_count,
                "match_percentage": match_percentage,
                "materials": materials_list,
            }

        return None

    except Exception as e:
        print(f"[MaterialDialog] Dataverse read failed for {rfp_id}, falling back to live matching: {e}")
        return None


@router.get("/rfp/{rfp_id}/materials")
async def get_rfp_materials(request: Request, rfp_id: str, company: str = None):
    """
    Extract materials from RFP Excel file and match with master file.
    Returns materials with match status for display in modal.
    Primary path: reads from stored Matched_Data JSON in Dataverse.
    Fallback: live-matches from Excel + master CSV (for old records).
    """
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        # PRIMARY PATH: Try stored Dataverse data
        stored_result = _try_materials_from_dataverse(rfp_id)
        if stored_result is not None:
            return JSONResponse(stored_result)

        # FALLBACK PATH: Live matching (for old records without new format)
        # Initialize GraphClient for master file download
        graph_client = GraphClient(
            get_setting("CLIENT_ID", ""), get_setting("CLIENT_SECRET", ""), get_setting("TENANT_ID", ""),
            get_setting("SHAREPOINT_HOSTNAME", ""), get_setting("SITE_PATH", ""), get_setting("DRIVE_NAME", "")
        )
        graph_client.auth()
        graph_client.resolve_site_and_drive()

        # Resolve company - either from query param or by searching
        selected_company = (company or "").strip()
        if not selected_company:
            # Try to find company from database or file system
            _, found_company = find_rfp_file_across_companies(rfp_id)
            if found_company:
                selected_company = found_company
            else:
                selected_company = get_setting("COMPANY_NAME", "")
                print(f"[WARN] Company not found for RFP {rfp_id}, using default: {selected_company}")

        # Find Excel file — download from SharePoint if not local
        excel_path = ensure_rfp_excel_from_sharepoint(rfp_id, selected_company, graph_client)

        if not excel_path or not os.path.exists(excel_path):
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found for RFP: {rfp_id} (company: {selected_company}). File not found locally or on SharePoint."
            )
        
        # Extract ALL materials from Excel (no intent filter — counts all items)
        materials_data = extract_materials_from_excel(excel_path, include_details=True, filter_by_intent=False)

        if not materials_data:
            return JSONResponse({
                "ok": True,
                "materials": [],
                "message": "No materials found in RFP file"
            })

        # Get cached master data (downloads only if needed)
        _output_dir = get_setting("OUTPUT_DIR", "")
        master_csv_local = os.path.join(_output_dir, "master_material.csv")
        master = get_cached_master_data(graph_client, master_csv_local)

        # Get cached keywords (downloads only if needed)
        keywords_csv_local = os.path.join(_output_dir, "unique_keywords.csv")
        keywords_list = get_cached_keywords(graph_client, keywords_csv_local)

        # Find 'material' column in master CSV
        master_col = find_column_name(master.columns, "material")
        if not master_col:
            raise HTTPException(
                status_code=500,
                detail="No 'material' column found in master CSV"
            )

        # Match materials using same logic as download time (exact code + keyword)
        materials_list = []
        for mat_data in materials_data:
            mat_code = mat_data["material_code"]
            name_text = mat_data.get("name", "")
            description_text = mat_data.get("description", "")

            # Method 1: Exact Material Code Match
            matched_rows = master[master[master_col].astype(str) == mat_code]
            is_matched = not matched_rows.empty
            match_method = "exact_code" if is_matched else None

            # Method 2: Keyword Matching (only if exact match failed)
            if not is_matched and keywords_list:
                name_keywords = extract_keywords_from_text(name_text)
                desc_keywords = extract_keywords_from_text(description_text)
                all_material_keywords = set(name_keywords + desc_keywords)

                for csv_keyword in keywords_list:
                    for mat_keyword in all_material_keywords:
                        if csv_keyword in mat_keyword or mat_keyword in csv_keyword:
                            is_matched = True
                            match_method = "keyword"
                            break
                    if is_matched:
                        break

            material_info = {
                "material_code": mat_code,
                "name": name_text,
                "description": description_text,
                "is_matched": is_matched,
                "match_method": match_method,  # "exact_code", "keyword", or None
                "selected": False,
                "reason": ""
            }

            # If matched, add additional info from master file
            if is_matched and not matched_rows.empty:
                matched_row = matched_rows.iloc[0]
                desc_col = find_column_name(master.columns, "description") or find_column_name(master.columns, "material description")
                if desc_col:
                    material_info["master_description"] = str(matched_row.get(desc_col, ""))
                else:
                    material_info["master_description"] = ""
                # Sanitize master_data: replace NaN/Infinity with None for JSON serialization
                raw_dict = matched_row.to_dict()
                material_info["master_data"] = {
                    k: (None if (isinstance(v, float) and (math.isnan(v) or math.isinf(v))) else v)
                    for k, v in raw_dict.items()
                }
            else:
                material_info["master_description"] = ""
                material_info["master_data"] = {}

            materials_list.append(material_info)

        # Count matches by method
        exact_code_matches = sum(1 for m in materials_list if m.get("match_method") == "exact_code")
        keyword_matches = sum(1 for m in materials_list if m.get("match_method") == "keyword")
        matched_count = sum(1 for m in materials_list if m["is_matched"])
        total_materials = len(materials_list)

        # Calculate matching percentage
        match_percentage = round((matched_count / total_materials * 100) if total_materials > 0 else 0, 1)

        # Write back Matched_Data to Dataverse in categorized format
        try:
            wb_exact = []
            wb_keyword = []
            wb_not_matched = []
            for m in materials_list:
                item = {
                    "material_code": m["material_code"],
                    "excel_name": m.get("name", ""),
                    "excel_description": m.get("description", ""),
                    "row_number": 0,
                    "column_name": "",
                }
                if m.get("match_method") == "exact_code":
                    item["material_description"] = m.get("master_description", "")
                    wb_exact.append(item)
                elif m.get("match_method") == "keyword":
                    item["material_description"] = m.get("master_description", "")
                    item["matched_keyword"] = m.get("matched_keyword", "")
                    wb_keyword.append(item)
                else:
                    wb_not_matched.append(item)

            categorized = {
                "rfp_id": rfp_id,
                "source_file": os.path.basename(excel_path) if excel_path else "",
                "rfp_end_date": "",
                "total_items": total_materials,
                "summary": {
                    "exact_match_count": len(wb_exact),
                    "keyword_match_count": len(wb_keyword),
                    "not_matched_count": len(wb_not_matched),
                    "match_percentage": match_percentage,
                },
                "exact_matches": wb_exact,
                "keyword_matches": wb_keyword,
                "not_matched": wb_not_matched,
            }
            matched_data_json = json.dumps(categorized)
            filter_val = rfp_id.replace("'", "''")
            existing = DATAVERSE.query_rows(
                get_setting("RFP_ACTIVITY_LOG_TABLE_API", "cr673_bahra_rfps_v2s"),
                filter_expr=f"RFP_ID eq '{filter_val}'",
                select="RFP_ID,Matched_Data",
                top=1,
                table_logical_name=get_setting("RFP_ACTIVITY_LOG_TABLE_LOGICAL", "cr673_bahra_rfps_v2"),
                use_display_names=True
            )
            rows_found = existing.get("value", []) if isinstance(existing, dict) else []
            if rows_found:
                existing_md = (rows_found[0].get("Matched_Data") or "").strip()
                if not existing_md:
                    # Resolve primary key for update
                    _logical = get_setting("RFP_ACTIVITY_LOG_TABLE_LOGICAL", "cr673_bahra_rfps_v2")
                    try:
                        _colmap = DATAVERSE.get_column_mapping(_logical)
                        _logical_to_display = {v: k for k, v in _colmap.items()}
                    except Exception:
                        _logical_to_display = {}
                    _pk_logical = f"{_logical}id"
                    _pk_display = _logical_to_display.get(_pk_logical)
                    record_id = (rows_found[0].get(_pk_display) if _pk_display else None) or rows_found[0].get(_pk_logical)
                    if record_id:
                        DATAVERSE.update_row(
                            get_setting("RFP_ACTIVITY_LOG_TABLE_API", "cr673_bahra_rfps_v2s"),
                            record_id,
                            {"Matched_Data": matched_data_json},
                            table_logical_name=_logical
                        )
                        print(f"[MaterialDialog] Wrote back Matched_Data for {rfp_id}")
        except Exception as wb_err:
            print(f"[MaterialDialog] Write-back failed for {rfp_id}: {wb_err}")

        return JSONResponse({
            "ok": True,
            "materials": materials_list,
            "rfp_id": rfp_id,
            "total_materials": total_materials,
            "matched_count": matched_count,
            "unmatched_count": total_materials - matched_count,
            "exact_code_matches": exact_code_matches,
            "keyword_matches": keyword_matches,
            "match_percentage": match_percentage
        })

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error processing materials: {str(e)}")

@router.get("/rfp/{rfp_id}/dynamic-form-structure")
async def get_dynamic_form_structure(request: Request, rfp_id: str, company: str = None):
    """
    Parse ORIGINAL Excel file and return dynamic form structure based on yellow cells.
    This generates the form fields that users need to fill.
    """
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        print(f"\n Generating dynamic form for RFP: {rfp_id}")

        # Resolve company - either from query param or by searching
        selected_company = (company or "").strip()
        if not selected_company:
            # Try to find company from database or file system
            _, found_company = find_rfp_file_across_companies(rfp_id)
            if found_company:
                selected_company = found_company
            else:
                selected_company = get_setting("COMPANY_NAME", "")
                print(f"[WARN] Company not found for RFP {rfp_id}, using default: {selected_company}")

        # Get ORIGINAL Excel file path (not unprotected) with company
        excel_path = get_rfp_excel_file_path(rfp_id, selected_company)

        if not os.path.exists(excel_path):
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found for RFP: {rfp_id}"
            )
        
        print(f" Parsing Excel: {excel_path}")
        
        # Parse Excel and generate form structure
        form_structure = parse_excel_for_dynamic_form(excel_path)
        
        print(f"[OK] Found {form_structure['total_fields']} yellow cells across {len(form_structure['sections'])} sheets")
        
        # Add RFP ID to response
        form_structure["rfp_id"] = rfp_id
        
        return JSONResponse({
            "ok": True,
            "form_structure": form_structure
        })
    
    except Exception as e:
        print(f"[ERROR] Error generating dynamic form: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating form: {str(e)}")


@router.get("/rfp/batch-match-percentages")
async def get_batch_match_percentages(request: Request, rfp_ids: str = Query(...), companies: str = Query(default="")):
    """
    Get match percentages for multiple RFPs in one request (optimized with caching).
    rfp_ids: comma-separated list of RFP IDs
    companies: JSON object mapping rfp_id -> company_name (optional)
    """
    if not request.session.get("user"):
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        rfp_id_list = [r.strip() for r in rfp_ids.split('|') if r.strip()]

        # Parse company mapping
        company_map = {}
        if companies:
            try:
                company_map = json.loads(companies)
            except Exception:
                company_map = {}

        if not rfp_id_list:
            return JSONResponse({
                "ok": True,
                "results": {}
            })

        # Initialize GraphClient once
        results = {}

        # Phase A: Collect cache hits
        uncached_ids = []
        for rfp_id in rfp_id_list:
            if rfp_id in _MATCH_PERCENTAGE_CACHE:
                cached = _MATCH_PERCENTAGE_CACHE[rfp_id]
                if cached.get("cache_version") == _MATCH_CACHE_VERSION:
                    results[rfp_id] = {
                        "match_percentage": cached["match_percentage"],
                        "total_materials": cached["total_materials"],
                        "matched_count": cached["matched_count"]
                    }
                    continue
            uncached_ids.append(rfp_id)

        # Phase B: Batch Dataverse query for all uncached RFPs at once
        fallback_ids = []
        if uncached_ids:
            dv_results = _batch_get_match_percentages_from_dataverse(uncached_ids)
            for rfp_id in uncached_ids:
                dv_result = dv_results.get(rfp_id)
                if dv_result is not None:
                    # Cache the Dataverse result
                    dv_result["source"] = "dataverse"
                    dv_result["cache_version"] = _MATCH_CACHE_VERSION
                    _MATCH_PERCENTAGE_CACHE[rfp_id] = dv_result
                    results[rfp_id] = {
                        "match_percentage": dv_result["match_percentage"],
                        "total_materials": dv_result["total_materials"],
                        "matched_count": dv_result["matched_count"]
                    }
                else:
                    fallback_ids.append(rfp_id)

        # Phase C: Parallel Excel fallback for remaining RFPs
        # Only init SharePoint + load master data when actually needed
        if fallback_ids:
            print(f"[BatchMatch] Excel fallback for {len(fallback_ids)} RFPs (parallel, max 5 workers)")

            graph_client = GraphClient(
                get_setting("CLIENT_ID", ""), get_setting("CLIENT_SECRET", ""), get_setting("TENANT_ID", ""),
                get_setting("SHAREPOINT_HOSTNAME", ""), get_setting("SITE_PATH", ""), get_setting("DRIVE_NAME", "")
            )
            graph_client.auth()
            graph_client.resolve_site_and_drive()

            # Load materials from Dataverse first, fallback to SharePoint CSV
            from services.master_data_service import get_all_materials_for_matching, get_all_keywords_for_matching
            dv_materials = get_all_materials_for_matching()
            if dv_materials:
                master = pd.DataFrame({"material": dv_materials})
                master_col = "material"
            else:
                _output_dir = get_setting("OUTPUT_DIR", "")
                master = get_cached_master_data(graph_client, os.path.join(_output_dir, "master_material.csv"))
                master_col = find_column_name(master.columns, "material")
                if not master_col:
                    raise HTTPException(status_code=500, detail="No 'material' column found in master CSV")

            keywords_list = get_all_keywords_for_matching()
            if not keywords_list:
                _output_dir = get_setting("OUTPUT_DIR", "")
                keywords_list = get_cached_keywords(graph_client, os.path.join(_output_dir, "unique_keywords.csv"))

            master_code_set = set(master[master_col].astype(str))

            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {
                    executor.submit(
                        calculate_match_percentage_optimized,
                        rfp_id, master, master_col, keywords_list,
                        company=company_map.get(rfp_id),
                        graph_client=graph_client,
                        master_code_set=master_code_set,
                        skip_dataverse=True
                    ): rfp_id
                    for rfp_id in fallback_ids
                }
                for future in as_completed(futures):
                    rfp_id = futures[future]
                    try:
                        result = future.result()
                        results[rfp_id] = {
                            "match_percentage": result["match_percentage"],
                            "total_materials": result["total_materials"],
                            "matched_count": result["matched_count"]
                        }
                    except Exception as e:
                        print(f"[BatchMatch] Error for {rfp_id}: {e}")
                        results[rfp_id] = {
                            "match_percentage": 0,
                            "total_materials": 0,
                            "matched_count": 0,
                            "error": str(e)
                        }

        print(f"[BatchMatch] Done: {len(results)} RFPs ({len(rfp_id_list) - len(uncached_ids)} cached, {len(uncached_ids) - len(fallback_ids)} from Dataverse, {len(fallback_ids)} from Excel)")

        return JSONResponse({
            "ok": True,
            "results": results
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@router.post("/submit-rfp-final")
async def submit_rfp_final(request: Request, user: dict = Depends(require_permission("rfp.submit"))):
    """
    Handle final RFP submission with all data and TDS file uploads.
    This endpoint:
    1. Saves TDS files to TDS-files folder
    2. Writes all submission data to Excel file
    3. Saves Excel file to rfp-upload-file folder
    """
    
    try:
        # Parse FormData
        form = await request.form()

        rfp_id = form.get("rfp_id")
        materials_data = form.get("materials_data")
        dynamic_fields_str = form.get("dynamic_fields")

        # Validate rfp_id
        if not rfp_id or not rfp_id.strip():
            raise HTTPException(status_code=400, detail="rfp_id is required")
        rfp_id = rfp_id.strip()

        # Validate rfp_id doesn't contain path traversal characters
        if ".." in rfp_id or "/" in rfp_id or "\\" in rfp_id:
            raise HTTPException(status_code=400, detail="Invalid rfp_id format")

        # Get company parameter - required for file path operations
        company = form.get("company", "").strip() if form.get("company") else ""

        # If company not provided, try to find it from database or file system
        if not company:
            from helpers.core_helper import find_rfp_file_across_companies, get_rfp_company_name
            # First try database lookup
            company = get_rfp_company_name(rfp_id)
            if not company:
                # Fall back to file system search
                _, found_company = find_rfp_file_across_companies(rfp_id)
                company = found_company

        if not company:
            # Default to COMPANY_NAME if still not found
            company = get_setting("COMPANY_NAME", "")
            print(f"   [WARN] Company not provided, using default: {company}")

        print(f"    Company: {company}")

        # Parse materials data from JSON string with error handling
        try:
            materials = json.loads(materials_data) if materials_data else []
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON in materials_data: {str(e)}")

        # Parse dynamic form fields from JSON string with error handling
        try:
            dynamic_fields = json.loads(dynamic_fields_str) if dynamic_fields_str else {}
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON in dynamic_fields: {str(e)}")

        # Collect all TDS files
        # Files come with keys like: tds_file_0, tds_file_1, OR material_0_file_0, material_1_file_0 (from material listing)
        tds_files = []
        tds_file_keys = []
        for key, value in form.items():
            if key.startswith("tds_file_") or (key.startswith("material_") and "_file_" in key):
                tds_files.append(value)
                tds_file_keys.append(key)

        print(f" Final RFP Submission Received:")
        print(f"   RFP ID: {rfp_id}")
        print(f"   Company: {company}")
        print(f"   Dynamic Fields: {len(dynamic_fields)} fields")
        print(f"   Materials: {len(materials)} items")
        print(f"   TDS Files: {len(tds_files)} files")
        if tds_files:
            print(f"   [Attachment] TDS File keys found: {tds_file_keys}")

        # Get folders - now with company parameter
        tds_folder = get_rfp_tds_folder_path(rfp_id, company)
        saved_excel_path = get_rfp_saved_excel_file_path(rfp_id, company)
        original_excel_path = get_rfp_excel_file_path(rfp_id, company)
        
        if not os.path.exists(original_excel_path):
            raise HTTPException(status_code=404, detail=f"Original Excel file not found for RFP: {rfp_id}")
        
        # 1. Save TDS files
        print(f"[Saved] Saving TDS files to: {tds_folder}")
        for upload_file in tds_files:
            # Read file content
            file_content = await upload_file.read()

            # Sanitize filename to prevent directory traversal
            safe_filename = os.path.basename(upload_file.filename)
            if not safe_filename or ".." in safe_filename:
                print(f"   [WARN] Skipping invalid filename: {upload_file.filename}")
                continue

            # Save to TDS-files folder
            file_path = os.path.join(tds_folder, safe_filename)
            with open(file_path, 'wb') as f:
                f.write(file_content)

            print(f"   [OK] Saved: {safe_filename}")
        
        # 2. Get the unprotected Excel file (not the original protected one)
        print(f" Looking for unprotected Excel file...")
        
        # Find the unprotected version
        base_name = os.path.basename(original_excel_path)
        name_without_ext, ext = os.path.splitext(base_name)
        unprotected_filename = f"{name_without_ext}_unprotected{ext}"
        unprotected_path = os.path.join(os.path.dirname(original_excel_path), unprotected_filename)
        
        # If unprotected version doesn't exist, use unprotect function
        if not os.path.exists(unprotected_path):
            print(f"[WARN] Unprotected file not found, creating it...")
            try:
                from helpers.unprotect_xls import unprotect_excel_file
                unprotected_path = unprotect_excel_file(original_excel_path, unprotected_path)
            except Exception as e:
                print(f"[WARN] Could not unprotect file: {e}. Using original file.")
                unprotected_path = original_excel_path
        
        print(f" Using file: {unprotected_path}")
        
        # 3. Copy unprotected file to rfp-upload-file folder (keep same format)
        import shutil
        print(f" Copying unprotected file to rfp-upload-file folder...")
        shutil.copy2(unprotected_path, saved_excel_path)
        
        # 4. Add submission details sheets - handle .xls and .xlsx differently
        print(f" Adding submission data to Excel...")
        
        is_old_format = saved_excel_path.endswith('.xls') and not saved_excel_path.endswith('.xlsx')
        
        if is_old_format:
            # For .xls files, use xlrd and xlwt to preserve format
            print(f"   Using xlrd/xlwt for .xls format...")
            try:
                import xlrd
                from xlwt import Workbook, easyxf
                from xlutils.copy import copy as xlutils_copy
                
                # Open the existing .xls file
                rb = xlrd.open_workbook(saved_excel_path, formatting_info=True)
                wb = xlutils_copy(rb)
                
                # Find and update existing sheets
                sheet_names = rb.sheet_names()
                print(f"    Found sheets: {sheet_names}")
                
                # Write dynamic form fields to their respective cells
                print(f"    Writing {len(dynamic_fields)} dynamic form fields...")
                for field_id, field_data in dynamic_fields.items():
                    try:
                        sheet_idx = field_data['sheet_index']
                        row = field_data['row']
                        col = field_data['col']
                        value = field_data['value']
                        sheet_name = field_data['sheet_name']
                        
                        # Get writable sheet
                        ws = wb.get_sheet(sheet_idx)
                        ws.write(row, col, value)
                        
                        print(f"      [OK] {sheet_name} - Cell ({row}, {col}): {value}")
                    except Exception as e:
                        print(f"      [WARN] Error writing field {field_id}: {e}")
                
                # 3. Update "Other Content" sheet with material information
                other_content_idx = None
                for idx, name in enumerate(sheet_names):
                    if 'other' in name.lower() and 'content' in name.lower():
                        other_content_idx = idx
                        break
                
                if other_content_idx is not None:
                    print(f"    Updating Other Content sheet with material details...")
                    ws_content = wb.get_sheet(other_content_idx)
                    
                    # Read existing sheet to find where to place data
                    content_sheet = rb.sheet_by_index(other_content_idx)
                    
                    # Find column indices by searching headers (map all columns dynamically)
                    header_row_idx = 0  # Usually headers are in first row
                    col_map = {}  # Maps: lowercase_header -> col_idx
                    
                    for col_idx in range(content_sheet.ncols):
                        try:
                            header_val = str(content_sheet.cell_value(header_row_idx, col_idx)).strip()
                            if header_val:
                                # Store original header and normalized version for matching
                                col_map[header_val] = col_idx
                                # Also store normalized version (without * and lowercase)
                                normalized_header = header_val.replace('*', '').strip().lower()
                                col_map[normalized_header] = col_idx
                        except:
                            continue
                    
                    print(f"    Found {len(col_map)} column mappings")
                    
                    # Find rows with material codes and update them directly using column mapping
                    for row_idx in range(1, content_sheet.nrows):
                        try:
                            # Look for material code in the row (usually in name column)
                            row_text = ""
                            for col_idx in range(min(5, content_sheet.ncols)):
                                cell_val = str(content_sheet.cell_value(row_idx, col_idx))
                                row_text += cell_val + " "
                            
                            # Check if any of our materials match this row
                            for mat in materials:
                                mat_code = mat.get('material_code', '')
                                if mat_code and mat_code in row_text:
                                    print(f"   [OK] Updating row {row_idx + 1} for material {mat_code}")
                                    
                                    # Get all dynamic fields from the material
                                    material_fields = mat.get('fields', {})
                                    
                                    # Write ALL material fields dynamically by matching field labels to column headers
                                    fields_written = 0
                                    for field_label, field_value in material_fields.items():
                                        if not field_value:  # Skip empty values
                                            continue
                                        
                                        # Normalize field label for matching
                                        normalized_label = field_label.replace('*', '').strip().lower()
                                        
                                        # Try to find matching column
                                        col_idx = None
                                        if field_label in col_map:
                                            col_idx = col_map[field_label]
                                        elif normalized_label in col_map:
                                            col_idx = col_map[normalized_label]
                                        
                                        if col_idx is not None:
                                            try:
                                                ws_content.write(row_idx, col_idx, str(field_value))
                                                print(f"      [OK] {field_label}: {field_value}")
                                                fields_written += 1
                                            except Exception as e:
                                                print(f"      [WARN] Error writing {field_label}: {e}")
                                        else:
                                            print(f"      [WARN] No column found for: {field_label}")
                                    
                                    print(f"    Material {mat_code}: {fields_written} fields written")
                                    break  # Found and updated, move to next row
                        except Exception as e:
                            continue
                
                # Save the modified workbook
                wb.save(saved_excel_path)
                print(f"[OK] Excel file (.xls) updated with user data: {saved_excel_path}")
                
            except ImportError as ie:
                print(f"[WARN] xlrd/xlwt/xlutils not installed: {ie}")
                print(f"   File copied but data not updated. Install: pip install xlrd xlwt xlutils")
            except Exception as e:
                print(f"[WARN] Error updating .xls file: {e}")
                import traceback
                traceback.print_exc()
                print(f"   File was copied to: {saved_excel_path}")
        
        else:
            # For .xlsx files, use openpyxl
            print(f"   Using openpyxl for .xlsx format...")
            try:
                from openpyxl import load_workbook
                
                # Load the workbook (preserves formatting)
                wb = load_workbook(saved_excel_path)
                
                print(f"    Found sheets: {wb.sheetnames}")
                
                # Write dynamic form fields to their respective cells
                print(f"    Writing {len(dynamic_fields)} dynamic form fields...")
                for field_id, field_data in dynamic_fields.items():
                    try:
                        sheet_name = field_data['sheet_name']
                        row = field_data['row'] + 1  # openpyxl is 1-indexed
                        col = field_data['col'] + 1  # openpyxl is 1-indexed
                        value = field_data['value']
                        
                        # Get sheet by name
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            ws.cell(row=row, column=col, value=value)
                            print(f"      [OK] {sheet_name} - Cell ({row}, {col}): {value}")
                        else:
                            print(f"      [WARN] Sheet '{sheet_name}' not found")
                    except Exception as e:
                        print(f"      [WARN] Error writing field {field_id}: {e}")
                
                # 3. Update "Other Content" sheet with material information
                content_sheet = None
                for sheet_name in wb.sheetnames:
                    if 'other' in sheet_name.lower() and 'content' in sheet_name.lower():
                        content_sheet = wb[sheet_name]
                        break
                
                if content_sheet:
                    print(f"    Updating Other Content sheet with material details...")
                    
                    # Find column indices by searching headers (map all columns dynamically)
                    header_row = 1  # Usually headers are in first row
                    col_map = {}  # Maps: header_text -> column_number
                    
                    for cell in content_sheet[header_row]:
                        if cell.value:
                            header_val = str(cell.value).strip()
                            # Store original header
                            col_map[header_val] = cell.column
                            # Also store normalized version (without * and lowercase)
                            normalized_header = header_val.replace('*', '').strip().lower()
                            col_map[normalized_header] = cell.column
                    
                    print(f"    Found {len(col_map)} column mappings")
                    
                    # Find rows with material codes and update them directly using column mapping
                    for row in content_sheet.iter_rows(min_row=2, max_row=content_sheet.max_row):
                        # Look for material code in the row
                        row_text = " ".join([str(cell.value) for cell in row[:5] if cell.value])
                        
                        # Check if any of our materials match this row
                        for mat in materials:
                            mat_code = mat.get('material_code', '')
                            if mat_code and mat_code in row_text:
                                print(f"   [OK] Updating row {row[0].row} for material {mat_code}")
                                
                                # Get all dynamic fields from the material
                                material_fields = mat.get('fields', {})
                                
                                # Write ALL material fields dynamically by matching field labels to column headers
                                fields_written = 0
                                for field_label, field_value in material_fields.items():
                                    if not field_value:  # Skip empty values
                                        continue
                                    
                                    # Normalize field label for matching
                                    normalized_label = field_label.replace('*', '').strip().lower()
                                    
                                    # Try to find matching column
                                    col_num = None
                                    if field_label in col_map:
                                        col_num = col_map[field_label]
                                    elif normalized_label in col_map:
                                        col_num = col_map[normalized_label]
                                    
                                    if col_num is not None:
                                        try:
                                            content_sheet.cell(row=row[0].row, column=col_num, value=str(field_value))
                                            print(f"      [OK] {field_label}: {field_value}")
                                            fields_written += 1
                                        except Exception as e:
                                            print(f"      [WARN] Error writing {field_label}: {e}")
                                    else:
                                        print(f"      [WARN] No column found for: {field_label}")
                                
                                print(f"    Material {mat_code}: {fields_written} fields written")
                                break  # Found and updated, move to next row
                
                # Save the workbook
                wb.save(saved_excel_path)
                wb.close()
                
                print(f"[OK] Excel file (.xlsx) updated with user data: {saved_excel_path}")
                
            except Exception as e:
                print(f"[WARN] Error updating .xlsx file: {e}")
                import traceback
                traceback.print_exc()
                print(f"   File was copied to: {saved_excel_path}")
        
        # Initialize status tracking
        sharepoint_status = "pending"
        auto_submit_status = "pending"

        # 4. Upload to SharePoint
        print(f"\n[Upload] Uploading files to SharePoint...")
        try:
            # Initialize SharePoint client
            graph_client = GraphClient(
                get_setting("CLIENT_ID", ""), get_setting("CLIENT_SECRET", ""), get_setting("TENANT_ID", ""),
                get_setting("SHAREPOINT_HOSTNAME", ""), get_setting("SITE_PATH", ""), get_setting("DRIVE_NAME", "")
            )
            graph_client.auth()
            graph_client.resolve_site_and_drive()
            
            # Upload saved Excel file to SharePoint rfp-upload-file folder
            if os.path.exists(saved_excel_path):
                sp_savedrfp_path = get_sharepoint_rfp_savedrfp_path(rfp_id, company)
                saved_excel_filename = os.path.basename(saved_excel_path)
                print(f"[Upload] Uploading Excel file to SharePoint: {sp_savedrfp_path}/{saved_excel_filename}")
                graph_client.upload_file_as(
                    saved_excel_path,
                    sp_savedrfp_path,
                    saved_excel_filename
                )
                print(f"[OK] Excel file uploaded to SharePoint successfully")

            # Upload TDS files to SharePoint TDS-files folder
            if os.path.exists(tds_folder):
                sp_tds_path = get_sharepoint_rfp_tds_path(rfp_id, company)
                tds_file_list = os.listdir(tds_folder)
                if tds_file_list:
                    print(f"[Upload] Uploading {len(tds_file_list)} TDS files to SharePoint...")
                    for tds_filename in tds_file_list:
                        tds_file_path = os.path.join(tds_folder, tds_filename)
                        if os.path.isfile(tds_file_path):
                            graph_client.upload_file_as(
                                tds_file_path,
                                sp_tds_path,
                                tds_filename
                            )
                            print(f"   [OK] Uploaded: {tds_filename}")
                    print(f"[OK] All TDS files uploaded to SharePoint successfully")
                else:
                    print(f"[INFO] No TDS files to upload")
            
            print(f"[OK] SharePoint upload completed successfully")
            sharepoint_status = "success"
        except Exception as e:
            print(f"[WARN] SharePoint upload error: {e}")
            import traceback
            traceback.print_exc()
            sharepoint_status = f"failed: {str(e)}"
            # Continue with the request but track the failure

        # 6. Trigger automatic RFP submission
        print(f"\n Triggering automatic RFP submission...")
        try:
            # Check if this specific RFP is already being submitted
            if _is_rfp_submitting(rfp_id):
                print(f"[WARN] RFP {rfp_id} is already being submitted, skipping duplicate trigger")
            elif not os.path.exists(saved_excel_path):
                print(f"[WARN] Saved Excel file not found, skipping automatic submission")
            else:
                # Add RFP to submitting set BEFORE creating task
                _add_submitting_rfp(rfp_id)
                print(f"[OK] RFP {rfp_id} marked as submitting")
                
                # Trigger submission automation in background with state management
                async def _submit_task():
                    try:
                        await run_automation_submit(rfp_id, company)
                        print(f"[OK] Automatic RFP submission completed for: {rfp_id}")
                    except Exception as e:
                        print(f"[ERROR] Automatic submission failed for {rfp_id}: {e}")
                        import traceback
                        traceback.print_exc()
                    finally:
                        _remove_submitting_rfp(rfp_id)
                        print(f"[OK] RFP {rfp_id} removed from submitting list")
                
                try:
                    asyncio.create_task(_submit_task())
                    print(f"[OK] Automatic RFP submission triggered in background for: {rfp_id}")
                except Exception as task_error:
                    # If task creation fails, remove from submitting set
                    _remove_submitting_rfp(rfp_id)
                    raise task_error
            auto_submit_status = "started"
        except Exception as e:
            print(f"[WARN] Failed to trigger automatic submission: {e}")
            import traceback
            traceback.print_exc()
            auto_submit_status = f"failed: {str(e)}"
            # Continue but track the failure

        # Build response with warnings if any operations failed
        warnings = []
        if "failed" in sharepoint_status:
            warnings.append(f"SharePoint upload {sharepoint_status}")
        if "failed" in auto_submit_status:
            warnings.append(f"Auto-submit {auto_submit_status}")

        return JSONResponse({
            "ok": True,
            "message": "RFP data saved successfully!" + (" (with warnings)" if warnings else ""),
            "rfp_id": rfp_id,
            "materials_count": len(materials),
            "tds_files_count": len(tds_files),
            "saved_excel_path": saved_excel_path,
            "sharepoint_upload": sharepoint_status,
            "auto_submit": auto_submit_status,
            "warnings": warnings if warnings else None
        })
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error saving RFP data: {str(e)}")
